"""HTML dashboard import helpers.

This service converts generated HTML reports into AppBI-native dashboards.
Deterministic parsing and validation define the available surface; composition
is resolved before charts and widgets are materialized inside that contract.
"""

from __future__ import annotations

import csv as csv_module
import html as html_module
import io
import json
import re
import secrets
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from io import BytesIO
from typing import Any, Dict, Iterable, List, Literal, Optional, Set, Tuple

import openpyxl
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.dependencies import require_edit_access, require_view_access
from app.core.logging import get_logger
from app.models import Dashboard
from app.models.dataset import Dataset, DatasetTable
from app.models.models import (
    Chart,
    ChartMetadata,
    ChartParameter,
    ChartType,
    DashboardChart,
    DataSource,
    DataSourceType,
)
from app.models.user import User
from app.schemas.dataset import DatasetCreate, TableCreate
from app.schemas.schemas import DataSourceCreate
from app.services.chart_semantic_service import with_chart_semantic_binding
from app.services.dashboard_service import (
    DEFAULT_DASHBOARD_PAGE,
    normalize_dashboard_theme_config,
    normalize_dashboard_widget_config,
)
from app.services.dataset_crud import DatasetCRUDService
from app.services.datasource_crud_service import DataSourceCRUDService
from app.services.dashboard_import_skill import compare_source_contract_to_dataset
from app.services.html_fragment_sanitizer import sanitize_html_fragment
from app.services.llm_client import LLMClient

logger = get_logger(__name__)

SourceMode = Literal["existing_dataset", "upload_excel"]
TargetMode = Literal["new_dashboard", "append_to_dashboard"]

_SUPPORTED_CHART_TYPES = {
    "BAR",
    "HORIZONTAL_BAR",
    "LINE",
    "PIE",
    "DONUT",
    "RADAR",
    "POLAR_AREA",
    "TIME_SERIES",
    "TABLE",
    "MATRIX",
    "AREA",
    "STACKED_BAR",
    "GROUPED_BAR",
    "BAR_LINE",
    "SCATTER",
    "BUBBLE",
    "HEATMAP",
    "TREEMAP",
    "FUNNEL",
    "GAUGE",
    "WATERFALL",
    "MAP_POINT",
    "MAP_REGION",
    "BOXPLOT",
    "BULLET",
    "SANKEY",
    "SUNBURST",
    "RIBBON",
    "TIMELINE",
    "WORD_CLOUD",
    "KPI",
    "PODIUM",
}
_BLOCK_ROLE_TO_CHART_TYPE = {
    "table": "TABLE",
    "kpi": "KPI",
    "chart": "BAR",
}
_BLOCK_KEYWORD_RE = re.compile(
    r"(chart|graph|plot|viz|visual|trend|timeseries|time series|share|mix|breakdown|composition|rank|top|bottom|kpi|metric|score|summary|table)",
    re.IGNORECASE,
)
_STYLE_PAIR_RE = re.compile(r"\s*([^:]+)\s*:\s*([^;]+)")
_SNAKE_RE = re.compile(r"[^a-z0-9]+")
_WHITESPACE_RE = re.compile(r"\s+")
_FIELD_TOKEN_RE = re.compile(r"[a-z0-9]+")

# Manual datasources created BY the wizard carry this name prefix; only those
# are eligible for automatic cleanup when their draft dataset goes away.
IMPORT_DATASOURCE_NAME_PREFIX = "[Dashboard Import]"
# A draft older than this was abandoned (tab closed / crash), so the
# client-side cancel never fired. Swept opportunistically — see
# purge_stale_import_drafts.
IMPORT_DRAFT_TTL_HOURS = 48


def build_ai_assist_meta(
    *,
    requested: bool,
    applied: bool,
    status: str,
    provider: Optional[str] = None,
    model: Optional[str] = None,
    message: Optional[str] = None,
) -> Dict[str, Any]:
    return {
        "requested": requested,
        "applied": applied,
        "status": status,
        "provider": provider,
        "model": model,
        "message": message,
    }


def _normalize_text(value: Any, *, max_len: int | None = None) -> str:
    text = _WHITESPACE_RE.sub(" ", str(value or "")).strip()
    if max_len and len(text) > max_len:
        return text[: max_len - 3].rstrip() + "..."
    return text


def _extract_gemini_text(response: Any) -> str:
    text = getattr(response, "text", "") or ""
    if text:
        return text

    for candidate in getattr(response, "candidates", []) or []:
        content = getattr(candidate, "content", None)
        for part in getattr(content, "parts", []) or []:
            part_text = getattr(part, "text", "") or ""
            if part_text:
                return part_text
    return ""


def _parse_ai_json(text: str) -> Dict[str, Any]:
    raw = str(text or "").strip()
    if not raw:
        raise ValueError("empty AI response")

    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw).strip()

    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        raw = raw[start : end + 1]

    raw = re.sub(r",\s*([}\]])", r"\1", raw)
    return json.loads(raw)


def _normalize_identifier(value: Any) -> str:
    lowered = _normalize_text(value).lower()
    return _SNAKE_RE.sub("_", lowered).strip("_")


def _field_tokens(value: Any) -> set[str]:
    return set(_FIELD_TOKEN_RE.findall(_normalize_identifier(value)))


# ── Column-name alias resolution ────────────────────────────────────────
# Claude authors plans in snake_case (e.g. ``math_score``) while uploaded
# CSVs frequently keep raw spaced headers (``math score``). The helpers below
# build a lookup keyed by a space/punctuation/case-insensitive fingerprint so
# plan identifiers can be rewritten to match the physical column names
# (quoted with double-quotes when they contain whitespace).
_IDENT_TOKEN_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*")
_ALNUM_ONLY_RE = re.compile(r"[^a-z0-9]+")


def _column_fingerprint(value: Any) -> str:
    """Collapse a name to lowercase alphanumerics only (e.g. ``Math Score`` → ``mathscore``)."""
    return _ALNUM_ONLY_RE.sub("", str(value or "").lower())


def _build_column_alias_map(columns: Iterable[Dict[str, Any]]) -> Dict[str, str]:
    """Build ``fingerprint -> real column name`` for a list of column dicts.

    When multiple real columns share the same fingerprint, the first one wins —
    this is almost never an issue in practice because physical column names
    inside a single table are unique.
    """
    mapping: Dict[str, str] = {}
    for col in columns or []:
        if not isinstance(col, dict):
            continue
        real = str(col.get("name") or "").strip()
        if not real:
            continue
        key = _column_fingerprint(real)
        if key and key not in mapping:
            mapping[key] = real
    return mapping


def _quote_identifier_if_needed(name: str, *, style: str = "sql") -> str:
    """Return ``name`` in a form that preserves non-identifier characters.

    ``style="sql"`` wraps with double-quotes (DuckDB/ANSI identifier quoting).
    ``style="expression"`` wraps with square brackets, matching
    ``TransformationCompiler``'s spreadsheet-style ``[col name]`` reference —
    double-quotes in calc-field expressions are rewritten into string literals
    and would break the SQL.
    """
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", name):
        return name
    if style == "expression":
        return "[" + name.replace("]", "]]") + "]"
    return '"' + name.replace('"', '""') + '"'


def _resolve_column_name(name: Any, alias_map: Dict[str, str]) -> Optional[str]:
    """Return the real column name matching *name* via fingerprint lookup."""
    if name is None:
        return None
    txt = str(name).strip()
    if not txt:
        return None
    fp = _column_fingerprint(txt)
    if not fp:
        return None
    return alias_map.get(fp)


def _rewrite_identifiers_in_expression(
    expression: str,
    alias_map: Dict[str, str],
    *,
    style: str = "sql",
) -> str:
    """Rewrite bare identifier tokens in a SQL-ish expression to real column names.

    String literals (single quoted; also double quoted in ``style="sql"`` mode)
    are preserved verbatim so values like ``'Male'`` are never mistakenly
    renamed. When the resolved column name contains whitespace or punctuation,
    it is emitted double-quoted (``style="sql"``) or bracket-wrapped
    (``style="expression"``) to match the downstream parser.
    """
    if not expression or not alias_map:
        return expression
    out: List[str] = []
    i = 0
    n = len(expression)
    while i < n:
        c = expression[i]
        if c == "'":
            # Single-quoted string literal (standard SQL).
            j = i + 1
            while j < n:
                if expression[j] == "'":
                    if j + 1 < n and expression[j + 1] == "'":
                        j += 2
                        continue
                    j += 1
                    break
                j += 1
            out.append(expression[i:j])
            i = j
            continue
        if c == '"' and style == "sql":
            # Double-quoted identifier — keep as-is in SQL mode. In expression
            # mode ``"x"`` is a string literal handled by the compiler, so we
            # leave the raw character for the compiler to rewrite.
            j = i + 1
            while j < n:
                if expression[j] == '"':
                    if j + 1 < n and expression[j + 1] == '"':
                        j += 2
                        continue
                    j += 1
                    break
                j += 1
            out.append(expression[i:j])
            i = j
            continue
        if c == "[" and style == "expression":
            # Preserve pre-existing ``[col]`` references verbatim.
            j = expression.find("]", i + 1)
            if j == -1:
                out.append(c)
                i += 1
                continue
            out.append(expression[i : j + 1])
            i = j + 1
            continue
        match = _IDENT_TOKEN_RE.match(expression, i)
        if match:
            tok = match.group(0)
            end = match.end()
            # Skip function calls: ``NAME(`` — even if a physical column happens
            # to share the name (e.g. ``Count``, ``Date``), we must not rewrite
            # the function keyword into a quoted identifier.
            k = end
            while k < n and expression[k] in " \t":
                k += 1
            is_function_call = k < n and expression[k] == "("
            real = None if is_function_call else alias_map.get(_column_fingerprint(tok))
            if real and real != tok:
                out.append(_quote_identifier_if_needed(real, style=style))
            else:
                out.append(tok)
            i = end
            continue
        out.append(c)
        i += 1
    return "".join(out)


def _parse_style_map(raw_style: Any) -> Dict[str, str]:
    if isinstance(raw_style, dict):
        return {
            _normalize_identifier(key): _normalize_text(value, max_len=120)
            for key, value in raw_style.items()
            if _normalize_identifier(key) and _normalize_text(value)
        }

    result: Dict[str, str] = {}
    for chunk in str(raw_style or "").split(";"):
        match = _STYLE_PAIR_RE.match(chunk)
        if not match:
            continue
        key = _normalize_identifier(match.group(1))
        value = _normalize_text(match.group(2), max_len=120)
        if key and value:
            result[key] = value
    return result


def _is_number_type(column_type: str) -> bool:
    return str(column_type or "").lower() in {
        "number",
        "integer",
        "float",
        "double",
        "decimal",
        "numeric",
        "bigint",
        "int",
    }


def _is_date_type(column_type: str) -> bool:
    return str(column_type or "").lower() in {
        "date",
        "datetime",
        "timestamp",
        "time",
    }


def _infer_preview_column_type(values: List[Any]) -> str:
    samples = [value for value in values if value is not None and str(value).strip() != ""][:20]
    if not samples:
        return "string"
    number_count = sum(1 for value in samples if _preview_value_is_number(value))
    if number_count == len(samples):
        return "number"
    date_count = sum(1 for value in samples if _preview_value_is_date_like(value))
    if date_count >= len(samples) * 0.8:
        return "date"
    return "string"


def _preview_value_is_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    try:
        float(str(value).replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def _preview_value_is_date_like(value: Any) -> bool:
    if hasattr(value, "isoformat"):
        return True
    return bool(re.match(r"^\d{2,4}[-/]\d{1,2}[-/]\d{1,4}", str(value).strip()))


def _coerce_preview_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return _normalize_text(value, max_len=500)


def parse_uploaded_source_sheets(
    *,
    file_bytes: bytes,
    filename: Optional[str],
) -> Dict[str, Dict[str, Any]]:
    lower_name = str(filename or "").lower()
    if lower_name.endswith(".csv"):
        return _parse_csv_source_sheets(file_bytes, filename=filename)
    return _parse_excel_source_sheets(file_bytes)


def _parse_excel_source_sheets(file_bytes: bytes) -> Dict[str, Dict[str, Any]]:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    try:
        sheets: Dict[str, Dict[str, Any]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            all_rows = list(worksheet.iter_rows(values_only=True))
            if not all_rows:
                sheets[sheet_name] = {"columns": [], "rows": []}
                continue

            header_row = [str(cell).strip() if cell is not None else "" for cell in all_rows[0]]
            headers = [header if header else f"col{i + 1}" for i, header in enumerate(header_row)]

            rows: List[Dict[str, Any]] = []
            for raw_row in all_rows[1:]:
                row_payload: Dict[str, Any] = {}
                has_value = False
                for index, header in enumerate(headers):
                    value = raw_row[index] if index < len(raw_row) else None
                    normalized_value = _coerce_preview_cell(value)
                    row_payload[header] = normalized_value
                    if normalized_value not in ("", None):
                        has_value = True
                if has_value:
                    rows.append(row_payload)

            columns = [
                {"name": header, "type": _infer_preview_column_type([row.get(header) for row in rows])}
                for header in headers
            ]
            sheets[sheet_name] = {"columns": columns, "rows": rows}

        return sheets
    finally:
        workbook.close()


def _parse_csv_source_sheets(
    file_bytes: bytes,
    *,
    filename: Optional[str],
) -> Dict[str, Dict[str, Any]]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv_module.DictReader(io.StringIO(text))
    rows = [dict(row) for row in reader]
    fieldnames = list(reader.fieldnames or [])
    columns = [
        {"name": header, "type": _infer_preview_column_type([row.get(header) for row in rows])}
        for header in fieldnames
    ]
    sheet_name = str(filename or "Sheet1").rsplit(".", 1)[0] or "Sheet1"
    return {sheet_name: {"columns": columns, "rows": rows}}


def _classify_block_role(block: Dict[str, Any]) -> str:
    explicit_role = _normalize_identifier(block.get("role"))
    if explicit_role in {"title", "section", "chart", "table", "kpi", "text"}:
        return explicit_role

    tag = str(block.get("tag") or "").lower()
    if tag == "table" or isinstance(block.get("table"), dict):
        return "table"
    if tag in {"h1", "h2", "h3"}:
        return "title"

    # Keywords come from what the author NAMED the block, never from its prose.
    # Searching the text let one word decide: a pull quote reading "Late orders
    # average 4.1 days beyond estimate" was classified as a KPI and laid out in
    # the metric strip.
    named_text = " ".join(
        [
            " ".join(str(item) for item in (block.get("classes") or [])),
            _normalize_text(block.get("id_attr")),
        ]
    ).lower()
    text = _normalize_text(block.get("text"))
    # What a metric LOOKS like: a small box that is mostly a figure. It is the
    # signal that survives generated HTML, which rarely has useful class names.
    if 0 < len(text) <= 48 and len(re.findall(r"\d", text)) >= 2:
        return "kpi"
    if "kpi" in named_text or "metric" in named_text or "score" in named_text:
        return "kpi"
    if _BLOCK_KEYWORD_RE.search(named_text + " " + _normalize_text(block.get("heading")).lower()):
        return "chart"
    return "text"


def _coerce_numberish(value: Any) -> Any:
    if isinstance(value, (int, float)) or value is None:
        return value
    text = _normalize_text(value)
    if not text:
        return None
    compact = text.replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", compact):
        try:
            return float(compact) if "." in compact else int(compact)
        except ValueError:
            return text
    return text


def _build_source_column_profile(name: str, col_type: str) -> Dict[str, Any]:
    return {
        "name": name,
        "type": str(col_type or "string").lower(),
        "tokens": sorted(_field_tokens(name)),
    }


def _score_field_match(label: str, column: Dict[str, Any]) -> float:
    label_norm = _normalize_identifier(label)
    if not label_norm:
        return 0.0
    column_name = str(column.get("name") or "")
    if label_norm == _normalize_identifier(column_name):
        return 1.0

    label_tokens = _field_tokens(label)
    column_tokens = set(column.get("tokens") or [])
    overlap = len(label_tokens & column_tokens)
    if label_tokens and column_tokens:
        jaccard = overlap / max(len(label_tokens | column_tokens), 1)
    else:
        jaccard = 0.0

    sequence = SequenceMatcher(None, label_norm, _normalize_identifier(column_name)).ratio()
    prefix = 0.15 if label_norm in _normalize_identifier(column_name) or _normalize_identifier(column_name) in label_norm else 0.0
    return max(jaccard + prefix, sequence * 0.9)


def _best_field_match(
    label: str,
    columns: List[Dict[str, Any]],
    *,
    numeric_only: bool = False,
    date_only: bool = False,
    exclude: Iterable[str] | None = None,
) -> Optional[str]:
    excluded = {str(item) for item in (exclude or [])}
    best_name: Optional[str] = None
    best_score = 0.0
    for column in columns:
        column_name = str(column.get("name") or "")
        if not column_name or column_name in excluded:
            continue
        if numeric_only and not _is_number_type(str(column.get("type") or "")):
            continue
        if date_only and not _is_date_type(str(column.get("type") or "")):
            continue
        score = _score_field_match(label, column)
        if score > best_score:
            best_name = column_name
            best_score = score
    if best_score >= 0.48:
        return best_name
    return None


def _unique_dashboard_name(db: Session, requested_name: str) -> str:
    base_name = _normalize_text(requested_name) or "Imported Dashboard"
    candidate = base_name
    suffix = 2
    while db.query(Dashboard).filter(Dashboard.name == candidate).first() is not None:
        candidate = f"{base_name} ({suffix})"
        suffix += 1
    return candidate


def _unique_chart_name(db: Session, owner_id: Any, requested_name: str) -> str:
    base_name = _normalize_text(requested_name) or "Imported Chart"
    candidate = base_name
    suffix = 2
    while True:
        conflict_query = db.query(Chart).filter(Chart.name == candidate)
        if owner_id is None:
            conflict_query = conflict_query.filter(Chart.owner_id.is_(None))
        else:
            conflict_query = conflict_query.filter(Chart.owner_id == owner_id)
        if conflict_query.first() is None:
            return candidate
        candidate = f"{base_name} ({suffix})"
        suffix += 1


def _looks_generic_import_page_name(value: Any) -> bool:
    normalized = _normalize_text(value, max_len=80).lower()
    if not normalized:
        return True
    if normalized in {"imported dashboard", "imported page"}:
        return True
    return bool(re.fullmatch(r"imported page \d+", normalized))


def _page_name_from_filename(filename: Any) -> str:
    raw_filename = str(filename or "").strip()
    if not raw_filename:
        return ""
    basename = re.split(r"[\\/]", raw_filename)[-1]
    stem = basename.rsplit(".", 1)[0]
    normalized = _normalize_text(re.sub(r"[_-]+", " ", stem), max_len=80)
    if not normalized:
        return ""
    return normalized[:1].upper() + normalized[1:]


def _resolve_batch_page_name(
    *,
    explicit_name: Any = None,
    document_title: Any = None,
    suggested_dashboard_name: Any = None,
    filename: Any = None,
    fallback_index: int,
) -> str:
    explicit = _normalize_text(explicit_name, max_len=80)
    if explicit:
        return explicit

    document_title_normalized = _normalize_text(document_title, max_len=80)
    suggested_name_normalized = _normalize_text(suggested_dashboard_name, max_len=80)
    for candidate in (document_title_normalized, suggested_name_normalized):
        if candidate and not _looks_generic_import_page_name(candidate):
            return candidate

    filename_name = _page_name_from_filename(filename)
    if filename_name:
        return filename_name

    return document_title_normalized or suggested_name_normalized or f"Imported Page {fallback_index}"


def _unique_page_name(requested_name: Any, used_names: Set[str]) -> str:
    base_name = _normalize_text(requested_name, max_len=80) or "Imported Page"
    candidate = base_name
    suffix = 2
    while candidate.lower() in used_names:
        suffix_text = f" ({suffix})"
        max_base_len = max(1, 80 - len(suffix_text))
        trimmed_base = base_name if len(base_name) <= max_base_len else base_name[:max_base_len].rstrip()
        candidate = f"{trimmed_base}{suffix_text}"
        suffix += 1
    used_names.add(candidate.lower())
    return candidate


def _generate_page_id() -> str:
    return f"page-import-{secrets.token_hex(4)}"


def _source_profile_from_rows(
    *,
    source_mode: SourceMode,
    dataset_id: Optional[int],
    dataset_name: Optional[str],
    dataset_table_id: Optional[int],
    dataset_table_name: Optional[str],
    columns: List[Dict[str, Any]],
    sample_rows: List[Dict[str, Any]],
    row_count: Optional[int] = None,
    uploaded_filename: Optional[str] = None,
) -> Dict[str, Any]:
    normalized_columns = [
        _build_source_column_profile(
            str(column.get("name") or column.get("key") or ""),
            str(column.get("type") or "string"),
        )
        for column in columns
        if str(column.get("name") or column.get("key") or "").strip()
    ]
    numeric_columns = [col["name"] for col in normalized_columns if _is_number_type(col["type"])]
    date_columns = [col["name"] for col in normalized_columns if _is_date_type(col["type"])]
    dimension_columns = [col["name"] for col in normalized_columns if col["name"] not in numeric_columns]

    return {
        "source_mode": source_mode,
        "dataset_id": dataset_id,
        "dataset_name": dataset_name,
        "dataset_table_id": dataset_table_id,
        "dataset_table_name": dataset_table_name,
        "uploaded_filename": uploaded_filename,
        "row_count": row_count,
        "columns": normalized_columns,
        "numeric_columns": numeric_columns,
        "date_columns": date_columns,
        "dimension_columns": dimension_columns,
        "sample_rows": sample_rows[:10],
    }


def _load_existing_source_profile(
    db: Session,
    *,
    current_user: User,
    dataset_table_id: int,
) -> Dict[str, Any]:
    db_table = db.query(DatasetTable).filter(DatasetTable.id == dataset_table_id).first()
    if not db_table:
        raise ValueError("Dataset table not found")

    dataset_obj = db.query(Dataset).filter(Dataset.id == db_table.dataset_id).first()
    if not dataset_obj:
        raise ValueError("Dataset not found")

    require_view_access(db, current_user, dataset_obj, "datasets")

    columns_cache = db_table.columns_cache if isinstance(db_table.columns_cache, dict) else {}
    cached_columns = columns_cache.get("columns") if isinstance(columns_cache, dict) else None
    sample_rows = list(db_table.sample_cache or [])

    columns: List[Dict[str, Any]]
    if isinstance(cached_columns, list) and cached_columns:
        columns = [
            {
                "name": str(col.get("name") or ""),
                "type": str(col.get("type") or "string"),
            }
            for col in cached_columns
            if str(col.get("name") or "").strip()
        ]
    elif sample_rows:
        first_row = sample_rows[0]
        columns = [
            {
                "name": str(key),
                "type": (
                    "number"
                    if isinstance(value, (int, float))
                    else "date"
                    if hasattr(value, "isoformat")
                    else "string"
                ),
            }
            for key, value in first_row.items()
        ]
    else:
        columns = []

    if not columns:
        raise ValueError(
            f"Dataset table '{db_table.display_name}' has no column metadata yet. "
            "Please preview the table first so column information is cached."
        )

    return _source_profile_from_rows(
        source_mode="existing_dataset",
        dataset_id=dataset_obj.id,
        dataset_name=dataset_obj.name,
        dataset_table_id=db_table.id,
        dataset_table_name=db_table.display_name,
        columns=columns,
        sample_rows=sample_rows[:10],
    )


def _load_existing_dataset_profiles(
    db: Session,
    *,
    current_user: User,
    dataset_id: int,
) -> Tuple[Dict[str, Any], Dict[str, Dict[str, Any]]]:
    """Load source profiles for ALL data tables in a dataset.

    Returns ``(primary_profile, all_profiles_dict)`` where keys in
    *all_profiles_dict* are the table display names (used as ``source_key``
    in the multi-source analyze/build flow).
    """
    dataset_obj = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset_obj:
        raise ValueError("Dataset not found")
    require_view_access(db, current_user, dataset_obj, "datasets")

    db_tables = (
        db.query(DatasetTable)
        .filter(
            DatasetTable.dataset_id == dataset_id,
            DatasetTable.source_kind != "generated_calendar",
        )
        .order_by(DatasetTable.id)
        .all()
    )
    if not db_tables:
        raise ValueError("Dataset has no data tables.")

    all_profiles: Dict[str, Dict[str, Any]] = {}
    for db_table in db_tables:
        columns_cache = db_table.columns_cache if isinstance(db_table.columns_cache, dict) else {}
        cached_columns = columns_cache.get("columns") if isinstance(columns_cache, dict) else None
        sample_rows = list(db_table.sample_cache or [])

        if isinstance(cached_columns, list) and cached_columns:
            columns = [
                {
                    "name": str(col.get("name") or ""),
                    "type": str(col.get("type") or "string"),
                }
                for col in cached_columns
                if str(col.get("name") or "").strip()
            ]
        elif sample_rows:
            first_row = sample_rows[0]
            columns = [
                {
                    "name": str(key),
                    "type": (
                        "number"
                        if isinstance(value, (int, float))
                        else "date"
                        if hasattr(value, "isoformat")
                        else "string"
                    ),
                }
                for key, value in first_row.items()
            ]
        else:
            columns = []

        if not columns:
            continue  # skip tables without column metadata

        source_key = db_table.display_name
        all_profiles[source_key] = _source_profile_from_rows(
            source_mode="existing_dataset",
            dataset_id=dataset_obj.id,
            dataset_name=dataset_obj.name,
            dataset_table_id=db_table.id,
            dataset_table_name=db_table.display_name,
            columns=columns,
            sample_rows=sample_rows[:10],
        )

    if not all_profiles:
        raise ValueError(
            "No tables with column metadata found in the dataset. "
            "Please preview tables first so column information is cached."
        )

    primary_profile = next(iter(all_profiles.values()))
    return primary_profile, all_profiles


def _load_uploaded_excel_source_profile(
    *,
    file_bytes: bytes,
    filename: Optional[str],
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    all_sheets = parse_uploaded_source_sheets(file_bytes=file_bytes, filename=filename)
    if not all_sheets:
        raise ValueError("Uploaded file does not contain any sheets or rows.")

    available_sheet_names = list(all_sheets.keys())
    selected_sheet_name = str(sheet_name or "").strip()
    if not selected_sheet_name or selected_sheet_name not in all_sheets:
        selected_sheet_name = available_sheet_names[0]

    sheet_data = all_sheets[selected_sheet_name]
    columns = list(sheet_data.get("columns") or [])
    rows = list(sheet_data.get("rows") or [])
    source_profile = _source_profile_from_rows(
        source_mode="upload_excel",
        dataset_id=None,
        dataset_name=None,
        dataset_table_id=None,
        dataset_table_name=selected_sheet_name,
        columns=columns,
        sample_rows=rows,
        row_count=len(rows),
        uploaded_filename=filename,
    )
    source_profile["available_sheets"] = available_sheet_names
    source_profile["selected_sheet_name"] = selected_sheet_name
    return source_profile


def _load_uploaded_single_file_multi_sheet_profiles(
    *,
    file_bytes: bytes,
    filename: Optional[str],
    primary_source_key: Optional[str] = None,
) -> Tuple[Dict[str, Dict[str, Any]], str]:
    """Build one source profile per sheet in a single uploaded workbook.

    Unlike :func:`_load_uploaded_multi_source_profiles`, the resulting source
    keys are the bare sheet names (e.g. ``customers``, ``deals``) instead of
    ``filename::sheet``, which matches authored HTML metadata that references
    sheets by name only.
    """
    all_sheets = parse_uploaded_source_sheets(file_bytes=file_bytes, filename=filename)
    if not all_sheets:
        raise ValueError("Uploaded file does not contain any sheets or rows.")

    all_profiles: Dict[str, Dict[str, Any]] = {}
    first_key: Optional[str] = None
    safe_filename = filename or "uploaded"
    for sheet_name, sheet_data in all_sheets.items():
        source_key = str(sheet_name).strip() or safe_filename
        if first_key is None:
            first_key = source_key
        columns = list(sheet_data.get("columns") or [])
        rows = list(sheet_data.get("rows") or [])
        profile = _source_profile_from_rows(
            source_mode="upload_excel",
            dataset_id=None,
            dataset_name=None,
            dataset_table_id=None,
            dataset_table_name=sheet_name,
            columns=columns,
            sample_rows=rows,
            row_count=len(rows),
            uploaded_filename=safe_filename,
        )
        profile["source_key"] = source_key
        profile["available_sheets"] = []
        all_profiles[source_key] = profile

    all_keys = list(all_profiles.keys())
    for profile in all_profiles.values():
        profile["available_sheets"] = all_keys

    resolved_primary = primary_source_key
    if not resolved_primary or resolved_primary not in all_profiles:
        resolved_primary = first_key or ""
    return all_profiles, resolved_primary


def _load_uploaded_multi_source_profiles(
    *,
    files: List[Tuple[bytes, Optional[str]]],
    primary_source_key: Optional[str] = None,
) -> Tuple[Dict[str, Dict[str, Any]], str]:
    """Build source profiles from multiple uploaded Excel/CSV files.

    Returns ``(profiles_by_key, resolved_primary_key)``.
    Key format: ``'{filename}::{sheet_name}'``.
    """
    all_profiles: Dict[str, Dict[str, Any]] = {}
    first_key: Optional[str] = None

    for file_bytes, filename in files:
        safe_filename = filename or "uploaded"
        all_sheets = parse_uploaded_source_sheets(file_bytes=file_bytes, filename=filename)
        for sheet_name, sheet_data in all_sheets.items():
            source_key = f"{safe_filename}::{sheet_name}"
            if first_key is None:
                first_key = source_key
            columns = list(sheet_data.get("columns") or [])
            rows = list(sheet_data.get("rows") or [])
            profile = _source_profile_from_rows(
                source_mode="upload_excel",
                dataset_id=None,
                dataset_name=None,
                dataset_table_id=None,
                dataset_table_name=sheet_name,
                columns=columns,
                sample_rows=rows,
                row_count=len(rows),
                uploaded_filename=safe_filename,
            )
            profile["source_key"] = source_key
            profile["available_sheets"] = []
            all_profiles[source_key] = profile

    all_keys = list(all_profiles.keys())
    for profile in all_profiles.values():
        profile["available_sheets"] = all_keys

    resolved_primary = primary_source_key
    if not resolved_primary or resolved_primary not in all_profiles:
        resolved_primary = first_key or ""
    return all_profiles, resolved_primary


def _build_ai_fix_source_profiles(
    *,
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
    derived_tables: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Dict[str, Dict[str, Any]]]:
    """Build the table schema map used by AI chart-fix prompts.

    The fixer needs schemas for both physical tables and any v1 derived tables.
    Without derived-table output columns in the prompt, AI may keep a derived
    source_key but suggest fields that only exist on the physical source.
    """

    profiles: Dict[str, Dict[str, Any]] = {}

    if isinstance(all_source_profiles, dict):
        for key, profile in all_source_profiles.items():
            if not key or not isinstance(profile, dict):
                continue
            copied = dict(profile)
            copied.setdefault("source_key", key)
            profiles[str(key)] = copied

    primary_key = str(
        source_profile.get("source_key")
        or source_profile.get("dataset_table_name")
        or source_profile.get("selected_sheet_name")
        or ""
    ).strip()
    if primary_key and primary_key not in profiles:
        copied = dict(source_profile)
        copied.setdefault("source_key", primary_key)
        profiles[primary_key] = copied

    for table in derived_tables or []:
        if not isinstance(table, dict):
            continue
        source_key = str(table.get("source_key") or "").strip()
        output_columns = list(table.get("output_columns") or [])
        if not source_key or not output_columns:
            continue

        normalized_columns: List[Dict[str, str]] = []
        numeric_columns: List[str] = []
        date_columns: List[str] = []
        dimension_columns: List[str] = []

        for raw_col in output_columns:
            if not isinstance(raw_col, dict):
                continue
            col_name = str(raw_col.get("name") or "").strip()
            col_type = str(raw_col.get("type") or "string").strip().lower()
            if not col_name:
                continue
            normalized_columns.append({"name": col_name, "type": col_type})
            if col_type in {"number", "numeric", "integer", "float", "double", "decimal"}:
                numeric_columns.append(col_name)
            elif col_type in {"date", "datetime", "timestamp"}:
                date_columns.append(col_name)
                dimension_columns.append(col_name)
            else:
                dimension_columns.append(col_name)

        if not normalized_columns:
            continue

        profiles[source_key] = {
            "source_key": source_key,
            "dataset_table_name": source_key,
            "dataset_name": source_profile.get("dataset_name"),
            "columns": normalized_columns,
            "numeric_columns": numeric_columns,
            "date_columns": date_columns,
            "dimension_columns": dimension_columns,
            "sample_rows": [],
            "row_count": None,
        }

    return profiles or None


def _plan_quality_score(plan: Dict[str, Any]) -> Tuple[int, int, float, int]:
    """Score a finalized plan for multi-source selection.  Higher is better."""
    fields_count = len(plan.get("source_fields_used") or [])
    type_bonus = 0 if plan.get("final_chart_type") == "TABLE" else 1
    confidence = plan.get("confidence", 0)
    warning_penalty = -len(plan.get("warnings") or [])
    return (fields_count, type_bonus, confidence, warning_penalty)


def _plain_html_excerpt(html_text: str) -> str:
    stripped = re.sub(r"<script[\s\S]*?</script>", " ", html_text, flags=re.IGNORECASE)
    stripped = re.sub(r"<style[\s\S]*?</style>", " ", stripped, flags=re.IGNORECASE)
    stripped = re.sub(r"<[^>]+>", " ", stripped)
    return _normalize_text(stripped, max_len=4000)


def _build_document_summary(
    *,
    html_text: str,
    html_summary: Dict[str, Any] | None,
) -> Dict[str, Any]:
    summary = dict(html_summary or {})
    blocks = summary.get("blocks")
    if not isinstance(blocks, list):
        blocks = []

    normalized_blocks: List[Dict[str, Any]] = []
    for index, raw_block in enumerate(blocks, start=1):
        if not isinstance(raw_block, dict):
            continue
        block_id = str(raw_block.get("id") or f"block-{index}")
        normalized_blocks.append(
            {
                "id": block_id,
                "order": int(raw_block.get("order") or index),
                "tag": str(raw_block.get("tag") or "div").lower(),
                "role": _classify_block_role(raw_block),
                "heading": _normalize_text(raw_block.get("heading"), max_len=160),
                "text": _normalize_text(raw_block.get("text"), max_len=900),
                "classes": [str(item) for item in (raw_block.get("classes") or []) if str(item).strip()],
                "id_attr": _normalize_text(raw_block.get("id_attr"), max_len=120),
                "style": _parse_style_map(raw_block.get("style")),
                "table": raw_block.get("table") if isinstance(raw_block.get("table"), dict) else None,
                # The block's appearance, frozen by the client with the
                # stylesheet's computed values inlined. Sanitized here rather
                # than at the point of use, so nothing downstream ever holds
                # raw source markup.
                "html": sanitize_html_fragment(str(raw_block.get("html") or ""))[0],
            }
        )

    if not normalized_blocks and html_text.strip():
        normalized_blocks.append(
            {
                "id": "block-1",
                "order": 1,
                "tag": "div",
                "role": "chart",
                "heading": "",
                "text": _plain_html_excerpt(html_text),
                "classes": [],
                "id_attr": "",
                "style": {},
                "table": None,
                "html": "",
            }
        )

    title = _normalize_text(summary.get("title") or summary.get("document_title"), max_len=180)
    if not title:
        title = next(
            (block.get("heading") or block.get("text") for block in normalized_blocks if block.get("role") == "title"),
            "",
        )
    if not title:
        title = "Imported Dashboard"

    # Labels of the source's own filter controls. A `<select>` is never a block
    # -- no heading, no text, no visual -- so before this every imported report
    # arrived with no slicers at all, however many filters the source had.
    filter_controls = [
        _normalize_text(item, max_len=60)
        for item in (summary.get("filterControls") or summary.get("filter_controls") or [])
        if _normalize_text(item, max_len=60)
    ][:8]

    return {
        "title": title,
        "blocks": sorted(normalized_blocks, key=lambda item: (item.get("order", 0), item.get("id", ""))),
        "html_excerpt": _plain_html_excerpt(html_text),
        "filter_controls": filter_controls,
    }


def _candidate_blocks(document_summary: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    candidates: List[Dict[str, Any]] = []
    ignored: List[Dict[str, Any]] = []
    for block in document_summary.get("blocks") or []:
        role = block.get("role")
        has_table = isinstance(block.get("table"), dict)
        text = _normalize_text(block.get("text"))
        heading = _normalize_text(block.get("heading"))
        if role in {"chart", "table", "kpi"} or has_table:
            candidates.append(block)
            continue
        if role == "title":
            ignored.append({"block_id": block["id"], "reason": "Document title block"})
            continue
        if len(text) < 40 and not heading:
            ignored.append({"block_id": block["id"], "reason": "Too little content"})
            continue
        if _BLOCK_KEYWORD_RE.search(f"{heading} {text}"):
            candidates.append({**block, "role": "chart"})
        else:
            ignored.append({"block_id": block["id"], "reason": "Narrative block not converted in chart-first import"})
    return candidates[:20], ignored


def _fallback_chart_kind(block: Dict[str, Any], source_profile: Dict[str, Any]) -> str:
    role = str(block.get("role") or "")
    if role == "table":
        return "TABLE"
    if role == "kpi":
        return "KPI"

    title_text = " ".join(
        filter(
            None,
            [
                _normalize_text(block.get("heading")),
                _normalize_text(block.get("text")),
            ],
        )
    ).lower()
    if any(token in title_text for token in ("trend", "monthly", "over time", "timeline", "daily", "weekly")) and source_profile.get("date_columns"):
        return "TIME_SERIES"
    if any(token in title_text for token in ("share", "mix", "composition", "%")):
        return "PIE"
    if any(token in title_text for token in ("kpi", "total", "avg", "average", "ratio", "score")):
        return "KPI"
    if any(token in title_text for token in ("top", "bottom", "rank", "ranking")):
        return "BAR"
    if source_profile.get("date_columns") and source_profile.get("numeric_columns"):
        return "TIME_SERIES"
    return "BAR" if source_profile.get("numeric_columns") else "TABLE"


_IDENTIFIER_NAME_SUFFIXES = ("_id", " id", "_key", " key")


def _is_identifier_like_column(name: Any) -> bool:
    """Mirror of the FE ``isIdentifierLikeField`` — a join-key / identifier
    column that must never be AUTO-SEEDED as a dimension/breakdown.

    On a modeled multi-table (galaxy) import target a bare key like
    ``product_id`` is ambiguous across the joined facts, so the semantic
    engine fails loudly ("Field 'product_id' xuất hiện ở nhiều bảng đã
    JOIN") — the backend mirror of DA7-B1/B3. Only the BLIND positional
    fallback is guarded; an explicit AI/label match is still honored.
    """
    bare = str(name or "").strip().lower()
    if not bare:
        return False
    return bare == "id" or bare.endswith(_IDENTIFIER_NAME_SUFFIXES) or "uuid" in bare


def _fallback_field_mapping(block: Dict[str, Any], source_profile: Dict[str, Any], chart_type: str) -> Dict[str, Any]:
    columns = list(source_profile.get("columns") or [])
    numeric_columns = list(source_profile.get("numeric_columns") or [])
    date_columns = list(source_profile.get("date_columns") or [])
    dimension_columns = list(source_profile.get("dimension_columns") or [])
    # Positional dimension/breakdown fallbacks must skip identifier/join-key
    # columns (DA7-B1/B3 backend sibling) — a bare key auto-seeded as a
    # grouping role fails loudly on a modeled multi-table target. Label/AI
    # matches above still use the full `columns`/`dimension_columns`.
    safe_dimension_columns = [name for name in dimension_columns if not _is_identifier_like_column(name)]
    table_summary = block.get("table") if isinstance(block.get("table"), dict) else {}
    headers = [str(item) for item in (table_summary.get("headers") or []) if str(item).strip()]
    label_context = " ".join(headers[:4]) or block.get("heading") or block.get("text") or ""

    primary_metric = _best_field_match(label_context, columns, numeric_only=True) or (numeric_columns[0] if numeric_columns else None)
    secondary_metric = None
    if primary_metric:
        secondary_metric = next((name for name in numeric_columns if name != primary_metric), None)

    primary_dimension = (
        _best_field_match(label_context, columns, date_only=True)
        or _best_field_match(label_context, columns)
        or (date_columns[0] if date_columns else None)
        or (safe_dimension_columns[0] if safe_dimension_columns else None)
    )

    breakdown = None
    if chart_type == "STACKED_BAR":
        breakdown = next((name for name in safe_dimension_columns if name != primary_dimension), None)

    mapping: Dict[str, Any] = {
        "dimension": primary_dimension,
        "timeField": primary_dimension if chart_type == "TIME_SERIES" and primary_dimension in date_columns else None,
        "metrics": ([{"field": primary_metric, "agg": "sum"}] if primary_metric else []),
        "breakdown": breakdown,
        "lineMetric": {"field": secondary_metric, "agg": "sum"} if chart_type == "BAR_LINE" and secondary_metric else None,
        "selectedColumns": [col.get("name") for col in columns[: min(len(columns), 8)] if col.get("name")],
        "scatterX": numeric_columns[0] if chart_type == "SCATTER" and len(numeric_columns) >= 1 else None,
        "scatterY": numeric_columns[1] if chart_type == "SCATTER" and len(numeric_columns) >= 2 else None,
    }

    if chart_type == "KPI" and not mapping["metrics"] and numeric_columns:
        mapping["metrics"] = [{"field": numeric_columns[0], "agg": "sum"}]
    if chart_type == "PIE" and not mapping["dimension"] and safe_dimension_columns:
        mapping["dimension"] = safe_dimension_columns[0]
    if chart_type == "TABLE" and headers:
        matched_headers: List[str] = []
        for header in headers:
            matched = _best_field_match(header, columns)
            if matched and matched not in matched_headers:
                matched_headers.append(matched)
        if matched_headers:
            mapping["selectedColumns"] = matched_headers
    return mapping


def _size_hint_from_block(block: Dict[str, Any], chart_type: str) -> str:
    style = block.get("style") or {}
    width_value = str(style.get("width") or style.get("grid_column") or "").lower()
    text = " ".join(
        [
            _normalize_text(block.get("heading")),
            _normalize_text(block.get("text")),
            " ".join(block.get("classes") or []),
        ]
    ).lower()
    if chart_type == "KPI":
        return "kpi"
    if chart_type == "TABLE":
        return "full"
    if "100%" in width_value or "full" in text or "wide" in text:
        return "full"
    if any(token in width_value for token in ("50%", "1 / span 6", "half")) or "half" in text:
        return "half"
    if any(token in text for token in ("small", "mini", "compact")):
        return "third"
    return "half"


def _layout_dimensions(chart_type: str, size_hint: str) -> Tuple[int, int]:
    if chart_type == "KPI":
        return 3, 3
    if chart_type == "TABLE":
        return 12, 6
    if size_hint == "full":
        return 12, 5
    if size_hint == "third":
        return 4, 4
    return 6, 4


def _normalize_declared_layout(raw_layout: Any) -> Optional[Dict[str, Any]]:
    """Keep valid authored grid/canvas geometry without inventing coordinates.

    A canvas tile may declare only ``xPx/yPx/wPx/hPx``. Treating that as a
    malformed grid layout used to erase the pixel geometry before the builder
    could persist it. Grid coordinates are therefore accepted as one complete
    group, while canvas coordinates are copied independently.
    """
    if not isinstance(raw_layout, dict) or not raw_layout:
        return None

    normalized: Dict[str, Any] = {}
    if all(key in raw_layout for key in ("x", "y", "w", "h")):
        try:
            x = int(raw_layout["x"])
            y = int(raw_layout["y"])
            w = max(1, min(12, int(raw_layout["w"])))
            h = max(1, int(raw_layout["h"]))
            if x >= 0 and y >= 0 and (x + w) <= 12:
                normalized.update({"x": x, "y": y, "w": w, "h": h})
        except (TypeError, ValueError):
            pass

    for key in ("xPx", "yPx", "wPx", "hPx"):
        value = raw_layout.get(key)
        if isinstance(value, bool) or value is None:
            continue
        try:
            number = float(value)
        except (TypeError, ValueError):
            continue
        if key in {"xPx", "yPx"} and number < 0:
            continue
        if key in {"wPx", "hPx"} and number <= 0:
            continue
        normalized[key] = int(number) if number.is_integer() else number

    z_value = raw_layout.get("z")
    if not isinstance(z_value, bool) and z_value is not None:
        try:
            normalized["z"] = int(z_value)
        except (TypeError, ValueError):
            pass
    if isinstance(raw_layout.get("styleConfigOverride"), dict):
        normalized["styleConfigOverride"] = dict(raw_layout["styleConfigOverride"])
    return normalized or None


def _assign_layouts(plans: List[Dict[str, Any]]) -> None:
    """Pack plans into a 12-column grid. Plans that already have a valid
    ``layout`` (provided by v1 metadata) are honored — their occupied cells
    become obstacles for newly-packed plans.
    """
    heights = [0] * 12
    # Seed the packing grid with plans that already supply a layout.
    for plan in plans:
        layout = plan.get("layout")
        if not isinstance(layout, dict):
            continue
        try:
            x = int(layout.get("x"))
            y = int(layout.get("y"))
            w = max(1, min(12, int(layout.get("w"))))
            h = max(1, int(layout.get("h")))
        except (TypeError, ValueError):
            plan["layout"] = None
            continue
        if x < 0 or y < 0 or (x + w) > 12:
            plan["layout"] = None
            continue
        for index in range(x, x + w):
            heights[index] = max(heights[index], y + h)

    for plan in plans:
        if isinstance(plan.get("layout"), dict):
            continue
        chart_type = str(plan.get("final_chart_type") or "BAR")
        size_hint = str(plan.get("size_hint") or "half")
        width, height = _layout_dimensions(chart_type, size_hint)
        best_x = 0
        best_y = 0
        best_score: Tuple[int, int] | None = None
        for x in range(0, 12 - width + 1):
            y = max(heights[x : x + width])
            score = (y, x)
            if best_score is None or score < best_score:
                best_score = score
                best_x = x
                best_y = y
        for index in range(best_x, best_x + width):
            heights[index] = best_y + height
        plan["layout"] = {"x": best_x, "y": best_y, "w": width, "h": height}


def _sanitize_metric(metric: Any, source_profile: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not isinstance(metric, dict):
        return None
    field = str(metric.get("field") or "").strip()
    if field not in source_profile.get("numeric_columns", []):
        return None
    agg = str(metric.get("agg") or "sum").lower()
    if agg not in {"sum", "avg", "count", "min", "max", "count_distinct"}:
        agg = "sum"
    return {"field": field, "agg": agg}


def _finalize_plan(
    *,
    block: Dict[str, Any],
    raw_plan: Dict[str, Any] | None,
    source_profile: Dict[str, Any],
) -> Dict[str, Any]:
    suggested_type = str((raw_plan or {}).get("suggested_chart_type") or (raw_plan or {}).get("final_chart_type") or _fallback_chart_kind(block, source_profile)).upper()
    original_chart_type = str((raw_plan or {}).get("original_chart_type") or _BLOCK_ROLE_TO_CHART_TYPE.get(str(block.get("role") or ""), "unknown")).upper()

    fallback_type = _fallback_chart_kind(block, source_profile)
    final_chart_type = suggested_type if suggested_type in _SUPPORTED_CHART_TYPES else fallback_type
    changed_chart_type = final_chart_type != suggested_type and bool(suggested_type)

    raw_mapping = (raw_plan or {}).get("field_mapping") if isinstance((raw_plan or {}).get("field_mapping"), dict) else {}
    fallback_mapping = _fallback_field_mapping(block, source_profile, final_chart_type)
    dimension = str(raw_mapping.get("dimension") or "").strip() if isinstance(raw_mapping, dict) else ""
    time_field = str(raw_mapping.get("timeField") or "").strip() if isinstance(raw_mapping, dict) else ""
    breakdown = str(raw_mapping.get("breakdown") or "").strip() if isinstance(raw_mapping, dict) else ""
    selected_columns = [
        str(item).strip()
        for item in (raw_mapping.get("selectedColumns") or [])
        if str(item).strip() in {col["name"] for col in source_profile.get("columns", [])}
    ] if isinstance(raw_mapping, dict) else []

    if dimension and dimension not in {col["name"] for col in source_profile.get("columns", [])}:
        dimension = ""
    if time_field and time_field not in source_profile.get("date_columns", []):
        time_field = ""
    if breakdown and breakdown not in {col["name"] for col in source_profile.get("columns", [])}:
        breakdown = ""

    metrics = [
        metric
        for metric in (_sanitize_metric(item, source_profile) for item in (raw_mapping.get("metrics") or []))
        if metric
    ] if isinstance(raw_mapping, dict) else []
    if not metrics:
        metrics = list(fallback_mapping.get("metrics") or [])

    line_metric = _sanitize_metric(raw_mapping.get("lineMetric"), source_profile) if isinstance(raw_mapping, dict) else None
    if final_chart_type == "BAR_LINE" and not line_metric:
        fallback_line = fallback_mapping.get("lineMetric")
        line_metric = fallback_line if isinstance(fallback_line, dict) else None

    scatter_x = str(raw_mapping.get("scatterX") or "").strip() if isinstance(raw_mapping, dict) else ""
    scatter_y = str(raw_mapping.get("scatterY") or "").strip() if isinstance(raw_mapping, dict) else ""
    if final_chart_type == "SCATTER":
        if scatter_x not in source_profile.get("numeric_columns", []):
            scatter_x = str(fallback_mapping.get("scatterX") or "")
        if scatter_y not in source_profile.get("numeric_columns", []):
            scatter_y = str(fallback_mapping.get("scatterY") or "")

    # Final contract validation per chart type.
    warnings: List[str] = []
    conversion_note = _normalize_text((raw_plan or {}).get("conversion_note"), max_len=240)

    if final_chart_type == "TABLE":
        selected_columns = selected_columns or list(fallback_mapping.get("selectedColumns") or [])
        if not selected_columns:
            selected_columns = [col["name"] for col in source_profile.get("columns", [])[: min(8, len(source_profile.get("columns", [])))]]
    elif final_chart_type == "KPI":
        if not metrics:
            final_chart_type = "TABLE"
            selected_columns = list(fallback_mapping.get("selectedColumns") or [])
            changed_chart_type = True
            conversion_note = conversion_note or "Converted to TABLE because no numeric field could power a KPI."
    elif final_chart_type == "TIME_SERIES":
        if not time_field:
            time_field = str(fallback_mapping.get("timeField") or "")
        if not time_field or not metrics:
            final_chart_type = "BAR" if metrics else "TABLE"
            changed_chart_type = True
            conversion_note = conversion_note or "Converted from TIME_SERIES because the selected source did not expose a valid date/value mapping."
    elif final_chart_type == "PIE":
        if not dimension:
            dimension = str(fallback_mapping.get("dimension") or "")
        if not dimension or not metrics:
            final_chart_type = "BAR" if metrics else "TABLE"
            changed_chart_type = True
            conversion_note = conversion_note or "Converted from PIE because the selected source did not expose both a label field and a numeric value."
    elif final_chart_type == "SCATTER":
        if not scatter_x or not scatter_y:
            final_chart_type = "TABLE"
            changed_chart_type = True
            conversion_note = conversion_note or "Converted from SCATTER because two numeric fields were not available."
    elif final_chart_type == "STACKED_BAR":
        if not dimension:
            dimension = str(fallback_mapping.get("dimension") or "")
        if not breakdown:
            breakdown = str(fallback_mapping.get("breakdown") or "")
        if not dimension or not breakdown or not metrics:
            final_chart_type = "BAR" if metrics else "TABLE"
            changed_chart_type = True
            conversion_note = conversion_note or "Converted from STACKED_BAR because the source could not supply dimension, stack field, and metric together."
    elif final_chart_type == "BAR_LINE":
        if not dimension:
            dimension = str(fallback_mapping.get("dimension") or "")
        if not metrics or not line_metric:
            final_chart_type = "BAR" if metrics else "TABLE"
            changed_chart_type = True
            conversion_note = conversion_note or "Converted from BAR_LINE because the source did not expose both bar and line measures."
    else:
        if final_chart_type != "TABLE" and not dimension:
            dimension = str(fallback_mapping.get("dimension") or "")
        if final_chart_type != "TABLE" and not metrics:
            final_chart_type = "TABLE"
            changed_chart_type = True
            conversion_note = conversion_note or f"Converted to TABLE because the source did not expose a numeric measure for {suggested_type or final_chart_type}."

    if changed_chart_type and not conversion_note:
        conversion_note = f"Requested {suggested_type or original_chart_type}, imported as {final_chart_type} to fit the native dashboard canvas."

    title = _normalize_text((raw_plan or {}).get("title") or block.get("heading") or block.get("text"), max_len=160) or f"Imported {final_chart_type.title()}"
    rationale = _normalize_text((raw_plan or {}).get("reasoning") or (raw_plan or {}).get("rationale"), max_len=260)
    confidence_raw = (raw_plan or {}).get("confidence")
    try:
        confidence = max(0.05, min(float(confidence_raw), 0.99))
    except (TypeError, ValueError):
        confidence = 0.62 if rationale else 0.5

    size_hint = str((raw_plan or {}).get("size_hint") or _size_hint_from_block(block, final_chart_type)).lower()
    if size_hint not in {"full", "half", "third", "kpi"}:
        size_hint = _size_hint_from_block(block, final_chart_type)

    source_fields_used = sorted(
        {
            *([dimension] if dimension else []),
            *([time_field] if time_field else []),
            *([breakdown] if breakdown else []),
            *([scatter_x] if scatter_x else []),
            *([scatter_y] if scatter_y else []),
            *([item.get("field") for item in metrics if item.get("field")]),
            *([line_metric.get("field")] if isinstance(line_metric, dict) and line_metric.get("field") else []),
            *(selected_columns or []),
        }
    )
    if not source_fields_used:
        warnings.append("No field mapping could be validated for this block.")

    role_config: Dict[str, Any] = {"metrics": metrics}
    if dimension:
        role_config["dimension"] = dimension
    if breakdown:
        role_config["breakdown"] = breakdown
    if final_chart_type == "TIME_SERIES" and time_field:
        role_config["timeField"] = time_field
        if not role_config.get("dimension"):
            role_config["dimension"] = time_field
    if final_chart_type == "BAR_LINE" and line_metric:
        role_config["lineMetric"] = line_metric
    if final_chart_type == "TABLE":
        role_config["selectedColumns"] = selected_columns
    if final_chart_type == "SCATTER":
        role_config["scatterX"] = scatter_x
        role_config["scatterY"] = scatter_y

    return {
        "block_id": block["id"],
        "order": block.get("order", 0),
        "title": title,
        "block_role": block.get("role"),
        "source_excerpt": _normalize_text(block.get("text"), max_len=280),
        "original_chart_type": original_chart_type,
        "requested_chart_type": suggested_type,
        "final_chart_type": final_chart_type,
        "changed_chart_type": changed_chart_type,
        "conversion_note": conversion_note or None,
        "rationale": rationale or None,
        "confidence": round(confidence, 2),
        "size_hint": size_hint,
        "source_fields_used": source_fields_used,
        "warnings": warnings,
        "role_config": role_config,
        "style_config": {"chartTitle": title},
    }


def _complete_json_with_import_provider(
    prompt: str,
    *,
    system_prompt: str,
    max_tokens: int = 1800,
) -> Optional[Dict[str, Any]]:
    provider = settings.html_import_ai_provider
    model = settings.html_import_ai_model
    if provider == "unavailable":
        return None

    # Legacy Gemini path — retained but NOT selected: html_import_ai_provider
    # now returns "openai" (or "unavailable"). Kept for reference only.
    if provider == "gemini":
        try:
            import google.generativeai as genai

            genai.configure(api_key=settings.GEMINI_API_KEY)
            client = genai.GenerativeModel(
                model_name=model,
                system_instruction=system_prompt,
                generation_config={
                    "temperature": 0.2,
                    "max_output_tokens": max_tokens,
                    "response_mime_type": "application/json",
                },
            )
            response = client.generate_content(prompt)
            text = getattr(response, "text", "") or ""
            if not text:
                for candidate in getattr(response, "candidates", []) or []:
                    content = getattr(candidate, "content", None)
                    for part in getattr(content, "parts", []) or []:
                        part_text = getattr(part, "text", "") or ""
                        if part_text:
                            text = part_text
                            break
                    if text:
                        break
            if not text:
                return None
            return _parse_ai_json(text)
        except Exception:
            logger.warning("Dashboard HTML import Gemini call failed", exc_info=True)
            if settings.active_api_keys:
                return LLMClient.complete_json(
                    prompt,
                    system=system_prompt,
                    model=settings.html_import_openrouter_model,
                    max_tokens=max_tokens,
                )
            return None

    # OpenAI path (default): LLMClient.complete_json targets api.openai.com and
    # uses OPENAI_API_KEY; model is gpt-4o-mini (settings.html_import_ai_model).
    return LLMClient.complete_json(prompt, system=system_prompt, model=model, max_tokens=max_tokens)


#: Words that appear in an expression without being a column: SQL functions,
#: operators and literals. Everything else an expression names has to be a real
#: column or the query will not run.
_EXPRESSION_RESERVED = {
    "abs", "and", "as", "avg", "between", "case", "cast", "ceil", "ceiling",
    "coalesce", "count", "current_date", "date", "date_diff", "day", "else",
    "end", "exp", "extract", "false", "floor", "greatest", "if", "ifnull", "in",
    "is", "least", "ln", "log", "max", "min", "mod", "month", "not", "null",
    "nullif", "or", "power", "round", "safe_cast", "safe_divide", "sqrt", "sum",
    "then", "true", "trunc", "when", "year",
}
_EXPRESSION_IDENT_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*")


def _expression_unknown_columns(expression: str, known_columns: Set[str]) -> List[str]:
    """Identifiers in an expression that are not columns and not SQL words."""
    known_lower = {str(c).lower() for c in known_columns}
    unknown: List[str] = []
    for token in _EXPRESSION_IDENT_RE.findall(expression or ""):
        lowered = token.lower()
        if lowered in _EXPRESSION_RESERVED or lowered in known_lower:
            continue
        if token not in unknown:
            unknown.append(token)
    return unknown


def _validate_calculated_fields(
    calculated_fields: List[Dict[str, Any]],
    source_columns: List[str],
    warnings_out: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """Validate AI-suggested calculated fields, keeping only safe ones.

    "Safe" has to include "references columns that exist". `all_known_columns`
    was assembled here and then never consulted, so an expression naming
    columns the source does not have -- `avg_delivery_days_current -
    avg_delivery_days_previous`, invented wholesale from a KPI card that showed
    a delta -- passed validation and was injected into the SELECT of EVERY
    chart. One hallucinated column therefore failed all ten charts of an import
    with `column "avg_delivery_days_current" does not exist`, and the Build
    button stayed disabled with no way for the author to see why.
    """
    from app.services.transformation_compiler import TransformationCompiler

    notes = warnings_out if warnings_out is not None else []
    valid: List[Dict[str, Any]] = []
    seen_names: set = set()
    all_known_columns = set(source_columns)

    for field in calculated_fields:
        if not isinstance(field, dict):
            continue
        name = str(field.get("name") or "").strip()
        expression = str(field.get("expression") or "").strip()
        if not name or not expression or name in seen_names:
            continue

        is_valid, error = TransformationCompiler.validate_expression(expression)
        if not is_valid:
            logger.warning("Skipping calculated field '%s': %s", name, error)
            notes.append(f"Calculated field '{name}' was dropped: {error}")
            continue

        # A field may reference the ones defined before it, so the known set
        # grows as we go -- but only with fields that themselves resolved.
        unknown = _expression_unknown_columns(expression, all_known_columns)
        if unknown:
            logger.warning(
                "Skipping calculated field '%s': unknown column(s) %s", name, ", ".join(unknown)
            )
            notes.append(
                f"Calculated field '{name}' was dropped: it refers to "
                + ", ".join(f"'{c}'" for c in unknown[:4])
                + ", which the source does not have."
            )
            continue

        seen_names.add(name)
        all_known_columns.add(name)
        valid.append({
            "name": name,
            "expression": expression,
            "label": str(field.get("label") or field.get("description") or name),
            "source_key": field.get("source_key"),
        })

    return valid


def _document_shape(
    document_summary: Dict[str, Any],
    candidates: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """The counts that decide what kind of report this is.

    Handed to the model because it was choosing the layout family from the
    document's language, where the answer is in its construction: a report with
    three KPI cards, one chart and a detail table is a brief however its
    headings are worded.
    """
    blocks = document_summary.get("blocks") or []
    roles = [str(b.get("role") or "") for b in blocks]
    prose = sum(
        len(_normalize_text(b.get("text")))
        for b in blocks
        if str(b.get("role") or "") not in {"chart", "table", "kpi"}
    )
    return {
        "kpi_blocks": roles.count("kpi"),
        "chart_blocks": roles.count("chart"),
        "table_blocks": roles.count("table"),
        "heading_blocks": roles.count("title"),
        "prose_characters": prose,
        "filter_controls": len(document_summary.get("filter_controls") or []),
        "chartable_blocks": len(candidates),
    }


def _ai_chart_plans(
    *,
    document_summary: Dict[str, Any],
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Tuple[Optional[List[Dict[str, Any]]], List[Dict[str, Any]], Dict[str, Any]]:
    """Return ``(charts, calculated_fields, ai_meta)``."""
    if settings.html_import_ai_provider == "unavailable":
        return None, [], build_ai_assist_meta(
            requested=True,
            applied=False,
            status="unavailable",
            message="No AI provider configured for HTML dashboard import.",
        )

    candidates, _ignored = _candidate_blocks(document_summary)
    if not candidates:
        return None, [], build_ai_assist_meta(
            requested=True,
            applied=False,
            status="failed",
            message="No chart-like blocks were detected in the HTML summary.",
        )

    source_col_names = [col.get("name") for col in (source_profile.get("columns") or [])]
    if all_source_profiles and len(all_source_profiles) > 1:
        source_col_names = list({
            col.get("name")
            for prof in all_source_profiles.values()
            for col in (prof.get("columns") or [])
        })

    # Build source_profile section for prompt — include all tables when multi-source
    use_multi_prompt = all_source_profiles and len(all_source_profiles) > 1
    if use_multi_prompt:
        source_section = {
            "dataset_name": source_profile.get("dataset_name"),
            "tables": [
                {
                    "source_key": key,
                    "dataset_table_name": prof.get("dataset_table_name"),
                    "columns": [
                        {"name": item.get("name"), "type": item.get("type")}
                        for item in (prof.get("columns") or [])
                    ],
                    "numeric_columns": prof.get("numeric_columns"),
                    "date_columns": prof.get("date_columns"),
                    "sample_rows": (prof.get("sample_rows") or [])[:3],
                }
                for key, prof in all_source_profiles.items()
            ],
        }
        multi_source_rules = [
            "The dataset has MULTIPLE tables. Each chart should use fields from ONE table only.",
            "For each chart, set source_key to the table name that best fits the chart's data needs.",
            "Calculated fields must reference columns from a single table; set source_key on the calculated_field too.",
        ]
    else:
        source_section = {
            "dataset_name": source_profile.get("dataset_name"),
            "dataset_table_name": source_profile.get("dataset_table_name"),
            "columns": [
                {"name": item.get("name"), "type": item.get("type")}
                for item in (source_profile.get("columns") or [])
            ],
            "numeric_columns": source_profile.get("numeric_columns"),
            "date_columns": source_profile.get("date_columns"),
            "sample_rows": source_profile.get("sample_rows"),
        }
        multi_source_rules = []

    prompt = json.dumps(
        {
            "task": (
                "Read this document as a COMPOSITION, then convert it. Work in three "
                "passes and report all three. (1) REGIONS: name the areas the page is "
                "actually built from. (2) BLOCKS: classify every block by what AppBI can "
                "do with it, without deciding chart types yet. (3) CHARTS: only for the "
                "blocks classified native_chart, produce the chart plan. Passes 1 and 2 "
                "exist because a report IS its arrangement: converting chart-first "
                "produced the same generic mosaic no matter what the source looked like. "
                "When the HTML references metrics that are not in the source columns, "
                "propose calculated_fields with SQL-safe expressions to derive them."
            ),
            "supported_chart_types": sorted(_SUPPORTED_CHART_TYPES),
            "rules": [
                "Stay inside the supported chart types list.",
                "Use fields from source_profile.columns OR from calculated_fields you define.",
                "If the HTML implies a metric not directly in the source (e.g. percentage, cumulative sum, profit = revenue - cost, remaining = budget - spent), create a calculated_field for it.",
                "Calculated field expressions must be simple SQL-safe math: +, -, *, /, ROUND(), COALESCE(), IF(condition, true_val, false_val). No SELECT/FROM/JOIN.",
                "Calculated field names must be valid SQL identifiers (no spaces, no special chars).",
                "If the HTML implies a chart AppBI does not support, choose the closest supported chart type and explain it in conversion_note.",
                "Every block gets exactly one classification in block_plan. Nothing is "
                "left out: a block you cannot use is classified, not omitted. Dropping "
                "blocks silently is what made imported reports lose their structure.",
                "classification meanings. native_chart = the block plots data AppBI can "
                "query. native_widget = a heading, a callout, a lead paragraph, a hero "
                "banner: structure a themeable widget can carry. slicer = a visible "
                "filter control (dropdown, segmented control, date range) that filters "
                "the rest of the page. html_fragment_preserved = visual richness with no "
                "AppBI equivalent (a gauge or illustration drawn in SVG, an avatar row, a "
                "progress rail, a styled badge grid); choose it ONLY when has_markup is "
                "true and nothing native fits, because a fragment is frozen: it will not "
                "follow the theme, translate, or respond to filters. "
                "unsupported_interaction = the block IS an interaction (a tab strip, a "
                "modal trigger, a live-refresh control, a drill-down); say what stops "
                "working in degraded_note.",
                "Choose layout_recipe BEFORE assigning any block to a region. The recipe "
                "decides which regions exist; regions do not add up to a recipe.",
                "Choose it from document_shape, not from the document's wording. The test "
                "that separates the two commonest answers: a console carries a GRID of "
                "charts (four or more); a brief carries one or two, with a detail table "
                "under them. Long prose beats both -- that is editorial.",
                "Set layout_recipe from how the document is BUILT, not its wording: "
                "console = KPI strip over a chart grid with a filter bar; "
                "brief = KPI strip plus one main chart and a detail table, read in a meeting; "
                "ops = a dense table/chart grid monitored continuously; "
                "editorial = long prose carrying a few wide visuals; "
                "stage = a wall of big numbers meant to be read from a distance.",
                "Region names to use: hero, kpi_strip, filter_rail, chart_mosaic, "
                "insight_column, detail_band, footer. Name only the regions the document "
                "actually has, in the order they appear.",
                "slicer_fields: list ONLY columns that a visible filter/dropdown/segmented "
                "control in the HTML is filtering by. Omit the key when there is no filter UI.",
                "Prefer TABLE when confidence is low or when the block behaves like a wide data table.",
                "KPI requires one numeric metric. TIME_SERIES requires one date field and one numeric metric. PIE requires one dimension and one numeric metric.",
                "BAR_LINE requires one dimension plus one bar metric and one line metric.",
                "SCATTER requires two numeric fields.",
                "PODIUM requires one dimension (name field) and one numeric metric; use it for top-N ranking blocks with medal/rank styling.",
                *multi_source_rules,
            ],
            "document_summary": {
                "title": document_summary.get("title"),
                "html_excerpt": document_summary.get("html_excerpt"),
                "filter_controls": document_summary.get("filter_controls") or [],
                "blocks": [
                    {
                        "id": block.get("id"),
                        "order": block.get("order"),
                        "role": block.get("role"),
                        "heading": block.get("heading"),
                        "text": _normalize_text(block.get("text"), max_len=240),
                        "classes": (block.get("classes") or [])[:8],
                        "style": block.get("style"),
                        "table": block.get("table"),
                        # Whether this block's appearance was captured. Without
                        # it html_fragment_preserved is not an option, and the
                        # model has to find something native or say the block is
                        # an interaction it cannot carry.
                        "has_markup": bool(block.get("html")),
                    }
                    # Composition needs the heading, note and control blocks too;
                    # a candidates-only list made "classify every block"
                    # impossible by construction.
                    for block in (document_summary.get("blocks") or [])[:40]
                ],
            },
            # The same structural facts the fallback heuristic uses. Without
            # them the model was choosing the family from the document's prose:
            # a report with three KPI cards, ONE chart and a detail table -- the
            # definition of a brief -- came back as "console", with reasoning
            # that described a brief.
            "document_shape": _document_shape(document_summary, candidates),
            "source_profile": source_section,
            "return_json_shape": {
                "dashboard_title": "string",
                # Pass 1. The composition, decided before any block is placed.
                # This is the piece that was missing: the old shape asked only
                # for charts, so the arrangement of the source was thrown away
                # and every import came out looking the same.
                "composition": {
                    "layout_recipe": "console|brief|ops|editorial|stage",
                    "reasoning": "what about the document's construction says this",
                    "regions": [
                        {
                            "name": "hero|kpi_strip|filter_rail|chart_mosaic|insight_column|detail_band|footer",
                            "block_ids": ["ids of the blocks that sit in this region, in order"],
                        }
                    ],
                },
                # Pass 2. EVERY block, classified. A block missing from this
                # list is a block the import silently loses.
                "block_plan": [
                    {
                        "block_id": "string",
                        "classification": "native_chart|native_widget|slicer|html_fragment_preserved|unsupported_interaction",
                        "widget_type": "section_header|callout|text|hero_strip | null (native_widget only)",
                        "degraded_note": "string | null: what the source block did that the imported one will not",
                    }
                ],
                # Columns any filter control in the source appears to drive, so
                # an imported report keeps its filter UX instead of arriving
                # with none.
                "slicer_fields": ["source column name"],
                "calculated_fields": [
                    {
                        "name": "ValidSQLIdentifier",
                        "expression": "col_a / col_b * 100",
                        "label": "Human readable label",
                        "source_key": "optional source_key for multi-source",
                    }
                ],
                "charts": [
                    {
                        "block_id": "string",
                        "title": "string",
                        "original_chart_type": "string",
                        "suggested_chart_type": "BAR|HORIZONTAL_BAR|LINE|PIE|TIME_SERIES|TABLE|AREA|STACKED_BAR|GROUPED_BAR|BAR_LINE|SCATTER|KPI|PODIUM",
                        "source_key": "table display_name (required for multi-table datasets)",
                        "field_mapping": {
                            "dimension": "string | null",
                            "timeField": "string | null",
                            "metrics": [{"field": "string", "agg": "sum|avg|count|min|max|count_distinct"}],
                            "breakdown": "string | null",
                            "lineMetric": {"field": "string", "agg": "sum|avg|count|min|max|count_distinct"} ,
                            "selectedColumns": ["string"],
                            "scatterX": "string | null",
                            "scatterY": "string | null",
                        },
                        "confidence": 0.0,
                        "size_hint": "full|half|third|kpi",
                        "reasoning": "string",
                        "conversion_note": "string | null",
                    }
                ],
            },
        },
        ensure_ascii=False,
    )

    system_prompt = (
        "You convert an HTML report into an AppBI dashboard. "
        "Composition first: decide what kind of report this is and what regions it is "
        "built from BEFORE you decide anything about individual charts, and classify "
        "every block -- a block you cannot convert is reported with a reason, never "
        "left out. "
        "When the raw data lacks a metric the HTML dashboard displays (like percentages, totals, differences, cumulative values), "
        "define it as a calculated_field with a SQL-safe expression before referencing it in charts. "
        "When the dataset has multiple tables, assign each chart to the most appropriate table via source_key. "
        "Be conservative. Stay inside the supported chart types. Prefer native widgets "
        "for structure; preserve sanitized static markup only when no native surface can "
        "carry the visual, and report unsupported interactions explicitly."
    )

    payload = _complete_json_with_import_provider(prompt, system_prompt=system_prompt, max_tokens=3000)
    if not isinstance(payload, dict):
        return None, [], build_ai_assist_meta(
            requested=True,
            applied=False,
            status="failed",
            provider=settings.html_import_ai_provider,
            model=settings.html_import_ai_model,
            message="AI provider did not return a usable chart plan payload.",
        )

    charts = payload.get("charts")
    if not isinstance(charts, list):
        return None, [], build_ai_assist_meta(
            requested=True,
            applied=False,
            status="failed",
            provider=settings.html_import_ai_provider,
            model=settings.html_import_ai_model,
            message="AI payload did not include a charts array.",
        )

    raw_calc_fields = payload.get("calculated_fields") or []
    if not isinstance(raw_calc_fields, list):
        raw_calc_fields = []
    calc_field_warnings: List[str] = []
    validated_calc_fields = _validate_calculated_fields(
        raw_calc_fields, source_col_names, warnings_out=calc_field_warnings
    )
    ai_meta = build_ai_assist_meta(
        requested=True,
        applied=True,
        status="applied",
        provider=settings.html_import_ai_provider,
        model=settings.html_import_ai_model,
        message="AI refined HTML block classification, field mapping, and calculated fields inside the native dashboard/chart contract.",
    )
    if calc_field_warnings:
        # Set AFTER build_ai_assist_meta, which REPLACES ai_meta wholesale --
        # writing this before it silently threw the warnings away.
        ai_meta["calculated_field_warnings"] = calc_field_warnings[:6]
    # Composition the model was asked for. Carried on ai_meta rather than as a
    # fourth return value so every existing caller keeps working; the analyzer
    # reads it back off the meta when it builds the dashboard contract.
    composition = payload.get("composition") if isinstance(payload.get("composition"), dict) else {}
    # `template_family` is the older key; the prompt now asks for it inside
    # composition as `layout_recipe`, and imports made by a previous version of
    # the prompt still round-trip through the old one.
    family = (
        _normalize_template_family(composition.get("layout_recipe"))
        or _normalize_template_family(payload.get("template_family"))
    )
    if family:
        ai_meta["template_family"] = family
    if composition.get("reasoning"):
        ai_meta["composition_reasoning"] = _normalize_text(composition.get("reasoning"), max_len=280)

    # Region assignment: which blocks the model says belong together. The
    # packer uses it to keep a source's grouping instead of re-deriving one
    # from block kinds alone.
    regions: List[Dict[str, Any]] = []
    for entry in (composition.get("regions") or []):
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or "").strip().lower()
        block_ids = [str(b).strip() for b in (entry.get("block_ids") or []) if str(b).strip()]
        if name and block_ids:
            regions.append({"name": name, "block_ids": block_ids[:24]})
    if regions:
        ai_meta["regions"] = regions[:8]

    # Pass 2: every block, classified. This is what stops a block from being
    # dropped in silence -- anything the model could not use is still named,
    # with a reason the author can read.
    block_plan: List[Dict[str, Any]] = []
    for entry in (payload.get("block_plan") or []):
        if not isinstance(entry, dict):
            continue
        block_id = str(entry.get("block_id") or "").strip()
        classification = str(entry.get("classification") or "").strip().lower()
        if not block_id or classification not in _BLOCK_CLASSIFICATIONS:
            continue
        widget_type = str(entry.get("widget_type") or "").strip().lower()
        block_plan.append({
            "block_id": block_id,
            "classification": classification,
            "widget_type": widget_type if widget_type in _AI_WIDGET_TYPES else "",
            "degraded_note": _normalize_text(entry.get("degraded_note"), max_len=200),
        })
    if block_plan:
        ai_meta["block_plan"] = block_plan[:40]

    slicer_fields = [
        str(v).strip() for v in (payload.get("slicer_fields") or [])
        if isinstance(v, (str, int, float)) and str(v).strip()
    ]
    if slicer_fields:
        ai_meta["slicer_fields"] = slicer_fields[:8]

    return charts, validated_calc_fields, ai_meta


_APPBI_META_RE = re.compile(
    r'<script[^>]+type\s*=\s*["\']application/appbi-dashboard["\'][^>]*>(.*?)</script>',
    re.DOTALL | re.IGNORECASE,
)

_APPBI_SOURCE_PLACEHOLDER_RE = re.compile(
    r"\{\{\s*source\s*:\s*([^}]+?)\s*\}\}"
)

class SourceContractMismatch(ValueError):
    """The file was written for data this dataset no longer has.

    Raised rather than warned: the alternative is importing a dashboard whose
    charts point at columns chosen by resemblance, which looks like it worked.
    """

    def __init__(self, findings: List[str]):
        self.findings = list(findings)
        super().__init__(" ".join(self.findings) or "Source contract does not match the selected dataset.")


APPBI_IMPORT_PLAN_V1_VERSION = "appbi-import/v1"
APPBI_IMPORT_PLAN_V2_VERSION = "appbi-import/v2"


def _extract_embedded_metadata(html_text: str) -> Optional[Dict[str, Any]]:
    """Extract ``<script type="application/appbi-dashboard">`` JSON from HTML.

    Returns the parsed dict or *None* when no valid metadata block is found.
    Accepts both the legacy shape and the AppBI Import Plan v1 shape.
    """
    match = _APPBI_META_RE.search(html_text)
    if not match:
        return None
    try:
        payload = json.loads(match.group(1))
        if isinstance(payload, dict) and (
            isinstance(payload.get("charts"), list)
            or isinstance(payload.get("pages"), list)
        ):
            return payload
    except (json.JSONDecodeError, ValueError):
        logger.warning("Found appbi-dashboard script tag but JSON is invalid.")
    return None


def _is_v1_metadata(embedded: Dict[str, Any]) -> bool:
    return str(embedded.get("version") or "").strip().lower() == APPBI_IMPORT_PLAN_V1_VERSION


def _is_v2_metadata(embedded: Dict[str, Any]) -> bool:
    return str(embedded.get("version") or "").strip().lower() == APPBI_IMPORT_PLAN_V2_VERSION


def _normalize_v2_pages(embedded: Dict[str, Any]) -> List[Dict[str, Any]]:
    raw_pages = embedded.get("pages") if isinstance(embedded.get("pages"), list) else []
    normalized_pages: List[Dict[str, Any]] = []
    for index, raw_page in enumerate(raw_pages, start=1):
        if not isinstance(raw_page, dict):
            continue
        charts = raw_page.get("charts") if isinstance(raw_page.get("charts"), list) else []
        page_key = _normalize_text(raw_page.get("page_key") or raw_page.get("id"), max_len=120) or f"page-{index}"
        title = _normalize_text(raw_page.get("title") or raw_page.get("name"), max_len=80) or f"Page {index}"
        normalized_pages.append(
            {
                "page_key": page_key,
                "title": title,
                "description": _normalize_text(raw_page.get("description"), max_len=400) or None,
                "charts": charts,
                "order": int(raw_page.get("order") or index),
            }
        )
    normalized_pages.sort(key=lambda item: (item.get("order", 0), item.get("page_key", "")))
    return normalized_pages


def _pick_default_v2_page(embedded: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    pages = _normalize_v2_pages(embedded)
    if not pages:
        return None

    dashboard_meta = embedded.get("dashboard") if isinstance(embedded.get("dashboard"), dict) else {}
    default_page_key = _normalize_text(dashboard_meta.get("default_page_key"), max_len=120)
    if default_page_key:
        for page in pages:
            if page.get("page_key") == default_page_key:
                return page
    return pages[0]


def _coerce_v2_page_to_v1_like_payload(
    *,
    embedded: Dict[str, Any],
    page: Dict[str, Any],
) -> Dict[str, Any]:
    dashboard_meta = dict(embedded.get("dashboard") or {})
    page_title = _normalize_text(page.get("title"), max_len=80) or "Imported Page"
    dashboard_meta["default_page_name"] = page_title
    return {
        "version": APPBI_IMPORT_PLAN_V2_VERSION,
        "strict": bool(embedded.get("strict", True)),
        "authoring_mode": str(embedded.get("authoring_mode") or "skill"),
        "dashboard": dashboard_meta,
        "source_contract": embedded.get("source_contract") if isinstance(embedded.get("source_contract"), dict) else {},
        "dataset_ops": embedded.get("dataset_ops") if isinstance(embedded.get("dataset_ops"), list) else [],
        "charts": list(page.get("charts") or []),
        "page_key": page.get("page_key"),
        "page_title": page_title,
        "page_description": page.get("description"),
        "page_count": len(_normalize_v2_pages(embedded)),
    }


def _analyze_from_v2_metadata_page(
    *,
    embedded: Dict[str, Any],
    page: Dict[str, Any],
    document_summary: Dict[str, Any],
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    page_payload = _coerce_v2_page_to_v1_like_payload(embedded=embedded, page=page)
    page_title = _normalize_text(page.get("title"), max_len=80) or document_summary.get("title") or "Imported Page"
    page_summary = dict(document_summary)
    page_summary["title"] = page_title

    result = _analyze_from_v1_metadata(
        embedded=page_payload,
        document_summary=page_summary,
        source_profile=source_profile,
        all_source_profiles=all_source_profiles,
    )
    result["document_title"] = page_title
    ai_meta = dict(result.get("ai_meta") or {})
    ai_meta["plan_version"] = APPBI_IMPORT_PLAN_V2_VERSION
    ai_meta["page_key"] = page.get("page_key")
    ai_meta["page_title"] = page_title
    ai_meta["page_count"] = page_payload.get("page_count")
    ai_meta["default_page_name"] = page_title
    result["ai_meta"] = ai_meta
    return result


def _finalize_plan_from_v1_metadata(
    *,
    raw: Dict[str, Any],
    order: int,
) -> Dict[str, Any]:
    """Build a finalized chart plan from a v1 ``charts[]`` entry.

    The v1 contract already uses the native runtime shape (``chart_type`` +
    ``role_config`` + ``base_filters`` + ``layout``), so we only need to
    normalize values and fall back to safe defaults where fields are absent.
    """
    chart_type_raw = str(raw.get("chart_type") or "TABLE").upper()
    final_chart_type = chart_type_raw if chart_type_raw in _SUPPORTED_CHART_TYPES else "TABLE"
    changed = final_chart_type != chart_type_raw
    conversion_note_raw = _normalize_text(raw.get("conversion_note"), max_len=280) or None
    if changed:
        conversion_note_raw = conversion_note_raw or (
            f"Requested {chart_type_raw}, imported as {final_chart_type}."
        )

    title = _normalize_text(raw.get("title"), max_len=160) or f"Imported {final_chart_type.title()}"
    description = _normalize_text(raw.get("description"), max_len=400) or None

    raw_role_config = raw.get("role_config") if isinstance(raw.get("role_config"), dict) else {}

    def _clean_metric(m: Any) -> Optional[Dict[str, str]]:
        if not isinstance(m, dict):
            return None
        field = str(m.get("field") or "").strip()
        if not field:
            return None
        agg = str(m.get("agg") or "sum").lower()
        if agg not in {"sum", "avg", "count", "min", "max", "count_distinct"}:
            agg = "sum"
        return {"field": field, "agg": agg}

    metrics = [m for m in (
        _clean_metric(m) for m in (raw_role_config.get("metrics") or [])
    ) if m]

    dimension = str(raw_role_config.get("dimension") or "").strip() or None
    time_field = str(raw_role_config.get("timeField") or "").strip() or None
    breakdown = str(raw_role_config.get("breakdown") or "").strip() or None
    scatter_x = str(raw_role_config.get("scatterX") or "").strip() or None
    scatter_y = str(raw_role_config.get("scatterY") or "").strip() or None
    selected_columns = [
        str(c).strip() for c in (raw_role_config.get("selectedColumns") or []) if str(c).strip()
    ]
    # Spec for TABLE charts uses `dimensions` (plural); accept it as an alias
    # for `selectedColumns` so authors can follow the role_config shape table
    # without falling back to the runtime-only field name.
    if not selected_columns:
        selected_columns = [
            str(c).strip() for c in (raw_role_config.get("dimensions") or []) if str(c).strip()
        ]
    line_metric = _clean_metric(raw_role_config.get("lineMetric"))

    role_config: Dict[str, Any] = {"metrics": metrics}
    if dimension:
        role_config["dimension"] = dimension
    if breakdown:
        role_config["breakdown"] = breakdown
    if final_chart_type == "TIME_SERIES" and time_field:
        role_config["timeField"] = time_field
        if not role_config.get("dimension"):
            role_config["dimension"] = time_field
    elif time_field:
        role_config["timeField"] = time_field
    if final_chart_type == "BAR_LINE" and line_metric:
        role_config["lineMetric"] = line_metric
    if final_chart_type == "TABLE" and selected_columns:
        role_config["selectedColumns"] = selected_columns
    if final_chart_type == "SCATTER":
        if scatter_x:
            role_config["scatterX"] = scatter_x
        if scatter_y:
            role_config["scatterY"] = scatter_y

    base_filters_raw = raw.get("base_filters") or []
    base_filters: List[Dict[str, Any]] = []
    for flt in base_filters_raw:
        if not isinstance(flt, dict):
            continue
        field = str(flt.get("field") or "").strip()
        op = str(flt.get("op") or flt.get("operator") or "eq").strip().lower()
        if not field:
            continue
        filter_entry: Dict[str, Any] = {"field": field, "op": op}
        if "value" in flt:
            filter_entry["value"] = flt.get("value")
        base_filters.append(filter_entry)

    layout = _normalize_declared_layout(raw.get("layout"))

    size_hint = str(raw.get("size_hint") or "").lower()
    if size_hint not in {"full", "half", "third", "kpi"}:
        if layout is not None and all(key in layout for key in ("x", "y", "w", "h")):
            lw = layout["w"]
            size_hint = (
                "kpi" if final_chart_type == "KPI" and lw <= 3
                else "third" if lw <= 4
                else "half" if lw <= 6
                else "full"
            )
        else:
            size_hint = (
                "kpi" if final_chart_type == "KPI"
                else "full" if final_chart_type == "TABLE"
                else "half"
            )

    style_config = dict(raw.get("style_config") or {})
    if "chartTitle" not in style_config:
        style_config["chartTitle"] = title

    source_fields_used = sorted({
        *(([dimension] if dimension else [])),
        *(([time_field] if time_field else [])),
        *(([breakdown] if breakdown else [])),
        *(([scatter_x] if scatter_x else [])),
        *(([scatter_y] if scatter_y else [])),
        *(m.get("field") for m in metrics if m.get("field")),
        *(([line_metric["field"]] if line_metric and line_metric.get("field") else [])),
        *(selected_columns or []),
    })

    try:
        order_val = int(raw.get("order") or order)
    except (TypeError, ValueError):
        order_val = order

    return {
        "block_id": str(raw.get("block_id") or f"v1-{order}"),
        "order": order_val,
        "title": title,
        "block_role": final_chart_type.lower() if final_chart_type in ("TABLE", "KPI") else "chart",
        "source_excerpt": description,
        "original_chart_type": chart_type_raw,
        "requested_chart_type": chart_type_raw,
        "final_chart_type": final_chart_type,
        "changed_chart_type": changed,
        "conversion_note": conversion_note_raw,
        "rationale": description,
        "confidence": 0.99,
        "size_hint": size_hint,
        "layout": layout,
        "source_fields_used": source_fields_used,
        "warnings": [],
        "role_config": role_config,
        "style_config": style_config,
        "source_key": str(raw.get("source_key") or "").strip() or None,
        "base_filters": base_filters,
    }


def _v1_dataset_ops_to_calc_fields_and_derived(
    dataset_ops: List[Any],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str]]:
    """Split v1 ``dataset_ops`` into legacy calc-fields and derived-tables lists."""
    calc_fields: List[Dict[str, Any]] = []
    derived_tables: List[Dict[str, Any]] = []
    warnings: List[str] = []
    for op in dataset_ops or []:
        if not isinstance(op, dict):
            continue
        kind = str(op.get("op") or "").strip().lower()
        if kind == "add_column":
            name = str(op.get("name") or "").strip()
            expression = str(op.get("expression") or "").strip()
            if not name or not expression:
                warnings.append("Skipped add_column op without name or expression.")
                continue
            calc_fields.append(
                {
                    "name": name,
                    "label": str(op.get("label") or name),
                    "expression": expression,
                    "source_key": str(op.get("source_key") or "").strip() or None,
                }
            )
        elif kind == "derived_table":
            src = str(op.get("source_key") or "").strip()
            sql_template = str(op.get("sql_template") or "").strip()
            inputs = [str(x).strip() for x in (op.get("inputs") or []) if str(x).strip()]
            if not src or not sql_template or not inputs:
                warnings.append(
                    f"Skipped derived_table '{src or '?'}' (missing source_key, sql_template, or inputs)."
                )
                continue
            derived_tables.append(
                {
                    "op": "derived_table",
                    "source_key": src,
                    "display_name": str(op.get("display_name") or src),
                    "inputs": inputs,
                    "sql_template": sql_template,
                    "output_columns": op.get("output_columns") or [],
                }
            )
        else:
            warnings.append(f"Unknown dataset op '{kind}' ignored.")
    return calc_fields, derived_tables, warnings


def _compare_source_contract(
    *,
    source_contract: Dict[str, Any],
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]],
    derived_source_keys: Iterable[str] = (),
) -> List[str]:
    """Compare ``source_contract.expected_source_keys`` to available profiles."""
    expected_raw = source_contract.get("expected_source_keys") or []
    expected = [str(k).strip() for k in expected_raw if str(k).strip()]
    if not expected:
        return []
    available: set[str] = set()
    if all_source_profiles:
        available.update(str(k) for k in all_source_profiles.keys())
    primary_key = (
        str(source_profile.get("source_key") or "").strip()
        or str(source_profile.get("display_name") or "").strip()
        or str(source_profile.get("selected_sheet_name") or "").strip()
    )
    if primary_key:
        available.add(primary_key)
    missing = [k for k in expected if k not in available]
    warnings: List[str] = []
    if missing:
        warnings.append(
            "Source contract expected the following source keys that were not "
            f"found in the uploaded data: {missing}. Re-upload the matching files or "
            "pick an existing dataset that contains these tables."
        )
    # Derived source keys must not collide with physical table names.
    collisions = [k for k in derived_source_keys if k in available]
    if collisions:
        warnings.append(
            f"Derived-table source keys collide with physical table names: {collisions}. "
            "The physical tables will take precedence; rename the derived keys to avoid confusion."
        )
    return warnings


def _resolve_source_placeholders(
    sql_template: str,
    *,
    source_key_to_alias: Dict[str, str],
) -> Tuple[Optional[str], List[str]]:
    missing: List[str] = []

    def _repl(match: re.Match[str]) -> str:
        key = match.group(1)
        alias = source_key_to_alias.get(key)
        if not alias:
            missing.append(key)
            return match.group(0)
        return alias

    resolved = _APPBI_SOURCE_PLACEHOLDER_RE.sub(_repl, sql_template)
    if missing:
        return None, sorted(set(missing))
    return resolved, []


def _build_source_key_alias_map(
    *,
    options: Iterable[Any],
    table_id_map: Dict[str, int],
) -> Dict[str, str]:
    """Build a source-key -> SQL alias lookup for derived-table resolution."""
    from app.services.dataset_table_sql_service import build_dataset_table_sql_alias

    alias_by_table_id: Dict[int, str] = {}
    alias_by_key: Dict[str, str] = {}

    for opt in options:
        table_id = getattr(opt, "table_id", None)
        alias = str(getattr(opt, "alias", "") or "").strip()
        display_name = str(getattr(opt, "display_name", "") or "").strip()
        if not alias:
            continue
        if isinstance(table_id, int):
            alias = build_dataset_table_sql_alias(table_id)
            alias_by_table_id[table_id] = alias
        if display_name:
            alias_by_key[display_name] = alias

    for source_key, table_id in (table_id_map or {}).items():
        alias = alias_by_table_id.get(table_id)
        normalized_source_key = str(source_key or "").strip()
        if normalized_source_key and alias:
            alias_by_key[normalized_source_key] = alias

    return alias_by_key


def _materialize_v1_derived_tables(
    db: Session,
    *,
    derived_tables: List[Dict[str, Any]],
    chart_plans: List[Dict[str, Any]],
    dataset_id: Optional[int],
    table_id_map: Dict[str, int],
) -> List[str]:
    """Rewrite chart plans bound to v1 derived tables to use ``queryMode=custom``.

    For each ``derived_table`` op, the SQL template is compiled against the
    dataset-table CTE namespace (``{{source:KEY}}`` -> ``dataset_table_<id>``) and
    every chart whose ``source_key`` matches the derived key is converted to a
    custom-SQL binding that produces the columns referenced in its role_config.
    """
    warnings: List[str] = []
    if not derived_tables or not chart_plans or dataset_id is None:
        if derived_tables and chart_plans and dataset_id is None:
            warnings.append(
                "Cannot materialize derived tables because no dataset is bound to this import."
            )
        return warnings

    from app.services.dataset_table_sql_service import (
        _indent_sql,
        build_dataset_table_live_query,
        get_dataset_table_reference_options,
        validate_and_clean_derived_query,
        DatasetTableSqlError,
    )

    dataset_obj = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if dataset_obj is None:
        warnings.append(
            f"Cannot materialize derived tables because dataset {dataset_id} was not found."
        )
        return warnings

    dataset_tables = (
        db.query(DatasetTable)
        .filter(
            DatasetTable.dataset_id == dataset_id,
            DatasetTable.source_kind != "generated_calendar",
        )
        .all()
    )
    table_by_id = {
        int(table.id): table
        for table in dataset_tables
        if getattr(table, "id", None) is not None
    }
    table_id_by_display_name = _build_table_id_map_from_dataset_tables(dataset_tables)

    options = get_dataset_table_reference_options(db, dataset_id)
    source_key_to_alias = _build_source_key_alias_map(
        options=options,
        table_id_map=table_id_map,
    )

    resolved_by_key: Dict[str, Dict[str, Any]] = {}
    for dt in derived_tables:
        source_key = str(dt.get("source_key") or "").strip()
        sql_template = str(dt.get("sql_template") or "").strip()
        inputs = [str(x).strip() for x in (dt.get("inputs") or []) if str(x).strip()]
        if not source_key or not sql_template or not inputs:
            continue
        missing_inputs = [k for k in inputs if k not in source_key_to_alias]
        if missing_inputs:
            warnings.append(
                f"Derived table '{source_key}' references input tables that are not "
                f"available in the dataset: {missing_inputs}. Charts bound to it "
                "will be skipped."
            )
            continue

        resolved_sql, missing_keys = _resolve_source_placeholders(
            sql_template,
            source_key_to_alias=source_key_to_alias,
        )
        if resolved_sql is None or missing_keys:
            warnings.append(
                f"Derived table '{source_key}' references unknown source keys in its SQL: "
                f"{missing_keys}. Charts bound to it will be skipped."
            )
            continue

        try:
            cleaned_sql = validate_and_clean_derived_query(resolved_sql)
        except DatasetTableSqlError as exc:
            warnings.append(
                f"Derived table '{source_key}' SQL failed validation: {exc}. "
                "Charts bound to it will be skipped."
            )
            continue

        ctes: List[str] = []
        resolved_datasource_id: int | None = None
        executable_sql = cleaned_sql
        cte_failed = False
        for input_key in inputs:
            alias = source_key_to_alias.get(input_key)
            if not alias:
                warnings.append(
                    f"Derived table '{source_key}' could not resolve a live alias for input '{input_key}'. "
                    "Charts bound to it will be skipped."
                )
                cte_failed = True
                break

            resolved_table_id = None
            if table_id_map:
                resolved_table_id = table_id_map.get(input_key)
            if resolved_table_id is None:
                resolved_table_id = table_id_by_display_name.get(input_key)
            if not isinstance(resolved_table_id, int):
                warnings.append(
                    f"Derived table '{source_key}' could not resolve dataset table for input '{input_key}'. "
                    "Charts bound to it will be skipped."
                )
                cte_failed = True
                break

            dependency_table = table_by_id.get(int(resolved_table_id))
            if dependency_table is None:
                warnings.append(
                    f"Derived table '{source_key}' references missing dataset table id {resolved_table_id} for input '{input_key}'. "
                    "Charts bound to it will be skipped."
                )
                cte_failed = True
                break

            try:
                live_datasource, live_sql = build_dataset_table_live_query(
                    db,
                    dataset_obj,
                    dependency_table,
                    required_datasource_id=resolved_datasource_id,
                )
            except DatasetTableSqlError as exc:
                warnings.append(
                    f"Derived table '{source_key}' could not build live SQL for input '{input_key}': {exc}. "
                    "Charts bound to it will be skipped."
                )
                cte_failed = True
                break

            resolved_datasource_id = int(live_datasource.id)
            ctes.append(f"{alias} AS (\n{_indent_sql(live_sql)}\n)")

        if cte_failed:
            continue

        if ctes:
            executable_sql = "WITH " + ",\n".join(ctes) + "\n" + cleaned_sql

        resolved_by_key[source_key] = {
            "sql": executable_sql,
            "primary_input": inputs[0],
        }

    if not resolved_by_key:
        return warnings

    for plan in chart_plans:
        plan_source_key = plan.get("source_key")
        if not plan_source_key or plan_source_key not in resolved_by_key:
            continue
        bundle = resolved_by_key[plan_source_key]
        # Rebind to the primary input so a real dataset_table_id can be resolved.
        plan["source_key"] = bundle["primary_input"]
        plan["query_mode"] = "custom"
        plan["custom_sql"] = bundle["sql"]
        # role_config already references the derived-table output columns,
        # which is exactly what custom SQL will emit.
        plan["custom_role_config"] = dict(plan.get("role_config") or {})

    return warnings


def _finalize_plan_from_metadata(
    *,
    raw: Dict[str, Any],
    source_profile: Dict[str, Any],
    order: int,
) -> Dict[str, Any]:
    """Build a finalized chart plan from embedded metadata (no AI needed).

    This mirrors ``_finalize_plan`` output but trusts the metadata values since
    they were authored by the skill-guided generation process.
    """
    suggested_type = str(raw.get("suggested_chart_type") or raw.get("chart_type") or "TABLE").upper()
    final_chart_type = suggested_type if suggested_type in _SUPPORTED_CHART_TYPES else "TABLE"
    changed = final_chart_type != suggested_type

    fm = raw.get("field_mapping") or {}
    dimension = str(fm.get("dimension") or "").strip() or None
    time_field = str(fm.get("timeField") or "").strip() or None
    breakdown = str(fm.get("breakdown") or "").strip() or None
    scatter_x = str(fm.get("scatterX") or "").strip() or None
    scatter_y = str(fm.get("scatterY") or "").strip() or None
    selected_columns = [str(c).strip() for c in (fm.get("selectedColumns") or []) if str(c).strip()]

    metrics = []
    for m in (fm.get("metrics") or []):
        if isinstance(m, dict) and m.get("field"):
            metrics.append({"field": str(m["field"]).strip(), "agg": str(m.get("agg") or "sum").lower()})

    line_metric = None
    if fm.get("lineMetric") and isinstance(fm["lineMetric"], dict) and fm["lineMetric"].get("field"):
        line_metric = {"field": str(fm["lineMetric"]["field"]).strip(), "agg": str(fm["lineMetric"].get("agg") or "sum").lower()}

    title = _normalize_text(raw.get("title"), max_len=160) or f"Imported {final_chart_type.title()}"

    # Build role_config matching _finalize_plan output
    role_config: Dict[str, Any] = {"metrics": metrics}
    if dimension:
        role_config["dimension"] = dimension
    if time_field:
        role_config["timeField"] = time_field
    if breakdown:
        role_config["breakdown"] = breakdown
    if line_metric:
        role_config["lineMetric"] = line_metric
    if selected_columns:
        role_config["selectedColumns"] = selected_columns
    if scatter_x:
        role_config["scatterX"] = scatter_x
    if scatter_y:
        role_config["scatterY"] = scatter_y

    source_fields_used = sorted({
        *(([dimension] if dimension else [])),
        *(([time_field] if time_field else [])),
        *(([breakdown] if breakdown else [])),
        *(([scatter_x] if scatter_x else [])),
        *(([scatter_y] if scatter_y else [])),
        *([m.get("field") for m in metrics if m.get("field")]),
        *(([line_metric["field"]] if line_metric and line_metric.get("field") else [])),
        *(selected_columns or []),
    })

    confidence = 0.95
    try:
        confidence = max(0.05, min(float(raw.get("confidence", 0.95)), 0.99))
    except (TypeError, ValueError):
        pass

    size_hint = str(raw.get("size_hint") or "half").lower()
    if size_hint not in {"full", "half", "third", "kpi"}:
        size_hint = "kpi" if final_chart_type == "KPI" else ("full" if final_chart_type == "TABLE" else "half")

    return {
        "block_id": str(raw.get("block_id") or f"meta-{order}"),
        "order": order,
        "title": title,
        "block_role": final_chart_type.lower() if final_chart_type in ("TABLE", "KPI") else "chart",
        "source_excerpt": _normalize_text(raw.get("reasoning"), max_len=280),
        "original_chart_type": str(raw.get("original_chart_type") or final_chart_type).upper(),
        "requested_chart_type": suggested_type,
        "final_chart_type": final_chart_type,
        "changed_chart_type": changed,
        "conversion_note": None if not changed else f"Requested {suggested_type}, imported as {final_chart_type}.",
        "rationale": _normalize_text(raw.get("reasoning"), max_len=260),
        "confidence": confidence,
        "size_hint": size_hint,
        "source_fields_used": source_fields_used,
        "warnings": [],
        "role_config": role_config,
        "style_config": {"chartTitle": title},
        "source_key": raw.get("source_key"),
    }


def _analyze_from_embedded_metadata(
    *,
    embedded: Dict[str, Any],
    document_summary: Dict[str, Any],
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Build the full analyze response from embedded ``application/appbi-dashboard`` metadata.

    Two shapes are supported:

    * **AppBI Import Plan v1** (``version == "appbi-import/v1"``) — the
      skill-generated native contract. Runtime-shaped ``role_config``,
      ``dataset_ops`` (``add_column`` + ``derived_table``), ``base_filters``
      and an explicit grid ``layout`` are trusted as-is.
    * **Legacy** — the older ``field_mapping`` + ``calculated_fields`` shape
      kept for backward compatibility with pre-skill dashboards.
    """
    if _is_v1_metadata(embedded):
        return _analyze_from_v1_metadata(
            embedded=embedded,
            document_summary=document_summary,
            source_profile=source_profile,
            all_source_profiles=all_source_profiles,
        )

    if _is_v2_metadata(embedded):
        default_page = _pick_default_v2_page(embedded)
        if default_page is None:
            return {
                "suggested_dashboard_name": _normalize_text(document_summary.get("title"), max_len=180) or "Imported Dashboard",
                "document_title": document_summary.get("title"),
                "source_profile": source_profile,
                "all_source_profiles": all_source_profiles,
                "chart_plans": [],
                "calculated_fields": [],
                "derived_tables": [],
                "ignored_blocks": [],
                "warnings": ["Embedded v2 metadata contained no importable pages."],
                "ai_meta": build_ai_assist_meta(
                    requested=False,
                    applied=False,
                    status="skipped",
                    message="Embedded application/appbi-dashboard v2 metadata contained no pages.",
                ),
            }
        return _analyze_from_v2_metadata_page(
            embedded=embedded,
            page=default_page,
            document_summary=document_summary,
            source_profile=source_profile,
            all_source_profiles=all_source_profiles,
        )

    # ── Legacy path (pre-v1 embedded metadata) ──────────────────────────
    raw_charts = embedded.get("charts") or []
    raw_calc_fields = embedded.get("calculated_fields") or []

    # Collect all known column names across profiles for calc-field validation
    all_col_names: List[str] = [col.get("name") for col in (source_profile.get("columns") or [])]
    if all_source_profiles:
        for prof in all_source_profiles.values():
            for col in (prof.get("columns") or []):
                if col.get("name") not in all_col_names:
                    all_col_names.append(col.get("name"))

    validated_calc_fields = _validate_calculated_fields(raw_calc_fields, all_col_names)

    # Enrich profiles with calculated field names
    def _enrich(profile: Dict[str, Any]) -> Dict[str, Any]:
        if not validated_calc_fields:
            return profile
        enriched = dict(profile)
        existing_cols = list(enriched.get("columns") or [])
        existing_names = {col.get("name") for col in existing_cols}
        for cf in validated_calc_fields:
            if cf["name"] not in existing_names:
                existing_cols.append({"name": cf["name"], "type": "number"})
        enriched["columns"] = existing_cols
        return enriched

    enriched_profile = _enrich(source_profile)

    finalized_plans: List[Dict[str, Any]] = []
    for idx, raw in enumerate(raw_charts, start=1):
        if not isinstance(raw, dict):
            continue
        plan = _finalize_plan_from_metadata(
            raw=raw,
            source_profile=enriched_profile,
            order=idx,
        )
        finalized_plans.append(plan)

    finalized_plans.sort(key=lambda p: (p.get("order", 0), p.get("block_id", "")))
    _legacy_contract = build_dashboard_contract(
        document_summary=document_summary,
        chart_plans=finalized_plans,
        embedded=embedded,
    )
    _legacy_ignored = _candidate_blocks(document_summary)[1]
    _legacy_block_warnings: List[str] = []
    _legacy_widgets = widgets_from_blocks(
        _legacy_ignored, document_summary, warnings_out=_legacy_block_warnings
    )
    _legacy_layout_warnings = compose_import_layout(
        chart_plans=finalized_plans,
        widgets=_legacy_widgets,
        family=_legacy_contract["template_family"],
    )

    title = _normalize_text(embedded.get("dashboard_title") or document_summary.get("title"), max_len=180) or "Imported Dashboard"

    _, ignored = _candidate_blocks(document_summary)

    warnings: List[str] = list(_legacy_layout_warnings) + _legacy_block_warnings
    if not finalized_plans:
        warnings.append("Embedded metadata contained no valid chart plans.")
    if any(p.get("changed_chart_type") for p in finalized_plans):
        warnings.append("One or more chart types from the metadata were adapted to supported AppBI types.")

    return {
        "suggested_dashboard_name": title,
        "document_title": document_summary.get("title"),
        "source_profile": source_profile,
        "all_source_profiles": all_source_profiles,
        "chart_plans": finalized_plans,
        "calculated_fields": validated_calc_fields,
        "derived_tables": [],
        "ignored_blocks": ignored,
        "warnings": warnings,
        **_legacy_contract,
        "widgets": _legacy_widgets,
        "ai_meta": build_ai_assist_meta(
            requested=False,
            applied=False,
            status="skipped",
            message="Chart plans loaded from embedded application/appbi-dashboard metadata. AI was not needed.",
        ),
    }


def _analyze_from_v1_metadata(
    *,
    embedded: Dict[str, Any],
    document_summary: Dict[str, Any],
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Build an analyze response from an AppBI Import Plan v1 payload."""
    dashboard_meta = embedded.get("dashboard") if isinstance(embedded.get("dashboard"), dict) else {}
    source_contract = embedded.get("source_contract") if isinstance(embedded.get("source_contract"), dict) else {}
    dataset_ops = embedded.get("dataset_ops") if isinstance(embedded.get("dataset_ops"), list) else []
    raw_charts = embedded.get("charts") if isinstance(embedded.get("charts"), list) else []

    # Per-source alias maps: source_key -> {fingerprint: real_column_name}.
    alias_maps_by_source: Dict[str, Dict[str, str]] = {}
    primary_alias_map = _build_column_alias_map(source_profile.get("columns") or [])
    primary_source_key = (
        str(source_profile.get("source_key") or "").strip()
        or str(source_profile.get("display_name") or "").strip()
        or str(source_profile.get("selected_sheet_name") or "").strip()
    )
    if primary_source_key:
        alias_maps_by_source[primary_source_key] = primary_alias_map
    if all_source_profiles:
        for key, prof in all_source_profiles.items():
            alias_maps_by_source[str(key)] = _build_column_alias_map(prof.get("columns") or [])

    union_alias_map: Dict[str, str] = {}
    fingerprint_sources: Dict[str, List[str]] = {}
    for source_key, mapping in alias_maps_by_source.items():
        for fp, real in mapping.items():
            fingerprint_sources.setdefault(fp, []).append(f"{source_key}.{real}")
            union_alias_map.setdefault(fp, real)
    collisions = {
        fp: refs for fp, refs in fingerprint_sources.items() if len(set(refs)) > 1
    }
    if collisions:
        logger.debug(
            "union_alias_map has %d fingerprint collisions across sources; charts "
            "with unresolved source_key may bind to the wrong column. Sample "
            "fingerprints: %s",
            len(collisions),
            list(collisions.keys())[:10],
        )
    if not union_alias_map:
        union_alias_map = primary_alias_map

    def _pick_alias_map(keys: Iterable[str]) -> Dict[str, str]:
        merged: Dict[str, str] = {}
        for key in keys:
            mapping = alias_maps_by_source.get(str(key))
            if mapping:
                for fp, real in mapping.items():
                    merged.setdefault(fp, real)
        return merged or union_alias_map

    calc_fields, derived_tables, op_warnings = _v1_dataset_ops_to_calc_fields_and_derived(dataset_ops)

    # Rewrite calc_field expressions so Claude's snake_case identifiers bind
    # to real column names. Calc-field expressions are consumed by
    # TransformationCompiler which uses ``[col name]`` syntax for quoted idents.
    for cf in calc_fields:
        expr = str(cf.get("expression") or "")
        if not expr:
            continue
        source_hint = cf.get("source_key")
        mapping = _pick_alias_map([source_hint] if source_hint else []) or union_alias_map
        cf["expression"] = _rewrite_identifiers_in_expression(expr, mapping, style="expression")

    # Rewrite derived_table SQL templates with the inputs' alias maps so
    # physical column references resolve against quoted real names.
    for dt in derived_tables:
        sql_template = str(dt.get("sql_template") or "")
        if not sql_template:
            continue
        mapping = _pick_alias_map(dt.get("inputs") or [])
        dt["sql_template"] = _rewrite_identifiers_in_expression(sql_template, mapping, style="sql")

    # Validate add_column expressions against the union of columns + calc fields.
    all_col_names: List[str] = [col.get("name") for col in (source_profile.get("columns") or [])]
    if all_source_profiles:
        for prof in all_source_profiles.values():
            for col in (prof.get("columns") or []):
                if col.get("name") not in all_col_names:
                    all_col_names.append(col.get("name"))
    validated_calc_fields = _validate_calculated_fields(calc_fields, all_col_names)

    # Register calculated column fingerprints so charts bound to a derived_table
    # whose role_config refers to those calc columns can still resolve.
    for cf in validated_calc_fields:
        fp = _column_fingerprint(cf.get("name") or "")
        if fp and fp not in union_alias_map:
            union_alias_map[fp] = cf["name"]

    # Register derived_table output columns so plans binding to derived sources
    # can use snake_case identifiers that match the SELECT aliases.
    for dt in derived_tables:
        mapping: Dict[str, str] = {}
        for oc in dt.get("output_columns") or []:
            if not isinstance(oc, dict):
                continue
            real = str(oc.get("name") or "").strip()
            if not real:
                continue
            fp = _column_fingerprint(real)
            if fp and fp not in mapping:
                mapping[fp] = real
        if mapping:
            alias_maps_by_source[str(dt.get("source_key"))] = mapping

    # Enrich profiles so downstream field checks see calculated columns.
    def _enrich(profile: Dict[str, Any]) -> Dict[str, Any]:
        if not validated_calc_fields:
            return profile
        enriched = dict(profile)
        existing_cols = list(enriched.get("columns") or [])
        existing_names = {col.get("name") for col in existing_cols}
        for cf in validated_calc_fields:
            if cf["name"] not in existing_names:
                existing_cols.append({"name": cf["name"], "type": "number"})
        enriched["columns"] = existing_cols
        return enriched

    _ = _enrich(source_profile)  # noqa: F841 (kept for symmetry with legacy path)

    known_columns_by_source: Dict[str, Set[str]] = {}

    def _register_known_columns(source_key: Any, columns: Iterable[Any]) -> None:
        key = str(source_key or "").strip()
        if not key:
            return
        bucket = known_columns_by_source.setdefault(key, set())
        for col in columns or []:
            if isinstance(col, dict):
                name = str(col.get("name") or "").strip()
            else:
                name = str(col or "").strip()
            if name:
                bucket.add(name)

    _register_known_columns(primary_source_key, source_profile.get("columns") or [])
    if all_source_profiles:
        for key, prof in all_source_profiles.items():
            _register_known_columns(key, prof.get("columns") or [])
    for cf in validated_calc_fields:
        calc_source_key = str(cf.get("source_key") or primary_source_key or "").strip()
        if calc_source_key:
            known_columns_by_source.setdefault(calc_source_key, set()).add(cf["name"])
    for dt in derived_tables:
        _register_known_columns(dt.get("source_key"), dt.get("output_columns") or [])

    finalized_plans: List[Dict[str, Any]] = []
    for idx, raw in enumerate(raw_charts, start=1):
        if not isinstance(raw, dict):
            continue
        finalized_plans.append(_finalize_plan_from_v1_metadata(raw=raw, order=idx))

    # Rewrite role_config / base_filters field references to the real column
    # names exposed by each chart's source (physical table OR derived_table).
    for plan in finalized_plans:
        plan_source_key = plan.get("source_key")
        mapping = (
            alias_maps_by_source.get(str(plan_source_key))
            if plan_source_key
            else None
        ) or union_alias_map
        _rewrite_plan_field_references(plan, mapping)

    field_resolution_warnings: List[str] = []
    for plan in finalized_plans:
        plan_source_key = str(plan.get("source_key") or primary_source_key or "").strip()
        known_columns = set(known_columns_by_source.get(plan_source_key) or set())
        if not known_columns and primary_source_key:
            known_columns = set(known_columns_by_source.get(primary_source_key) or set())
        if not known_columns:
            continue
        unresolved = [
            f"{label}: {value}"
            for label, value in _iter_plan_field_references(plan)
            if value not in known_columns
        ]
        if unresolved:
            warning = (
                f"Chart '{plan.get('block_id')}' references fields not found in source "
                f"'{plan_source_key or 'unknown'}': {unresolved}. Chart may fail at runtime."
            )
            field_resolution_warnings.append(warning)
            logger.warning(
                "v1 import block_id=%s has unresolved fields %s (source_key=%s). "
                "Chart may fail at runtime.",
                plan.get("block_id"),
                unresolved,
                plan_source_key or None,
            )

    finalized_plans.sort(key=lambda p: (p.get("order", 0), p.get("block_id", "")))

    # v1 metadata may declare explicit grid coordinates. Those are the author's
    # report and are never overwritten; anything they left unplaced is composed
    # by the template's recipe around them.
    _v1_contract = build_dashboard_contract(
        document_summary=document_summary,
        chart_plans=finalized_plans,
        embedded=embedded,
    )
    _v1_ignored_preview, = (_candidate_blocks(document_summary)[1],)
    _v1_block_warnings: List[str] = []
    # A declared plan owns its non-chart tiles. Falling through to inference
    # when the plan HAS widgets produced each heading twice -- once as declared
    # and once re-derived from the markup -- and the re-derived copies landed on
    # top of the charts.
    _declared_widgets = widgets_from_plan(embedded.get("widgets"))
    _v1_widgets = _declared_widgets or widgets_from_blocks(
        _v1_ignored_preview, document_summary, warnings_out=_v1_block_warnings
    )
    _v1_layout_warnings = compose_import_layout(
        chart_plans=finalized_plans,
        widgets=_v1_widgets,
        family=_v1_contract["template_family"],
    )

    dash_title = (
        _normalize_text(dashboard_meta.get("title"), max_len=180)
        or _normalize_text(embedded.get("dashboard_title"), max_len=180)
        or _normalize_text(document_summary.get("title"), max_len=180)
        or "Imported Dashboard"
    )
    dash_description = _normalize_text(dashboard_meta.get("description"), max_len=400) or None

    _, ignored = _candidate_blocks(document_summary)

    # Phase 3: the source contract.
    #
    # A file authored against the skill declares the dataset it was written for
    # and the columns it uses. If those columns are gone, the honest answer is
    # to stop: re-mapping a reference to whatever looks similar is exactly the
    # guessing the declarative path exists to remove, and it silently produces a
    # dashboard that reads the wrong numbers.
    _contract_tables = [
        {"display_name": key, "columns": profile.get("columns") or []}
        for key, profile in (all_source_profiles or {}).items()
    ] or [{
        "display_name": str(source_profile.get("dataset_table_name") or ""),
        "columns": source_profile.get("columns") or [],
    }]
    _contract_ok, _contract_findings = compare_source_contract_to_dataset(
        source_contract=source_contract, tables=_contract_tables
    )
    if not _contract_ok:
        raise SourceContractMismatch(_contract_findings)

    contract_warnings = list(_contract_findings) + _compare_source_contract(
        source_contract=source_contract,
        source_profile=source_profile,
        all_source_profiles=all_source_profiles,
        derived_source_keys=[dt["source_key"] for dt in derived_tables],
    )

    warnings: List[str] = list(_v1_layout_warnings) + _v1_block_warnings
    warnings.extend(op_warnings)
    warnings.extend(contract_warnings)
    warnings.extend(field_resolution_warnings)
    if not finalized_plans:
        warnings.append("Embedded v1 metadata contained no valid chart plans.")
    if any(p.get("changed_chart_type") for p in finalized_plans):
        warnings.append(
            "One or more chart types from the metadata were adapted to supported AppBI types."
        )
    if len(validated_calc_fields) != len(calc_fields):
        warnings.append(
            "Some add_column operations were rejected by the expression validator and "
            "will be skipped."
        )

    ai_meta = build_ai_assist_meta(
        requested=False,
        applied=False,
        status="skipped",
        message=(
            "Chart plans loaded from embedded application/appbi-dashboard "
            "metadata (AppBI Import Plan v1). AI was not needed."
        ),
    )
    ai_meta["plan_version"] = APPBI_IMPORT_PLAN_V1_VERSION
    ai_meta["authoring_mode"] = str(embedded.get("authoring_mode") or "skill")
    ai_meta["dashboard_description"] = dash_description
    ai_meta["default_page_name"] = (
        _normalize_text(dashboard_meta.get("default_page_name"), max_len=80) or None
    )
    ai_meta["source_contract"] = source_contract or {}
    ai_meta["derived_table_count"] = len(derived_tables)

    return {
        "suggested_dashboard_name": dash_title,
        "document_title": document_summary.get("title"),
        "source_profile": source_profile,
        "all_source_profiles": all_source_profiles,
        "chart_plans": finalized_plans,
        "calculated_fields": validated_calc_fields,
        "derived_tables": derived_tables,
        "ignored_blocks": ignored,
        "warnings": warnings,
        "ai_meta": ai_meta,
        # A round-trip must not redesign the author's report: anything the
        # metadata declares is honoured verbatim here.
        **_v1_contract,
        "widgets": _v1_widgets,
        # Declared filters. The v1 path never built slicers at all, so a plan
        # that asked for `["order_status"]` imported with an empty filter bar
        # reading "No slicers added" -- and a report without its filters is a
        # screenshot. Names are still resolved against real columns: a declared
        # slicer that matches no column is dropped rather than guessed at.
        "slicers": slicers_from_source(
            ai_slicer_fields=[
                str(field).strip()
                for field in (dashboard_meta.get("slicers") or [])
                if str(field).strip()
            ],
            document_summary=document_summary,
            source_profile=source_profile,
            all_source_profiles=all_source_profiles,
            # The plan said what the filters are. Harvesting more from the
            # markup on top of that produced a slicer nobody asked for.
            declared=True,
        ),
    }


def _rewrite_plan_field_references(plan: Dict[str, Any], alias_map: Dict[str, str]) -> None:
    """In-place rewrite of all field identifiers in a chart plan's runtime config."""
    if not alias_map:
        return

    def _resolve(name: Any) -> Any:
        real = _resolve_column_name(name, alias_map)
        return real if real else name

    role_config = plan.get("role_config")
    if isinstance(role_config, dict):
        for key in ("dimension", "timeField", "breakdown", "scatterX", "scatterY"):
            val = role_config.get(key)
            if isinstance(val, str) and val:
                role_config[key] = _resolve(val)
        metrics = role_config.get("metrics")
        if isinstance(metrics, list):
            for metric in metrics:
                if isinstance(metric, dict) and metric.get("field"):
                    metric["field"] = _resolve(metric["field"])
        line_metric = role_config.get("lineMetric")
        if isinstance(line_metric, dict) and line_metric.get("field"):
            line_metric["field"] = _resolve(line_metric["field"])
        selected = role_config.get("selectedColumns")
        if isinstance(selected, list):
            role_config["selectedColumns"] = [_resolve(c) for c in selected if c]

    base_filters = plan.get("base_filters")
    if isinstance(base_filters, list):
        for flt in base_filters:
            if isinstance(flt, dict) and flt.get("field"):
                flt["field"] = _resolve(flt["field"])

    fields_used = plan.get("source_fields_used")
    if isinstance(fields_used, list):
        plan["source_fields_used"] = sorted({_resolve(f) for f in fields_used if f})


def _iter_plan_field_references(plan: Dict[str, Any]) -> List[Tuple[str, str]]:
    """Return labeled field references used by a chart plan."""
    refs: List[Tuple[str, str]] = []
    role_config = plan.get("role_config")
    if isinstance(role_config, dict):
        for key in ("dimension", "timeField", "breakdown", "scatterX", "scatterY"):
            value = role_config.get(key)
            if isinstance(value, str) and value:
                refs.append((f"role_config.{key}", value))
        metrics = role_config.get("metrics")
        if isinstance(metrics, list):
            for index, metric in enumerate(metrics):
                if isinstance(metric, dict):
                    field = str(metric.get("field") or "").strip()
                    if field:
                        refs.append((f"role_config.metrics[{index}]", field))
        line_metric = role_config.get("lineMetric")
        if isinstance(line_metric, dict):
            field = str(line_metric.get("field") or "").strip()
            if field:
                refs.append(("role_config.lineMetric", field))
        selected = role_config.get("selectedColumns")
        if isinstance(selected, list):
            for index, value in enumerate(selected):
                field = str(value or "").strip()
                if field:
                    refs.append((f"role_config.selectedColumns[{index}]", field))

    base_filters = plan.get("base_filters")
    if isinstance(base_filters, list):
        for index, flt in enumerate(base_filters):
            if isinstance(flt, dict):
                field = str(flt.get("field") or "").strip()
                if field:
                    refs.append((f"base_filters[{index}]", field))

    return refs


def _build_table_id_map_from_dataset_tables(dataset_tables: Iterable[Any]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for dataset_table in dataset_tables:
        display_name = str(getattr(dataset_table, "display_name", "") or "").strip()
        table_id = getattr(dataset_table, "id", None)
        if display_name and isinstance(table_id, int):
            mapping[display_name] = table_id
    return mapping




# Layout families the import can target. These mirror the FE layout templates
# one-for-one; an import that picks a family is picking the same composition a
# user would pick from the theme menu.
# ── Layout recipes ──────────────────────────────────────────────────────────
#
# A template is not a palette and not a bag of tokens: it is a COMPOSITION. But
# a composition is not a seating chart either.
#
# The first version of this sorted every block into a contiguous band by kind --
# all the KPIs, then all the charts, then all the notes. That reads well on a
# document that was already built that way and destroys every other one: the
# three section headers of a five-section report ended up stacked on top of each
# other at rows 0, 2 and 4, introducing nothing, and the KPI cards that opened
# the source finished up below the charts they were meant to summarise.
#
# So the split is: the TEMPLATE owns shape, the DOCUMENT owns order.
#
#   * `lead` bands are hoisted above the body. Only the things that are a
#     header by nature belong here -- a hero, a KPI strip. This is what makes a
#     `stage` report open with a wall of numbers whatever order the source used.
#   * `flow` gives each kind its shape -- how many sit across, how tall they are,
#     whether they span the full width -- and the body is laid out in the
#     source's own order using those shapes. A section header therefore still
#     sits directly above the tiles it introduces.
#
# The grid is 12 columns wide and one unit of `h` is about 96 px tall.
#
# That number is not arbitrary and it is not 80: the runtime grid is 36 columns
# (`GRID_FINER = 3` in `frontend/src/lib/dashboard-pages.ts`), and a layout
# written in 12-column coordinates is upscaled x3 when it is read. So `h` here
# buys three runtime rows, and `h: 10` renders as a 944 px tile -- which is how
# a four-chart report came out five thousand pixels tall, mostly white.
#
# Measured on a live grid: rendered height = 96 * h - 16.
#   h=1 ->  80px   a slim band
#   h=2 -> 176px   a KPI card or a note
#   h=4 -> 368px   a chart you can actually read
#   h=7 -> 656px   an oversized, across-the-room visual
#
# Charts stay individually draggable afterwards: the recipe decides where things
# START, not where they are locked.

#: Grid width every recipe is expressed in.
RECIPE_GRID_COLS = 12

#: Every widget kind an import may persist. The runtime renderer must have a
#: case for each one -- a kind that is not here is silently rewritten to
#: "chart", which then renders as a load failure.
#: Colorways the theme catalog offers. Kept beside the templates because the
#: two travel together: a declared plan names one of each.
IMPORT_COLORWAYS = {
    "indigo", "emerald", "coral", "ocean", "amber", "slate", "midnight", "graphite",
}

IMPORT_WIDGET_TYPES = {
    "chart",
    "text",
    "image",
    "countdown",
    "shape",
    "parameter_switcher",
    "section_header",
    "callout",
    "hero_strip",
    "html_fragment",
}

#: `lead` band: kinds hoisted to the top, `cols` across, `h` tall, at most `max`.
#: `flow`: per-kind shape for the body. `full` spans all 12 columns.
LAYOUT_RECIPES: Dict[str, Dict[str, Any]] = {
    # The default SaaS console: headline numbers under a filter bar, then the
    # visuals that explain them, two across.
    "console": {
        "filter_dock": "top",
        "lead": [
            {"name": "hero", "kinds": ["hero_strip"], "cols": 1, "h": 3, "max": 1, "full": True},
            {"name": "kpi_strip", "kinds": ["kpi"], "cols": 4, "h": 2, "max": 8},
        ],
        "flow": {
            "section_header": {"cols": 1, "h": 1, "full": True},
            "hero_strip": {"cols": 1, "h": 3, "full": True},
            "kpi": {"cols": 4, "h": 2},
            "chart": {"cols": 2, "h": 4},
            "html_fragment": {"cols": 2, "h": 4},
            "callout": {"cols": 2, "h": 2},
            "text": {"cols": 2, "h": 2},
            "table": {"cols": 1, "h": 5, "full": True},
        },
    },
    # A board pack read in a meeting: filters out of the way in a left rail, a
    # KPI row, then one chart at a time carrying the argument, then the detail
    # people ask for.
    "brief": {
        "filter_dock": "left",
        "lead": [
            {"name": "kpi_row", "kinds": ["kpi"], "cols": 4, "h": 2, "max": 8},
        ],
        "flow": {
            "section_header": {"cols": 1, "h": 1, "full": True},
            "hero_strip": {"cols": 1, "h": 3, "full": True},
            "kpi": {"cols": 4, "h": 2},
            "chart": {"cols": 1, "h": 6, "full": True},
            "html_fragment": {"cols": 1, "h": 5, "full": True},
            "callout": {"cols": 1, "h": 2, "full": True},
            "text": {"cols": 1, "h": 2, "full": True},
            "table": {"cols": 1, "h": 6, "full": True},
        },
    },
    # A wall display: maximum information per pixel, filters compact on a rail,
    # everything packed three across.
    "ops": {
        "filter_dock": "left",
        "lead": [
            {"name": "status_row", "kinds": ["kpi"], "cols": 6, "h": 2, "max": 12},
        ],
        "flow": {
            "section_header": {"cols": 1, "h": 1, "full": True},
            "hero_strip": {"cols": 1, "h": 2, "full": True},
            "kpi": {"cols": 6, "h": 2},
            "chart": {"cols": 3, "h": 4},
            "html_fragment": {"cols": 3, "h": 4},
            "table": {"cols": 3, "h": 4},
            "callout": {"cols": 3, "h": 2},
            "text": {"cols": 3, "h": 2},
        },
    },
    # An article: narrative sections and the wide visuals that illustrate them,
    # read top to bottom in the order it was written.
    "editorial": {
        "filter_dock": "drawer",
        "lead": [
            {"name": "hero", "kinds": ["hero_strip"], "cols": 1, "h": 4, "max": 1, "full": True},
            # A data story opens with its headline figures under the standfirst;
            # left to flow inline they end up one per row, each a third of the
            # width, with prose between them.
            {"name": "figures", "kinds": ["kpi"], "cols": 3, "h": 2, "max": 6},
        ],
        "flow": {
            "section_header": {"cols": 1, "h": 2, "full": True},
            "hero_strip": {"cols": 1, "h": 4, "full": True},
            "kpi": {"cols": 3, "h": 2},
            "chart": {"cols": 1, "h": 5, "full": True},
            "html_fragment": {"cols": 1, "h": 4, "full": True},
            "callout": {"cols": 2, "h": 2},
            "text": {"cols": 1, "h": 2, "full": True},
            "table": {"cols": 1, "h": 5, "full": True},
        },
    },
    # Read from across a room: a wall of big numbers, then a couple of
    # oversized visuals. Nothing small, nothing dense.
    "stage": {
        "filter_dock": "top",
        "lead": [
            {"name": "hero", "kinds": ["hero_strip"], "cols": 1, "h": 4, "max": 1, "full": True},
            {"name": "metric_wall", "kinds": ["kpi"], "cols": 3, "h": 4, "max": 9},
        ],
        "flow": {
            "section_header": {"cols": 1, "h": 2, "full": True},
            "hero_strip": {"cols": 1, "h": 4, "full": True},
            "kpi": {"cols": 3, "h": 4},
            "chart": {"cols": 1, "h": 7, "full": True},
            "html_fragment": {"cols": 1, "h": 5, "full": True},
            "callout": {"cols": 1, "h": 2, "full": True},
            "text": {"cols": 1, "h": 2, "full": True},
            "table": {"cols": 2, "h": 5},
        },
    },
}

#: Shape used for a kind no recipe mentions, so a new widget kind lands
#: somewhere reasonable instead of nowhere.
_FALLBACK_SHAPE = {"cols": 2, "h": 8}


def _block_kind(item: Dict[str, Any]) -> str:
    """Which recipe kind a planned block belongs to.

    Charts declare a chart type; widgets declare a widget type. Both fold into
    the one vocabulary the recipes speak, so a single packer can place them.
    """
    wt = str(item.get("widget_type") or "").strip().lower()
    if wt and wt != "chart":
        # A KPI card that could not be bound to a column is still a KPI card.
        # Shaped as a plain callout it came out six columns wide and five rows
        # tall -- four of them filling the top of the report as fat boxes, where
        # the source had a tidy strip of four.
        if wt in {"callout", "text", "html_fragment"} and str(item.get("_source_role") or "") == "kpi":
            return "kpi"
        known = {"section_header", "callout", "text", "hero_strip", "html_fragment"}
        return wt if wt in known else "text"
    ct = str(item.get("final_chart_type") or item.get("chart_type") or "").upper()
    if ct == "KPI":
        return "kpi"
    if ct in {"TABLE", "MATRIX"}:
        return "table"
    return "chart"


def apply_layout_recipe(items: List[Dict[str, Any]], family: str) -> List[str]:
    """Place every block: template shapes, document order.

    Charts and widgets go through ONE pass. They used to be packed separately --
    `_assign_layouts` saw only the charts -- so every widget reached the builder
    with no layout at all and the build defaulted it to x=0, y=0, stacking every
    section header on top of each other and on top of the first chart.

    Returns human-readable warnings; layouts are written onto the items in
    place. Nothing is dropped.
    """
    recipe = LAYOUT_RECIPES.get(family) or LAYOUT_RECIPES["console"]
    flow: Dict[str, Any] = recipe["flow"]
    warnings: List[str] = []

    # The source's own reading order. A report's structure IS its order, which
    # is what keeps a section header attached to the charts it introduces.
    ordered = sorted(items, key=lambda it: (int(it.get("order") or 0), str(it.get("block_id") or "")))

    # Pass 1: hoist the header bands. A KPI strip is a header whatever position
    # the source gave it, and interleaved KPIs read as scattered noise.
    lead_bands: List[Tuple[Dict[str, Any], List[Dict[str, Any]]]] = []
    hoisted: Set[int] = set()
    for band in recipe.get("lead") or []:
        members: List[Dict[str, Any]] = []
        for item in ordered:
            if id(item) in hoisted or len(members) >= int(band["max"]):
                continue
            if _block_kind(item) in band["kinds"]:
                members.append(item)
                hoisted.add(id(item))
        if members:
            lead_bands.append((band, members))

    y = 0
    for band, members in lead_bands:
        cols = 1 if band.get("full") else max(1, int(band["cols"]))
        span = RECIPE_GRID_COLS // cols
        height = int(band["h"])
        for index, item in enumerate(members):
            column = index % cols
            if column == 0 and index > 0:
                y += height
            item["layout"] = {"x": column * span, "y": y, "w": span, "h": height}
            item["_region"] = band["name"]
        y += height

    # Pass 2: the body, in the source's order, wearing the template's shapes.
    # `column` and `row_height` track the run being filled; a run ends when the
    # shape changes, when the row is full, or at a full-width block.
    column = 0
    row_height = 0
    current_shape: Optional[str] = None

    for item in ordered:
        if id(item) in hoisted:
            continue
        kind = _block_kind(item)
        shape = flow.get(kind) or _FALLBACK_SHAPE
        cols = 1 if shape.get("full") else max(1, int(shape["cols"]))
        span = RECIPE_GRID_COLS // cols
        height = int(shape["h"])
        signature = f"{kind}:{cols}:{height}"

        # A section header always opens a new row: it introduces what follows,
        # so it must not share a line with the tail of what came before.
        starts_row = (
            kind == "section_header"
            or current_shape != signature
            or column + span > RECIPE_GRID_COLS
        )
        if starts_row and row_height:
            y += row_height
            column = 0
            row_height = 0

        item["layout"] = {"x": column, "y": y, "w": span, "h": height}
        item["_region"] = kind
        column += span
        row_height = max(row_height, height)
        current_shape = signature
        if column >= RECIPE_GRID_COLS:
            y += row_height
            column = 0
            row_height = 0
            current_shape = None

    if not flow:
        warnings.append(f"The '{family}' layout defines no shapes; blocks were laid out at a default size.")
    return warnings


#: The layout templates, named EXACTLY as the frontend catalog names them
#: (`frontend/src/lib/dashboard-theme-catalog.ts`). The id written into
#: `theme_config.templateId` is looked up there by `expandThemeIdentity`, so a
#: name that only exists on one side silently resolves to nothing and the
#: imported report lands on stock defaults — which is what "presentation"
#: (backend) vs "stage" (frontend) was doing to every AI-classified deck.
IMPORT_TEMPLATE_FAMILIES = ("console", "brief", "ops", "editorial", "stage")

#: Words the model (or a hand-written HTML) may reasonably use for a template.
#: Accepting synonyms costs nothing; silently dropping a valid answer because it
#: said "presentation" instead of "stage" costs the whole layout.
_TEMPLATE_ALIASES = {
    "presentation": "stage",
    "deck": "stage",
    "executive": "brief",
    "executive_brief": "brief",
    "executive-brief": "brief",
    "operations": "ops",
    "command_center": "ops",
    "article": "editorial",
    "report": "editorial",
    "saas": "console",
    "dashboard": "console",
}

#: Where each template puts the filters. Kept here rather than inferred so an
#: imported report and a hand-themed one land on identical layouts.
_FAMILY_DOCK = {
    "console": "top",
    "brief": "left",
    "ops": "left",
    "editorial": "drawer",
    "stage": "top",
}


#: What the model may say about a block. Five outcomes, and every block gets
#: exactly one -- the point is that "we could not use it" is a REPORTED outcome
#: rather than an omission.
_BLOCK_CLASSIFICATIONS = {
    "native_chart",
    "native_widget",
    "slicer",
    "html_fragment_preserved",
    "unsupported_interaction",
}

#: Widget kinds the model is allowed to name for a native_widget block.
_AI_WIDGET_TYPES = {"section_header", "callout", "text", "hero_strip"}


def _normalize_template_family(value: Any) -> Optional[str]:
    name = str(value or "").strip().lower().replace(" ", "_")
    name = _TEMPLATE_ALIASES.get(name, name)
    return name if name in IMPORT_TEMPLATE_FAMILIES else None


def choose_template_family(document_summary: Dict[str, Any], chart_plans: List[Dict[str, Any]]) -> str:
    """Pick the layout family a source HTML most resembles.

    Deliberately structural, not lexical: it reads how the document is BUILT —
    how many KPIs sit above the charts, whether there is a table-heavy grid,
    whether the copy runs long — because that is what the family encodes. A
    keyword match on the title would just be a different way of guessing.
    """
    kinds = [str(p.get("final_chart_type") or "").upper() for p in chart_plans]
    # KPIs are counted from the DOCUMENT, not only from the charts.
    #
    # A KPI card the planner cannot bind becomes a widget, so it leaves
    # `chart_plans` entirely -- and a brief is defined by having a KPI row, so a
    # document with three headline figures and one main chart came out as a
    # console the moment those figures stopped being charts. What the source has
    # does not change because of what AppBI managed to do with it.
    charted_ids = {str(p.get("block_id")) for p in chart_plans}
    kpi_blocks = {
        str(b.get("id"))
        for b in (document_summary.get("blocks") or [])
        if str(b.get("role") or "") == "kpi"
    }
    kpi_count = sum(1 for k in kinds if k == "KPI") + len(kpi_blocks - charted_ids)
    table_count = sum(1 for k in kinds if k in {"TABLE", "MATRIX"})
    chart_count = max(len(kinds) - kpi_count - table_count, 0)
    prose = sum(
        len(_normalize_text(b.get("text")))
        for b in (document_summary.get("blocks") or [])
        if str(b.get("role") or "") not in {"chart", "table", "kpi"}
    )

    # Long-form copy carrying a few visuals is an article, not a console.
    if prose > 1200 and chart_count <= 4:
        return "editorial"
    # Tables dominating means someone is working a list, not reading a summary.
    if table_count >= 2 and table_count >= chart_count:
        return "ops"
    # A dense grid with little narrative is an operations wall.
    if chart_count + table_count >= 10 and prose < 400:
        return "ops"
    # A KPI strip WITH a table is a board pack -- someone will read the detail.
    # Checked before "stage" because the table is the deciding signal: without
    # it the same shape is a wall display, with it it is a summary.
    #
    # The chart count is part of the test, not decoration. A brief carries ONE
    # chart making the argument; a console carries a grid of them. Without this
    # clause a 4-KPI/4-chart/1-table SaaS console came out as a brief, which
    # then packed its four charts into the recipe's single main_chart slot.
    if kpi_count >= 3 and table_count >= 1 and chart_count <= 3:
        return "brief"
    # A handful of big numbers and one or two visuals reads from across a room.
    if kpi_count >= 3 and chart_count <= 2:
        return "stage"
    return "console"


def _detect_dark_mode(document_summary: Dict[str, Any]) -> bool:
    """Does the source read as a dark report?

    Only the coarse signal is used — an explicit dark background on a block or
    the word in a class name. AppBI paints from its own colorways, so all this
    decides is WHICH colorway, never a literal colour lifted from the HTML.
    """
    for block in (document_summary.get("blocks") or [])[:8]:
        style = block.get("style") or {}
        bg = str(style.get("background") or style.get("background-color") or "").strip().lower()
        m = re.match(r"^#([0-9a-f]{6})$", bg)
        if m:
            r, g, b = (int(m.group(1)[i:i + 2], 16) for i in (0, 2, 4))
            if (0.2126 * r + 0.7152 * g + 0.0722 * b) < 90:
                return True
        classes = " ".join(block.get("classes") or []).lower()
        if "dark" in classes or "night" in classes:
            return True
    return False


def _slicer_blocks(document_summary: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Blocks that look like filter controls in the source."""
    out: List[Dict[str, Any]] = []
    for block in document_summary.get("blocks") or []:
        hay = " ".join([
            _normalize_text(block.get("heading")),
            " ".join(block.get("classes") or []),
            _normalize_text(block.get("id_attr")),
        ]).lower()
        if _FILTERISH_RE.search(hay):
            out.append(block)
    return out


_FILTERISH_RE = re.compile(
    r"(filter|slicer|facet|segment|control|dropdown|select|date-?range|timeframe|"
    r"period|toolbar|bo-?loc|khoang-?thoi-?gian)",
    re.IGNORECASE,
)




# ── HTML fragment preservation ──────────────────────────────────────────────

# The fragment sanitizer lives in its own module: the persistence layer needs
# it too and cannot import this one back.
#: Markup that carries meaning a heading or a paragraph cannot: a gauge drawn
#: in SVG, an avatar row, a progress rail, an illustration.
_FRAGMENT_RICH_RE = re.compile(r"<(svg|img|table|ul|ol|figure)\b", re.IGNORECASE)
#: Below this the fragment is a wrapper around a sentence, and a text widget
#: says the same thing while staying themeable, searchable and translatable.
#:
#: Two, not four. A KPI card is three elements -- a label, a figure, a delta --
#: and at four they failed this test, had too little text to become a text
#: widget, and were dropped with "held no content AppBI could carry". Two of a
#: brief's three headline figures went that way while the third, which the model
#: happened to classify, survived; the same document lost different blocks on
#: different runs.
_FRAGMENT_MIN_ELEMENTS = 2


def _fragment_is_worth_preserving(fragment: str, text: str) -> bool:
    """Whether a block earns an html_fragment rather than a native widget.

    Preserving is not free: a fragment is inert markup frozen at import time.
    It does not follow the dashboard theme, cannot be translated, and will not
    reflow the way a native widget does. So it is the answer only when there is
    something in the block that no native widget can express.
    """
    if not fragment:
        return False
    if _FRAGMENT_RICH_RE.search(fragment):
        return True
    element_count = fragment.count("<") - fragment.count("</")
    if element_count < _FRAGMENT_MIN_ELEMENTS:
        return False
    # A wall of prose is prose, however many divs it is wrapped in.
    return len(text) < 400


def split_plans_by_classification(
    chart_plans: List[Dict[str, Any]],
    block_plan: Optional[List[Dict[str, Any]]],
    block_roles: Optional[Dict[str, str]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Let the model's classification decide what a planned block becomes.

    The classification used to be applied only to blocks the chart planner had
    already given up on, which made it decorative: a block the model called a
    heading or a preserved fragment still went through as a chart, because the
    planner had produced SOMETHING for it.

    That is how four KPI cards -- classified `native_widget`, and holding a
    figure with no column behind it -- were laid out as four full-width detail
    tables. The planner's fallback for "cannot bind this" is TABLE, and a table
    is the widest, tallest tile there is, so an unbindable block became the most
    prominent thing on the page.

    Returns `(charts, reclassified_blocks)`. The reclassified entries keep
    their block id so the widget builder can pick them up.
    """
    planned = {
        str(entry.get("block_id")): entry
        for entry in (block_plan or [])
        if isinstance(entry, dict) and entry.get("block_id")
    }
    roles = {str(k): str(v) for k, v in (block_roles or {}).items()}

    def _planner_gave_up(plan: Dict[str, Any]) -> bool:
        """A KPI card rendered as a table is the planner admitting defeat.

        TABLE is its fallback for "I cannot bind this", and a table is also the
        widest, tallest tile in the grid -- so an unbindable block became the
        most prominent thing on the page.

        This check does not consult the model at all, and that is the point:
        classification varies run to run, so the same document imported as four
        tidy cards one time and four full-size tables the next.
        """
        return (
            roles.get(str(plan.get("block_id"))) == "kpi"
            and str(plan.get("final_chart_type") or "").upper() in {"TABLE", "MATRIX"}
        )

    if not planned:
        forced = [p for p in chart_plans if _planner_gave_up(p)]
        if not forced:
            return chart_plans, []
        kept = [p for p in chart_plans if p not in forced]
        return kept, [{"block_id": p.get("block_id"), "entry": None, "plan": p} for p in forced]

    charts: List[Dict[str, Any]] = []
    reclassified: List[Dict[str, Any]] = []

    for plan in chart_plans:
        entry = planned.get(str(plan.get("block_id")))
        classification = str((entry or {}).get("classification") or "")
        if _planner_gave_up(plan):
            reclassified.append({"block_id": plan.get("block_id"), "entry": entry, "plan": plan})
            continue
        if classification in {"", "native_chart", "slicer"}:
            # `slicer` blocks stay in the chart list on purpose: the slicer
            # itself is built from resolved COLUMNS, and dropping the chart
            # would lose a real visual over a control we may not resolve.
            charts.append(plan)
            continue

        # A chart the planner had to guess at is worth less than the model's
        # read of the block; a chart it was confident about is worth more.
        if float(plan.get("confidence") or 0) >= 0.75 and classification != "unsupported_interaction":
            charts.append(plan)
            continue

        reclassified.append({"block_id": plan.get("block_id"), "entry": entry, "plan": plan})

    return charts, reclassified


def _section_header_config(heading: str, text: str) -> Dict[str, Any]:
    """Title and, only when it says something else, a subtitle.

    A block that IS a heading has the same string for both, and setting them
    both rendered the section as "Revenue & fulfilment" with "REVENUE &
    FULFILMENT" repeated underneath it in small caps.
    """
    title = heading or text[:80]
    subtitle = text[:120] if heading else ""
    if subtitle.strip().lower() == title.strip().lower():
        subtitle = ""
    return {"title": title, "subtitle": subtitle}


def _widget_own_words(config: Any) -> str:
    """The first thing a widget's config actually says, for use as its title."""
    if not isinstance(config, dict):
        return ""
    for key in ("title", "headline", "text", "template", "subtitle", "subhead"):
        value = config.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def widgets_from_plan(declared: Any) -> List[Dict[str, Any]]:
    """Non-chart tiles the document DECLARED, in the shape the packer expects.

    When a plan declares its widgets, they are the answer -- inference must not
    run alongside them. Running both produced a document's headings twice: once
    from the plan and once re-derived from the markup, landing a full-width
    section header on top of the first KPI at (0, 0).

    A declared widget with a declared layout keeps it exactly; one without gets
    placed by the recipe like anything else.
    """
    out: List[Dict[str, Any]] = []
    for index, raw in enumerate(declared or [], start=1):
        if not isinstance(raw, dict):
            continue
        widget_type = str(raw.get("widget_type") or "").strip().lower()
        if widget_type not in IMPORT_WIDGET_TYPES or widget_type == "chart":
            continue
        config = raw.get("widget_config")
        entry: Dict[str, Any] = {
            "block_id": str(raw.get("block_id") or f"widget-{index}"),
            "order": int(raw.get("order") or index),
            "widget_type": widget_type,
            "widget_config": dict(config) if isinstance(config, dict) else {},
            # The title names the TILE, in the preview and on the built
            # dashboard. Falling straight to the kind produced a tile called
            # "Callout" sitting above the sentence it was carrying, which tells
            # a reader nothing they could not already see.
            "title": _normalize_text(raw.get("title"), max_len=160)
            or _normalize_text(_widget_own_words(config), max_len=160)
            or widget_type.replace("_", " ").title(),
            "size_hint": "full" if widget_type in {"section_header", "hero_strip"} else "half",
        }
        layout = _normalize_declared_layout(raw.get("layout"))
        if layout:
            entry["layout"] = layout
        out.append(entry)
    return out


def widgets_from_blocks(
    ignored_blocks: List[Dict[str, Any]],
    document_summary: Dict[str, Any],
    block_plan: Optional[List[Dict[str, Any]]] = None,
    warnings_out: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """Turn the blocks a chart-first import throws away into real widgets.

    A heading, a callout and a lead paragraph are what tell a reader how a
    report is organised. Dropping them with "narrative block not converted"
    kept the numbers and discarded the structure, which is most of why an
    imported dashboard read as a bag of loose charts.

    `block_plan` is the model's classification of every block. When it has an
    opinion about a block, that opinion wins: it read the block's own markup and
    can tell a status badge from a heading, which the class-name regexes below
    cannot. The regexes remain as the answer for every block the model did not
    reach -- and for imports running with AI assist off.

    Preference order is native-first. A native widget follows the theme,
    translates and reflows; a preserved fragment does none of those, so it is
    the answer only when nothing native can carry the block. Nothing is dropped
    in silence: blocks that yield no widget are appended to `warnings_out`.
    """
    by_id = {str(b.get("id")): b for b in (document_summary.get("blocks") or [])}
    planned = {
        str(entry.get("block_id")): entry
        for entry in (block_plan or [])
        if isinstance(entry, dict) and entry.get("block_id")
    }
    notes = warnings_out if warnings_out is not None else []
    out: List[Dict[str, Any]] = []

    for entry in ignored_blocks or []:
        block = by_id.get(str(entry.get("block_id")))
        if not block:
            continue
        heading = _normalize_text(block.get("heading"), max_len=120)
        text = _normalize_text(block.get("text"), max_len=600)
        tag = str(block.get("tag") or "").lower()
        classes = " ".join(block.get("classes") or []).lower()
        fragment = str(block.get("html") or "")
        plan = planned.get(str(block.get("id"))) or {}
        classification = str(plan.get("classification") or "")
        degraded: List[str] = []
        if plan.get("degraded_note"):
            degraded.append(str(plan["degraded_note"]))

        widget_type = ""
        widget_config: Dict[str, Any] = {}

        if classification == "unsupported_interaction":
            # Named, not converted. The report is honest about the hole instead
            # of quietly rendering one block fewer than the source had.
            label = heading or text[:60] or str(block.get("id"))
            note = plan.get("degraded_note") or "the interaction it provided does not exist in AppBI"
            notes.append(f"Block '{label}' was not imported: {note}.")
            continue

        if classification == "html_fragment_preserved" and fragment:
            widget_type = "html_fragment"
            widget_config = {"html": fragment}
        elif classification == "native_widget" and plan.get("widget_type"):
            widget_type = str(plan["widget_type"])
            # A KPI card the planner could not bind is the case the preserved
            # fragment exists for: AppBI has no live figure to draw, and the
            # source has a finished one -- a small caps label, a 27px number, a
            # coloured delta. Demoted to a callout it rendered as a single line
            # of 11px grey text in a 160px box, which is a worse answer than the
            # source's own design.
            if (
                widget_type in {"callout", "text"}
                and str(block.get("role") or "") in {"kpi", "chart"}
                and fragment
            ):
                widget_type = "html_fragment"
            if widget_type == "section_header":
                widget_config = _section_header_config(heading, text)
            elif widget_type == "callout":
                widget_config = {"text": text or heading, "tone": "accent"}
            elif widget_type == "hero_strip":
                widget_config = {"headline": heading or text[:60], "subhead": text[:140] if heading else ""}
            elif widget_type == "html_fragment":
                widget_config = {"html": fragment}
            else:
                # Anything this chain has no branch for is copy. The branch
                # above is not optional: without it the html_fragment chosen a
                # few lines earlier fell through to here and was rewritten back
                # to text, so the preserved-card path looked wired up and never
                # once produced a fragment.
                widget_type = "text"
                widget_config = {"template": text or heading}

        if not widget_type:
            # No usable opinion from the model: fall back to structure.
            if tag in {"h1", "h2", "h3"} or heading:
                widget_type = "section_header"
                widget_config = _section_header_config(heading, text)
            elif re.search(r"(callout|note|alert|tip|warning|insight)", classes):
                widget_type = "callout"
                widget_config = {"text": text, "tone": "accent"}
            elif _fragment_is_worth_preserving(fragment, text):
                widget_type = "html_fragment"
                widget_config = {"html": fragment}
            elif len(text) >= 40:
                widget_type = "text"
                widget_config = {"template": text}
            else:
                label = heading or text[:60] or str(block.get("id"))
                notes.append(f"Block '{label}' held no content AppBI could carry and was skipped.")
                continue

        if widget_type in {"callout", "text", "hero_strip"} and str(block.get("role") or "") in {"kpi", "chart"}:
            # The block showed a NUMBER. Carrying it as copy keeps the report
            # looking like its source, but the figure is now a quotation, not a
            # measurement, and a reader has no other way to tell.
            degraded.append(
                "The figures in this block are copied from the source document. They do not "
                "refresh and they ignore the dashboard's filters."
            )

        if widget_type == "html_fragment":
            # Said plainly so an author knows what to rebuild natively if the
            # frozen copy is not good enough.
            degraded.append(
                "Preserved as static markup: any hover, tooltip or click behaviour in "
                "the original block no longer runs, and it will not follow the dashboard "
                "theme or respond to filters."
            )
            widget_config["degraded"] = degraded
        elif degraded:
            widget_config["degraded"] = degraded

        out.append({
            "block_id": block.get("id"),
            "order": block.get("order", 0),
            # What the block WAS in the source, kept so the packer can give a
            # demoted KPI card the shape of a KPI rather than of a note.
            "_source_role": str(block.get("role") or ""),
            "widget_type": widget_type,
            "widget_config": widget_config,
            "title": heading or widget_type.replace("_", " ").title(),
            # A header spans the row it introduces; a note is half width.
            "size_hint": "full" if widget_type in {"section_header", "hero_strip"} else "half",
        })
    return out



def slicers_from_source(
    *,
    ai_slicer_fields: List[str],
    document_summary: Dict[str, Any],
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
    declared: bool = False,
) -> List[Dict[str, Any]]:
    """Recreate the source's filter controls as real slicer entries.

    A report without its filters is a screenshot.

    Two things this gets wrong if you let it. The first is scope: a dataset has
    many tables, and resolving only against the PRIMARY one silently dropped
    every declared filter that lived elsewhere -- a plan asking to filter by
    `payment_type` (Order Payments) or `review_score` (Order Reviews) built a
    dashboard with an empty filter bar and no explanation.

    The second is invention. Candidate names are harvested from filter-ish
    blocks as well, which is a reasonable guess for a document that declared
    nothing -- but applied to a plan that DID declare its filters it produced a
    slicer nobody asked for: a report declaring `seller_state` came out
    filtering by `order_status`. So `declared` turns the guessing off.
    """
    # Every column the dataset offers, not just the primary table's. A declared
    # filter names a real column; which table holds it is not the author's
    # problem.
    columns: List[Dict[str, Any]] = list(source_profile.get("columns") or [])
    seen_names = {str(c.get("name") or "") for c in columns}
    for profile in (all_source_profiles or {}).values():
        for column in (profile.get("columns") or []):
            name = str(column.get("name") or "")
            if name and name not in seen_names:
                seen_names.add(name)
                columns.append(column)
    if not columns:
        return []

    # The controls the source actually rendered come first: they are evidence,
    # where a filter-ish BLOCK is an inference and the model's list is a guess.
    candidates: List[str] = list(ai_slicer_fields or [])
    if not declared:
        candidates.extend(document_summary.get("filter_controls") or [])
        for block in _slicer_blocks(document_summary):
            for part in (block.get("heading"), block.get("text")):
                label = _normalize_text(part, max_len=60)
                if label:
                    candidates.append(label)

    seen: Set[str] = set()
    out: List[Dict[str, Any]] = []
    for raw in candidates:
        matched = _resolve_column_name(raw, _build_column_alias_map(columns)) or _best_field_match(raw, columns)
        if not matched or matched in seen:
            continue
        # An identifier makes a useless slicer: thousands of distinct values and
        # no meaning to a reader.
        if _is_identifier_like_column(matched):
            continue
        seen.add(matched)
        col_type = next((str(c.get("type") or "") for c in columns if c.get("name") == matched), "")
        out.append({
            "id": f"slicer-{len(out) + 1}",
            "field": matched,
            "label": matched,
            "type": "date" if _is_date_type(col_type) else "number" if _is_number_type(col_type) else "dropdown",
            "operator": "date_between" if _is_date_type(col_type) else "in",
            "value": None if _is_date_type(col_type) else [],
            "scope": "all",
        })
        if len(out) >= 4:  # a filter bar past four controls stops being scannable
            break
    return out


def build_dashboard_contract(
    *,
    document_summary: Dict[str, Any],
    chart_plans: List[Dict[str, Any]],
    embedded: Optional[Dict[str, Any]] = None,
    ai_family: Optional[str] = None,
) -> Dict[str, Any]:
    """The dashboard-level half of an import.

    Charts alone do not make a report: what makes one recognisable is where the
    filters sit, how the sections are grouped, and which look it wears. When the
    source declares those (AppBI metadata) they are honoured verbatim — a
    round-trip must not redesign the author's report. When it does not, they are
    DERIVED from the document's structure and expressed in AppBI's own token
    vocabulary, never as CSS lifted from the page.
    """
    dash_meta = (embedded or {}).get("dashboard") if isinstance((embedded or {}).get("dashboard"), dict) else {}

    declared_theme = None
    for source in (embedded or {}, dash_meta):
        candidate = source.get("theme_config") if isinstance(source, dict) else None
        if isinstance(candidate, dict) and candidate:
            declared_theme = candidate
            break

    declared_dock = None
    for source in (embedded or {}, dash_meta):
        candidate = source.get("slicer_cluster_layout") if isinstance(source, dict) else None
        if isinstance(candidate, dict) and candidate:
            declared_dock = candidate
            break

    # Declared metadata first, then the model's read of the document, then the
    # structural heuristic. The model sees the source's own language — headings,
    # captions, how it addresses its reader — which the heuristic cannot.
    family = (
        _normalize_template_family((embedded or {}).get("template_family"))
        or _normalize_template_family(dash_meta.get("template_family"))
        or _normalize_template_family(ai_family)
        or choose_template_family(document_summary, chart_plans)
    )

    if declared_theme is not None:
        theme_config = dict(declared_theme)
    else:
        # Expressed as an identity, not a pile of tokens: the FE resolves
        # templateId/colorwayId through the same catalog the theme menu uses, so
        # an imported report is indistinguishable from one themed by hand.
        # A declared colorway is a choice; the dark-mode sniff is a guess for
        # documents that made none. Ignoring the choice meant a plan asking for
        # `ocean` imported as `indigo`, with nothing to say why.
        declared_colorway = _normalize_identifier(
            (embedded or {}).get("colorway") or dash_meta.get("colorway")
        )
        theme_config = {
            "templateId": family,
            "colorwayId": (
                declared_colorway
                if declared_colorway in IMPORT_COLORWAYS
                else ("graphite" if _detect_dark_mode(document_summary) else "indigo")
            ),
        }

    if declared_dock is not None:
        slicer_cluster_layout = dict(declared_dock)
    else:
        dock = _FAMILY_DOCK.get(family, "top")
        slicer_cluster_layout = {
            "position": dock,
            "direction": "vertical" if dock in {"left", "right", "drawer"} else "horizontal",
        }

    layout_mode = str((embedded or {}).get("layout_mode") or dash_meta.get("layout_mode") or "grid").strip().lower()
    if layout_mode not in {"grid", "canvas"}:
        layout_mode = "grid"
    canvas_config = (embedded or {}).get("canvas_config") or dash_meta.get("canvas_config")

    return {
        "theme_config": theme_config,
        "layout_mode": layout_mode,
        "canvas_config": canvas_config if isinstance(canvas_config, dict) else None,
        "slicer_cluster_layout": slicer_cluster_layout,
        "template_family": family,
    }



def _grid_rect(item: Dict[str, Any]) -> Optional[Dict[str, int]]:
    layout = item.get("layout")
    if not isinstance(layout, dict) or not all(key in layout for key in ("x", "y", "w", "h")):
        return None
    try:
        rect = {key: int(layout[key]) for key in ("x", "y", "w", "h")}
    except (TypeError, ValueError):
        return None
    if (
        rect["x"] < 0
        or rect["y"] < 0
        or rect["w"] <= 0
        or rect["h"] <= 0
        or rect["x"] + rect["w"] > RECIPE_GRID_COLS
    ):
        return None
    return rect


def _rects_overlap(a: Dict[str, int], b: Dict[str, int]) -> bool:
    return (
        a["x"] < b["x"] + b["w"]
        and b["x"] < a["x"] + a["w"]
        and a["y"] < b["y"] + b["h"]
        and b["y"] < a["y"] + a["h"]
    )


def _place_generated_layouts_around_authored(
    authored: List[Dict[str, Any]],
    generated: List[Dict[str, Any]],
) -> List[str]:
    """Keep recipe shapes while moving generated tiles out of authored space."""
    warnings: List[str] = []
    occupied: List[Tuple[str, Dict[str, int]]] = []
    for item in authored:
        rect = _grid_rect(item)
        if rect is None:
            # Canvas-only geometry is valid and must remain untouched. It does
            # not participate in the grid collision pass.
            continue
        block_id = str(item.get("block_id") or "authored tile")
        collisions = [other_id for other_id, other in occupied if _rects_overlap(rect, other)]
        if collisions:
            warnings.append(
                f"Authored tile '{block_id}' overlaps {', '.join(collisions)}; "
                "the declared coordinates were preserved."
            )
        occupied.append((block_id, rect))

    moved = 0
    ordered = sorted(
        generated,
        key=lambda item: (
            (_grid_rect(item) or {}).get("y", 0),
            (_grid_rect(item) or {}).get("x", 0),
            int(item.get("order") or 0),
        ),
    )
    for item in ordered:
        desired = _grid_rect(item)
        if desired is None:
            continue
        preferred_x = desired["x"]
        x_candidates = [preferred_x] + [
            x for x in range(0, RECIPE_GRID_COLS - desired["w"] + 1)
            if x != preferred_x
        ]
        y = desired["y"]
        while True:
            placed: Optional[Dict[str, int]] = None
            for x in x_candidates:
                candidate = {**desired, "x": x, "y": y}
                if not any(_rects_overlap(candidate, rect) for _, rect in occupied):
                    placed = candidate
                    break
            if placed is not None:
                break
            y += 1

        if placed["x"] != desired["x"] or placed["y"] != desired["y"]:
            moved += 1
        item["layout"] = {**dict(item.get("layout") or {}), **placed}
        occupied.append((str(item.get("block_id") or "generated tile"), placed))

    if moved:
        warnings.append(f"Moved {moved} generated tile(s) to avoid authored layout coordinates.")
    return warnings


def compose_import_layout(
    *,
    chart_plans: List[Dict[str, Any]],
    widgets: List[Dict[str, Any]],
    family: str,
) -> List[str]:
    """Lay out an imported report through its template's recipe.

    Replaces the old `_assign_layouts` path for import. That one packed blocks
    into a 12-column grid purely by a per-block "size hint", which is why every
    imported report came out as the same undifferentiated mosaic however the
    source was built -- the hint described one tile, and nothing described the
    REPORT. A recipe does: named regions, in order, each with a shape.

    Three rules the packer keeps:

    * A layout DECLARED by the source (AppBI metadata) is never overwritten. A
      round-trip must reproduce the author's report, not redesign it.
    * Charts and widgets are packed TOGETHER. Packing them separately is what
      left every widget at x=0,y=0, stacking each section header on top of the
      first chart.
    * The source's own order governs the body. The recipe supplies the SHAPE
      of each kind and hoists the header bands; it does not re-sort the report.
      An earlier version sorted every block into a band by kind, which stacked
      a report's three section headers on top of one another at the very top,
      introducing nothing.
    """
    # Declared-ness is a property of a TILE, not of charts. Splitting only the
    # charts meant a declared widget layout was handed to the recipe and
    # overwritten: a plan that put its section header at y=2 got it packed to
    # y=0, on top of the first KPI.
    everything = list(chart_plans) + list(widgets)
    declared = [it for it in everything if isinstance(it.get("layout"), dict) and it["layout"]]
    free = [it for it in everything if not (isinstance(it.get("layout"), dict) and it["layout"])]

    warnings = apply_layout_recipe(free, family)

    if declared and free:
        warnings.extend(_place_generated_layouts_around_authored(declared, free))
    return warnings


def analyze_dashboard_html_import(
    *,
    html_text: str,
    html_summary: Dict[str, Any] | None,
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    document_summary = _build_document_summary(html_text=html_text, html_summary=html_summary)

    # ── Fast path: embedded metadata from skill-generated HTML ──────────
    embedded = _extract_embedded_metadata(html_text)
    if embedded is not None:
        return _analyze_from_embedded_metadata(
            embedded=embedded,
            document_summary=document_summary,
            source_profile=source_profile,
            all_source_profiles=all_source_profiles,
        )

    # ── Standard path: AI-assisted analysis ─────────────────────────────
    candidates, ignored = _candidate_blocks(document_summary)

    ai_plans, calculated_fields, ai_meta = _ai_chart_plans(
        document_summary=document_summary,
        source_profile=source_profile,
        all_source_profiles=all_source_profiles,
    )
    ai_lookup = {
        str(item.get("block_id")): item
        for item in (ai_plans or [])
        if isinstance(item, dict) and str(item.get("block_id") or "").strip()
    }

    # Enrich source profiles with calculated field names so _finalize_plan accepts them
    def _enrich_profile(profile: Dict[str, Any], calc_fields: List[Dict[str, Any]]) -> Dict[str, Any]:
        if not calc_fields:
            return profile
        enriched = dict(profile)
        existing_cols = list(enriched.get("columns") or [])
        existing_names = {col.get("name") for col in existing_cols}
        for cf in calc_fields:
            if cf["name"] not in existing_names:
                existing_cols.append({"name": cf["name"], "type": "number"})
        enriched["columns"] = existing_cols
        existing_numeric = list(enriched.get("numeric_columns") or [])
        for cf in calc_fields:
            if cf["name"] not in existing_numeric:
                existing_numeric.append(cf["name"])
        enriched["numeric_columns"] = existing_numeric
        return enriched

    enriched_profile = _enrich_profile(source_profile, calculated_fields)
    enriched_all = None
    if all_source_profiles:
        enriched_all = {k: _enrich_profile(v, calculated_fields) for k, v in all_source_profiles.items()}

    use_multi = enriched_all and len(enriched_all) > 1
    finalized_plans: List[Dict[str, Any]] = []
    for block in candidates:
        if use_multi:
            best_plan: Optional[Dict[str, Any]] = None
            for key, profile in enriched_all.items():
                plan = _finalize_plan(
                    block=block,
                    raw_plan=ai_lookup.get(block["id"]),
                    source_profile=profile,
                )
                plan["source_key"] = key
                if best_plan is None or _plan_quality_score(plan) > _plan_quality_score(best_plan):
                    best_plan = plan
            if best_plan is not None:
                finalized_plans.append(best_plan)
        else:
            plan = _finalize_plan(
                block=block,
                raw_plan=ai_lookup.get(block["id"]),
                source_profile=enriched_profile,
            )
            finalized_plans.append(plan)

    finalized_plans = [plan for plan in finalized_plans if plan.get("source_fields_used") or plan.get("final_chart_type") == "TABLE"]
    finalized_plans.sort(key=lambda item: (item.get("order", 0), item.get("block_id", "")))

    # Classify, THEN compose, THEN place. Each step's input is the previous
    # step's output:
    #
    #   1. classification decides which blocks are charts at all,
    #   2. the surviving chart mix decides the template family,
    #   3. the family's recipe decides where everything goes.
    #
    # Getting this order wrong is not cosmetic. Computing the family first --
    # as this did -- meant counting four unbindable KPI cards as four TABLEs,
    # and a table-heavy count picks a table-heavy template for a document that
    # has one table in it.
    block_warnings: List[str] = []
    finalized_plans, reclassified = split_plans_by_classification(
        finalized_plans,
        (ai_meta or {}).get("block_plan"),
        block_roles={
            str(b.get("id")): str(b.get("role") or "")
            for b in (document_summary.get("blocks") or [])
        },
    )

    contract = build_dashboard_contract(
        document_summary=document_summary,
        chart_plans=finalized_plans,
        embedded=None,
        ai_family=(ai_meta or {}).get("template_family"),
    )

    # A candidate the planner never returned a plan for is the third way a
    # block can disappear, and it was the quietest: it is not in `ignored`
    # (the analyzer thought it WAS chartable) and it is not in the plans (the
    # model skipped it), so nothing downstream ever saw it again. Two of a
    # brief's three headline figures went missing exactly this way.
    planned_ids = {str(p.get("block_id")) for p in finalized_plans}
    planned_ids.update(str(item["block_id"]) for item in reclassified)
    unplanned = [
        {"block_id": block.get("id")}
        for block in candidates
        if str(block.get("id")) not in planned_ids
    ]

    widgets = widgets_from_blocks(
        ignored + unplanned + [{"block_id": item["block_id"]} for item in reclassified],
        document_summary,
        block_plan=(ai_meta or {}).get("block_plan"),
        warnings_out=block_warnings,
    )
    if reclassified:
        block_warnings.append(
            f"{len(reclassified)} block(s) the chart planner could not bind to data were kept "
            "as their own kind of content instead of being forced into a table. Figures shown "
            "in them come from the source document and do not refresh."
        )
    layout_warnings = compose_import_layout(
        chart_plans=finalized_plans,
        widgets=widgets,
        family=contract["template_family"],
    )

    warnings: List[str] = list(layout_warnings) + block_warnings
    warnings.extend((ai_meta or {}).get("calculated_field_warnings") or [])
    if not finalized_plans:
        warnings.append("No chart plans could be validated from the imported HTML and selected source.")
    if any(plan.get("changed_chart_type") for plan in finalized_plans):
        warnings.append("One or more HTML visuals were adapted to the closest supported AppBI chart type.")
    if widgets:
        warnings.append(
            f"Recovered {len(widgets)} non-chart block(s) — section headings, notes and copy — "
            f"as widgets, laid out with the '{contract['template_family']}' template."
        )

    return {
        "suggested_dashboard_name": _normalize_text(document_summary.get("title"), max_len=180) or "Imported Dashboard",
        "document_title": document_summary.get("title"),
        "source_profile": source_profile,
        "all_source_profiles": all_source_profiles,
        "chart_plans": finalized_plans,
        "calculated_fields": calculated_fields,
        "ignored_blocks": ignored,
        "warnings": warnings,
        "ai_meta": ai_meta,
        # The composition derived above: which template, where the filters live,
        # what the non-chart blocks became. Without it an import produced charts
        # and nothing else — no filter placement, no density, no hierarchy.
        **contract,
        "widgets": widgets,
        "slicers": slicers_from_source(
            ai_slicer_fields=(ai_meta or {}).get("slicer_fields") or [],
            document_summary=document_summary,
            source_profile=source_profile,
            # A filter column can live in any table of the dataset, not only
            # the one the primary chart happens to read from.
            all_source_profiles=all_source_profiles,
        ),
    }


def analyze_dashboard_html_import_batch(
    *,
    documents: List[Dict[str, Any]],
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Analyze multiple HTML documents without changing the single-document contract.

    Each input document normally maps to one output page, but embedded
    ``appbi-import/v2`` metadata may expand one HTML file into multiple page
    analyses inside the batch envelope.
    """
    analyzed_documents: List[Dict[str, Any]] = []
    used_page_names: Set[str] = set()

    for index, raw_document in enumerate(documents or [], start=1):
        if not isinstance(raw_document, dict):
            raise ValueError(f"documents[{index - 1}] must be an object.")

        document_id = _normalize_text(raw_document.get("document_id"), max_len=120) or f"document-{index}"
        html_text = str(raw_document.get("html_text") or "").strip()
        if not html_text:
            raise ValueError(f"documents[{index - 1}].html_text is required.")

        html_summary = raw_document.get("html_summary")
        if html_summary is not None and not isinstance(html_summary, dict):
            raise ValueError(f"documents[{index - 1}].html_summary must be an object when provided.")

        embedded = _extract_embedded_metadata(html_text)
        if embedded and _is_v2_metadata(embedded):
            base_filename = _normalize_text(raw_document.get("filename"), max_len=260) or None
            base_page_name = _normalize_text(raw_document.get("page_name"), max_len=80)
            document_summary = _build_document_summary(
                html_text=html_text,
                html_summary=html_summary if isinstance(html_summary, dict) else None,
            )
            pages = _normalize_v2_pages(embedded)
            if not pages:
                raise ValueError(f"documents[{index - 1}] contains appbi-import/v2 metadata but no pages[].")

            for page_index, page in enumerate(pages, start=1):
                page_analysis = _analyze_from_v2_metadata_page(
                    embedded=embedded,
                    page=page,
                    document_summary=document_summary,
                    source_profile=source_profile,
                    all_source_profiles=all_source_profiles,
                )
                page_name = _unique_page_name(
                    _resolve_batch_page_name(
                        explicit_name=_normalize_text(page.get("title"), max_len=80) or base_page_name,
                        document_title=page_analysis.get("document_title"),
                        suggested_dashboard_name=page_analysis.get("suggested_dashboard_name"),
                        filename=base_filename,
                        fallback_index=len(analyzed_documents) + 1,
                    ),
                    used_page_names,
                )
                analyzed_documents.append(
                    {
                        "document_id": f"{document_id}::{_normalize_text(page.get('page_key'), max_len=120) or page_index}",
                        "filename": base_filename,
                        "page_name": page_name,
                        "analysis": page_analysis,
                    }
                )
            continue

        analysis = analyze_dashboard_html_import(
            html_text=html_text,
            html_summary=html_summary,
            source_profile=source_profile,
            all_source_profiles=all_source_profiles,
        )
        suggested_page_name = _unique_page_name(
            _resolve_batch_page_name(
                explicit_name=raw_document.get("page_name"),
                document_title=analysis.get("document_title"),
                suggested_dashboard_name=analysis.get("suggested_dashboard_name"),
                filename=raw_document.get("filename"),
                fallback_index=index,
            ),
            used_page_names,
        )
        analyzed_documents.append(
            {
                "document_id": document_id,
                "filename": _normalize_text(raw_document.get("filename"), max_len=260) or None,
                "page_name": suggested_page_name,
                "analysis": analysis,
            }
        )

    if not analyzed_documents:
        raise ValueError("At least one HTML document is required for batch analyze.")

    first_analysis = analyzed_documents[0]["analysis"]
    return {
        "suggested_dashboard_name": first_analysis.get("suggested_dashboard_name") or "Imported Dashboard",
        "document_count": len(analyzed_documents),
        "documents": analyzed_documents,
    }


def delete_import_draft_dataset(db: Session, dataset_obj: Dataset) -> None:
    """Delete ONE draft dataset created by the HTML-import wizard.

    Flushes but does NOT commit — the caller owns the transaction so a draft
    purge can ride along with whatever else it is doing (e.g. deleting the
    datasource the draft points at).

    Also removes the wizard's own throw-away manual datasource once nothing
    else references it. NOTE: the Dataset row must be deleted EXPLICITLY —
    ``dataset_tables.datasource_id`` is ``ON DELETE CASCADE``, so dropping the
    datasource alone wipes the tables and leaves the Dataset behind as an
    invisible empty shell.
    """
    table_ids = [t.id for t in dataset_obj.tables]
    datasource_ids = {t.datasource_id for t in dataset_obj.tables if t.datasource_id}

    if table_ids:
        # Drafts should have no charts but clean up defensively so the dataset
        # can be deleted without leaving charts pointing at nothing.
        db.query(Chart).filter(Chart.dataset_table_id.in_(table_ids)).delete(
            synchronize_session=False
        )
    db.delete(dataset_obj)
    db.flush()

    for ds_id in datasource_ids:
        ds_row = db.query(DataSource).filter(DataSource.id == ds_id).first()
        if not ds_row:
            continue
        if ds_row.type != DataSourceType.MANUAL:
            continue
        if not str(ds_row.name or "").startswith(IMPORT_DATASOURCE_NAME_PREFIX):
            continue
        still_in_use = (
            db.query(DatasetTable).filter(DatasetTable.datasource_id == ds_id).first()
        )
        if still_in_use is None:
            db.delete(ds_row)
    db.flush()


def purge_stale_import_drafts(
    db: Session,
    *,
    datasource_id: Optional[int] = None,
    older_than_hours: Optional[int] = IMPORT_DRAFT_TTL_HOURS,
) -> List[int]:
    """Delete abandoned wizard drafts and return the ids that were removed.

    Drafts are hidden from every listing, so an abandoned one is invisible yet
    still counts as a "dataset using this source" — which silently blocks
    ``DELETE /datasources/{id}`` forever. The wizard only cancels its draft
    from the browser (modal close / beforeunload), so a crashed tab or a killed
    session always leaks one.

    ``datasource_id`` restricts the sweep to drafts touching that source, and
    ``older_than_hours=None`` disables the age filter (used when the source is
    being deleted anyway — age is irrelevant then). Never commits.
    """
    query = db.query(Dataset).filter(Dataset.is_draft.is_(True))
    if older_than_hours is not None:
        cutoff = datetime.utcnow() - timedelta(hours=older_than_hours)
        # created_at is nullable on legacy rows — treat NULL as stale.
        query = query.filter(
            (Dataset.created_at.is_(None)) | (Dataset.created_at < cutoff)
        )
    if datasource_id is not None:
        query = query.filter(
            Dataset.id.in_(
                db.query(DatasetTable.dataset_id).filter(
                    DatasetTable.datasource_id == datasource_id
                )
            )
        )

    purged: List[int] = []
    for dataset_obj in query.all():
        dataset_id = dataset_obj.id
        # SAVEPOINT per draft: one undeletable draft (unexpected FK) must not
        # poison the caller's transaction or abort the rest of the sweep.
        savepoint = db.begin_nested()
        try:
            delete_import_draft_dataset(db, dataset_obj)
            savepoint.commit()
        except Exception:
            savepoint.rollback()
            logger.warning("Could not purge stale import draft %s", dataset_id, exc_info=True)
            continue
        purged.append(dataset_id)
    if purged:
        logger.info("Purged stale dashboard-import drafts: %s", purged)
    return purged


def create_manual_dataset_from_excel_source(
    db: Session,
    *,
    current_user: User,
    file_bytes: bytes,
    filename: Optional[str],
    requested_name: Optional[str],
    selected_sheet_name: Optional[str] = None,
) -> Tuple[int, int]:
    all_sheets = parse_uploaded_source_sheets(file_bytes=file_bytes, filename=filename)
    if not all_sheets:
        raise ValueError("Uploaded Excel source does not contain usable data rows.")

    available_sheet_names = list(all_sheets.keys())
    primary_sheet_name = str(selected_sheet_name or "").strip()
    if not primary_sheet_name or primary_sheet_name not in all_sheets:
        primary_sheet_name = available_sheet_names[0]

    if not any((sheet_payload.get("rows") or []) for sheet_payload in all_sheets.values()):
        raise ValueError("Uploaded Excel source does not contain usable data rows.")

    base_name = _normalize_text(requested_name) or _normalize_text(primary_sheet_name) or "Imported Data"
    datasource_name = f"{IMPORT_DATASOURCE_NAME_PREFIX} {base_name}"
    data_source = DataSourceCRUDService.create(
        db,
        DataSourceCreate(
            name=datasource_name,
            type="manual",
            config={"sheets": all_sheets},
        ),
        owner_id=current_user.id,
    )

    dataset = DatasetCRUDService.create_dataset(
        db,
        DatasetCreate(name=base_name),
        owner_id=current_user.id,
    )

    primary_table_id: Optional[int] = None
    for sheet_name, sheet_data in all_sheets.items():
        if not (sheet_data.get("columns") or []):
            continue
        table = DatasetCRUDService.add_table_to_dataset(
            db,
            dataset.id,
            TableCreate(
                datasource_id=data_source.id,
                source_kind="physical_table",
                source_table_name=sheet_name,
                display_name=sheet_name,
            ),
        )
        if table is None:
            raise ValueError(f"Failed to attach sheet '{sheet_name}' as a dataset table.")
        if sheet_name == primary_sheet_name:
            primary_table_id = table.id

    if primary_table_id is None:
        raise ValueError("Failed to resolve a primary sheet/table for chart creation.")
    return dataset.id, primary_table_id


def create_manual_dataset_from_multi_excel_source(
    db: Session,
    *,
    current_user: User,
    files: List[Tuple[bytes, Optional[str]]],
    requested_name: Optional[str],
) -> Tuple[int, Dict[str, int]]:
    """Create a dataset from multiple Excel/CSV files.

    Returns ``(dataset_id, source_key_to_table_id_map)``.
    """
    all_sheet_data: Dict[str, Dict[str, Any]] = {}
    source_key_to_safe_name: Dict[str, str] = {}

    for file_bytes, filename in files:
        safe_filename = filename or "uploaded"
        sheets = parse_uploaded_source_sheets(file_bytes=file_bytes, filename=filename)
        for sheet_name, sheet_data in sheets.items():
            source_key = f"{safe_filename}::{sheet_name}"
            # DuckDB cannot handle '::' in table names — use ' - ' for storage
            safe_storage_name = f"{safe_filename} - {sheet_name}"
            all_sheet_data[source_key] = sheet_data
            source_key_to_safe_name[source_key] = safe_storage_name

    if not all_sheet_data:
        raise ValueError("Uploaded files do not contain usable data rows.")
    if not any((data.get("rows") or []) for data in all_sheet_data.values()):
        raise ValueError("Uploaded files do not contain usable data rows.")

    # Build config with DuckDB-safe keys (no '::')
    safe_config_sheets = {
        source_key_to_safe_name[k]: v for k, v in all_sheet_data.items()
    }

    base_name = _normalize_text(requested_name) or "Imported Data"
    datasource_name = f"{IMPORT_DATASOURCE_NAME_PREFIX} {base_name}"
    data_source = DataSourceCRUDService.create(
        db,
        DataSourceCreate(
            name=datasource_name,
            type="manual",
            config={"sheets": safe_config_sheets},
        ),
        owner_id=current_user.id,
    )

    dataset = DatasetCRUDService.create_dataset(
        db,
        DatasetCreate(name=base_name),
        owner_id=current_user.id,
    )

    sheet_name_counts: Dict[str, int] = {}
    for source_key in all_sheet_data:
        _, sheet_name = source_key.rsplit("::", 1) if "::" in source_key else ("", source_key)
        sheet_name_counts[sheet_name] = sheet_name_counts.get(sheet_name, 0) + 1

    table_id_map: Dict[str, int] = {}
    for source_key, sheet_data in all_sheet_data.items():
        if not (sheet_data.get("columns") or []):
            continue
        if "::" in source_key:
            filename_part, sheet_name = source_key.rsplit("::", 1)
        else:
            filename_part, sheet_name = "", source_key
        if sheet_name_counts.get(sheet_name, 0) > 1 and filename_part:
            display_name = f"{sheet_name} ({filename_part})"
        else:
            display_name = sheet_name

        safe_name = source_key_to_safe_name[source_key]
        table = DatasetCRUDService.add_table_to_dataset(
            db,
            dataset.id,
            TableCreate(
                datasource_id=data_source.id,
                source_kind="physical_table",
                source_table_name=safe_name,
                display_name=display_name,
            ),
        )
        if table is None:
            raise ValueError(f"Failed to attach '{display_name}' as a dataset table.")
        table_id_map[source_key] = table.id

    if not table_id_map:
        raise ValueError("No usable sheets found in uploaded files.")
    return dataset.id, table_id_map


def _build_chart_config(plan: Dict[str, Any], dataset_id: Optional[int]) -> Dict[str, Any]:
    """Materialize a chart plan into the persisted chart config shape.

    Honors optional fields produced by the manual-edit flow:
      - query_mode: 'generated' | 'custom'
      - custom_sql: SQL text for custom mode
      - custom_role_config: role config bound to the custom SQL output
      - base_filters: base filter list applied to all runtime queries
    Missing fields fall back to the legacy generated-only behavior so that
    plans produced before this was added continue to build correctly.
    """
    role_config = dict(plan.get("role_config") or {})
    custom_role_config = dict(plan.get("custom_role_config") or {}) or {"metrics": []}

    raw_query_mode = str(plan.get("query_mode") or "generated").strip().lower()
    custom_sql = str(plan.get("custom_sql") or "").strip()
    query_mode = "custom" if (raw_query_mode == "custom" and custom_sql) else "generated"

    base_filters_raw = plan.get("base_filters") or []
    base_filters = [item for item in base_filters_raw if isinstance(item, dict)]

    if query_mode == "custom":
        active_role_config = custom_role_config or role_config
    else:
        active_role_config = role_config

    config: Dict[str, Any] = {
        "dataset_id": dataset_id,
        "filters": list(base_filters),
        "baseFilters": list(base_filters),
        "chartType": plan.get("final_chart_type"),
        "queryMode": query_mode,
        "roleConfig": active_role_config,
        "generatedRoleConfig": role_config,
        "customRoleConfig": custom_role_config,
        "styleConfig": dict(plan.get("style_config") or {}),
    }
    if custom_sql:
        config["customSql"] = custom_sql
    return config


def _apply_calculated_fields_to_tables(
    db: Session,
    *,
    calculated_fields: List[Dict[str, Any]],
    table_id_map: Dict[str, int],
    default_table_id: Optional[int],
) -> None:
    """Inject ``add_column`` transformation steps into dataset tables for AI-suggested calculated fields."""
    from app.services.transformation_compiler import TransformationCompiler

    # Group fields by target table
    table_fields: Dict[int, List[Dict[str, Any]]] = {}
    for cf in calculated_fields:
        if not isinstance(cf, dict):
            continue
        name = str(cf.get("name") or "").strip()
        expression = str(cf.get("expression") or "").strip()
        if not name or not expression:
            continue

        is_valid, _ = TransformationCompiler.validate_expression(expression)
        if not is_valid:
            continue

        source_key = cf.get("source_key")
        if source_key and source_key in table_id_map:
            tid = table_id_map[source_key]
        elif default_table_id:
            tid = default_table_id
        else:
            # Apply to all tables as fallback
            for tid_val in (table_id_map.values() if table_id_map else []):
                table_fields.setdefault(tid_val, []).append(cf)
            continue
        table_fields.setdefault(tid, []).append(cf)

    for table_id, fields in table_fields.items():
        db_table = db.query(DatasetTable).filter(DatasetTable.id == table_id).first()
        if not db_table:
            continue

        existing_transforms = list(db_table.transformations or [])
        existing_calc_names = {
            t.get("params", {}).get("newField")
            for t in existing_transforms
            if t.get("type") == "add_column"
        }

        for cf in fields:
            field_name = cf["name"]
            if field_name in existing_calc_names:
                continue
            existing_transforms.append({
                "type": "add_column",
                "enabled": True,
                "params": {
                    "newField": field_name,
                    "expression": cf["expression"],
                },
            })
            existing_calc_names.add(field_name)

            # Update columns_cache with the new calculated column
            # columns_cache is stored as {"columns": [...], ...} dict format
            raw_cache = db_table.columns_cache
            if isinstance(raw_cache, dict):
                cols_list = list(raw_cache.get("columns") or [])
            elif isinstance(raw_cache, list):
                cols_list = list(raw_cache)
            else:
                cols_list = []
            if not any(c.get("name") == field_name for c in cols_list):
                cols_list.append({
                    "name": field_name,
                    "type": "number",
                    "nullable": True,
                    "is_calculated": True,
                    "description": cf.get("label") or field_name,
                })
                if isinstance(raw_cache, dict):
                    db_table.columns_cache = {**raw_cache, "columns": cols_list}
                else:
                    db_table.columns_cache = {"columns": cols_list}

        db_table.transformations = existing_transforms
        db.flush()


def build_dashboard_from_import(
    db: Session,
    *,
    current_user: User,
    analysis: Dict[str, Any],
    source_mode: SourceMode,
    dataset_table_id: Optional[int],
    dataset_id: Optional[int] = None,
    source_bytes: Optional[bytes],
    source_filename: Optional[str],
    source_files: Optional[List[Tuple[bytes, Optional[str]]]] = None,
    selected_sheet_name: Optional[str],
    dashboard_name: Optional[str],
    target_mode: TargetMode,
    target_dashboard_id: Optional[int],
    included_block_ids: List[str],
) -> Dict[str, Any]:
    chart_plans = [plan for plan in (analysis.get("chart_plans") or []) if isinstance(plan, dict)]
    if included_block_ids:
        selected_ids = {str(item) for item in included_block_ids}
        chart_plans = [plan for plan in chart_plans if str(plan.get("block_id")) in selected_ids]
    if not chart_plans:
        raise ValueError("No chart plans were selected for build.")

    resolved_dataset_table_id = dataset_table_id
    resolved_dataset_id: Optional[int] = None
    table_id_map: Dict[str, int] = {}

    if source_mode == "existing_dataset":
        if dataset_id is not None:
            # Multi-table mode: use all non-calendar tables from the dataset
            dataset_obj = db.query(Dataset).filter(Dataset.id == dataset_id).first()
            if not dataset_obj:
                raise ValueError("Dataset not found.")
            require_view_access(db, current_user, dataset_obj, "datasets")
            resolved_dataset_id = dataset_obj.id

            db_tables = (
                db.query(DatasetTable)
                .filter(
                    DatasetTable.dataset_id == dataset_id,
                    DatasetTable.source_kind != "generated_calendar",
                )
                .order_by(DatasetTable.id)
                .all()
            )
            if not db_tables:
                raise ValueError("Dataset has no data tables.")
            table_id_map = _build_table_id_map_from_dataset_tables(db_tables)
            resolved_dataset_table_id = db_tables[0].id
        elif resolved_dataset_table_id is not None:
            # Legacy single-table mode
            db_table = db.query(DatasetTable).filter(DatasetTable.id == resolved_dataset_table_id).first()
            if not db_table:
                raise ValueError("Dataset table not found.")
            dataset_obj = db.query(Dataset).filter(Dataset.id == db_table.dataset_id).first()
            if not dataset_obj:
                raise ValueError("Dataset not found.")
            require_view_access(db, current_user, dataset_obj, "datasets")
            resolved_dataset_id = dataset_obj.id
        else:
            raise ValueError("dataset_id or dataset_table_id is required when building from an existing source.")
    else:
        dataset_permission = (current_user.permissions or {}).get("datasets", "none")
        if dataset_permission not in {"edit", "full"}:
            raise ValueError("Creating a temporary dataset from Excel requires datasets edit permission.")

        if source_files and len(source_files) > 0:
            resolved_dataset_id, table_id_map = create_manual_dataset_from_multi_excel_source(
                db,
                current_user=current_user,
                files=source_files,
                requested_name=dashboard_name or analysis.get("suggested_dashboard_name"),
            )
            resolved_dataset_table_id = next(iter(table_id_map.values())) if table_id_map else None
        elif source_bytes:
            resolved_dataset_id, resolved_dataset_table_id = create_manual_dataset_from_excel_source(
                db,
                current_user=current_user,
                file_bytes=source_bytes,
                filename=source_filename,
                requested_name=dashboard_name or analysis.get("suggested_dashboard_name"),
                selected_sheet_name=selected_sheet_name or analysis.get("source_profile", {}).get("selected_sheet_name"),
            )
            if resolved_dataset_id is not None:
                db_tables = (
                    db.query(DatasetTable)
                    .filter(
                        DatasetTable.dataset_id == resolved_dataset_id,
                        DatasetTable.source_kind != "generated_calendar",
                    )
                    .order_by(DatasetTable.id)
                    .all()
                )
                table_id_map = _build_table_id_map_from_dataset_tables(db_tables)
        else:
            raise ValueError("Uploaded Excel source is no longer available. Please analyze again.")

    if resolved_dataset_table_id is None and not table_id_map:
        raise ValueError("No dataset table is available for chart creation.")

    # Apply AI-suggested calculated fields as transformation steps
    calc_fields = analysis.get("calculated_fields") or []
    if calc_fields and isinstance(calc_fields, list):
        _apply_calculated_fields_to_tables(
            db,
            calculated_fields=calc_fields,
            table_id_map=table_id_map,
            default_table_id=resolved_dataset_table_id,
        )

    # AppBI Import Plan v1: materialize derived_table ops into per-chart
    # customSql bindings so charts can read pre-aggregated data without
    # needing new DatasetTable rows.
    derived_tables = analysis.get("derived_tables") or []
    if derived_tables and isinstance(derived_tables, list):
        derived_warnings = _materialize_v1_derived_tables(
            db,
            derived_tables=derived_tables,
            chart_plans=chart_plans,
            dataset_id=resolved_dataset_id,
            table_id_map=table_id_map,
        )
        if derived_warnings:
            logger.info(
                "Derived-table materialization produced %d warnings for dashboard import.",
                len(derived_warnings),
            )

    # Phase 3: validate that chart source_keys actually resolve to a table.
    if table_id_map:
        unresolved = sorted({
            str(plan.get("source_key"))
            for plan in chart_plans
            if plan.get("source_key") and str(plan.get("source_key")) not in table_id_map
        })
        if unresolved:
            logger.warning(
                "Import contains charts referencing source_keys not in the dataset: %s",
                unresolved,
            )

    created_charts: List[Chart] = []
    type_changes: List[Dict[str, Any]] = []

    # Split plans by widget_type. Plans without widget_type (or widget_type="chart")
    # follow the legacy chart-creation path. Non-chart widgets (text/image/countdown/
    # shape/parameter_switcher) skip Chart creation and only get a DashboardChart row
    # with widget_type + widget_config set, chart_id=NULL. The migration
    # 20260501_0001 made dashboard_charts.chart_id nullable for exactly this case.
    def _plan_widget_type(plan: Dict[str, Any]) -> str:
        """The widget kind this plan builds, or "chart".

        Unknown kinds fall back to "chart", and that fallback is a trap worth
        naming: a `section_header` missing from this set was persisted as a
        chart with `chart_id = NULL`, so the renderer went looking for a chart
        that was never there and the tile read "Failed to load chart" -- three
        of them across the top of an imported report, with the correct
        `widget_config` sitting right there in the row.

        So this set is not a nicety. Any widget kind the importer can EMIT has
        to be listed here, and `frontend/scripts/check-theme-presets.mjs` fails
        the build if the renderer has no case for one of them.
        """
        wt = str(plan.get("widget_type") or "chart").strip().lower()
        return wt if wt in IMPORT_WIDGET_TYPES else "chart"

    def _layout_number(layout: Dict[str, Any], key: str) -> float | None:
        value = layout.get(key)
        if isinstance(value, bool) or value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _materialize_dashboard_layout(
        raw_layout: Dict[str, Any],
        *,
        page_id: str,
        default_w: int,
        default_h: int,
        custom_title: str | None = None,
    ) -> Dict[str, Any]:
        layout = dict(raw_layout or {})
        materialized: Dict[str, Any] = {
            "x": int(_layout_number(layout, "x") or 0),
            "y": int(_layout_number(layout, "y") or 0),
            "w": int(_layout_number(layout, "w") or default_w),
            "h": int(_layout_number(layout, "h") or default_h),
            "pageId": page_id,
        }
        if custom_title:
            materialized["custom_title"] = custom_title

        # Preserve canvas-mode geometry from MCP/HTML plans. Earlier builds only
        # persisted x/y/w/h, so a pixel-perfect canvas prompt silently fell back
        # to grid-like placement.
        for key in ("xPx", "yPx", "wPx", "hPx"):
            value = _layout_number(layout, key)
            if value is not None:
                materialized[key] = value
        z_value = _layout_number(layout, "z")
        if z_value is not None:
            materialized["z"] = int(z_value)
        if isinstance(layout.get("styleConfigOverride"), dict):
            materialized["styleConfigOverride"] = layout["styleConfigOverride"]
        return materialized

    chart_plans_only = [p for p in chart_plans if _plan_widget_type(p) == "chart"]
    # `analysis["widgets"]` carries the section headers / callouts / text the
    # analyzer recovered from blocks a chart-first import used to discard. They
    # are the source's visual hierarchy; without them the import reproduces the
    # numbers and loses the report. They join the same widget path plans already
    # use (chart_id NULL), so nothing downstream needs to know the difference.
    contract_widgets = [w for w in (analysis.get("widgets") or []) if isinstance(w, dict)]
    widget_plans = [p for p in chart_plans if _plan_widget_type(p) != "chart"] + contract_widgets

    for index, plan in enumerate(chart_plans_only, start=1):
        # Prefer user-edited title/description from the manual editor, falling back
        # to the AI-generated ones captured during analysis.
        raw_user_name = _normalize_text(plan.get("chart_name"), max_len=160)
        display_title = (
            raw_user_name
            or _normalize_text(plan.get("title"), max_len=160)
            or f"Imported Chart {index}"
        )
        internal_name = _unique_chart_name(db, current_user.id, display_title)
        user_description = _normalize_text(plan.get("chart_description"), max_len=1024)
        if user_description:
            chart_description = user_description
        else:
            chart_description_parts = [
                _normalize_text(plan.get("rationale"), max_len=400),
                _normalize_text(plan.get("conversion_note"), max_len=280),
                _normalize_text(plan.get("source_excerpt"), max_len=320),
            ]
            chart_description = "\n\n".join(part for part in chart_description_parts if part)

        chart_table_id = resolved_dataset_table_id
        if table_id_map:
            plan_source_key = plan.get("source_key")
            if plan_source_key and plan_source_key in table_id_map:
                chart_table_id = table_id_map[plan_source_key]

        # Manual edit may have switched the chart to a different table within the
        # same dataset. The override is an absolute dataset_table_id; we verify it
        # belongs to the active dataset before trusting it.
        table_override = plan.get("dataset_table_id_override")
        if isinstance(table_override, int) and table_override > 0:
            override_table = db.query(DatasetTable).filter(DatasetTable.id == table_override).first()
            if override_table and (
                resolved_dataset_id is None or override_table.dataset_id == resolved_dataset_id
            ):
                chart_table_id = override_table.id

        chart_config = with_chart_semantic_binding(
            db,
            chart_table_id,
            _build_chart_config(plan, resolved_dataset_id),
            auto_generate=True,
        )
        db_chart = Chart(
            name=internal_name,
            description=chart_description or None,
            dataset_table_id=chart_table_id,
            chart_type=ChartType(str(plan.get("final_chart_type") or "TABLE").upper()),
            config=chart_config,
            owner_id=current_user.id,
        )
        db.add(db_chart)
        db.flush()
        created_charts.append(db_chart)

        if plan.get("changed_chart_type"):
            type_changes.append(
                {
                    "block_id": plan.get("block_id"),
                    "title": display_title,
                    "from": plan.get("requested_chart_type") or plan.get("original_chart_type"),
                    "to": plan.get("final_chart_type"),
                    "note": plan.get("conversion_note"),
                }
            )

        metadata_tags = [
            f"import-html:{plan.get('block_role')}",
            f"import-html:{str(plan.get('final_chart_type') or '').lower()}",
        ]
        db.add(
            ChartMetadata(
                chart_id=db_chart.id,
                domain="html_import",
                intent=str(plan.get("final_chart_type") or "").lower(),
                metrics=[item.get("field") for item in (plan.get("role_config", {}).get("metrics") or []) if item.get("field")],
                dimensions=[field for field in [plan.get("role_config", {}).get("dimension"), plan.get("role_config", {}).get("timeField")] if field],
                tags=[tag for tag in metadata_tags if tag],
            )
        )

    ai_meta = analysis.get("ai_meta") if isinstance(analysis.get("ai_meta"), dict) else {}
    import_page_name = (
        _normalize_text(ai_meta.get("default_page_name"), max_len=80)
        or _normalize_text(
            analysis.get("document_title")
            or dashboard_name
            or analysis.get("suggested_dashboard_name"),
            max_len=80,
        )
        or "Imported Page"
    )
    dashboard_obj: Dashboard
    page_id: str

    if target_mode == "append_to_dashboard":
        if target_dashboard_id is None:
            raise ValueError("target_dashboard_id is required when appending to a dashboard.")
        dashboard_obj = db.query(Dashboard).filter(Dashboard.id == target_dashboard_id).first()
        if not dashboard_obj:
            raise ValueError("Target dashboard not found.")
        require_edit_access(db, current_user, dashboard_obj, "dashboards")
        pages = list(dashboard_obj.pages_config or [DEFAULT_DASHBOARD_PAGE])
        page_id = _generate_page_id()
        pages.append({"id": page_id, "name": import_page_name})
        dashboard_obj.pages_config = pages
    else:
        resolved_dashboard_name = _unique_dashboard_name(
            db,
            dashboard_name or analysis.get("suggested_dashboard_name") or "Imported Dashboard",
        )
        page_id = _generate_page_id()
        # Plan can declare canvas mode at the dashboard level; default stays "grid"
        # so legacy plans without these fields keep their original behavior.
        plan_layout_mode = str(analysis.get("layout_mode") or "grid").strip().lower()
        if plan_layout_mode not in {"grid", "canvas"}:
            plan_layout_mode = "grid"
        plan_canvas_config = analysis.get("canvas_config") if isinstance(analysis.get("canvas_config"), dict) else None
        plan_theme_config = normalize_dashboard_theme_config(
            analysis.get("theme_config") if isinstance(analysis.get("theme_config"), dict) else None
        )
        dashboard_kwargs: Dict[str, Any] = {
            "name": resolved_dashboard_name,
            "description": "Imported from HTML dashboard layout",
            "owner_id": current_user.id,
            "filters_config": [],
            "public_filters_config": [],
            "pages_config": [{"id": page_id, "name": import_page_name}],
            "layout_mode": plan_layout_mode,
        }
        if plan_canvas_config is not None:
            dashboard_kwargs["canvas_config"] = plan_canvas_config
        if plan_theme_config is not None:
            dashboard_kwargs["theme_config"] = plan_theme_config
        # Filter placement is part of the imported report, not an afterthought:
        # a source built around a left rail must not land as a top bar. Stored
        # on its own column, so it has to be carried across explicitly.
        plan_dock = analysis.get("slicer_cluster_layout")
        if isinstance(plan_dock, dict) and plan_dock:
            dashboard_kwargs["slicer_cluster_layout"] = plan_dock
        plan_slicers = analysis.get("slicers")
        if isinstance(plan_slicers, list) and plan_slicers:
            dashboard_kwargs["slicers_config"] = plan_slicers
        dashboard_obj = Dashboard(**dashboard_kwargs)
        db.add(dashboard_obj)
        db.flush()

    # First, persist DashboardChart rows for real charts (legacy path).
    for chart_obj, plan in zip(created_charts, chart_plans_only):
        layout = dict(plan.get("layout") or {})
        widget_type = _plan_widget_type(plan)  # "chart"
        widget_config = normalize_dashboard_widget_config(
            widget_type,
            plan.get("widget_config") if isinstance(plan.get("widget_config"), dict) else None,
        )
        custom_title = _normalize_text(plan.get("title"), max_len=160) or chart_obj.name
        dashboard_chart = DashboardChart(
            dashboard_id=dashboard_obj.id,
            chart_id=chart_obj.id,
            widget_type=widget_type,
            widget_config=widget_config,
            layout=_materialize_dashboard_layout(
                layout,
                page_id=page_id,
                default_w=6,
                default_h=4,
                custom_title=custom_title,
            ),
            parameters={},
        )
        db.add(dashboard_chart)

    # Then persist DashboardChart rows for non-chart widgets (text/image/etc).
    # No Chart row exists for these, so chart_id stays NULL.
    for plan in widget_plans:
        layout = dict(plan.get("layout") or {})
        widget_type = _plan_widget_type(plan)
        widget_config = normalize_dashboard_widget_config(
            widget_type,
            plan.get("widget_config") if isinstance(plan.get("widget_config"), dict) else {},
        )
        widget_title = _normalize_text(plan.get("title"), max_len=160) or widget_type.title()
        dashboard_chart = DashboardChart(
            dashboard_id=dashboard_obj.id,
            chart_id=None,
            widget_type=widget_type,
            widget_config=widget_config,
            layout=_materialize_dashboard_layout(
                layout,
                page_id=page_id,
                default_w=6,
                default_h=4,
                custom_title=widget_title,
            ),
            parameters={},
        )
        db.add(dashboard_chart)

    db.commit()
    db.refresh(dashboard_obj)

    # Auto-generate the full semantic model (with FK joins detection) for the
    # target dataset. Without this step, charts created via build only get a
    # bare-bones SemanticModel/Views (sync_dataset_model_structure path) and
    # cross-table joins are NOT discovered — users must manually click "regenerate
    # model" later. We trigger the full pipeline once per build so the data model
    # auto-connects after import.
    if resolved_dataset_id is not None:
        try:
            from app.services.dataset_model_service import generate_dataset_model
            generate_dataset_model(db, int(resolved_dataset_id), force=False)
            db.commit()
        except Exception:
            # Auto-join detection is best-effort — failing here must not roll
            # back the dashboard import. The user can still call generate_model
            # manually if joins are missing.
            logger.exception(
                "Auto generate_dataset_model after import failed for dataset %s",
                resolved_dataset_id,
            )
            db.rollback()
            db.refresh(dashboard_obj)

    from app.services.dashboard_service import DashboardService

    hydrated_dashboard = DashboardService.get_by_id(db, dashboard_obj.id)
    return {
        "dashboard": hydrated_dashboard,
        "dashboard_id": dashboard_obj.id,
        "created_chart_count": len(created_charts),
        "created_widget_count": len(widget_plans),
        "layout_mode": dashboard_obj.layout_mode,
        "type_changes": type_changes,
        "page_id": page_id,
        "page_name": import_page_name,
        "dataset_id": resolved_dataset_id,
        "dataset_table_id": resolved_dataset_table_id,
        "dataset_table_ids": table_id_map if table_id_map else None,
    }


def build_dashboard_from_import_batch(
    db: Session,
    *,
    current_user: User,
    documents: List[Dict[str, Any]],
    source_mode: SourceMode,
    dataset_table_id: Optional[int],
    dataset_id: Optional[int] = None,
    source_bytes: Optional[bytes],
    source_filename: Optional[str],
    source_files: Optional[List[Tuple[bytes, Optional[str]]]] = None,
    selected_sheet_name: Optional[str],
    dashboard_name: Optional[str],
    target_mode: TargetMode,
    target_dashboard_id: Optional[int],
) -> Dict[str, Any]:
    """Build multiple HTML analyses into multiple dashboard pages.

    The first document uses the existing single-document build flow unchanged.
    Subsequent documents append pages into the dashboard created or targeted by
    the first document, preserving the legacy contract and behavior.
    """
    normalized_documents = [item for item in (documents or []) if isinstance(item, dict)]
    if not normalized_documents:
        raise ValueError("At least one analyzed document is required for batch build.")

    page_results: List[Dict[str, Any]] = []
    current_target_mode: TargetMode = target_mode
    current_target_dashboard_id = target_dashboard_id
    effective_source_mode: SourceMode = source_mode
    effective_dataset_id = dataset_id
    effective_dataset_table_id = dataset_table_id
    effective_source_bytes = source_bytes
    effective_source_filename = source_filename
    effective_source_files = source_files
    final_result: Optional[Dict[str, Any]] = None
    used_page_names: Set[str] = set()

    for index, raw_document in enumerate(normalized_documents, start=1):
        analysis = raw_document.get("analysis")
        if not isinstance(analysis, dict):
            raise ValueError(f"documents[{index - 1}].analysis must be an object.")

        document_id = _normalize_text(raw_document.get("document_id"), max_len=120) or f"document-{index}"
        page_name = _unique_page_name(
            _resolve_batch_page_name(
                explicit_name=raw_document.get("page_name"),
                document_title=analysis.get("document_title"),
                suggested_dashboard_name=analysis.get("suggested_dashboard_name"),
                filename=raw_document.get("filename"),
                fallback_index=index,
            ),
            used_page_names,
        )
        included_block_ids = raw_document.get("included_block_ids") or []
        if not isinstance(included_block_ids, list) or any(not isinstance(item, str) for item in included_block_ids):
            raise ValueError(f"documents[{index - 1}].included_block_ids must be a string array when provided.")

        available_plans = [plan for plan in (analysis.get("chart_plans") or []) if isinstance(plan, dict)]
        if not available_plans:
            logger.info(
                "Skipping batch HTML document %s because it contains no chart_plans.",
                document_id,
            )
            continue
        if included_block_ids:
            selected_ids = {str(item) for item in included_block_ids}
            if not any(str(plan.get("block_id")) in selected_ids for plan in available_plans):
                logger.info(
                    "Skipping batch HTML document %s because no selected blocks remain after filtering.",
                    document_id,
                )
                continue

        per_document_analysis = dict(analysis)
        per_document_analysis["document_title"] = page_name

        build_result = build_dashboard_from_import(
            db,
            current_user=current_user,
            analysis=per_document_analysis,
            source_mode=effective_source_mode,
            dataset_table_id=effective_dataset_table_id,
            dataset_id=effective_dataset_id,
            source_bytes=effective_source_bytes,
            source_filename=effective_source_filename,
            source_files=effective_source_files,
            selected_sheet_name=selected_sheet_name,
            dashboard_name=dashboard_name if index == 1 else page_name,
            target_mode=current_target_mode,
            target_dashboard_id=current_target_dashboard_id,
            included_block_ids=included_block_ids,
        )
        final_result = build_result
        current_target_mode = "append_to_dashboard"
        current_target_dashboard_id = build_result["dashboard_id"]
        effective_source_mode = "existing_dataset"
        effective_dataset_id = build_result.get("dataset_id")
        effective_dataset_table_id = build_result.get("dataset_table_id")
        effective_source_bytes = None
        effective_source_filename = None
        effective_source_files = None

        page_results.append(
            {
                "document_id": document_id,
                "filename": _normalize_text(raw_document.get("filename"), max_len=260) or None,
                "page_id": build_result["page_id"],
                "page_name": build_result["page_name"],
                "created_chart_count": build_result["created_chart_count"],
                "type_changes": build_result.get("type_changes") or [],
            }
        )

    if final_result is None:
        raise ValueError("No batch document produced any chart plans to build.")
    total_chart_count = sum(int(item.get("created_chart_count") or 0) for item in page_results)
    return {
        "dashboard": final_result["dashboard"],
        "dashboard_id": final_result["dashboard_id"],
        "created_chart_count": total_chart_count,
        "pages": page_results,
        "dataset_id": final_result.get("dataset_id"),
        "dataset_table_id": final_result.get("dataset_table_id"),
        "dataset_table_ids": final_result.get("dataset_table_ids"),
    }


# ---------------------------------------------------------------------------
# Validate chart plans — dry-run queries before building
# ---------------------------------------------------------------------------


class _VirtualTable:
    """Lightweight wrapper injecting extra transformations without touching the ORM object."""

    def __init__(self, real_table: DatasetTable, extra_transformations: List[Dict[str, Any]]):
        self._real = real_table
        self.transformations = list(real_table.transformations or []) + extra_transformations

    def __getattr__(self, name: str) -> Any:
        return getattr(self._real, name)


def validate_chart_plans(
    db: Session,
    *,
    current_user: User,
    dataset_id: int,
    chart_plans: List[Dict[str, Any]],
    calculated_fields: List[Dict[str, Any]],
    derived_tables: Optional[List[Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """Dry-run each chart plan's query against the dataset datasource and return per-block results.

    Returns a list of ``{"block_id": str, "status": "ok"|"error", "error": str|None}``.
    """
    from app.models.models import DataSource
    from app.services.datasource_service import DataSourceConnectionService
    from app.services.live_query_service import (
        _dialect_for_ds_type,
        build_live_agg_query,
    )

    tables = (
        db.query(DatasetTable)
        .filter(
            DatasetTable.dataset_id == dataset_id,
            DatasetTable.source_kind != "generated_calendar",
        )
        .all()
    )
    if not tables:
        return [
            {"block_id": p.get("block_id", ""), "status": "error", "error": "No data tables found in dataset."}
            for p in chart_plans
        ]

    table_map: Dict[str, DatasetTable] = {t.display_name: t for t in tables}
    default_table = tables[0]

    datasource = db.query(DataSource).filter(DataSource.id == default_table.datasource_id).first()
    if not datasource:
        return [
            {"block_id": p.get("block_id", ""), "status": "error", "error": "DataSource not found."}
            for p in chart_plans
        ]

    ds_type = datasource.type if isinstance(datasource.type, str) else datasource.type.value
    dialect = _dialect_for_ds_type(ds_type)

    prepared_plans: List[Dict[str, Any]] = [dict(plan) for plan in (chart_plans or [])]
    if derived_tables:
        derived_warnings = _materialize_v1_derived_tables(
            db,
            derived_tables=derived_tables,
            chart_plans=prepared_plans,
            dataset_id=dataset_id,
            table_id_map=_build_table_id_map_from_dataset_tables(tables),
        )
        if derived_warnings:
            logger.info(
                "Validation materialized derived tables with %d warning(s).",
                len(derived_warnings),
            )

    # Group calculated fields by source_key so we inject them into the right table
    calc_by_source: Dict[str, List[Dict[str, Any]]] = {}
    for cf in calculated_fields or []:
        sk = cf.get("source_key") or "__default__"
        calc_by_source.setdefault(sk, []).append(cf)

    def _extra_transforms_for(table_display_name: str) -> List[Dict[str, Any]]:
        items = list(calc_by_source.get(table_display_name, []))
        items.extend(calc_by_source.get("__default__", []))
        return [
            {
                "type": "add_column",
                "enabled": True,
                "params": {"newField": cf["name"], "expression": cf["expression"]},
            }
            for cf in items
        ]

    results: List[Dict[str, Any]] = []
    for plan in prepared_plans:
        block_id = plan.get("block_id", "")
        try:
            source_key = plan.get("source_key")
            real_table = table_map.get(source_key, default_table) if source_key else default_table

            # Manual edit may have picked a different table within the dataset.
            override_id = plan.get("dataset_table_id_override")
            if isinstance(override_id, int) and override_id > 0:
                override_real = next((t for t in tables if t.id == override_id), None)
                if override_real is not None:
                    real_table = override_real

            chart_type = plan.get("final_chart_type", "TABLE")
            role_config = dict(plan.get("role_config") or {})
            custom_role_config = dict(plan.get("custom_role_config") or {})
            raw_query_mode = str(plan.get("query_mode") or "generated").strip().lower()
            custom_sql = str(plan.get("custom_sql") or "").strip()
            use_custom = raw_query_mode == "custom" and bool(custom_sql)

            if use_custom:
                # Custom SQL path: wrap the user SQL as the base table and apply the
                # chart-level aggregation on top, mirroring the runtime flow in
                # LiveQueryService.execute_chart_query_from_sql.
                from app.services.live_query_service import validate_select_only

                validate_select_only(custom_sql)
                base_table = f"({custom_sql.rstrip(';').strip()}) AS _appbi_live"
                active_role_config = custom_role_config or role_config
                agg_sql, _ = build_live_agg_query(
                    base_table, chart_type, active_role_config, [], dialect, limit_override=5,
                )
                logger.info(
                    "validate_chart_plan (custom SQL) block_id=%s table=%s agg_sql=%s",
                    block_id, real_table.display_name, agg_sql[:300],
                )
            else:
                virtual = _VirtualTable(real_table, _extra_transforms_for(real_table.display_name))
                from app.services.dataset_relation_service import resolve_dataset_table_relation
                base_plan = resolve_dataset_table_relation(datasource, virtual)
                base_table = f"({base_plan.sql}) AS _appbi_live"
                agg_sql, _ = build_live_agg_query(
                    base_table, chart_type, role_config, [], dialect, limit_override=5,
                )
                logger.info(
                    "validate_chart_plan block_id=%s source_key=%s table=%s base_sql=%s agg_sql=%s",
                    block_id, source_key, real_table.display_name, base_plan.sql[:200], agg_sql[:300],
                )

            DataSourceConnectionService.execute_query(ds_type, datasource.config, agg_sql, limit=5)
            results.append({"block_id": block_id, "status": "ok", "error": None})
        except Exception as exc:
            logger.error("validate_chart_plan FAILED block_id=%s source_key=%s error=%s", block_id, source_key, str(exc))
            results.append({"block_id": block_id, "status": "error", "error": str(exc)})

    return results


# ---------------------------------------------------------------------------
# AI fix a single broken chart plan
# ---------------------------------------------------------------------------


def ai_fix_chart_plan(
    *,
    chart_plan: Dict[str, Any],
    error_message: str,
    source_profile: Dict[str, Any],
    all_source_profiles: Optional[Dict[str, Dict[str, Any]]] = None,
) -> Optional[Dict[str, Any]]:
    """Ask the configured AI provider to repair a chart plan that failed validation.

    Returns the corrected plan dict (same shape) or ``None`` on failure.
    """
    if settings.html_import_ai_provider == "unavailable":
        return None

    # Build column inventory — include ALL tables so AI can reassign source_key
    plan_source_key = chart_plan.get("source_key")
    current_profile = source_profile
    if plan_source_key and all_source_profiles and plan_source_key in all_source_profiles:
        current_profile = all_source_profiles[plan_source_key]

    # Build per-table column maps
    tables_info: Dict[str, Any] = {}
    if all_source_profiles and len(all_source_profiles) > 1:
        for tbl_key, prof in all_source_profiles.items():
            tables_info[tbl_key] = {
                "columns": [{"name": c.get("name"), "type": c.get("type")} for c in (prof.get("columns") or [])],
                "numeric_columns": prof.get("numeric_columns", []),
                "date_columns": prof.get("date_columns", []),
            }
    else:
        tbl_key = plan_source_key or "__default__"
        tables_info[tbl_key] = {
            "columns": [{"name": c.get("name"), "type": c.get("type")} for c in (current_profile.get("columns") or [])],
            "numeric_columns": current_profile.get("numeric_columns", []),
            "date_columns": current_profile.get("date_columns", []),
        }

    has_multi_tables = len(tables_info) > 1

    rules = [
        "Return the FULL fixed chart plan in the same shape as chart_plan above.",
        "field_mapping is derived from role_config — only return role_config.",
        "If a referenced field does not exist in the current table, first check if another table has those fields and change source_key to that table.",
        "Only if no table has the fields, replace them with the closest matching column from the current table.",
        "If no reasonable fix is possible, change final_chart_type to TABLE and use selectedColumns from available_columns.",
        "Keep block_id and title unchanged.",
        "Return JSON only.",
    ]
    if has_multi_tables:
        rules.insert(2, f"The dataset has multiple tables: {list(tables_info.keys())}. You MAY change source_key to reassign the chart to the correct table.")

    prompt = json.dumps(
        {
            "task": (
                "A chart plan failed validation against the data source. "
                "Fix the chart plan so the query succeeds. "
                "You may change role_config fields, aggregations, chart type, or source_key (table assignment)."
            ),
            "supported_chart_types": sorted(_SUPPORTED_CHART_TYPES),
            "error_message": error_message,
            "chart_plan": {
                "block_id": chart_plan.get("block_id"),
                "title": chart_plan.get("title"),
                "final_chart_type": chart_plan.get("final_chart_type"),
                "role_config": chart_plan.get("role_config"),
                "source_key": chart_plan.get("source_key"),
            },
            "tables": tables_info,
            "role_config_shapes": {
                "BAR|HORIZONTAL_BAR|LINE|PIE|AREA|STACKED_BAR|GROUPED_BAR": {"dimension": "field_name", "metrics": [{"field": "field_name", "agg": "sum|count|avg|min|max"}], "breakdown": "optional_field"},
                "KPI": {"metrics": [{"field": "field_name", "agg": "sum|count|avg|min|max"}]},
                "TABLE": {"selectedColumns": ["col1", "col2"], "metrics": []},
                "SCATTER": {"scatterX": "field_name", "scatterY": "field_name", "metrics": []},
                "BAR_LINE": {"dimension": "field_name", "metrics": [{"field": "f", "agg": "sum"}], "lineMetric": {"field": "f", "agg": "avg"}},
                "TIME_SERIES": {"timeField": "date_field", "metrics": [{"field": "f", "agg": "sum"}]},
            },
            "rules": rules,
            "return_json_shape": {
                "block_id": "string",
                "title": "string",
                "final_chart_type": "string",
                "role_config": {},
                "source_key": "string | null",
                "fix_note": "string (brief explanation of what was changed)",
            },
        },
        ensure_ascii=False,
    )

    system_prompt = (
        "You are an expert at fixing broken BI chart configurations. "
        "Return only a single JSON object with the corrected chart plan. "
        "Do not add commentary outside the JSON."
    )

    result = _complete_json_with_import_provider(prompt, system_prompt=system_prompt, max_tokens=1200)
    logger.info("ai_fix_chart_plan block_id=%s AI_INPUT=%s", chart_plan.get("block_id"), prompt[:500])
    logger.info("ai_fix_chart_plan block_id=%s AI_OUTPUT=%s", chart_plan.get("block_id"), json.dumps(result, ensure_ascii=False)[:500] if result else "None")
    if not isinstance(result, dict) or "role_config" not in result:
        return None

    # Merge the fix back onto the original plan so callers get a complete object
    fixed = dict(chart_plan)
    fixed["final_chart_type"] = result.get("final_chart_type", chart_plan.get("final_chart_type"))
    fixed["role_config"] = result["role_config"]
    fixed["fix_note"] = result.get("fix_note")
    # Allow AI to reassign to a different table
    if result.get("source_key"):
        fixed["source_key"] = result["source_key"]
    # Re-derive source_fields_used from the new role_config
    rc = fixed["role_config"]
    fields: List[str] = []
    for key in ("dimension", "timeField", "breakdown", "scatterX", "scatterY"):
        val = rc.get(key)
        if val:
            fields.append(val)
    for m in rc.get("metrics") or []:
        if isinstance(m, dict) and m.get("field"):
            fields.append(m["field"])
    lm = rc.get("lineMetric")
    if isinstance(lm, dict) and lm.get("field"):
        fields.append(lm["field"])
    for col in rc.get("selectedColumns") or []:
        fields.append(col)
    fixed["source_fields_used"] = list(dict.fromkeys(fields))
    return fixed


# ---------------------------------------------------------------------------
# Verbatim dashboard snapshot (Duplicate + Export/Import round-trip)
#
# Unlike the fuzzy HTML analyzer above, a "dashboard snapshot" captures the
# LIVE dashboard structure exactly — each chart's chart_type + config +
# dataset_table_id and each tile's layout — and rebuilds it row-for-row with
# NO column matching and NO AI. This single core powers three features:
#   * Duplicate   = serialize_dashboard_snapshot -> rebuild_dashboard_from_snapshot (in-process)
#   * Export HTML = serialize_dashboard_snapshot -> export_dashboard_html
#   * Import HTML = _extract_embedded_metadata -> rebuild_dashboard_from_snapshot
# Scope is same-instance: the embedded dataset_table_ids must still resolve.
# Cross-instance imports get a clear 422 rather than a silent mis-binding.
# ---------------------------------------------------------------------------

APPBI_SNAPSHOT_VERSION = "appbi-snapshot/v1"


def is_dashboard_snapshot(embedded: Optional[Dict[str, Any]]) -> bool:
    """True when an extracted embedded-metadata payload is an appbi snapshot."""
    return (
        isinstance(embedded, dict)
        and str(embedded.get("version") or "").strip().lower() == APPBI_SNAPSHOT_VERSION
    )


def _serialize_chart_for_snapshot(chart: Chart) -> Dict[str, Any]:
    """Capture a Chart definition verbatim (rendering + semantic surface)."""
    chart_type = chart.chart_type.value if isinstance(chart.chart_type, ChartType) else str(chart.chart_type)
    payload: Dict[str, Any] = {
        "name": chart.name,
        "description": chart.description,
        "chart_type": chart_type,
        "dataset_table_id": chart.dataset_table_id,
        "config": chart.config or {},
    }

    meta = getattr(chart, "chart_meta", None)
    if meta is not None:
        payload["meta"] = {
            "domain": meta.domain,
            "intent": meta.intent,
            "metrics": meta.metrics or [],
            "dimensions": meta.dimensions or [],
            "tags": meta.tags or [],
        }

    params = getattr(chart, "parameters", None) or []
    if params:
        payload["parameters"] = [
            {
                "parameter_name": p.parameter_name,
                "parameter_type": p.parameter_type,
                "column_mapping": p.column_mapping,
                "default_value": p.default_value,
                "description": p.description,
            }
            for p in params
        ]
    return payload


def serialize_dashboard_snapshot(dashboard: Dashboard) -> Dict[str, Any]:
    """Serialize a live dashboard into a portable, verbatim snapshot dict.

    The dashboard must be loaded with its ``dashboard_charts`` (and each
    tile's ``chart``) — ``DashboardService.get_by_id`` already eager-loads
    that graph. No DB access happens here.
    """
    tiles: List[Dict[str, Any]] = []
    for dc in dashboard.dashboard_charts or []:
        widget_type = str(dc.widget_type or "chart").strip().lower() or "chart"
        tile: Dict[str, Any] = {
            "dashboard_chart_id": dc.id,
            "widget_type": widget_type,
            "widget_config": dc.widget_config or {},
            "parameters": dc.parameters or {},
            "layout": dc.layout or {},
        }
        if widget_type == "chart" and dc.chart is not None:
            tile["chart"] = _serialize_chart_for_snapshot(dc.chart)
        else:
            tile["chart"] = None
        tiles.append(tile)

    return {
        "version": APPBI_SNAPSHOT_VERSION,
        "kind": "dashboard-snapshot",
        "source_dashboard_id": dashboard.id,
        "dashboard": {
            "name": dashboard.name,
            "description": dashboard.description,
            "pages_config": dashboard.pages_config or [DEFAULT_DASHBOARD_PAGE],
            "layout_mode": dashboard.layout_mode or "grid",
            "canvas_config": dashboard.canvas_config or None,
            "theme_config": dashboard.theme_config or None,
            "filters_config": dashboard.filters_config or [],
            "slicers_config": dashboard.slicers_config or [],
            "slicer_cluster_layout": dashboard.slicer_cluster_layout or None,
            "public_filters_config": dashboard.public_filters_config or [],
        },
        "charts": tiles,
    }


def export_dashboard_html(dashboard: Dashboard) -> str:
    """Render a self-contained, re-importable HTML document for a dashboard.

    The machine-readable snapshot lives in a
    ``<script type="application/appbi-dashboard">`` tag (the same tag the
    importer scans). The visible body is a lightweight, readable outline of
    the pages and tiles — this is a round-trip carrier, not a pixel render.
    """
    snapshot = serialize_dashboard_snapshot(dashboard)
    snapshot_json = json.dumps(snapshot, ensure_ascii=False, indent=2)

    def esc(value: Any) -> str:
        return html_module.escape(str(value if value is not None else ""))

    pages = snapshot["dashboard"]["pages_config"] or [DEFAULT_DASHBOARD_PAGE]
    tiles = snapshot["charts"]
    tiles_by_page: Dict[str, List[Dict[str, Any]]] = {}
    for tile in tiles:
        layout = tile.get("layout") or {}
        page_id = str(layout.get("pageId") or (pages[0].get("id") if pages else "page-1"))
        tiles_by_page.setdefault(page_id, []).append(tile)

    body_sections: List[str] = []
    for page in pages:
        page_id = str(page.get("id") or "page-1")
        page_name = esc(page.get("name") or page_id)
        rows: List[str] = []
        for tile in tiles_by_page.get(page_id, []):
            layout = tile.get("layout") or {}
            chart = tile.get("chart")
            if chart:
                title = esc((layout.get("custom_title")) or chart.get("name") or "Chart")
                kind = esc(chart.get("chart_type") or "CHART")
                binding = f"dataset_table #{esc(chart.get('dataset_table_id'))}"
            else:
                title = esc((tile.get("widget_config") or {}).get("title") or tile.get("widget_type") or "Widget")
                kind = esc(str(tile.get("widget_type") or "widget").upper())
                binding = "—"
            pos = (
                f"x{esc(layout.get('x', 0))} y{esc(layout.get('y', 0))} "
                f"w{esc(layout.get('w', 0))} h{esc(layout.get('h', 0))}"
            )
            rows.append(
                "<tr>"
                f"<td>{title}</td><td><code>{kind}</code></td>"
                f"<td>{binding}</td><td class='pos'>{pos}</td>"
                "</tr>"
            )
        table = (
            "<table><thead><tr><th>Tile</th><th>Type</th>"
            "<th>Data binding</th><th>Position</th></tr></thead>"
            f"<tbody>{''.join(rows) or '<tr><td colspan=4>(empty)</td></tr>'}</tbody></table>"
        )
        body_sections.append(f"<section><h2>{page_name}</h2>{table}</section>")

    name = esc(snapshot["dashboard"]["name"])
    description = esc(snapshot["dashboard"]["description"])
    return (
        "<!DOCTYPE html>\n"
        '<html lang="en">\n<head>\n<meta charset="utf-8" />\n'
        f"<title>{name} — AppBI dashboard</title>\n"
        "<style>"
        "body{font-family:system-ui,-apple-system,'Segoe UI',sans-serif;margin:2rem;color:#1e293b;background:#f8fafc}"
        "h1{margin:0 0 .25rem}p.desc{color:#64748b;margin:.25rem 0 1.5rem}"
        ".note{background:#eef2ff;border:1px solid #c7d2fe;border-radius:8px;padding:.75rem 1rem;color:#3730a3;font-size:.9rem;margin-bottom:1.5rem}"
        "section{margin-bottom:1.5rem}h2{font-size:1.05rem;border-bottom:2px solid #e2e8f0;padding-bottom:.35rem}"
        "table{border-collapse:collapse;width:100%;font-size:.9rem}"
        "th,td{border:1px solid #e2e8f0;padding:.4rem .6rem;text-align:left}"
        "th{background:#f1f5f9}code{background:#f1f5f9;padding:.1rem .35rem;border-radius:4px}"
        "td.pos{font-variant-numeric:tabular-nums;color:#64748b}"
        "</style>\n</head>\n<body>\n"
        f"<h1>{name}</h1>\n"
        f"<p class='desc'>{description}</p>\n"
        "<div class='note'>This file was exported from AppBI and can be re-imported to create a new "
        "dashboard. The full definition is stored in the embedded snapshot below — you may edit "
        "values inside that script tag, but keep it valid JSON.</div>\n"
        f"{''.join(body_sections)}\n"
        f'<script type="application/appbi-dashboard">\n{snapshot_json}\n</script>\n'
        "</body>\n</html>\n"
    )


def rebuild_dashboard_from_snapshot(
    db: Session,
    *,
    snapshot: Dict[str, Any],
    current_user: User,
    name_override: Optional[str] = None,
) -> Dashboard:
    """Rebuild a dashboard row-for-row from a verbatim snapshot.

    No analyzer, no column matching: charts are recreated from the embedded
    ``chart_type`` + ``config`` + ``dataset_table_id``. Each referenced
    dataset table must exist in this instance and be viewable by the caller;
    otherwise a ValueError (surfaced as 422) lists the unresolved ids.
    """
    if not is_dashboard_snapshot(snapshot):
        raise ValueError("Not an AppBI dashboard snapshot (missing appbi-snapshot/v1 version).")

    dash_meta = snapshot.get("dashboard") if isinstance(snapshot.get("dashboard"), dict) else {}
    tiles = snapshot.get("charts") if isinstance(snapshot.get("charts"), list) else []

    # --- Validate every referenced dataset table up-front (fail closed). ---
    referenced_ids: Set[int] = set()
    for tile in tiles:
        chart = tile.get("chart") if isinstance(tile, dict) else None
        if isinstance(chart, dict) and isinstance(chart.get("dataset_table_id"), int):
            referenced_ids.add(int(chart["dataset_table_id"]))

    missing_ids: List[int] = []
    forbidden_ids: List[int] = []
    for table_id in sorted(referenced_ids):
        table = db.query(DatasetTable).filter(DatasetTable.id == table_id).first()
        if table is None:
            missing_ids.append(table_id)
            continue
        dataset = db.query(Dataset).filter(Dataset.id == table.dataset_id).first()
        if dataset is None:
            missing_ids.append(table_id)
            continue
        try:
            require_view_access(db, current_user, dataset, "datasets")
        except Exception:
            forbidden_ids.append(table_id)

    if missing_ids or forbidden_ids:
        parts: List[str] = []
        if missing_ids:
            parts.append(f"not found in this instance: {missing_ids}")
        if forbidden_ids:
            parts.append(f"not accessible to you: {forbidden_ids}")
        raise ValueError(
            "Cannot import snapshot — referenced dataset tables " + "; ".join(parts) + "."
        )

    # --- Create the dashboard shell. ---
    requested_name = _normalize_text(name_override or dash_meta.get("name"), max_len=200) or "Imported Dashboard"
    resolved_name = _unique_dashboard_name(db, requested_name)

    layout_mode = str(dash_meta.get("layout_mode") or "grid").strip().lower()
    if layout_mode not in {"grid", "canvas"}:
        layout_mode = "grid"

    pages_config = dash_meta.get("pages_config") if isinstance(dash_meta.get("pages_config"), list) else None
    if not pages_config:
        pages_config = [dict(DEFAULT_DASHBOARD_PAGE)]

    dashboard_kwargs: Dict[str, Any] = {
        "name": resolved_name,
        "description": _normalize_text(dash_meta.get("description"), max_len=1024) or None,
        "owner_id": current_user.id,
        "pages_config": pages_config,
        "layout_mode": layout_mode,
        "filters_config": dash_meta.get("filters_config") if isinstance(dash_meta.get("filters_config"), list) else [],
        "slicers_config": dash_meta.get("slicers_config") if isinstance(dash_meta.get("slicers_config"), list) else [],
        "public_filters_config": dash_meta.get("public_filters_config") if isinstance(dash_meta.get("public_filters_config"), list) else [],
    }
    if isinstance(dash_meta.get("slicer_cluster_layout"), dict):
        dashboard_kwargs["slicer_cluster_layout"] = dash_meta["slicer_cluster_layout"]
    if isinstance(dash_meta.get("canvas_config"), dict):
        dashboard_kwargs["canvas_config"] = dash_meta["canvas_config"]
    theme_config = normalize_dashboard_theme_config(
        dash_meta.get("theme_config") if isinstance(dash_meta.get("theme_config"), dict) else None
    )
    if theme_config is not None:
        dashboard_kwargs["theme_config"] = theme_config

    dashboard_obj = Dashboard(**dashboard_kwargs)
    db.add(dashboard_obj)
    db.flush()

    # --- Recreate each tile verbatim. ---
    for tile in tiles:
        if not isinstance(tile, dict):
            continue
        widget_type = str(tile.get("widget_type") or "chart").strip().lower() or "chart"
        layout = tile.get("layout") if isinstance(tile.get("layout"), dict) else {}
        chart_spec = tile.get("chart") if isinstance(tile.get("chart"), dict) else None

        if widget_type == "chart" and chart_spec is not None:
            try:
                chart_type = ChartType(str(chart_spec.get("chart_type") or "TABLE").upper())
            except ValueError:
                chart_type = ChartType.TABLE
            chart_name = _unique_chart_name(
                db,
                current_user.id,
                _normalize_text(chart_spec.get("name"), max_len=255) or "Imported Chart",
            )
            new_chart = Chart(
                name=chart_name,
                description=_normalize_text(chart_spec.get("description"), max_len=1024) or None,
                dataset_table_id=chart_spec.get("dataset_table_id"),
                chart_type=chart_type,
                config=chart_spec.get("config") or {},
                owner_id=current_user.id,
            )
            db.add(new_chart)
            db.flush()

            meta = chart_spec.get("meta") if isinstance(chart_spec.get("meta"), dict) else None
            if meta:
                db.add(
                    ChartMetadata(
                        chart_id=new_chart.id,
                        domain=meta.get("domain"),
                        intent=meta.get("intent"),
                        metrics=meta.get("metrics") or [],
                        dimensions=meta.get("dimensions") or [],
                        tags=meta.get("tags") or [],
                    )
                )
            for p in chart_spec.get("parameters") or []:
                if not isinstance(p, dict) or not p.get("parameter_name") or not p.get("parameter_type"):
                    continue
                db.add(
                    ChartParameter(
                        chart_id=new_chart.id,
                        parameter_name=str(p["parameter_name"]),
                        parameter_type=str(p["parameter_type"]),
                        column_mapping=p.get("column_mapping"),
                        default_value=p.get("default_value"),
                        description=p.get("description"),
                    )
                )

            db.add(
                DashboardChart(
                    dashboard_id=dashboard_obj.id,
                    chart_id=new_chart.id,
                    widget_type="chart",
                    widget_config=tile.get("widget_config") if isinstance(tile.get("widget_config"), dict) else None,
                    layout=layout,
                    parameters=tile.get("parameters") if isinstance(tile.get("parameters"), dict) else {},
                )
            )
        else:
            # Non-chart widget (text/image/shape/countdown/parameter_switcher).
            db.add(
                DashboardChart(
                    dashboard_id=dashboard_obj.id,
                    chart_id=None,
                    widget_type=widget_type,
                    widget_config=normalize_dashboard_widget_config(
                        widget_type,
                        tile.get("widget_config") if isinstance(tile.get("widget_config"), dict) else {},
                    ),
                    layout=layout,
                    parameters={},
                )
            )

    db.commit()
    db.refresh(dashboard_obj)

    from app.services.dashboard_service import DashboardService

    return DashboardService.get_by_id(db, dashboard_obj.id)
