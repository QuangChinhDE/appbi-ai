"""Excel layout helpers shared by template import/export services."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Constants ──────────────────────────────────────────────────────────

# Full-screen canvas: blocks scale to ~1200 px wide with small margin
CANVAS_W = 1200
MARGIN = 24
USABLE_W = CANVAS_W - 2 * MARGIN

# Excel → pixel conversion factors
DEFAULT_COL_WIDTH_CHARS = 8.43  # Excel default column width (characters)
DEFAULT_ROW_HEIGHT_PT = 15.0     # Excel default row height (points)
CHAR_TO_PX = 7.5                 # 1 character ≈ 7.5 px at default font
PT_TO_PX = 1.333                 # 1 pt ≈ 1.333 px at 96 DPI

# Content-aware sizing constants (matching frontend text-xs = 12px font)
FONT_PX = 12                     # font-size in the table renderer
CHAR_WIDTH_PX = 7.0              # average character width at 12px
LINE_HEIGHT_PX = 18              # line-height for text-xs
CELL_PAD_Y = 8                   # py-1 = 4px top + 4px bottom
MIN_ROW_HEIGHT = LINE_HEIGHT_PX + CELL_PAD_Y  # minimum height for one line
MIN_COL_WIDTH_PX = 40            # minimum column width (px)

GAP_BETWEEN_BLOCKS = 8          # px gap between stacked blocks


# ── Helpers ────────────────────────────────────────────────────────────

def _col_width_px(ws: Worksheet, col_idx: int) -> float:
    """Return the width in pixels for a 1-based column index."""
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    if dim and dim.width is not None and dim.width > 0:
        return dim.width * CHAR_TO_PX
    return DEFAULT_COL_WIDTH_CHARS * CHAR_TO_PX


def _row_height_px(ws: Worksheet, row_idx: int) -> float:
    """Return the height in pixels for a 1-based row index."""
    dim = ws.row_dimensions.get(row_idx)
    if dim and dim.height is not None and dim.height > 0:
        return dim.height * PT_TO_PX
    return DEFAULT_ROW_HEIGHT_PT * PT_TO_PX


def _color_hex(color) -> Optional[str]:
    """Extract a #RRGGBB hex string from an openpyxl Color object, or None."""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb and isinstance(color.rgb, str):
        rgb = color.rgb
        # openpyxl sometimes returns AARRGGBB (8 chars)
        if len(rgb) == 8:
            rgb = rgb[2:]
        if len(rgb) == 6 and rgb != "000000":
            return f"#{rgb}"
    return None


def _border_sides(border) -> Optional[Dict[str, bool]]:
    """Extract a simple per-side border map from an openpyxl Border."""
    if border is None:
        return None

    sides: Dict[str, bool] = {}
    for side_name in ("top", "right", "bottom", "left"):
        side = getattr(border, side_name, None)
        if side is not None and getattr(side, "style", None):
            sides[side_name] = True

    return sides or None


def _cell_value_str(cell: Cell) -> str:
    """Return the display string for a cell value."""
    if cell.value is None:
        return ""
    return str(cell.value)


def _estimate_cell_height(text: str, col_width_px: float) -> float:
    """
    Estimate the rendered height of a cell given its text and column width.
    Accounts for text wrapping and explicit newlines.
    """
    if not text:
        return MIN_ROW_HEIGHT

    # Available width for text (subtract cell horizontal padding ~16px)
    avail_w = max(col_width_px - 16, CHAR_WIDTH_PX)
    chars_per_line = max(1, int(avail_w / CHAR_WIDTH_PX))

    total_lines = 0
    for line in text.split("\n"):
        line_len = len(line.strip()) or 1
        wrapped = max(1, -(-line_len // chars_per_line))  # ceil division
        total_lines += wrapped

    return total_lines * LINE_HEIGHT_PX + CELL_PAD_Y


def _normalize_widths_to_total(
    widths: List[float],
    total_width: float,
    min_width: float = MIN_COL_WIDTH_PX,
) -> List[float]:
    """Scale column widths so their sum matches the available width."""
    if not widths:
        return widths

    total_width = max(float(total_width), min_width * len(widths))
    current_total = sum(widths)
    if current_total <= 0:
        even_width = round(total_width / len(widths), 1)
        return [even_width] * len(widths)

    normalized = [max(min_width, (width / current_total) * total_width) for width in widths]
    current_total = sum(normalized)
    overflow = current_total - total_width

    if overflow > 0:
        shrinkable = [i for i, width in enumerate(normalized) if width > min_width]
        while overflow > 0.05 and shrinkable:
            per_col = overflow / len(shrinkable)
            next_shrinkable = []
            reduced = 0.0
            for i in shrinkable:
                available = normalized[i] - min_width
                delta = min(available, per_col)
                normalized[i] -= delta
                reduced += delta
                if normalized[i] - min_width > 0.05:
                    next_shrinkable.append(i)
            if reduced <= 0:
                break
            overflow -= reduced
            shrinkable = next_shrinkable
    elif overflow < -0.05:
        deficit = -overflow
        growable = [i for i in range(len(normalized))]
        per_col = deficit / len(growable)
        for i in growable:
            normalized[i] += per_col

    normalized = [round(width, 1) for width in normalized]
    remainder = round(total_width - sum(normalized), 1)
    if normalized and abs(remainder) >= 0.1:
        normalized[-1] = round(max(min_width, normalized[-1] + remainder), 1)

    return normalized


def _build_merge_map(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """
    Build a dict mapping (row, col) → (rowSpan, colSpan) for every
    top-left anchor of a merged-cell range.  Also build a set of
    cells that are "hidden" by merges (non-anchor cells).
    Returns (anchor_map, hidden_set).
    """
    anchor_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    hidden: set = set()
    for rng in ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        row_span = max_r - min_r + 1
        col_span = max_c - min_c + 1
        anchor_map[(min_r, min_c)] = (row_span, col_span)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    hidden.add((r, c))
    return anchor_map, hidden


def _is_row_empty(ws: Worksheet, row_idx: int, min_col: int, max_col: int) -> bool:
    """Check if a row has no values in the used columns."""
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        if cell.value is not None and str(cell.value).strip():
            return False
    return True


# ── Section splitter ───────────────────────────────────────────────────

def _split_into_sections(
    ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int,
) -> List[Tuple[int, int]]:
    """
    Split the used range into contiguous sections separated by empty rows.
    Returns list of (start_row, end_row) tuples (1-based, inclusive).
    """
    sections: List[Tuple[int, int]] = []
    current_start: Optional[int] = None

    for r in range(min_row, max_row + 1):
        empty = _is_row_empty(ws, r, min_col, max_col)
        if not empty:
            if current_start is None:
                current_start = r
        else:
            if current_start is not None:
                sections.append((current_start, r - 1))
                current_start = None

    # Close last section
    if current_start is not None:
        sections.append((current_start, max_row))

    return sections


# ── Main parser ────────────────────────────────────────────────────────

def _auto_widen_columns(
    ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int,
    col_widths: List[float],
) -> None:
    """
    In-place adjustment: compute the max text length per column and widen
    narrow columns at the expense of wider ones to minimize wrapping.
    Keeps total width constant.
    """
    num_cols = max_col - min_col + 1
    # Compute max char count per column (ignoring merged cells)
    max_chars = [0] * num_cols
    for r in range(min_row, max_row + 1):
        for ci in range(num_cols):
            c = min_col + ci
            cell = ws.cell(row=r, column=c)
            val = str(cell.value) if cell.value is not None else ""
            # Use the longest single line
            for line in val.split("\n"):
                max_chars[ci] = max(max_chars[ci], len(line))

    # Compute desired width per column = max_chars * CHAR_WIDTH_PX + padding
    desired = [max(MIN_COL_WIDTH_PX, chars * CHAR_WIDTH_PX + 16) for chars in max_chars]

    total_current = sum(col_widths)
    total_desired = sum(desired)

    if total_desired <= 0:
        return

    # Scale desired to fit total width
    ratio = total_current / total_desired
    for i in range(num_cols):
        col_widths[i] = round(max(MIN_COL_WIDTH_PX, desired[i] * ratio), 1)


