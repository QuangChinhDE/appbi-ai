"""
Smart Excel Structure Detector — V1 (Rule-based, no AI).

Analyzes an uploaded Excel file and auto-detects:
  - Header zone (company name, report title, exchange rate)
  - Column group rows (merged header cells)
  - Data header row (column names)
  - Data rows + subtotal/group rows
  - Footer zone (notes, signature boxes)

Returns an AnalysisResult dict that the frontend can preview / edit
before confirming template creation.
"""

from __future__ import annotations

from collections import Counter
import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.services.excel_parser import (
    _build_merge_map,
    _color_hex,
    _col_width_px,
    _border_sides,
    _is_row_empty,
    MIN_COL_WIDTH_PX,
)

# ── Constants ─────────────────────────────────────────────────────────

_SIGNATURE_KEYWORDS = {
    "ky ten", "chu ky", "ký tên", "chữ ký", "signature",
    "nguoi lap", "người lập", "prepared by", "made by",
    "ke toan", "kế toán", "accountant",
    "giam doc", "giám đốc", "director",
    "truong phong", "trưởng phòng", "manager",
    "phe duyet", "phê duyệt", "approved",
    "xac nhan", "xác nhận", "confirmed",
}

_META_KEYWORDS = {
    "ty gia", "tỷ giá", "exchange", "rate",
    "ky bao cao", "kỳ báo cáo", "period",
    "ngay", "ngày", "date", "thang", "tháng", "nam", "năm",
}

_FOOTER_KEYWORDS = {
    "ghi chú", "ghi chu", "lưu ý", "luu y", "note", "notes",
}

_TITLE_KEYWORDS = {
    "báo giá", "bao gia", "quotation", "quote",
}

_CURRENCY_PATTERNS = re.compile(
    r'\b(USD|KIP|VND|LAK|THB|EUR|JPY|CNY|đồng|dong)\b', re.IGNORECASE,
)


# ── Row classification ────────────────────────────────────────────────

@dataclass
class _RowInfo:
    row_idx: int
    non_empty_count: int = 0
    numeric_count: int = 0
    bold_count: int = 0
    non_empty_bold_count: int = 0
    merge_spans: list = field(default_factory=list)
    max_merge_span: int = 0
    max_text_len: int = 0
    bg_color: Optional[str] = None
    total_visible_cells: int = 0
    is_empty: bool = True
    numeric_ratio: float = 0.0
    fill_ratio: float = 0.0
    all_bold: bool = False
    texts: list = field(default_factory=list)


def _classify_row(
    ws: Worksheet,
    r: int,
    min_col: int,
    max_col: int,
    anchor_map: Dict,
    hidden_set: set,
    num_cols: int,
) -> _RowInfo:
    """Compute metrics for a single row."""
    info = _RowInfo(row_idx=r)
    for c in range(min_col, max_col + 1):
        if (r, c) in hidden_set:
            continue
        cell = ws.cell(row=r, column=c)
        text = str(cell.value).strip() if cell.value is not None else ""
        info.total_visible_cells += 1

        if text:
            info.non_empty_count += 1
            info.max_text_len = max(info.max_text_len, len(text))
            info.texts.append(text)
        if _is_numeric_text(text):
            info.numeric_count += 1
        font = cell.font
        if font and font.bold:
            info.bold_count += 1
            if text:
                info.non_empty_bold_count += 1
        if (r, c) in anchor_map:
            _, cs = anchor_map[(r, c)]
            if cs > 1:
                info.merge_spans.append(cs)
                info.max_merge_span = max(info.max_merge_span, cs)
        if info.bg_color is None and cell.fill:
            bg = _color_hex(cell.fill.fgColor)
            if bg:
                info.bg_color = bg

    info.is_empty = info.non_empty_count == 0
    info.numeric_ratio = (
        info.numeric_count / info.non_empty_count
        if info.non_empty_count > 0
        else 0.0
    )
    info.fill_ratio = info.non_empty_count / max(num_cols, 1)
    info.all_bold = (
        info.non_empty_bold_count > 0
        and info.non_empty_bold_count == info.non_empty_count
    )
    return info


# ── Number / type helpers ─────────────────────────────────────────────

def _is_numeric_text(text: str) -> bool:
    """Check if text looks like a number (handles commas, spaces, parens)."""
    if not text:
        return False
    cleaned = text.replace(",", "").replace(" ", "").replace("\u00a0", "")
    cleaned = cleaned.strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = cleaned[1:-1]
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
    if cleaned in ("-", "—", "–", ""):
        return False
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _infer_column_type(
    values: List[str],
) -> Tuple[str, str, Optional[str], bool]:
    """
    Sample values and infer (format, align, suffix, highlight_negative).

    Returns: (format, align, suffix, highlight_negative)
    """
    samples = [v for v in values if v and v not in ("-", "—", "–", "0")][:20]
    if not samples:
        return "text", "left", None, False

    numeric_samples = [v for v in samples if _is_numeric_text(v)]
    num_count = len(numeric_samples)
    pct_count = sum(1 for v in numeric_samples if "%" in v)
    has_negative = any(
        v.startswith("-") or v.startswith("(")
        for v in numeric_samples
    )

    # Percentage
    if numeric_samples and pct_count >= len(numeric_samples) * 0.5:
        return "percentage", "right", "%", has_negative

    # Numeric
    if num_count >= len(samples) * 0.6:
        # Check decimal
        decimal_count = sum(
            1 for v in samples
            if _is_numeric_text(v) and "." in v.replace(",", "")
        )
        fmt = "decimal" if decimal_count > len(samples) * 0.3 else "integer"

        # Detect currency suffix from values
        suffix = None
        all_text = " ".join(samples)
        m = _CURRENCY_PATTERNS.search(all_text)
        if m:
            suffix = m.group(1).upper()

        return fmt, "right", suffix, has_negative

    return "text", "left", None, False


def _to_snake_case(label: str) -> str:
    """Convert label to snake_case key. Delegates to identifier_utils for consistency."""
    from app.services.identifier_utils import normalize_identifier
    return normalize_identifier(label, fallback="col")


def _luminance(hex_color: str) -> float:
    """Compute relative luminance of a hex color."""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) != 6:
        return 0.5
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) / 255


def _contains_keyword(texts: List[str], keywords: set) -> bool:
    """Check if any text in list contains a keyword."""
    joined = " ".join(texts).lower()
    return any(kw in joined for kw in keywords)


def _normalize_search_text(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text or "")
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_only).strip().lower()


def _joined_row_text(info: _RowInfo) -> str:
    return " ".join(str(text).strip() for text in info.texts if str(text).strip()).strip()


def _is_major_break_row(info: _RowInfo) -> bool:
    row_text = _joined_row_text(info)
    if not row_text:
        return False

    normalized = _normalize_search_text(row_text)
    if any(keyword in normalized for keyword in _FOOTER_KEYWORDS):
        return True
    if normalized.startswith(("ii.", "iii.", "iv.", "v.", "vi.", "vii.", "viii.", "ix.", "x.")):
        return True
    if "tong ket" in normalized:
        return True
    return False


def _extract_explicit_title(row_text: str) -> Optional[str]:
    """Extract a business-facing title from a header row when a title keyword is present."""
    compact = re.sub(r"\s+", " ", row_text or "").strip().rstrip(":- ")
    if not compact:
        return None

    lowered = compact.lower()
    for keyword in _TITLE_KEYWORDS:
        idx = lowered.find(keyword)
        if idx >= 0:
            title = compact[idx:].strip().rstrip(":- ")
            if title:
                return title[0].upper() + title[1:]
    return None


def _effective_row_width(row_infos: Dict[int, _RowInfo]) -> int:
    non_empty_counts = sorted(
        info.non_empty_count
        for info in row_infos.values()
        if not info.is_empty and info.non_empty_count > 0
    )
    if not non_empty_counts:
        return 0
    return max(non_empty_counts[len(non_empty_counts) // 2], non_empty_counts[-1] // 2, 4)


def _header_min_cells(row_infos: Dict[int, _RowInfo], num_cols: int) -> int:
    effective_width = _effective_row_width(row_infos)
    baseline = effective_width if effective_width > 0 else max(num_cols, 4)
    return max(3, min(14, int(round(baseline * 0.45))))


def _data_min_cells(row_infos: Dict[int, _RowInfo], num_cols: int) -> int:
    effective_width = _effective_row_width(row_infos)
    baseline = effective_width if effective_width > 0 else max(num_cols, 4)
    return max(4, min(18, int(round(baseline * 0.45))))


def _looks_like_data_anchor(info: _RowInfo, min_cells: int) -> bool:
    if info.is_empty or info.non_empty_count < min_cells:
        return False
    if info.numeric_ratio >= 0.45:
        return True
    return info.numeric_count >= max(2, min_cells // 2) and info.non_empty_count >= min_cells


def _looks_like_header_candidate(info: _RowInfo, min_cells: int) -> bool:
    if info.is_empty or info.non_empty_count < min_cells:
        return False
    if info.numeric_ratio >= 0.45 or info.max_text_len > 80:
        return False
    return True


def _best_effort_header_row(
    row_infos: Dict[int, _RowInfo],
    min_row: int,
    max_row: int,
    num_cols: int,
) -> Optional[int]:
    rows = [r for r in sorted(row_infos.keys()) if min_row <= r <= max_row and not row_infos[r].is_empty]
    if not rows:
        return None

    header_min = _header_min_cells(row_infos, num_cols)
    data_min = _data_min_cells(row_infos, num_cols)

    for anchor_index, row_idx in enumerate(rows):
        anchor_info = row_infos[row_idx]
        if not _looks_like_data_anchor(anchor_info, data_min):
            continue

        scored_candidates: List[Tuple[int, int]] = []
        for prev_index in range(max(0, anchor_index - 4), anchor_index):
            candidate_row = rows[prev_index]
            candidate_info = row_infos[candidate_row]
            if not _looks_like_header_candidate(candidate_info, header_min):
                continue

            distance = anchor_index - prev_index
            score = candidate_info.non_empty_count * 4
            score += candidate_row
            score -= min(candidate_info.max_merge_span, 12) * 2
            score -= int(candidate_info.numeric_ratio * 20)
            score -= max(distance - 1, 0) * 3
            if candidate_info.max_merge_span <= 2:
                score += 8
            if distance == 1:
                score += 12
            scored_candidates.append((score, candidate_row))

        if scored_candidates:
            scored_candidates.sort()
            return scored_candidates[-1][1]

    top_rows = rows[: min(12, len(rows))]
    fallback_candidates = [
        row_idx
        for row_idx in top_rows
        if _looks_like_header_candidate(row_infos[row_idx], max(2, header_min - 2))
    ]
    if not fallback_candidates:
        return None

    return max(
        fallback_candidates,
        key=lambda row_idx: (
            row_infos[row_idx].non_empty_count,
            -row_infos[row_idx].max_merge_span,
            row_idx,
        ),
    )


def _best_effort_title_row(
    row_infos: Dict[int, _RowInfo],
    min_row: int,
    max_row: int,
) -> Optional[int]:
    candidate_rows = [
        row_idx
        for row_idx in range(min_row, max_row + 1)
        if row_idx in row_infos and not row_infos[row_idx].is_empty
    ]
    if not candidate_rows:
        return None

    explicit_rows = []
    for row_idx in candidate_rows:
        row_text = _joined_row_text(row_infos[row_idx])
        if _extract_explicit_title(row_text):
            explicit_rows.append(row_idx)
    if explicit_rows:
        return explicit_rows[0]

    return max(
        candidate_rows,
        key=lambda row_idx: (
            row_infos[row_idx].max_merge_span * 2 + (1 if row_infos[row_idx].all_bold else 0),
            -row_infos[row_idx].numeric_ratio,
            -row_infos[row_idx].non_empty_count,
            -row_idx,
        ),
    )


# ── Main analyzer ─────────────────────────────────────────────────────

def analyze_excel_structure(
    file_bytes: bytes,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Analyze an Excel file and return detected structure.

    Returns a dict suitable for JSON serialization with keys:
        header_lines, report_title, report_meta, column_groups,
        columns, group_by_column, show_subtotals, footer_lines,
        signature_count, signature_labels, theme,
        recommended_table_schema, data_preview, total_data_rows,
        confidence, sheet_names, analyzed_sheet
    """

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    all_sheet_names = wb.sheetnames

    if sheet_name and sheet_name in all_sheet_names:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1
    num_cols = max_col - min_col + 1

    anchor_map, hidden_set = _build_merge_map(ws)

    # ── Phase 1: Classify every row ──────────────────────────────────
    row_infos: Dict[int, _RowInfo] = {}
    for r in range(min_row, max_row + 1):
        row_infos[r] = _classify_row(ws, r, min_col, max_col, anchor_map, hidden_set, num_cols)

    # ── Phase 2: Find data header row ────────────────────────────────
    data_header_row = _find_data_header_row(row_infos, min_row, max_row, num_cols)

    if data_header_row is None:
        # Fallback: cannot detect structure, return minimal result
        return _minimal_result(ws, all_sheet_names, sheet_name, row_infos, min_row, max_row, min_col, max_col, num_cols)

    # ── Phase 3: Header zone (above data header) ─────────────────────
    header_lines, report_title, report_meta, column_groups = _detect_header_zone(
        ws, row_infos, anchor_map, min_row, data_header_row, min_col, max_col, num_cols,
    )

    # ── Phase 4: Columns from data header row ────────────────────────
    columns, col_indices = _detect_columns(
        ws, data_header_row, min_col, max_col, anchor_map, hidden_set, row_infos,
    )

    # ── Phase 5: Data zone + subtotals ───────────────────────────────
    data_start = data_header_row + 1
    data_end, group_by_column, show_subtotals, subtotal_bg = _detect_data_zone(
        ws, row_infos, data_start, max_row, columns, min_col,
    )

    # ── Phase 6: Footer zone ─────────────────────────────────────────
    footer_lines, signature_count, signature_labels = _detect_footer(
        ws, row_infos, data_end + 1, max_row, min_col, max_col,
    )

    # ── Phase 7: Column type inference from data ─────────────────────
    _enrich_columns_with_data(ws, columns, col_indices, data_start, data_end)

    # ── Phase 8: Column groups mapping ───────────────────────────────
    mapped_groups = _map_column_groups_to_columns(column_groups, col_indices, min_col, columns)

    # ── Phase 9: Theme extraction ────────────────────────────────────
    theme = _extract_theme(row_infos, data_header_row, subtotal_bg)

    # ── Phase 10: Recommended table schema ───────────────────────────
    table_schema = [
        {
            "name": c["key"],
            "display_name": c["label"],
            "type": "number" if c["format"] in ("integer", "decimal", "percentage") else "string",
        }
        for c in columns
    ]

    # ── Phase 11: Data preview (first 5 rows) ────────────────────────
    data_preview = _extract_data_preview(ws, columns, col_indices, data_start, min(data_start + 4, data_end))

    # ── Phase 12: Confidence score ───────────────────────────────────
    total_data_rows = max(0, data_end - data_start + 1)
    confidence = _compute_confidence(
        data_header_row is not None,
        total_data_rows,
        len(header_lines),
        len(mapped_groups),
        show_subtotals,
        signature_count,
        columns,
    )

    return {
        "header_lines": header_lines,
        "report_title": report_title,
        "report_meta": report_meta,
        "column_groups": mapped_groups,
        "columns": columns,
        "group_by_column": group_by_column,
        "show_subtotals": show_subtotals,
        "footer_lines": footer_lines,
        "signature_count": signature_count,
        "signature_labels": signature_labels,
        "theme": theme,
        "recommended_table_schema": table_schema,
        "data_preview": data_preview,
        "total_data_rows": total_data_rows,
        "confidence": round(confidence, 2),
        "sheet_names": all_sheet_names,
        "analyzed_sheet": sheet_name,
    }


# ── Phase 2: Data header row detection ────────────────────────────────

def _find_data_header_row(
    row_infos: Dict[int, _RowInfo],
    min_row: int,
    max_row: int,
    num_cols: int,
) -> Optional[int]:
    """Find the row where the leaf-level column labels live."""
    rows = sorted(row_infos.keys())
    header_min = _header_min_cells(row_infos, num_cols)
    data_min = _data_min_cells(row_infos, num_cols)

    anchored_candidate = _best_effort_header_row(row_infos, min_row, max_row, num_cols)
    if anchored_candidate is not None:
        return anchored_candidate

    for i, r in enumerate(rows):
        info = row_infos[r]
        if info.is_empty:
            continue

        if (
            info.non_empty_count >= header_min
            and info.max_text_len <= 80
            and info.numeric_ratio < 0.35
        ):
            for j in range(i + 1, min(i + 4, len(rows))):
                next_info = row_infos[rows[j]]
                if next_info.is_empty:
                    continue
                if _looks_like_data_anchor(next_info, data_min):
                    return r
                if _looks_like_header_candidate(next_info, max(2, header_min - 2)):
                    break
                break

    for r in rows:
        info = row_infos[r]
        if info.non_empty_count >= header_min and info.numeric_ratio < 0.45 and not info.is_empty:
            return r
    return None


# ── Phase 3: Header zone ─────────────────────────────────────────────

def _detect_header_zone(
    ws: Worksheet,
    row_infos: Dict[int, _RowInfo],
    anchor_map: Dict,
    min_row: int,
    data_header_row: int,
    min_col: int,
    max_col: int,
    num_cols: int,
) -> Tuple[List[Dict], str, Optional[str], List[Dict]]:
    """Detect header lines, title, meta, and column group rows above data header."""
    header_lines: List[Dict] = []
    report_title = ""
    report_meta: Optional[str] = None
    column_groups_raw: List[Dict] = []

    if data_header_row <= min_row:
        return header_lines, report_title, report_meta, column_groups_raw

    # Scan rows above data header, bottom-up to find column groups first
    rows_above = [r for r in range(min_row, data_header_row) if r in row_infos and not row_infos[r].is_empty]

    # Column group rows: immediately above data header, have partial merges
    col_group_rows = []
    for r in reversed(rows_above):
        info = row_infos[r]
        has_partial_merges = (
            len(info.merge_spans) >= 2
            and all(s <= num_cols * 0.7 for s in info.merge_spans)
        )
        if has_partial_merges and info.non_empty_count >= 2:
            col_group_rows.insert(0, r)
        else:
            break  # Stop once we hit non-column-group rows

    # Extract column groups
    for gr in col_group_rows:
        for c in range(min_col, max_col + 1):
            if (gr, c) in anchor_map:
                rs, cs = anchor_map[(gr, c)]
                if cs >= 2:
                    cell = ws.cell(row=gr, column=c)
                    text = str(cell.value).strip() if cell.value is not None else ""
                    if text:
                        column_groups_raw.append({
                            "label": text,
                            "start_col": c,
                            "col_span": cs,
                        })

    # Header rows: everything above column group rows
    header_rows = [r for r in rows_above if r not in col_group_rows]

    explicit_title_row = None
    explicit_title = None
    for r in header_rows:
        info = row_infos[r]
        row_text = " ".join(info.texts).strip()
        extracted = _extract_explicit_title(row_text)
        if extracted:
            explicit_title_row = r
            explicit_title = extracted

    # Find title fallback: row with widest merge or largest text that's bold
    best_title_row = None
    best_title_score = 0
    for r in header_rows:
        info = row_infos[r]
        score = info.max_merge_span * 2 + (1 if info.all_bold else 0)
        if score > best_title_score and info.non_empty_count > 0:
            best_title_score = score
            best_title_row = r

    for r in header_rows:
        info = row_infos[r]
        if info.is_empty:
            continue

        row_text = " ".join(info.texts).strip()
        if not row_text:
            continue

        # Prefer explicit title rows like "bao gia ..." over company-name rows.
        if r == explicit_title_row and explicit_title:
            report_title = explicit_title
            continue

        # Check if this is the fallback title row
        if not explicit_title and r == best_title_row and best_title_score > 2:
            report_title = row_text
            continue

        # Check if meta (exchange rate, period)
        if _contains_keyword(info.texts, _META_KEYWORDS) and report_meta is None:
            report_meta = row_text
            continue

        # Regular header line
        right_text = None
        if len(info.texts) >= 2:
            # Detect 2-column header: left text + right text
            left_texts = []
            right_texts = []
            mid = (min_col + max_col) // 2
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                text = str(cell.value).strip() if cell.value is not None else ""
                if text:
                    if c <= mid:
                        left_texts.append(text)
                    else:
                        right_texts.append(text)
            if left_texts and right_texts:
                row_text = " ".join(left_texts)
                right_text = " ".join(right_texts)

        font_size = "base"
        if info.max_merge_span >= num_cols * 0.5:
            font_size = "lg"
        elif info.non_empty_count <= 2:
            font_size = "sm"

        header_lines.append({
            "text": row_text,
            "right_text": right_text,
            "align": "left",
            "bold": info.all_bold,
            "font_size": font_size,
        })

    # If no explicit title found, use first bold header line
    if not report_title and header_lines:
        for i, hl in enumerate(header_lines):
            if hl.get("bold"):
                report_title = hl["text"]
                header_lines.pop(i)
                break

    return header_lines, report_title, report_meta, column_groups_raw


def _resolve_header_cell_text(
    ws: Worksheet,
    row_idx: int,
    col_idx: int,
    anchor_map: Dict,
) -> str:
    cell = ws.cell(row=row_idx, column=col_idx)
    if cell.value not in (None, ""):
        return str(cell.value).strip()

    for (anchor_row, anchor_col), (_, col_span) in anchor_map.items():
        if anchor_row != row_idx:
            continue
        if anchor_col <= col_idx <= anchor_col + col_span - 1:
            anchor_cell = ws.cell(row=anchor_row, column=anchor_col)
            if anchor_cell.value not in (None, ""):
                return str(anchor_cell.value).strip()
    return ""


# ── Phase 4: Column detection ────────────────────────────────────────

def _detect_columns(
    ws: Worksheet,
    data_header_row: int,
    min_col: int,
    max_col: int,
    anchor_map: Dict,
    hidden_set: set,
    row_infos: Optional[Dict[int, _RowInfo]] = None,
) -> Tuple[List[Dict], List[int]]:
    """Extract column definitions from the data header row."""
    columns: List[Dict] = []
    col_indices: List[int] = []  # 1-based Excel column indices

    for c in range(min_col, max_col + 1):
        header_parts: List[str] = []
        for header_row in range(max(1, data_header_row - 3), data_header_row + 1):
            if row_infos is not None:
                row_info = row_infos.get(header_row)
                if row_info is None or row_info.is_empty or row_info.numeric_ratio >= 0.45:
                    continue
                if header_row < data_header_row and row_info.non_empty_count <= 2 and row_info.max_merge_span >= 4:
                    continue

            text = _resolve_header_cell_text(ws, header_row, c, anchor_map).strip()
            if not text or _is_numeric_text(text):
                continue
            if not header_parts or header_parts[-1] != text:
                header_parts.append(text)

        if not header_parts and (data_header_row, c) in hidden_set:
            continue

        display_label = header_parts[-1] if header_parts else ""
        if not display_label:
            cell = ws.cell(row=data_header_row, column=c)
            display_label = str(cell.value).strip() if cell.value is not None else ""
        if not display_label:
            continue

        key_source = " / ".join(header_parts) if header_parts else display_label
        key = _to_snake_case(key_source)
        # Ensure unique keys
        existing_keys = [col["key"] for col in columns]
        if key in existing_keys:
            key = f"{key}_{len(columns) + 1}"
        if not key:
            key = f"col_{len(columns) + 1}"

        width = _col_width_px(ws, c)

        columns.append({
            "label": display_label,
            "key": key,
            "inferred_type": "text",
            "width_px": round(width, 1),
            "align": "left",
            "format": "text",
            "suffix": None,
            "bold": False,
            "highlight_negative": False,
            "source_col_idx": c,
        })
        col_indices.append(c)

    return columns, col_indices


# ── Phase 5: Data zone + subtotal detection ──────────────────────────

def _detect_data_zone(
    ws: Worksheet,
    row_infos: Dict[int, _RowInfo],
    data_start: int,
    max_row: int,
    columns: List[Dict],
    min_col: int,
) -> Tuple[int, Optional[str], bool, Optional[str]]:
    """
    Detect data end, group-by column, subtotals.
    Returns: (data_end_row, group_by_column, show_subtotals, subtotal_bg)
    """
    data_end = data_start
    subtotal_rows: List[int] = []
    consecutive_empty = 0
    expected_fill = len(columns)
    seen_data_rows = 0

    for r in range(data_start, max_row + 1):
        info = row_infos.get(r)
        if info is None or info.is_empty:
            consecutive_empty += 1
            if consecutive_empty >= 2:
                break
            continue
        consecutive_empty = 0

        if seen_data_rows >= 3 and _is_major_break_row(info):
            break

        # Check if this looks like a data or subtotal row
        if info.non_empty_count >= expected_fill * 0.3:
            data_end = r
            seen_data_rows += 1
            # Subtotal: bold row with different pattern
            if (
                info.all_bold
                and info.non_empty_count < expected_fill * 0.8
                and info.numeric_ratio > 0.3
            ):
                subtotal_rows.append(r)
        elif info.fill_ratio < 0.15:
            # Very sparse row — likely end of data
            break
        else:
            data_end = r

    # Determine group-by column from subtotal rows
    group_by_column: Optional[str] = None
    show_subtotals = len(subtotal_rows) > 0
    subtotal_bg: Optional[str] = None

    if subtotal_rows and columns:
        subtotal_bg = row_infos.get(subtotal_rows[0], _RowInfo(0)).bg_color
        # Group-by: find the leftmost text column that has value in subtotal rows
        for col in columns:
            ci = col["source_col_idx"]
            cell = ws.cell(row=subtotal_rows[0], column=ci)
            text = str(cell.value).strip() if cell.value is not None else ""
            if text and not _is_numeric_text(text):
                group_by_column = col["key"]
                break

    if group_by_column:
        group_col = next((col for col in columns if col["key"] == group_by_column), None)
        if group_col:
            source_col_idx = group_col["source_col_idx"]
            distinct_values = set()
            non_empty_values = 0
            for r in range(data_start, data_end + 1):
                value = ws.cell(row=r, column=source_col_idx).value
                text = str(value).strip() if value is not None else ""
                if not text or _is_numeric_text(text):
                    continue
                non_empty_values += 1
                distinct_values.add(text)

            if non_empty_values == 0 or len(distinct_values) >= max(3, int(non_empty_values * 0.8)):
                group_by_column = None

    return data_end, group_by_column, show_subtotals, subtotal_bg


# ── Phase 6: Footer detection ────────────────────────────────────────

def _detect_footer(
    ws: Worksheet,
    row_infos: Dict[int, _RowInfo],
    footer_start: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> Tuple[List[str], int, List[str]]:
    """Detect footer lines and signature slots."""
    footer_lines: List[str] = []
    signature_count = 0
    signature_labels: List[str] = []
    capture_notes = False
    note_row_max_span = max(1, max_col - min_col)

    for r in range(footer_start, max_row + 1):
        info = row_infos.get(r)
        if info is None or info.is_empty:
            if capture_notes:
                capture_notes = False
            continue

        row_text = _joined_row_text(info)
        normalized_row = _normalize_search_text(row_text)
        is_footer_marker = any(keyword in normalized_row for keyword in _FOOTER_KEYWORDS)

        # Check for signature keywords
        if _contains_keyword(info.texts, _SIGNATURE_KEYWORDS):
            # These texts are likely signature labels
            for text in info.texts:
                text_lower = text.lower()
                if any(kw in text_lower for kw in _SIGNATURE_KEYWORDS):
                    signature_labels.append(text)
                    signature_count += 1
            capture_notes = False
            continue

        # Check for bordered empty cells (signature boxes)
        bordered_empty = 0
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            text = str(cell.value).strip() if cell.value is not None else ""
            borders = _border_sides(cell.border)
            if not text and borders and borders.get("bottom"):
                bordered_empty += 1

        if bordered_empty >= 2 and info.non_empty_count == 0:
            if not signature_count:
                signature_count = bordered_empty
            capture_notes = False
            continue

        if is_footer_marker:
            capture_notes = True
            if row_text:
                footer_lines.append(row_text)
            continue

        if capture_notes:
            is_note_continuation = (
                info.non_empty_count <= 1
                or info.max_merge_span >= note_row_max_span
            )
            if is_note_continuation and row_text:
                footer_lines.append(row_text)
                continue
            capture_notes = False

    return footer_lines, signature_count, signature_labels


# ── Phase 7: Enrich columns with data samples ────────────────────────

def _enrich_columns_with_data(
    ws: Worksheet,
    columns: List[Dict],
    col_indices: List[int],
    data_start: int,
    data_end: int,
) -> None:
    """Sample data rows to infer column types, formats, suffixes."""
    sample_end = min(data_start + 19, data_end)

    for i, col in enumerate(columns):
        ci = col_indices[i]
        values: List[str] = []
        for r in range(data_start, sample_end + 1):
            cell = ws.cell(row=r, column=ci)
            text = str(cell.value).strip() if cell.value is not None else ""
            values.append(text)

        fmt, align, suffix, highlight_neg = _infer_column_type(values)
        col["format"] = fmt
        col["inferred_type"] = fmt
        col["align"] = align
        col["highlight_negative"] = highlight_neg
        if suffix:
            col["suffix"] = suffix

        # Also check header label for currency hints
        if not col["suffix"]:
            m = _CURRENCY_PATTERNS.search(col["label"])
            if m:
                col["suffix"] = m.group(1).upper()


# ── Phase 8: Map column groups ───────────────────────────────────────

def _map_column_groups_to_columns(
    raw_groups: List[Dict],
    col_indices: List[int],
    min_col: int,
    columns: List[Dict],
) -> List[Dict]:
    """Map raw column groups (Excel col ranges) to detected column indices."""
    mapped: List[Dict] = []
    for g in raw_groups:
        start_excel = g["start_col"]
        end_excel = start_excel + g["col_span"] - 1

        # Find which detected columns fall in this range
        start_idx = None
        span = 0
        for i, ci in enumerate(col_indices):
            if start_excel <= ci <= end_excel:
                if start_idx is None:
                    start_idx = i
                span += 1

        if start_idx is not None and span > 0:
            mapped.append({
                "label": g["label"],
                "start_col_idx": start_idx,
                "span": span,
            })

    return mapped


# ── Phase 9: Theme extraction ────────────────────────────────────────

def _extract_theme(
    row_infos: Dict[int, _RowInfo],
    data_header_row: int,
    subtotal_bg: Optional[str],
) -> Dict[str, str]:
    """Extract theme colors from the data header row and subtotal rows."""
    header_info = row_infos.get(data_header_row)
    header_bg = header_info.bg_color if header_info else None

    if not header_bg:
        header_bg = "#073763"

    header_text = "#ffffff" if _luminance(header_bg) < 0.5 else "#000000"

    theme = {
        "header_bg": header_bg,
        "header_text": header_text,
        "group_bg": "#c9daf8",
        "group_text": "#073763",
        "subtotal_bg": subtotal_bg or "#dbeafe",
        "subtotal_text": "#1e40af",
        "accent_color": header_bg,
    }
    return theme


# ── Phase 11: Data preview ───────────────────────────────────────────

def _extract_data_preview(
    ws: Worksheet,
    columns: List[Dict],
    col_indices: List[int],
    data_start: int,
    data_end: int,
) -> List[Dict[str, Any]]:
    """Extract first N data rows for preview."""
    preview: List[Dict[str, Any]] = []
    for r in range(data_start, data_end + 1):
        row_data: Dict[str, Any] = {}
        for i, col in enumerate(columns):
            ci = col_indices[i]
            cell = ws.cell(row=r, column=ci)
            val = cell.value
            if val is None:
                row_data[col["key"]] = ""
            elif isinstance(val, (int, float)):
                row_data[col["key"]] = val
            else:
                row_data[col["key"]] = str(val).strip()
        preview.append(row_data)
    return preview


# ── Phase 12: Confidence score ───────────────────────────────────────

def _compute_confidence(
    header_found: bool,
    total_data_rows: int,
    num_header_lines: int,
    num_col_groups: int,
    show_subtotals: bool,
    signature_count: int,
    columns: List[Dict],
) -> float:
    """Compute confidence score 0.0 - 1.0."""
    score = 0.0
    if header_found:
        score += 0.30
    if total_data_rows >= 3:
        score += 0.20
    if num_header_lines > 0:
        score += 0.10
    if num_col_groups > 0:
        score += 0.10
    if show_subtotals:
        score += 0.10
    if signature_count > 0:
        score += 0.10
    num_typed = sum(1 for c in columns if c["format"] != "text")
    if num_typed >= len(columns) * 0.3:
        score += 0.10
    return min(score, 1.0)


def _coerce_numberish_value(value: Any) -> Optional[Any]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if text in ("", "-", "—", "–"):
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    is_percentage = text.endswith("%")
    if is_percentage:
        text = text[:-1]

    cleaned = text.replace(",", "").replace(" ", "").replace("\xa0", "")
    if not cleaned:
        return None

    try:
        number: Any = float(cleaned) if "." in cleaned else int(cleaned)
    except ValueError:
        return None

    if negative:
        number = -number
    if is_percentage:
        number = number / 100
    return number


# ── CSV analyzer ─────────────────────────────────────────────────────

_CSV_TITLE_KEYWORDS = {
    "bang", "bao gia", "quotation", "quote", "report", "payroll",
}

_CSV_HELPER_KEYWORDS = {
    "cong thuc", "formula", "co dinh", "fixed",
}

_CSV_AUXILIARY_HEADER_KEYWORDS = {
    "bao cao", "chi tiet", "kiem tra", "chia cong theo bhxh",
    "ghi chu", "phat hien", "so bi trung", "dac biet",
}


def _normalize_text_value(text: str) -> str:
    text = (text or "").replace("Đ", "D").replace("đ", "d")
    nfkd = unicodedata.normalize("NFKD", text or "")
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", ascii_only).strip().lower()


def _is_csv_ignored_header_text(text: str) -> bool:
    normalized = _normalize_text_value(text)
    if not normalized:
        return False
    if normalized.startswith("#"):
        return True
    if any(keyword in normalized for keyword in _CSV_HELPER_KEYWORDS):
        return True
    return any(keyword in normalized for keyword in _CSV_AUXILIARY_HEADER_KEYWORDS)


def _csv_row_text(row: List[str]) -> str:
    return ", ".join(cell.strip() for cell in row if cell and cell.strip())


def _csv_primary_row_text(row: List[str], max_gap: int = 3) -> str:
    indexed = [(idx, cell.strip()) for idx, cell in enumerate(row) if cell and cell.strip()]
    if not indexed:
        return ""

    parts = [indexed[0][1]]
    last_idx = indexed[0][0]
    for idx, text in indexed[1:]:
        if idx - last_idx > max_gap:
            break
        parts.append(text)
        last_idx = idx

    return ", ".join(parts)


def _classify_csv_row(row: List[str]) -> Dict[str, Any]:
    indexed_non_empty = [
        (idx, str(cell).strip())
        for idx, cell in enumerate(row)
        if str(cell).strip()
    ]
    non_empty = [cell for _, cell in indexed_non_empty]
    normalized = [_normalize_text_value(cell) for cell in non_empty]
    counts = Counter(normalized)
    helper_hits = sum(
        1
        for text in normalized
        if any(keyword in text for keyword in _CSV_HELPER_KEYWORDS)
    )
    numeric_count = sum(1 for cell in non_empty if _is_numeric_text(cell))
    return {
        "non_empty_count": len(non_empty),
        "numeric_count": numeric_count,
        "numeric_ratio": (numeric_count / len(non_empty)) if non_empty else 0.0,
        "max_text_len": max((len(cell) for cell in non_empty), default=0),
        "texts": non_empty,
        "helper_ratio": (helper_hits / len(non_empty)) if non_empty else 0.0,
        "top_repeat_ratio": (max(counts.values()) / len(non_empty)) if counts else 0.0,
        "first_non_empty_idx": indexed_non_empty[0][0] if indexed_non_empty else None,
        "last_non_empty_idx": indexed_non_empty[-1][0] if indexed_non_empty else None,
    }


def _looks_like_csv_data_row(info: Dict[str, Any], width: int) -> bool:
    min_cells = max(4, int(width * 0.45))
    return (
        info["non_empty_count"] >= min_cells
        and (info["numeric_count"] >= 2 or info["numeric_ratio"] >= 0.2)
    )


def _find_csv_data_start(row_infos: List[Dict[str, Any]], width: int) -> int:
    for idx, info in enumerate(row_infos):
        if not _looks_like_csv_data_row(info, width):
            continue
        future_data = sum(
            1
            for next_info in row_infos[idx + 1: idx + 4]
            if _looks_like_csv_data_row(next_info, width)
        )
        if future_data >= 1:
            return idx

    for idx, info in enumerate(row_infos):
        if info["non_empty_count"] >= max(3, int(width * 0.4)):
            return min(idx + 1, len(row_infos) - 1)
    return 0


def _select_csv_header_rows(row_infos: List[Dict[str, Any]], data_start_idx: int) -> List[int]:
    candidates: List[int] = []
    blank_streak = 0

    for idx in range(data_start_idx - 1, -1, -1):
        info = row_infos[idx]
        if info["non_empty_count"] == 0:
            blank_streak += 1
            if candidates and blank_streak >= 2:
                break
            continue

        blank_streak = 0
        candidates.insert(0, idx)
        if len(candidates) >= 6:
            break

    filtered = [
        idx
        for idx in candidates
        if row_infos[idx]["non_empty_count"] > 1
        and row_infos[idx]["helper_ratio"] < 0.5
    ]

    if not filtered:
        return []

    anchor_rows = filtered[-3:]
    anchor_start = min(
        row_infos[idx]["first_non_empty_idx"]
        for idx in anchor_rows
        if row_infos[idx]["first_non_empty_idx"] is not None
    )
    anchor_density = max(row_infos[idx]["non_empty_count"] for idx in anchor_rows)
    density_floor = max(4, int(anchor_density * 0.45))

    return [
        idx
        for idx in filtered
        if (
            row_infos[idx]["first_non_empty_idx"] is not None
            and row_infos[idx]["first_non_empty_idx"] <= anchor_start + 12
        )
        or row_infos[idx]["non_empty_count"] >= density_floor
    ]


def _expand_csv_header_row(row: List[str], lower_rows: List[List[str]]) -> List[str]:
    filled = [str(cell).strip() for cell in row]
    last_text = ""
    last_text_idx = -1

    for idx, text in enumerate(filled):
        if text:
            last_text = text
            last_text_idx = idx
            continue
        if not last_text:
            continue

        next_non_empty = next((j for j in range(idx + 1, len(filled)) if filled[j]), None)
        if next_non_empty is None:
            continue
        if last_text_idx >= 0 and (next_non_empty - last_text_idx) > 3:
            continue

        has_lower_content = any(
            idx < len(lower_row) and str(lower_row[idx]).strip()
            for lower_row in lower_rows
        )
        if has_lower_content:
            filled[idx] = last_text

    return filled


def _pick_csv_title_and_header_lines(
    all_rows: List[List[str]],
    row_infos: List[Dict[str, Any]],
    first_header_idx: int,
) -> Tuple[str, List[Dict[str, Any]]]:
    preamble_rows = [idx for idx in range(first_header_idx) if row_infos[idx]["non_empty_count"] > 0]
    header_anchor_start = row_infos[first_header_idx].get("first_non_empty_idx")
    row_texts = {
        idx: _csv_primary_row_text(all_rows[idx]) or _csv_row_text(all_rows[idx])
        for idx in preamble_rows
    }

    title_idx: Optional[int] = None
    for idx in reversed(preamble_rows):
        normalized = _normalize_text_value(row_texts[idx])
        if any(keyword in normalized for keyword in _CSV_TITLE_KEYWORDS):
            title_idx = idx
            break

    if title_idx is None:
        for idx in reversed(preamble_rows):
            if len(row_texts[idx]) >= 10:
                title_idx = idx
                break

    title_suffix_idx: Optional[int] = None
    report_title = ""
    if title_idx is not None:
        report_title = row_texts[title_idx]
        following_rows = [idx for idx in preamble_rows if idx > title_idx]
        for idx in following_rows:
            normalized = _normalize_text_value(row_texts[idx])
            if _contains_keyword([normalized], _META_KEYWORDS):
                report_title = f"{report_title} - {row_texts[idx]}"
                title_suffix_idx = idx
                break

    header_lines: List[Dict[str, Any]] = []
    for idx in preamble_rows:
        if idx == title_idx or idx == title_suffix_idx:
            continue
        text = row_texts[idx]
        if not text:
            continue
        if (
            header_anchor_start is not None
            and row_infos[idx].get("first_non_empty_idx") is not None
            and row_infos[idx]["first_non_empty_idx"] > header_anchor_start + 12
            and row_infos[idx]["non_empty_count"] <= 8
            and text.lstrip()[:1].isdigit()
        ):
            continue
        normalized = _normalize_text_value(text)
        if any(keyword in normalized for keyword in _CSV_TITLE_KEYWORDS):
            continue
        header_lines.append({
            "text": text,
            "right_text": None,
            "align": "left",
            "bold": False,
            "font_size": "base",
        })

    return report_title, header_lines

def analyze_csv_structure(
    file_bytes: bytes,
    filename: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Analyze a CSV file with support for title rows and multi-row table headers.
    """
    import csv
    import io

    # Decode
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        sheet_name = Path(filename).stem if filename else "data"
        return {
            "header_lines": [], "report_title": "", "report_meta": None,
            "column_groups": [], "columns": [], "group_by_column": None,
            "show_subtotals": False, "footer_lines": [],
            "signature_count": 0, "signature_labels": [],
            "theme": _default_theme(),
            "recommended_table_schema": [], "data_preview": [],
            "total_data_rows": 0, "confidence": 0.1,
            "sheet_names": [sheet_name], "analyzed_sheet": sheet_name,
        }

    sheet_name = Path(filename).stem if filename else "data"
    width = max((len(row) for row in all_rows), default=0)
    row_infos = [_classify_csv_row(row) for row in all_rows]

    data_start_idx = _find_csv_data_start(row_infos, width)
    header_row_indices = _select_csv_header_rows(row_infos, data_start_idx)

    if not header_row_indices:
        header_row_indices = [max(0, data_start_idx - 1)]

    report_title, header_zone_lines = _pick_csv_title_and_header_lines(
        all_rows,
        row_infos,
        header_row_indices[0],
    )

    expanded_header_rows: List[List[str]] = []
    for pos, row_idx in enumerate(header_row_indices):
        lower_rows = [all_rows[idx] for idx in header_row_indices[pos + 1:]]
        expanded_header_rows.append(_expand_csv_header_row(all_rows[row_idx], lower_rows))

    # Extract columns by combining header fragments from top -> bottom.
    columns: List[Dict] = []
    for ci in range(width):
        parts: List[str] = []
        for row in expanded_header_rows:
            label_part = row[ci].strip() if ci < len(row) else ""
            if not label_part:
                continue
            if _is_csv_ignored_header_text(label_part):
                continue
            if not parts or parts[-1] != label_part:
                parts.append(label_part)

        label = " - ".join(parts).strip(" -")
        if not label:
            continue

        key = _to_snake_case(label) or f"col_{ci + 1}"
        existing_keys = [c["key"] for c in columns]
        if key in existing_keys:
            key = f"{key}_{ci + 1}"
        columns.append({
            "label": label, "key": key, "inferred_type": "text",
            "width_px": 120, "align": "left", "format": "text",
            "suffix": None, "bold": False, "highlight_negative": False,
            "source_col_idx": ci,
        })

    # Sample data rows for type inference
    data_rows = all_rows[data_start_idx:]
    for col in columns:
        ci = col["source_col_idx"]
        values = []
        for row in data_rows[:20]:
            if ci < len(row):
                values.append(row[ci].strip())
        fmt, align, suffix, hl_neg = _infer_column_type(values)
        col["format"] = fmt
        col["inferred_type"] = fmt
        col["align"] = align
        col["highlight_negative"] = hl_neg
        if suffix:
            col["suffix"] = suffix

    # Data preview
    preview: List[Dict[str, Any]] = []
    for row in data_rows[:5]:
        row_dict: Dict[str, Any] = {}
        for col in columns:
            ci = col["source_col_idx"]
            val = row[ci].strip() if ci < len(row) else ""
            if _is_numeric_text(val):
                try:
                    cleaned = val.replace(",", "").replace(" ", "").replace("\xa0", "")
                    row_dict[col["key"]] = float(cleaned) if "." in cleaned else int(cleaned)
                except ValueError:
                    row_dict[col["key"]] = val
            else:
                row_dict[col["key"]] = val
        preview.append(row_dict)

    table_schema = [
        {"name": c["key"], "display_name": c["label"],
         "type": "number" if c["format"] in ("integer", "decimal", "percentage") else "string"}
        for c in columns
    ]

    total_data_rows = len(data_rows)
    confidence = 0.35
    if report_title:
        confidence += 0.1
    if header_zone_lines:
        confidence += 0.1
    if len(header_row_indices) > 1:
        confidence += 0.1
    if total_data_rows >= 3:
        confidence += 0.2
    if len(columns) >= 3:
        confidence += 0.2
    num_typed = sum(1 for c in columns if c["format"] != "text")
    if num_typed >= len(columns) * 0.3:
        confidence += 0.1

    return {
        "header_lines": header_zone_lines,
        "report_title": report_title,
        "report_meta": None,
        "column_groups": [],
        "columns": columns,
        "group_by_column": None,
        "show_subtotals": False,
        "footer_lines": [],
        "signature_count": 0,
        "signature_labels": [],
        "theme": _default_theme(),
        "recommended_table_schema": table_schema,
        "data_preview": preview,
        "total_data_rows": total_data_rows,
        "confidence": round(min(confidence, 1.0), 2),
        "sheet_names": [sheet_name],
        "analyzed_sheet": sheet_name,
    }


def extract_excel_import_sheet_data(
    file_bytes: bytes,
    sheet_name: Optional[str] = None,
) -> Tuple[str, Dict[str, Any]]:
    """Extract structured rows for import-confirm using the same Excel heuristics as analyze."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        min_row = ws.min_row or 1
        max_row = ws.max_row or 1
        min_col = ws.min_column or 1
        max_col = ws.max_column or 1
        num_cols = max_col - min_col + 1

        anchor_map, hidden_set = _build_merge_map(ws)
        row_infos: Dict[int, _RowInfo] = {}
        for r in range(min_row, max_row + 1):
            row_infos[r] = _classify_row(ws, r, min_col, max_col, anchor_map, hidden_set, num_cols)

        data_header_row = _find_data_header_row(row_infos, min_row, max_row, num_cols)
        if data_header_row is None:
            return sheet_name, {"columns": [], "rows": []}

        columns, col_indices = _detect_columns(ws, data_header_row, min_col, max_col, anchor_map, hidden_set, row_infos)
        data_start = data_header_row + 1
        data_end, _, _, _ = _detect_data_zone(ws, row_infos, data_start, max_row, columns, min_col)
        _enrich_columns_with_data(ws, columns, col_indices, data_start, data_end)

        rows = _extract_data_preview(ws, columns, col_indices, data_start, data_end)
        schema = [
            {
                "name": col["key"],
                "type": "number" if col["format"] in ("integer", "decimal", "percentage") else "string",
            }
            for col in columns
        ]
        numeric_keys = {col["name"] for col in schema if col["type"] == "number"}
        for row in rows:
            for key in numeric_keys:
                row[key] = _coerce_numberish_value(row.get(key))
        return sheet_name, {"columns": schema, "rows": rows}
    finally:
        wb.close()


def extract_csv_import_sheet_data(
    file_bytes: bytes,
    filename: Optional[str] = None,
) -> Tuple[str, Dict[str, Any]]:
    """Extract structured CSV rows for import-confirm using the same header heuristics as analyze."""
    import csv
    import io

    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1")

    all_rows = list(csv.reader(io.StringIO(text)))
    sheet_name = Path(filename).stem if filename else "Sheet1"
    if not all_rows:
        return sheet_name, {"columns": [], "rows": []}

    width = max((len(row) for row in all_rows), default=0)
    row_infos = [_classify_csv_row(row) for row in all_rows]
    data_start_idx = _find_csv_data_start(row_infos, width)
    header_row_indices = _select_csv_header_rows(row_infos, data_start_idx)
    if not header_row_indices:
        header_row_indices = [max(0, data_start_idx - 1)]

    expanded_header_rows: List[List[str]] = []
    for pos, row_idx in enumerate(header_row_indices):
        lower_rows = [all_rows[idx] for idx in header_row_indices[pos + 1:]]
        expanded_header_rows.append(_expand_csv_header_row(all_rows[row_idx], lower_rows))

    columns: List[Dict[str, Any]] = []
    for ci in range(width):
        parts: List[str] = []
        for row in expanded_header_rows:
            label_part = row[ci].strip() if ci < len(row) else ""
            if not label_part:
                continue
            if _is_csv_ignored_header_text(label_part):
                continue
            if not parts or parts[-1] != label_part:
                parts.append(label_part)

        label = " - ".join(parts).strip(" -")
        if not label:
            continue

        key = _to_snake_case(label) or f"col_{ci + 1}"
        existing_keys = [c["key"] for c in columns]
        if key in existing_keys:
            key = f"{key}_{ci + 1}"

        columns.append({
            "label": label,
            "key": key,
            "format": "text",
            "source_col_idx": ci,
        })

    data_rows = all_rows[data_start_idx:]
    for col in columns:
        ci = col["source_col_idx"]
        values = [row[ci].strip() for row in data_rows[:20] if ci < len(row)]
        fmt, _, _, _ = _infer_column_type(values)
        col["format"] = fmt

    rows: List[Dict[str, Any]] = []
    for row in data_rows:
        row_dict: Dict[str, Any] = {}
        has_value = False
        for col in columns:
            ci = col["source_col_idx"]
            raw_val = row[ci].strip() if ci < len(row) else ""
            if col["format"] in ("integer", "decimal", "percentage"):
                value = _coerce_numberish_value(raw_val)
            else:
                value = raw_val

            row_dict[col["key"]] = value
            if value not in ("", None):
                has_value = True

        if has_value:
            rows.append(row_dict)

    schema = [
        {
            "name": col["key"],
            "type": "number" if col["format"] in ("integer", "decimal", "percentage") else "string",
        }
        for col in columns
    ]
    return sheet_name, {"columns": schema, "rows": rows}


def _default_theme() -> Dict[str, str]:
    return {
        "header_bg": "#073763", "header_text": "#ffffff",
        "group_bg": "#c9daf8", "group_text": "#073763",
        "subtotal_bg": "#dbeafe", "subtotal_text": "#1e40af",
        "accent_color": "#073763",
    }


# ── Fallback: Minimal result ─────────────────────────────────────────

def _minimal_result(
    ws: Worksheet,
    all_sheet_names: List[str],
    sheet_name: str,
    row_infos: Dict[int, _RowInfo],
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    num_cols: int,
) -> Dict[str, Any]:
    """Return a minimal result when structure detection fails."""
    anchor_map, hidden_set = _build_merge_map(ws)
    header_row = _best_effort_header_row(row_infos, min_row, max_row, num_cols)
    columns: List[Dict[str, Any]] = []
    col_indices: List[int] = []
    header_lines: List[Dict[str, Any]] = []
    report_title = ""
    report_meta: Optional[str] = None
    data_preview: List[Dict[str, Any]] = []
    total_data_rows = 0

    if header_row is not None:
        columns, col_indices = _detect_columns(ws, header_row, min_col, max_col, anchor_map, hidden_set, row_infos)
        if columns:
            data_start = header_row + 1
            data_end, _, _, _ = _detect_data_zone(ws, row_infos, data_start, max_row, columns, min_col)
            _enrich_columns_with_data(ws, columns, col_indices, data_start, data_end)
            data_preview = _extract_data_preview(ws, columns, col_indices, data_start, min(data_start + 4, data_end))
            total_data_rows = max(0, data_end - data_start + 1)

        title_row = _best_effort_title_row(row_infos, min_row, header_row - 1) if header_row > min_row else None
        if title_row is not None:
            title_text = _joined_row_text(row_infos[title_row])
            report_title = _extract_explicit_title(title_text) or title_text

        for r in range(min_row, header_row):
            info = row_infos.get(r)
            if info is None or info.is_empty or info.numeric_ratio >= 0.45:
                continue

            row_text = _joined_row_text(info)
            if not row_text:
                continue
            if title_row is not None and r == title_row:
                continue

            if report_meta is None and _contains_keyword(info.texts, _META_KEYWORDS):
                report_meta = row_text
                continue

            header_lines.append({
                "text": row_text,
                "right_text": None,
                "align": "left",
                "bold": info.all_bold,
                "font_size": "sm" if info.non_empty_count <= 2 else "base",
            })

    if not columns:
        for r in range(min_row, max_row + 1):
            info = row_infos.get(r)
            if info and not info.is_empty:
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    text = str(cell.value).strip() if cell.value is not None else ""
                    if text:
                        key = _to_snake_case(text) or f"col_{len(columns)+1}"
                        columns.append({
                            "label": text, "key": key, "inferred_type": "text",
                            "width_px": 120, "align": "left", "format": "text",
                            "suffix": None, "bold": False, "highlight_negative": False,
                            "source_col_idx": c,
                        })
                        col_indices.append(c)
                break

    return {
        "header_lines": header_lines,
        "report_title": report_title,
        "report_meta": report_meta,
        "column_groups": [],
        "columns": columns,
        "group_by_column": None,
        "show_subtotals": False,
        "footer_lines": [],
        "signature_count": 0,
        "signature_labels": [],
        "theme": {"header_bg": "#073763", "header_text": "#ffffff",
                  "group_bg": "#c9daf8", "group_text": "#073763",
                  "subtotal_bg": "#dbeafe", "subtotal_text": "#1e40af",
                  "accent_color": "#073763"},
        "recommended_table_schema": [
            {
                "name": c["key"],
                "display_name": c["label"],
                "type": "number" if c["format"] in ("integer", "decimal", "percentage") else "string",
            }
            for c in columns
        ],
        "data_preview": data_preview,
        "total_data_rows": total_data_rows,
        "confidence": 0.25 if columns else 0.1,
        "sheet_names": all_sheet_names,
        "analyzed_sheet": sheet_name,
    }
