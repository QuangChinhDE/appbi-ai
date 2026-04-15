"""
ExcelExportService — render a ReportTemplate with live dataset data into an .xlsx file.

Flow:
  1. Walk template blocks → collect all DataFieldBinding references
  2. Group bindings by (dataset_id, table_id) source
  3. Fetch live rows from each source via LiveQueryService.execute_preview_query
  4. Resolve cell values (aggregate or row-by-row expansion)
  5. Write resolved blocks into an openpyxl workbook and return bytes
"""
from __future__ import annotations

import statistics
from datetime import date as dt_date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.models import DataSource, Dataset
from app.services import DatasetCRUDService
from app.services.dataset_calendar_service import is_generated_calendar_table
from app.services.dataset_table_sql_service import (
    DatasetTableSqlError,
    build_live_proxy_table_for_dataset_table,
    is_derived_table,
)
from app.services.live_query_service import LiveQueryService

logger = get_logger(__name__)

# Characters-per-pixel constant (mirrors excel_parser.py)
_CHAR_TO_PX = 7.5
_EXPORT_ROW_LIMIT = 5000


def _sheet_border(borders: dict[str, Any] | None) -> Border | None:
    if not borders:
        return None

    thin = Side(style="thin", color="000000")
    return Border(
        top=thin if borders.get("top") else Side(),
        right=thin if borders.get("right") else Side(),
        bottom=thin if borders.get("bottom") else Side(),
        left=thin if borders.get("left") else Side(),
    )


# ── Binding traversal (mirrors collectValueBindings in use-template-preview-data.ts) ──


def _collect_bindings_from_value(value: Any, out: list[dict]) -> None:
    if isinstance(value, list):
        for item in value:
            _collect_bindings_from_value(item, out)
    elif isinstance(value, dict) and value.get("type") == "field":
        out.append(value)


def _collect_bindings_from_block(block: dict, out: list[dict]) -> None:
    btype = block.get("type", "")
    cfg = block.get("config", {}) or {}

    if btype == "text":
        _collect_bindings_from_value(cfg.get("content"), out)
        return

    if btype == "table":
        for row in cfg.get("rows", []):
            for cell in row.get("cells", []):
                _collect_bindings_from_value(cell.get("value"), out)


def _source_key(binding: dict) -> str:
    return f"{binding['datasetId']}:{binding['tableId']}"


def _get_unique_sources(blocks: list[dict]) -> list[tuple[int, int]]:
    bindings: list[dict] = []
    for block in blocks:
        _collect_bindings_from_block(block, bindings)

    seen: set[str] = set()
    sources: list[tuple[int, int]] = []
    for b in bindings:
        key = _source_key(b)
        if key not in seen:
            seen.add(key)
            sources.append((int(b["datasetId"]), int(b["tableId"])))
    return sources


# ── Data fetching ──────────────────────────────────────────────────────────────


def _fetch_source_rows(
    db: Session,
    dataset_id: int,
    table_id: int,
    source_filters: list[dict],
) -> list[dict]:
    """Fetch rows from a single (dataset_id, table_id) with optional filters."""
    dataset_obj: Dataset | None = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset_obj:
        logger.warning("export: dataset %s not found", dataset_id)
        return []

    db_table = DatasetCRUDService.get_table_by_id(db, table_id)
    if not db_table or db_table.dataset_id != dataset_id:
        logger.warning("export: table %s not found in dataset %s", table_id, dataset_id)
        return []

    datasource: DataSource | None = None
    target_table = db_table

    if is_generated_calendar_table(db_table) or is_derived_table(db_table):
        try:
            datasource, target_table = build_live_proxy_table_for_dataset_table(
                db, dataset_obj, db_table
            )
        except DatasetTableSqlError as exc:
            logger.warning("export: proxy table error for table %s: %s", table_id, exc)
            return []
    else:
        datasource = db.query(DataSource).filter(DataSource.id == db_table.datasource_id).first()
        if not datasource:
            logger.warning("export: datasource not found for table %s", table_id)
            return []

    try:
        result = LiveQueryService.execute_preview_query(
            datasource=datasource,
            db_table=target_table,
            limit=_EXPORT_ROW_LIMIT,
            offset=0,
            filters=source_filters or None,
        )
        return result.get("rows", [])
    except Exception as exc:
        logger.warning("export: query failed for table %s: %s", table_id, exc)
        return []


# ── Filter mapping ─────────────────────────────────────────────────────────────


def _build_filters_by_source(
    template_filters: list[dict],
    active_filters: list[dict],
) -> dict[str, list[dict]]:
    """Map active_filter values onto their template filter definitions, group by source key."""
    # Index template filter defs by id
    filter_defs = {f["id"]: f for f in (template_filters or []) if "id" in f}

    result: dict[str, list[dict]] = {}
    for af in active_filters:
        fid = af.get("filterId") or af.get("filter_id")
        value = af.get("value")
        if not fid or value is None or value == "":
            continue

        fdef = filter_defs.get(fid)
        if not fdef:
            continue

        key = f"{fdef['datasetId']}:{fdef['tableId']}"
        if key not in result:
            result[key] = []

        result[key].append({
            "field": fdef["column"],
            "operator": fdef.get("operator", "eq"),
            "value": value,
        })

    return result


# ── Value resolution (mirrors aggregateBindingValue + resolveValue in TS) ──────


def _aggregate(rows: list[dict], column: str, agg: str) -> Any:
    if agg == "count":
        return len(rows)
    values = []
    for row in rows:
        v = row.get(column)
        try:
            values.append(float(v))
        except (TypeError, ValueError):
            pass

    if not values:
        return None

    if agg == "sum":
        return sum(values)
    if agg == "avg":
        return sum(values) / len(values)
    if agg == "min":
        return min(values)
    if agg == "max":
        return max(values)
    if agg == "last":
        return rows[-1].get(column) if rows else None
    # first / default
    return rows[0].get(column) if rows else None


def _resolve_binding(binding: dict, source_rows: list[dict], row_context: dict | None) -> Any:
    column = binding.get("column", "")
    agg = binding.get("agg")

    if agg:
        return _aggregate(source_rows, column, agg)

    row = row_context if row_context is not None else (source_rows[0] if source_rows else {})
    return row.get(column)


def _to_cell_value(raw: Any):
    """Convert a raw Python value to something suitable for an openpyxl cell."""
    if raw is None:
        return ""
    if isinstance(raw, (dt_date, datetime)):
        return raw
    if isinstance(raw, float):
        # Round floats that are whole numbers
        if raw == int(raw):
            return int(raw)
        return raw
    return raw


def _resolve_value(value: Any, source_map: dict[str, list[dict]], row_context: dict | None) -> Any:
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        if value.get("type") == "field":
            key = _source_key(value)
            rows = source_map.get(key, [])
            return _to_cell_value(_resolve_binding(value, rows, row_context))
        if value.get("type") == "formula":
            # Formulas are not evaluated — return expression as string
            return value.get("expression", "")
    return ""


def _is_repeating_row(row: dict) -> str | None:
    """Return source key if row should repeat once per data row, else None."""
    sources: set[str] = set()
    for cell in row.get("cells", []):
        v = cell.get("value")
        if isinstance(v, dict) and v.get("type") == "field" and not v.get("agg"):
            sources.add(_source_key(v))
    return list(sources)[0] if len(sources) == 1 else None


# ── Block resolution ──────────────────────────────────────────────────────────


def _resolve_table_block(block: dict, source_map: dict[str, list[dict]]) -> list[list[dict]]:
    """
    Return a list of resolved rows.
    Each resolved row is a list of cell dicts with 'value' replaced by scalar.
    """
    cfg = block.get("config", {}) or {}
    resolved_rows: list[list[dict]] = []

    for row in cfg.get("rows", []):
        repeat_key = None if row.get("isHeader") else _is_repeating_row(row)

        if repeat_key:
            data_rows = source_map.get(repeat_key, [])
            expand_over: list[dict | None] = data_rows if data_rows else [None]
        else:
            expand_over = [None]

        for data_row in expand_over:
            resolved_cells = []
            for cell in row.get("cells", []):
                resolved_cells.append({
                    **cell,
                    "value": _resolve_value(cell.get("value"), source_map, data_row),
                })
            resolved_rows.append(resolved_cells)

    return resolved_rows


def _resolve_text_block(block: dict, source_map: dict[str, list[dict]]) -> str:
    content = (block.get("config") or {}).get("content", "")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for seg in content:
            if isinstance(seg, str):
                parts.append(seg)
            else:
                parts.append(str(_resolve_value(seg, source_map, None) or ""))
        return "".join(parts)
    return ""


# ── openpyxl workbook builder ─────────────────────────────────────────────────

_HEX_COLORS = {
    "white": "FFFFFF", "black": "000000",
    "red": "FF0000", "blue": "0000FF", "green": "008000",
    "yellow": "FFFF00", "gray": "808080", "grey": "808080",
    "lightblue": "ADD8E6", "lightgreen": "90EE90", "lightyellow": "FFFFE0",
}


def _css_to_hex(css_color: str | None) -> str | None:
    if not css_color:
        return None
    c = css_color.strip().lower()
    if c.startswith("#"):
        h = c[1:]
        if len(h) == 3:
            h = "".join(ch * 2 for ch in h)
        return h.upper() if len(h) == 6 else None
    return _HEX_COLORS.get(c)


def _write_table_block(ws, block: dict, source_map: dict[str, list[dict]], start_row: int) -> int:
    """Write a table block starting at start_row; return next available row index."""
    cfg = block.get("config", {}) or {}
    column_widths: list[float] = cfg.get("columnWidths") or []
    resolved_rows = _resolve_table_block(block, source_map)

    if not resolved_rows:
        return start_row

    num_cols = max(len(r) for r in resolved_rows) if resolved_rows else 0

    # Set column widths from template
    for ci, px_w in enumerate(column_widths):
        col_letter = get_column_letter(ci + 1)
        ws.column_dimensions[col_letter].width = max(8.0, px_w / _CHAR_TO_PX)

    current_row = start_row
    for row_cells in resolved_rows:
        col_cursor = 1
        ci = 0
        while ci < len(row_cells):
            cell_def = row_cells[ci]
            value = cell_def.get("value", "")
            col_span: int = cell_def.get("colSpan") or 1
            is_hidden: bool = cell_def.get("hidden") or False

            if is_hidden:
                ci += 1
                col_cursor += 1
                continue

            xl_cell = ws.cell(row=current_row, column=col_cursor, value=value)

            # Formatting
            is_bold = bool(cell_def.get("bold"))
            is_italic = bool(cell_def.get("italic"))
            if is_bold or is_italic:
                xl_cell.font = Font(bold=is_bold, italic=is_italic)

            align = cell_def.get("align", "left")
            xl_cell.alignment = Alignment(
                horizontal=align,
                vertical="center",
                wrap_text=True,
            )

            bg = _css_to_hex(cell_def.get("bg"))
            if bg:
                xl_cell.fill = PatternFill(fill_type="solid", fgColor=bg)

            border = _sheet_border(cell_def.get("borders"))
            if border:
                xl_cell.border = border

            # Merge columns
            if col_span > 1:
                end_col = col_cursor + col_span - 1
                ws.merge_cells(
                    start_row=current_row,
                    start_column=col_cursor,
                    end_row=current_row,
                    end_column=end_col,
                )
                col_cursor = end_col + 1
            else:
                col_cursor += 1

            ci += 1

        current_row += 1

    return current_row


def _write_title_block(ws, block: dict, num_cols: int, start_row: int) -> int:
    cfg = block.get("config", {}) or {}
    text = cfg.get("text") or ""
    subtitle = cfg.get("subtitle") or ""

    if text:
        cell = ws.cell(row=start_row, column=1, value=text)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if num_cols > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
        start_row += 1

    if subtitle:
        cell = ws.cell(row=start_row, column=1, value=subtitle)
        cell.font = Font(size=10, color="595959")
        cell.alignment = Alignment(horizontal="center")
        if num_cols > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
        start_row += 1

    return start_row


def _write_text_block(ws, block: dict, source_map: dict[str, list[dict]], num_cols: int, start_row: int) -> int:
    text = _resolve_text_block(block, source_map)
    if text:
        cell = ws.cell(row=start_row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True)
        if num_cols > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_cols)
        start_row += 1
    return start_row


def _write_signature_block(ws, block: dict, num_cols: int, start_row: int) -> int:
    cfg = block.get("config", {}) or {}
    columns = cfg.get("columns") or []
    if not columns:
        return start_row + 2

    # Write titles in a row
    col_width = max(1, num_cols // len(columns)) if columns else 1
    for i, col_def in enumerate(columns):
        title = col_def.get("title", "") if isinstance(col_def, dict) else str(col_def)
        xl_col = i * col_width + 1
        cell = ws.cell(row=start_row, column=xl_col, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    start_row += 1
    # Blank line for actual signature
    start_row += 2
    # Subtitle row
    for i, col_def in enumerate(columns):
        subtitle = col_def.get("subtitle", "") if isinstance(col_def, dict) else ""
        if subtitle:
            xl_col = i * col_width + 1
            cell = ws.cell(row=start_row, column=xl_col, value=subtitle)
            cell.alignment = Alignment(horizontal="center")

    return start_row + 1


# ── Public API ────────────────────────────────────────────────────────────────


def export_template_to_excel(
    db: Session,
    template,
    active_filters: list[dict],
) -> bytes:
    """
    Render *template* with live dataset data and return xlsx bytes.

    Parameters
    ----------
    db:
        SQLAlchemy session.
    template:
        ReportTemplate ORM object (has .blocks, .filters, .name).
    active_filters:
        List of ``{"filterId": str, "value": Any}`` dicts from the request.
    """
    blocks: list[dict] = template.blocks or []
    template_filters: list[dict] = template.filters or []

    # ── 1. Collect unique sources ──
    sources = _get_unique_sources(blocks)

    # ── 2. Build filter map ──
    filters_by_source = _build_filters_by_source(template_filters, active_filters)

    # ── 3. Fetch rows per source ──
    source_map: dict[str, list[dict]] = {}
    for dataset_id, table_id in sources:
        key = f"{dataset_id}:{table_id}"
        source_filters = filters_by_source.get(key, [])
        rows = _fetch_source_rows(db, dataset_id, table_id, source_filters)
        source_map[key] = rows
        logger.info(
            "export: fetched %d rows for source %s (filters: %d)",
            len(rows), key, len(source_filters),
        )

    # ── 4. Determine max columns across table blocks ──
    max_cols = 1
    for block in blocks:
        if block.get("type") == "table":
            cfg = block.get("config", {}) or {}
            max_cols = max(max_cols, cfg.get("columns") or 1)

    # ── 5. Build workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = (template.name or "Report")[:31]  # Excel sheet name limit

    # Sort blocks top-to-bottom by layout.y
    sorted_blocks = sorted(blocks, key=lambda b: (b.get("layout") or {}).get("y", 0))

    current_row = 1
    for block in sorted_blocks:
        btype = block.get("type", "")

        if btype == "table":
            current_row = _write_table_block(ws, block, source_map, current_row)
        elif btype == "title":
            current_row = _write_title_block(ws, block, max_cols, current_row)
        elif btype == "text":
            current_row = _write_text_block(ws, block, source_map, max_cols, current_row)
        elif btype == "signature":
            current_row = _write_signature_block(ws, block, max_cols, current_row)
        elif btype in ("spacer", "image"):
            current_row += 1  # spacer gap

        # 1-row gap between blocks
        current_row += 1

    # ── 6. Serialize to bytes ──
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
