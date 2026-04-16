"""AI-assisted refinement for smart template import analysis.

This layer runs after deterministic parsing and only maps workbook context into
existing template-analysis fields. It must not invent new columns or change the
core data extraction path.
"""

from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from app.core.config import settings
from app.core.logging import get_logger
from app.services.excel_structure_detector import (
    _build_merge_map,
    _classify_row,
    _detect_columns,
    _detect_data_zone,
    _find_data_header_row,
)
from app.services.llm_client import LLMClient

logger = get_logger(__name__)

_ALLOWED_ALIGN = {"left", "center", "right"}
_ALLOWED_FONT_SIZES = {"sm", "base", "lg", "xl"}
_MAX_CELL_TEXT = 900
_MAX_ZONE_ROWS = 12
_THEME_KEYS = {
    "header_bg", "header_text", "group_bg", "group_text",
    "subtotal_bg", "subtotal_text", "accent_color",
}


def build_ai_assist_meta(
    *,
    requested: bool,
    applied: bool,
    status: str,
    provider: Optional[str] = None,
    model: Optional[str] = None,
    message: Optional[str] = None,
) -> Dict[str, Any]:
    return {
        "requested": requested,
        "applied": applied,
        "status": status,
        "provider": provider,
        "model": model,
        "message": message,
    }


def refine_import_analysis_with_ai(
    *,
    file_bytes: bytes,
    analysis: Dict[str, Any],
    filename: Optional[str],
    sheet_name: Optional[str],
    is_csv: bool,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    if is_csv:
        return analysis, build_ai_assist_meta(
            requested=True,
            applied=False,
            status="unsupported",
            message="AI layout refinement currently targets Excel workbooks. CSV still uses rule-based analysis.",
        )

    if not settings.template_import_ai_available:
        return analysis, build_ai_assist_meta(
            requested=True,
            applied=False,
            status="unavailable",
            message="Gemini refinement is not configured on the backend runtime yet.",
        )

    provider_candidates = _provider_candidates()
    if not provider_candidates:
        return analysis, build_ai_assist_meta(
            requested=True,
            applied=False,
            status="unavailable",
            message="No Gemini provider is configured for template import AI.",
        )

    try:
        context = _build_excel_refinement_context(file_bytes, analysis, sheet_name)
        failures: List[str] = []
        for provider, model in provider_candidates:
            try:
                payload = _call_import_ai(context, provider=provider, model=model)
                if not isinstance(payload, dict):
                    raise ValueError("AI returned no structured result.")

                refined = _merge_ai_overrides(analysis, payload, context)
                suffix = ""
                if failures:
                    suffix = f" Fallback applied after: {' | '.join(failures)}"
                return refined, build_ai_assist_meta(
                    requested=True,
                    applied=True,
                    status="applied",
                    provider=provider,
                    model=model,
                    message="Rule-based parse refined by Gemini mapping into template fields." + suffix,
                )
            except Exception as exc:
                failure = f"{provider}:{exc}"
                failures.append(failure)
                logger.warning("template import AI refinement failed with %s (%s)", provider, exc)

        return analysis, build_ai_assist_meta(
            requested=True,
            applied=False,
            status="failed",
            provider=provider_candidates[0][0],
            model=provider_candidates[0][1],
            message=f"AI refinement failed, fallbacking to rule-based output: {' | '.join(failures)}",
        )
    except Exception as exc:
        logger.warning("template import AI refinement failed: %s", exc)
        return analysis, build_ai_assist_meta(
            requested=True,
            applied=False,
            status="failed",
            provider=provider_candidates[0][0] if provider_candidates else None,
            model=provider_candidates[0][1] if provider_candidates else None,
            message=f"AI refinement failed, fallbacking to rule-based output: {exc}",
        )


def _provider_candidates() -> List[Tuple[str, str]]:
    candidates: List[Tuple[str, str]] = []
    if settings.GEMINI_API_KEY.strip():
        candidates.append(("gemini", settings.GEMINI_IMPORT_MODEL.strip() or "gemini-2.5-flash-lite"))
    if settings.active_api_keys:
        candidates.append(("openrouter-gemini", settings.OPENROUTER_GEMINI_IMPORT_MODEL.strip() or "google/gemini-2.5-flash-lite"))
    return candidates


def _call_import_ai(context: Dict[str, Any], *, provider: str, model: str) -> Optional[Dict[str, Any]]:
    system_prompt = (
        "You refine a deterministic Excel-to-template analysis for a report-template builder. "
        "The table extraction is already handled by code. Only improve how workbook presentation maps into existing template fields. "
        "Never invent columns, rows, signature slots, or business text that are not present in the provided workbook context. "
        "Do not change dataset shape. Only map fields compatible with the existing template module. "
        "Always return valid JSON."
    )

    user_prompt = json.dumps(
        {
            "task": "Refine the initial template format inside the current TemplateDefinition v3 surface: report_title, report_title_style, report_meta, header_lines, footer_lines, signature_count, signature_labels, column_groups, group_by_column, show_subtotals, theme, and column presentation overrides.",
            "rules": [
                "Keep deterministic columns unchanged: do not add, remove, rename, or re-key columns.",
                "You may refine column presentation only: width_px, align, bold, visible, suffix, highlight_negative, and label wording.",
                "Prefer returning row references for header/footer mapping instead of repeating long text blocks verbatim.",
                "Header rows must exclude the main report title when that title is captured in report_title or report_title_row.",
                "Footer rows should include notes/conditions only, not table data rows.",
                "group_by_column must be null or one of the provided column keys.",
                "column_groups must use start_col_idx/span relative to the provided detected columns array.",
                "report_title_style should improve the initial visual placement of the main title within the current builder, not invent a freeform layout.",
                "If a field is uncertain, keep it null or empty rather than inventing.",
            ],
            "return_json_shape": {
                "report_title": "string | null",
                "report_title_row": 0,
                "report_title_style": {
                    "align": "left|center|right",
                    "bold": True,
                    "font_size": "sm|base|lg|xl",
                },
                "report_meta": "string | null",
                "header_row_refs": [
                    {
                        "row": 0,
                        "align": "left|center|right",
                        "bold": True,
                        "font_size": "sm|base|lg|xl",
                    }
                ],
                "footer_row_refs": [
                    {
                        "row": 0,
                        "align": "left|center|right",
                        "bold": False,
                        "font_size": "sm|base|lg|xl",
                    }
                ],
                "signature_count": 0,
                "signature_labels": ["string"],
                "column_groups": [
                    {"label": "string", "start_col_idx": 0, "span": 0}
                ],
                "group_by_column": "string | null",
                "show_subtotals": True,
                "theme": {
                    "header_bg": "#000000",
                    "header_text": "#ffffff",
                    "group_bg": "#eeeeee",
                    "group_text": "#111111",
                    "subtotal_bg": "#f5f5f5",
                    "subtotal_text": "#111111",
                    "accent_color": "#000000",
                },
                "column_overrides": [
                    {
                        "key": "string",
                        "label": "string | null",
                        "width_px": 120,
                        "align": "left|center|right",
                        "bold": False,
                        "visible": True,
                        "suffix": "string | null",
                        "highlight_negative": False,
                    }
                ],
            },
            "context": context,
        },
        ensure_ascii=False,
    )

    if provider == "gemini":
        return _call_direct_gemini(user_prompt, system_prompt, model)
    return LLMClient.complete_json(user_prompt, system=system_prompt, model=model, max_tokens=1400)


def _call_direct_gemini(prompt: str, system_prompt: str, model: str) -> Optional[Dict[str, Any]]:
    try:
        import google.generativeai as genai
    except Exception as exc:  # pragma: no cover - depends on runtime package availability
        raise RuntimeError("google-generativeai package is not available") from exc

    genai.configure(api_key=settings.GEMINI_API_KEY)
    client = genai.GenerativeModel(
        model_name=model,
        system_instruction=system_prompt,
        generation_config={
            "temperature": 0.2,
            "max_output_tokens": 1400,
            "response_mime_type": "application/json",
        },
    )
    response = client.generate_content(prompt)
    text = _extract_gemini_text(response)
    if not text:
        return None
    return _parse_ai_json(text)


def _extract_gemini_text(response: Any) -> str:
    text = getattr(response, "text", "") or ""
    if text:
        return text

    for candidate in getattr(response, "candidates", []) or []:
        content = getattr(candidate, "content", None)
        for part in getattr(content, "parts", []) or []:
            part_text = getattr(part, "text", "") or ""
            if part_text:
                return part_text
    return ""


def _parse_ai_json(text: str) -> Dict[str, Any]:
    raw = str(text or "").strip()
    if not raw:
        raise ValueError("empty AI response")

    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\s*```$", "", raw).strip()

    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        raw = raw[start : end + 1]

    raw = re.sub(r",\s*([}\]])", r"\1", raw)
    return json.loads(raw)


def _build_excel_refinement_context(
    file_bytes: bytes,
    analysis: Dict[str, Any],
    sheet_name: Optional[str],
) -> Dict[str, Any]:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        min_row = worksheet.min_row or 1
        max_row = worksheet.max_row or 1
        min_col = worksheet.min_column or 1
        max_col = worksheet.max_column or 1
        num_cols = max_col - min_col + 1

        anchor_map, hidden_set = _build_merge_map(worksheet)
        row_infos = {
            row_idx: _classify_row(worksheet, row_idx, min_col, max_col, anchor_map, hidden_set, num_cols)
            for row_idx in range(min_row, max_row + 1)
        }
        data_header_row = _find_data_header_row(row_infos, min_row, max_row, num_cols)

        columns: List[Dict[str, Any]] = []
        data_end = max_row
        if data_header_row is not None:
            columns, _ = _detect_columns(worksheet, data_header_row, min_col, max_col, anchor_map, hidden_set)
            data_end, _, _, _ = _detect_data_zone(worksheet, row_infos, data_header_row + 1, max_row, columns, min_col)

        header_row_ids = [row_idx for row_idx in range(min_row, (data_header_row or min_row)) if not row_infos[row_idx].is_empty]
        footer_row_ids = [row_idx for row_idx in range(data_end + 1, max_row + 1) if not row_infos[row_idx].is_empty]

        return {
            "sheet_name": worksheet.title,
            "sheet_names": workbook.sheetnames,
            "dimensions": {
                "min_row": min_row,
                "max_row": max_row,
                "min_col": min_col,
                "max_col": max_col,
            },
            "deterministic_analysis": {
                "report_title": analysis.get("report_title"),
                "report_meta": analysis.get("report_meta"),
                "header_lines": analysis.get("header_lines", []),
                "column_groups": analysis.get("column_groups", []),
                "columns": [
                    {
                        "index": idx,
                        "key": column.get("key"),
                        "label": column.get("label"),
                        "width_px": column.get("width_px"),
                        "align": column.get("align"),
                        "bold": column.get("bold"),
                        "format": column.get("format"),
                        "suffix": column.get("suffix"),
                        "highlight_negative": column.get("highlight_negative"),
                    }
                    for idx, column in enumerate(analysis.get("columns", []))
                ],
                "group_by_column": analysis.get("group_by_column"),
                "show_subtotals": analysis.get("show_subtotals"),
                "footer_lines": [_trim_text(_normalize_optional_string(line) or "", 400) for line in analysis.get("footer_lines", [])],
                "signature_count": analysis.get("signature_count", 0),
                "signature_labels": analysis.get("signature_labels", []),
                "theme": analysis.get("theme", {}),
                "confidence": analysis.get("confidence"),
            },
            "header_candidate_rows": _summarize_rows(worksheet, header_row_ids[:_MAX_ZONE_ROWS], min_col, max_col, anchor_map, hidden_set),
            "data_header_row": _summarize_rows(worksheet, [data_header_row] if data_header_row is not None else [], min_col, max_col, anchor_map, hidden_set),
            "footer_candidate_rows": _summarize_rows(worksheet, footer_row_ids[:_MAX_ZONE_ROWS], min_col, max_col, anchor_map, hidden_set),
            "data_preview": analysis.get("data_preview", [])[:5],
        }
    finally:
        workbook.close()


def _summarize_rows(
    worksheet: Any,
    row_ids: List[int],
    min_col: int,
    max_col: int,
    anchor_map: Dict[Any, Any],
    hidden_set: set,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    mid_col_idx = max(0, (max_col - min_col) // 2)
    for row_idx in row_ids:
        cells: List[Dict[str, Any]] = []
        left_parts: List[str] = []
        right_parts: List[str] = []
        max_merge_span = 1
        all_bold = True
        for col_idx in range(min_col, max_col + 1):
            if (row_idx, col_idx) in hidden_set:
                continue
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = "" if cell.value is None else str(cell.value).strip()
            if not value:
                continue
            _, merge_span = anchor_map.get((row_idx, col_idx), (1, 1))
            align = getattr(cell.alignment, "horizontal", None)
            relative_col = col_idx - min_col
            max_merge_span = max(max_merge_span, merge_span)
            is_bold = bool(getattr(cell.font, "bold", False))
            all_bold = all_bold and is_bold
            if relative_col <= mid_col_idx:
                left_parts.append(_trim_text(value, 240))
            else:
                right_parts.append(_trim_text(value, 240))
            cells.append(
                {
                    "col": relative_col,
                    "text": _trim_text(value, _MAX_CELL_TEXT),
                    "merge_span": merge_span,
                    "bold": is_bold,
                    "align": align,
                }
            )
        if cells:
            row_text = " ".join(part for part in left_parts + right_parts if part).strip()
            rows.append({
                "row": row_idx,
                "text": row_text,
                "left_text": " ".join(left_parts).strip() or row_text,
                "right_text": " ".join(right_parts).strip() or None,
                "cell_count": len(cells),
                "max_merge_span": max_merge_span,
                "all_bold": all_bold,
                "cells": cells,
            })
    return rows


def _merge_ai_overrides(base: Dict[str, Any], overrides: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(base)
    column_keys = {str(col.get("key", "")) for col in base.get("columns", []) if col.get("key")}
    column_count = len(base.get("columns", []))
    header_candidates = context.get("header_candidate_rows", [])
    footer_candidates = context.get("footer_candidate_rows", [])

    title = _normalize_optional_string(overrides.get("report_title"))
    if title:
        merged["report_title"] = title
    elif "report_title_row" in overrides:
        title_from_row = _resolve_title_from_row_ref(header_candidates, overrides.get("report_title_row"))
        if title_from_row:
            merged["report_title"] = title_from_row

    if "report_meta" in overrides:
        meta = _normalize_optional_string(overrides.get("report_meta"))
        merged["report_meta"] = meta

    if "report_title_style" in overrides:
        merged["report_title_style"] = _normalize_title_style(overrides.get("report_title_style"))

    if "header_row_refs" in overrides:
        merged["header_lines"] = _resolve_row_refs(header_candidates, overrides.get("header_row_refs"), default_font_size="base")
    elif "header_lines" in overrides:
        merged["header_lines"] = _normalize_header_lines(overrides.get("header_lines"))

    if "footer_row_refs" in overrides:
        merged["footer_lines"] = _resolve_row_refs(footer_candidates, overrides.get("footer_row_refs"), default_font_size="sm")
    elif "footer_lines" in overrides:
        merged["footer_lines"] = _normalize_footer_lines(overrides.get("footer_lines"))

    if "theme" in overrides:
        merged["theme"] = _normalize_theme(overrides.get("theme"), merged.get("theme"))

    if "column_overrides" in overrides:
        merged["columns"] = _merge_column_overrides(merged.get("columns", []), overrides.get("column_overrides"))

    if "column_groups" in overrides:
        merged["column_groups"] = _normalize_column_groups(overrides.get("column_groups"), column_count)

    if "group_by_column" in overrides:
        group_by = overrides.get("group_by_column")
        merged["group_by_column"] = str(group_by) if group_by in column_keys else None

    if isinstance(overrides.get("show_subtotals"), bool):
        merged["show_subtotals"] = overrides["show_subtotals"]

    if "signature_labels" in overrides or "signature_count" in overrides:
        labels = [text for text in [_normalize_optional_string(item) for item in overrides.get("signature_labels", []) or []] if text]
        raw_count = overrides.get("signature_count")
        count = raw_count if isinstance(raw_count, int) and raw_count >= 0 else len(labels)
        if labels and count < len(labels):
            count = len(labels)
        merged["signature_labels"] = labels
        merged["signature_count"] = count

    return merged


def _normalize_optional_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_header_lines(value: Any) -> List[Dict[str, Any]]:
    if not isinstance(value, list):
        return []

    result: List[Dict[str, Any]] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        text = _normalize_optional_string(item.get("text"))
        right_text = _normalize_optional_string(item.get("right_text"))
        if not text and not right_text:
            continue
        align = str(item.get("align") or "left").lower()
        font_size = str(item.get("font_size") or "base").lower()
        result.append(
            {
                "text": text or "",
                "right_text": right_text,
                "align": align if align in _ALLOWED_ALIGN else "left",
                "bold": bool(item.get("bold", False)),
                "font_size": font_size if font_size in _ALLOWED_FONT_SIZES else "base",
            }
        )
    return result


def _normalize_title_style(value: Any) -> Dict[str, Any]:
    if not isinstance(value, dict):
        return {}

    normalized: Dict[str, Any] = {}
    align = str(value.get("align") or "").lower()
    if align in _ALLOWED_ALIGN:
        normalized["align"] = align
    font_size = str(value.get("font_size") or value.get("fontSize") or "").lower()
    if font_size in _ALLOWED_FONT_SIZES:
        normalized["font_size"] = font_size
    if "bold" in value:
        normalized["bold"] = bool(value.get("bold"))
    return normalized


def _resolve_title_from_row_ref(candidates: List[Dict[str, Any]], value: Any) -> Optional[str]:
    if not isinstance(value, int):
        return None
    candidate = next((row for row in candidates if row.get("row") == value), None)
    if not candidate:
        return None
    return _normalize_optional_string(candidate.get("text"))


def _resolve_row_refs(
    candidates: List[Dict[str, Any]],
    value: Any,
    *,
    default_font_size: str,
) -> List[Dict[str, Any]]:
    if not isinstance(value, list):
        return []

    by_row = {
        int(row.get("row")): row
        for row in candidates
        if isinstance(row, dict) and isinstance(row.get("row"), int)
    }
    lines: List[Dict[str, Any]] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        row_id = item.get("row")
        if not isinstance(row_id, int) or row_id not in by_row:
            continue
        row = by_row[row_id]
        align = str(item.get("align") or "left").lower()
        font_size = str(item.get("font_size") or default_font_size).lower()
        text = _normalize_optional_string(row.get("left_text")) or _normalize_optional_string(row.get("text")) or ""
        right_text = _normalize_optional_string(row.get("right_text"))
        if not text and not right_text:
            continue
        lines.append({
            "text": text,
            "right_text": right_text,
            "align": align if align in _ALLOWED_ALIGN else "left",
            "bold": bool(item.get("bold", row.get("all_bold", False))),
            "font_size": font_size if font_size in _ALLOWED_FONT_SIZES else default_font_size,
        })
    return lines


def _normalize_footer_lines(value: Any) -> List[Dict[str, Any] | str]:
    if not isinstance(value, list):
        return []

    lines: List[Dict[str, Any] | str] = []
    for item in value:
        if isinstance(item, dict):
            line = _normalize_header_lines([item])
            if line:
                lines.append(line[0])
            continue

        text = _normalize_optional_string(item)
        if text:
            lines.append(text)
    return lines


def _normalize_theme(value: Any, base_theme: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    theme = dict(base_theme or {})
    if not isinstance(value, dict):
        return theme

    for key, raw in value.items():
        normalized_key = str(key or "").strip().lower()
        if normalized_key not in _THEME_KEYS:
            continue
        text = _normalize_optional_string(raw)
        if text:
            theme[normalized_key] = text
    return theme


def _merge_column_overrides(columns: List[Dict[str, Any]], value: Any) -> List[Dict[str, Any]]:
    if not isinstance(value, list):
        return columns

    by_key = {
        str(col.get("key")): dict(col)
        for col in columns
        if col.get("key")
    }
    ordered_keys = [str(col.get("key")) for col in columns if col.get("key")]

    for item in value:
        if not isinstance(item, dict):
            continue
        key = str(item.get("key") or "").strip()
        if not key or key not in by_key:
            continue

        current = by_key[key]
        label = _normalize_optional_string(item.get("label"))
        if label:
            current["label"] = label

        width = item.get("width_px")
        if isinstance(width, (int, float)):
            current["width_px"] = round(max(60.0, min(float(width), 720.0)), 1)

        align = str(item.get("align") or "").lower()
        if align in _ALLOWED_ALIGN:
            current["align"] = align

        if "bold" in item:
            current["bold"] = bool(item.get("bold"))
        if "visible" in item:
            current["visible"] = bool(item.get("visible"))
        if "highlight_negative" in item:
            current["highlight_negative"] = bool(item.get("highlight_negative"))

        suffix = item.get("suffix")
        if suffix is None:
            pass
        else:
            current["suffix"] = _normalize_optional_string(suffix)

        by_key[key] = current

    return [by_key[key] for key in ordered_keys if key in by_key]


def _normalize_column_groups(value: Any, column_count: int) -> List[Dict[str, Any]]:
    if not isinstance(value, list):
        return []

    groups: List[Dict[str, Any]] = []
    for item in value:
        if not isinstance(item, dict):
            continue
        label = _normalize_optional_string(item.get("label"))
        start = item.get("start_col_idx")
        span = item.get("span")
        if not label or not isinstance(start, int) or not isinstance(span, int):
            continue
        if start < 0 or span <= 0 or start >= column_count:
            continue
        safe_span = min(span, column_count - start)
        if safe_span <= 0:
            continue
        groups.append({
            "label": label,
            "start_col_idx": start,
            "span": safe_span,
        })
    return groups


def _trim_text(value: str, max_len: int) -> str:
    text = str(value or "").strip()
    if len(text) <= max_len:
        return text
    return text[: max_len - 1].rstrip() + "…"