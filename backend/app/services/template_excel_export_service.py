"""Server-side Excel export for TemplateDefinition v3 runtime."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import DataSource, Dataset
from app.services.dataset_calendar_service import is_generated_calendar_table
from app.services.dataset_crud import DatasetCRUDService
from app.services.dataset_table_sql_service import (
    DatasetTableSqlError,
    build_live_proxy_table_for_dataset_table,
    is_derived_table,
)
from app.services.live_query_service import LiveQueryService
from app.services.template_document_runtime_service import build_template_document_runtime_preview
from app.services.template_document_schema import is_template_document_definition, normalize_template_document
from app.services.template_runtime_service import (
    build_template_preview_table_proxy,
    resolve_template_runtime_filters,
)

_EXPORT_LIMIT = 5000
_DEFAULT_THEME = {
    "headerBg": "#073763",
    "headerText": "#ffffff",
    "groupBg": "#c9daf8",
    "groupText": "#073763",
    "subtotalBg": "#dbeafe",
    "subtotalText": "#1e40af",
}

_DOCUMENT_SHEET_MAX_COLUMNS = 8


def _hex_color(value: str | None, fallback: str) -> str:
    raw = str(value or fallback).strip().lstrip("#")
    if len(raw) == 3:
        raw = "".join(ch * 2 for ch in raw)
    return raw.upper() if len(raw) == 6 else fallback.lstrip("#").upper()


def _format_display_value(value: Any, fmt: str | None, suffix: str | None) -> str:
    if value is None or value == "":
        return "-"

    format_name = str(fmt or "text").lower()
    number_value: float | None = None
    try:
        number_value = float(value)
    except (TypeError, ValueError):
        number_value = None

    if format_name == "integer" and number_value is not None:
        rendered = f"{round(number_value):,}"
    elif format_name == "decimal" and number_value is not None:
        rendered = f"{number_value:,.2f}"
    elif format_name == "percentage" and number_value is not None:
        rendered = f"{round(number_value * 100)}%"
    else:
        rendered = str(value)

    return f"{rendered} {suffix}".strip() if suffix else rendered


def _raw_column_value(column: Dict[str, Any], row: Dict[str, Any]) -> Any:
    if column.get("expression"):
        return row.get(column.get("key"))
    source_column = column.get("sourceColumn") or column.get("key")
    return row.get(source_column)


def _numeric_column_value(column: Dict[str, Any], row: Dict[str, Any]) -> float:
    raw = _raw_column_value(column, row)
    try:
        return float(raw)
    except (TypeError, ValueError):
        return 0.0


def _group_rows(
    rows: List[Dict[str, Any]],
    group_by_key: str | None,
    columns: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    if not group_by_key:
        return [{"label": None, "rows": rows, "count": len(rows)}]

    group_column = next((col for col in columns if col.get("key") == group_by_key), None)
    source_column = (group_column or {}).get("sourceColumn") or group_by_key

    groups: Dict[str, List[Dict[str, Any]]] = {}
    for row in rows:
        key = str(row.get(source_column) or "Other")
        groups.setdefault(key, []).append(row)

    return [
        {"label": label, "rows": grouped_rows, "count": len(grouped_rows)}
        for label, grouped_rows in groups.items()
    ]


def _build_group_header_rows(
    visible_columns: List[Dict[str, Any]],
    column_groups: List[Dict[str, Any]],
) -> List[List[Dict[str, Any]]]:
    if not column_groups:
        return []

    normalized_groups = []
    for group in column_groups:
        level = int(group.get("level") or 1)
        normalized_groups.append({**group, "level": max(1, level)})

    rows: List[List[Dict[str, Any]]] = []
    for level in sorted({int(group["level"]) for group in normalized_groups}):
        groups_at_level = [group for group in normalized_groups if int(group["level"]) == level]
        group_map: Dict[str, Dict[str, Any]] = {}
        for group in groups_at_level:
            for column_id in group.get("columnIds") or []:
                group_map[str(column_id)] = group

        cells: List[Dict[str, Any]] = []
        index = 0
        has_group = False
        while index < len(visible_columns):
            column = visible_columns[index]
            group = group_map.get(str(column.get("id")))
            if group:
                group_column_ids = {str(column_id) for column_id in group.get("columnIds") or []}
                span = 0
                while index + span < len(visible_columns) and str(visible_columns[index + span].get("id")) in group_column_ids:
                    span += 1
                cells.append({"label": str(group.get("label") or ""), "span": max(1, span)})
                has_group = True
                index += max(1, span)
            else:
                cells.append({"label": "", "span": 1})
                index += 1

        if has_group:
            rows.append(cells)

    return rows


def _build_appendix_definition(
    definition: Dict[str, Any],
    section: Dict[str, Any],
) -> Dict[str, Any] | None:
    column_map = {
        str(column.get("key")): column
        for column in list(definition.get("columns") or [])
        if column.get("key")
    }
    section_columns = []
    for column_key in section.get("columnKeys") or []:
        column = column_map.get(str(column_key))
        if column:
            section_columns.append({**column, "visible": True})

    if not section_columns:
        return None

    section_column_ids = {str(column.get("id")) for column in section_columns if column.get("id")}
    section_column_groups = []
    for group in list(definition.get("columnGroups") or []):
        group_column_ids = [
            str(column_id)
            for column_id in group.get("columnIds") or []
            if str(column_id) in section_column_ids
        ]
        if group_column_ids:
            section_column_groups.append({**group, "columnIds": group_column_ids})

    return {
        **definition,
        "layout": "table",
        "columns": section_columns,
        "groupBy": section.get("groupBy"),
        "showSubtotals": section.get("showSubtotals", False),
        "columnGroups": section_column_groups or None,
        "appendixSections": None,
    }


def _resolve_runtime_source(db: Session, dataset_id: int, table_id: int):
    dataset_obj = db.query(Dataset).filter(Dataset.id == dataset_id).first()
    if not dataset_obj:
        raise ValueError("Dataset not found")

    db_table = DatasetCRUDService.get_table_by_id(db, table_id)
    if not db_table or db_table.dataset_id != dataset_id:
        raise ValueError("Table not found in this dataset")

    datasource: DataSource | None = None
    target_table = db_table
    if is_generated_calendar_table(db_table) or is_derived_table(db_table):
        try:
            datasource, target_table = build_live_proxy_table_for_dataset_table(
                db,
                dataset_obj,
                db_table,
            )
        except DatasetTableSqlError as exc:
            raise ValueError(str(exc)) from exc
    else:
        datasource = db.query(DataSource).filter(DataSource.id == db_table.datasource_id).first()
        if not datasource:
            raise ValueError("Datasource not found")

    return datasource, target_table


def _fetch_runtime_rows(
    db: Session,
    template: Any,
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[str]]:
    definition = template.blocks if isinstance(template.blocks, dict) else {}
    data_source = definition.get("dataSource") or {}
    dataset_id = int(data_source.get("datasetId") or 0)
    table_id = int(data_source.get("tableId") or 0)
    columns = list(definition.get("columns") or [])

    if dataset_id <= 0 or table_id <= 0:
        return definition, [], []

    datasource, target_table = _resolve_runtime_source(db, dataset_id, table_id)
    preview_table, formula_errors = build_template_preview_table_proxy(target_table, columns)
    if formula_errors:
        details = "; ".join(f"{item['key']}: {item['error']}" for item in formula_errors)
        raise ValueError(f"Some template formulas are invalid. {details}")

    runtime_filters = resolve_template_runtime_filters(
        template.filters or [],
        [],
        dataset_id=dataset_id,
        table_id=table_id,
    )
    result = LiveQueryService.execute_preview_query(
        datasource=datasource,
        db_table=preview_table,
        limit=_EXPORT_LIMIT,
        offset=0,
        filters=runtime_filters or None,
    )
    rows = list(result.get("rows") or [])
    output_columns = [str(column) for column in list(result.get("columns") or [])]
    return definition, rows, output_columns


def _merge_row(ws, row_index: int, start_column: int, end_column: int, value: Any):
    cell = ws.cell(row=row_index, column=start_column, value=value)
    if end_column > start_column:
        ws.merge_cells(start_row=row_index, start_column=start_column, end_row=row_index, end_column=end_column)
    return cell


def _write_header(ws, definition: Dict[str, Any], total_columns: int, row_index: int) -> int:
    header = definition.get("header") or {}
    header_lines = list(header.get("lines") or [])
    for item in header_lines:
        text = str((item or {}).get("text") or "")
        right_text = str((item or {}).get("rightText") or "")
        merged = text if not right_text else f"{text}    {right_text}".strip()
        cell = _merge_row(ws, row_index, 1, total_columns, merged)
        cell.font = Font(bold=bool((item or {}).get("bold")))
        cell.alignment = Alignment(horizontal=str((item or {}).get("align") or "left"), vertical="center")
        row_index += 1

    title = str(header.get("title") or "").strip()
    if title:
        cell = _merge_row(ws, row_index, 1, total_columns, title)
        cell.font = Font(bold=header.get("titleBold", True), size=14)
        cell.alignment = Alignment(horizontal=str(header.get("titleAlign") or "center"), vertical="center")
        row_index += 1

    meta = str(header.get("meta") or "").strip()
    if meta:
        cell = _merge_row(ws, row_index, 1, total_columns, meta)
        cell.font = Font(italic=True, size=10)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        row_index += 1

    return row_index


def _write_footer(ws, definition: Dict[str, Any], total_columns: int, row_index: int) -> int:
    footer = definition.get("footer") or {}
    for item in footer.get("lines") or []:
        cell = _merge_row(ws, row_index, 1, total_columns, str((item or {}).get("text") or ""))
        cell.font = Font(bold=bool((item or {}).get("bold")))
        cell.alignment = Alignment(horizontal=str((item or {}).get("align") or "left"), vertical="center")
        row_index += 1

    signature_slots = max(0, int(footer.get("signatureSlots") or 0))
    labels = list(footer.get("signatureLabels") or [])
    if signature_slots > 0:
        row_index += 1
        slot_width = max(1, total_columns // max(1, signature_slots))
        for index in range(signature_slots):
            start_column = 1 + index * slot_width
            end_column = min(total_columns, start_column + slot_width - 1)
            _merge_row(ws, row_index, start_column, end_column, "____________________")
            label_cell = _merge_row(ws, row_index + 1, start_column, end_column, labels[index] if index < len(labels) else f"Signature {index + 1}")
            label_cell.alignment = Alignment(horizontal="center")
        row_index += 2

    return row_index


def _write_table_section(
    ws,
    definition: Dict[str, Any],
    rows: List[Dict[str, Any]],
    row_index: int,
    *,
    section_title: str | None = None,
    section_description: str | None = None,
) -> Tuple[int, int]:
    columns = list(definition.get("columns") or [])
    visible_columns = [column for column in columns if column.get("visible", True) is not False]
    theme = {**_DEFAULT_THEME, **(definition.get("theme") or {})}
    total_columns = max(1, len(visible_columns) + 1)

    if section_title:
        section_cell = _merge_row(ws, row_index, 1, total_columns, section_title)
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=_hex_color(theme.get("sectionBg") or theme.get("groupBg"), _DEFAULT_THEME["groupBg"]),
        )
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        row_index += 1
    if section_description:
        description_cell = _merge_row(ws, row_index, 1, total_columns, section_description)
        description_cell.font = Font(italic=True, size=10)
        description_cell.alignment = Alignment(horizontal="left", vertical="center")
        row_index += 1

    if not visible_columns:
        cell = _merge_row(ws, row_index, 1, total_columns, "No columns configured.")
        cell.alignment = Alignment(horizontal="center")
        return row_index + 1, total_columns

    for excel_column, column in enumerate(visible_columns, start=2):
        width = max(12, int((column.get("width") or 120) / 9))
        ws.column_dimensions[get_column_letter(excel_column)].width = width
    ws.column_dimensions["A"].width = 6

    for header_row in _build_group_header_rows(visible_columns, list(definition.get("columnGroups") or [])):
        current_column = 2
        for item in header_row:
            start_column = current_column
            end_column = current_column + max(1, int(item["span"])) - 1
            cell = _merge_row(ws, row_index, start_column, end_column, item["label"])
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=_hex_color(theme.get("headerBg"), _DEFAULT_THEME["headerBg"]),
            )
            current_column = end_column + 1
        row_index += 1

    header_fill = PatternFill(fill_type="solid", fgColor=_hex_color(theme.get("headerBg"), _DEFAULT_THEME["headerBg"]))
    header_font = Font(bold=True, color=_hex_color(theme.get("headerText"), _DEFAULT_THEME["headerText"]))
    ws.cell(row=row_index, column=1, value="#").font = header_font
    ws.cell(row=row_index, column=1).fill = header_fill
    for column_index, column in enumerate(visible_columns, start=2):
        cell = ws.cell(row=row_index, column=column_index, value=str(column.get("label") or column.get("key") or "Column"))
        cell.font = header_font
        cell.fill = header_fill
    row_index += 1

    grouped_rows = _group_rows(rows, definition.get("groupBy"), columns)
    subtotal_fill = PatternFill(fill_type="solid", fgColor=_hex_color(theme.get("subtotalBg"), _DEFAULT_THEME["subtotalBg"]))
    group_fill = PatternFill(fill_type="solid", fgColor=_hex_color(theme.get("groupBg"), _DEFAULT_THEME["groupBg"]))
    show_subtotals = definition.get("showSubtotals") is not False

    for group in grouped_rows:
        if group.get("label"):
            group_cell = _merge_row(ws, row_index, 1, total_columns, str(group["label"]))
            group_cell.font = Font(bold=True, color=_hex_color(theme.get("groupText"), _DEFAULT_THEME["groupText"]))
            group_cell.fill = group_fill
            row_index += 1

        group_rows = list(group.get("rows") or [])
        for item_index, row in enumerate(group_rows, start=1):
            ws.cell(row=row_index, column=1, value=item_index)
            for column_index, column in enumerate(visible_columns, start=2):
                value = _format_display_value(
                    _raw_column_value(column, row),
                    column.get("format"),
                    column.get("suffix"),
                )
                ws.cell(row=row_index, column=column_index, value=value)
            row_index += 1

        if show_subtotals and group_rows:
            subtotal_label = f"Subtotal ({len(group_rows)} rows)"
            subtotal_cell = ws.cell(row=row_index, column=1, value=subtotal_label)
            subtotal_cell.font = Font(bold=True)
            subtotal_cell.fill = subtotal_fill
            for column_index, column in enumerate(visible_columns, start=2):
                if column.get("type") in {"formula", "subtotal"} or column.get("format") in {"integer", "decimal"}:
                    total = sum(_numeric_column_value(column, row) for row in group_rows)
                    rendered = _format_display_value(total, column.get("format"), column.get("suffix"))
                    cell = ws.cell(row=row_index, column=column_index, value=rendered)
                    cell.font = Font(bold=True)
                    cell.fill = subtotal_fill
                else:
                    ws.cell(row=row_index, column=column_index, value="")
                    ws.cell(row=row_index, column=column_index).fill = subtotal_fill
            row_index += 1

    if not rows:
        cell = _merge_row(ws, row_index, 1, total_columns, "No data rows.")
        cell.alignment = Alignment(horizontal="center")
        row_index += 1

    return row_index, total_columns


def _write_table_layout(ws, definition: Dict[str, Any], rows: List[Dict[str, Any]], row_index: int) -> int:
    visible_columns = [
        column for column in list(definition.get("columns") or []) if column.get("visible", True) is not False
    ]
    footer_total_columns = max(1, len(visible_columns) + 1)
    row_index = _write_header(ws, definition, footer_total_columns, row_index)
    row_index, main_total_columns = _write_table_section(ws, definition, rows, row_index)
    footer_total_columns = max(footer_total_columns, main_total_columns)

    for section in list(definition.get("appendixSections") or []):
        appendix_definition = _build_appendix_definition(definition, section)
        if not appendix_definition:
            continue
        row_index += 2
        row_index, appendix_total_columns = _write_table_section(
            ws,
            appendix_definition,
            rows,
            row_index,
            section_title=str(section.get("title") or "Appendix"),
            section_description=str(section.get("description") or "").strip() or None,
        )
        footer_total_columns = max(footer_total_columns, appendix_total_columns)

    return _write_footer(ws, definition, footer_total_columns, row_index)


def _write_card_layout(ws, definition: Dict[str, Any], rows: List[Dict[str, Any]], row_index: int) -> int:
    columns = list(definition.get("columns") or [])
    card_config = definition.get("cardConfig") or {}
    title_key = card_config.get("titleColumn")
    subtitle_keys = list(card_config.get("subtitleColumns") or [])
    deduction_keys = set(card_config.get("deductionColumns") or [])
    total_label = card_config.get("totalLabel")
    total_columns = 4

    row_index = _write_header(ws, definition, total_columns, row_index)
    title_column = next((column for column in columns if column.get("key") == title_key), None)
    subtitle_columns = [column for column in columns if column.get("key") in subtitle_keys]
    body_columns = [column for column in columns if column.get("visible", True) is not False and column.get("key") not in {title_key, *subtitle_keys}]

    if not rows:
        _merge_row(ws, row_index, 1, total_columns, "No data rows.")
        return _write_footer(ws, definition, total_columns, row_index + 1)

    for card_index, row in enumerate(rows, start=1):
        title = _format_display_value(_raw_column_value(title_column or {"key": "title"}, row), None, None) if title_column else f"Row {card_index}"
        subtitle = " | ".join(
            _format_display_value(_raw_column_value(column, row), column.get("format"), column.get("suffix"))
            for column in subtitle_columns
        )

        title_cell = _merge_row(ws, row_index, 1, total_columns, title)
        title_cell.font = Font(bold=True, size=12)
        row_index += 1
        if subtitle:
            subtitle_cell = _merge_row(ws, row_index, 1, total_columns, subtitle)
            subtitle_cell.font = Font(italic=True, size=10)
            row_index += 1

        for column in body_columns:
            ws.cell(row=row_index, column=1, value=str(column.get("label") or column.get("key") or "Column"))
            value = _format_display_value(_raw_column_value(column, row), column.get("format"), column.get("suffix"))
            if column.get("key") in deduction_keys and value not in {"-", ""} and not str(value).startswith("-"):
                value = f"-{value}"
            _merge_row(ws, row_index, 2, total_columns, value)
            row_index += 1

        if total_label:
            subtotal_values = [
                _format_display_value(_raw_column_value(column, row), column.get("format"), column.get("suffix"))
                for column in columns if column.get("type") == "subtotal"
            ]
            ws.cell(row=row_index, column=1, value=str(total_label))
            _merge_row(ws, row_index, 2, total_columns, " + ".join(subtotal_values))
            row_index += 1

        row_index += 1

    return _write_footer(ws, definition, total_columns, row_index)


def _write_cross_tab_layout(ws, definition: Dict[str, Any], rows: List[Dict[str, Any]], row_index: int) -> int:
    columns = list(definition.get("columns") or [])
    config = definition.get("crossTabConfig") or {}
    pivot_column = next((column for column in columns if column.get("key") == config.get("pivotColumn")), None)
    value_column = next((column for column in columns if column.get("key") == config.get("valueColumn")), None)
    row_columns = [column for column in columns if column.get("key") in set(config.get("rowColumns") or [])]
    total_columns = max(3, len(row_columns) + 1)

    row_index = _write_header(ws, definition, total_columns, row_index)
    if not pivot_column or not value_column or not row_columns:
        _merge_row(ws, row_index, 1, total_columns, "Cross-tab is not fully configured.")
        return _write_footer(ws, definition, total_columns, row_index + 1)

    pivot_source = pivot_column.get("sourceColumn") or pivot_column.get("key")
    value_source = value_column.get("sourceColumn") or value_column.get("key")
    pivot_values = sorted({str(row.get(pivot_source) or "") for row in rows})

    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        row_key = "||".join(str(row.get((column.get("sourceColumn") or column.get("key")) or "")) for column in row_columns)
        entry = grouped.setdefault(row_key, {"rowData": row, "values": {}, "total": 0.0})
        pivot_value = str(row.get(pivot_source) or "")
        numeric_value = _numeric_column_value({"key": value_source}, row)
        entry["values"][pivot_value] = entry["values"].get(pivot_value, 0.0) + numeric_value
        entry["total"] += numeric_value

    pivot_rows = list(grouped.values())
    col_totals = {pivot_value: 0.0 for pivot_value in pivot_values}
    for entry in pivot_rows:
        for pivot_value in pivot_values:
            col_totals[pivot_value] += float(entry["values"].get(pivot_value, 0.0))

    total_columns = len(row_columns) + len(pivot_values) + (1 if config.get("showRowTotal", True) else 0)
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    column_index = 1
    for column in row_columns:
        cell = ws.cell(row=row_index, column=column_index, value=str(column.get("label") or column.get("key") or "Column"))
        cell.fill = header_fill
        cell.font = header_font
        column_index += 1
    for pivot_value in pivot_values:
        cell = ws.cell(row=row_index, column=column_index, value=pivot_value)
        cell.fill = header_fill
        cell.font = header_font
        column_index += 1
    if config.get("showRowTotal", True):
        cell = ws.cell(row=row_index, column=column_index, value="Total")
        cell.fill = header_fill
        cell.font = header_font
    row_index += 1

    for entry in pivot_rows:
        column_index = 1
        for column in row_columns:
            ws.cell(
                row=row_index,
                column=column_index,
                value=str(entry["rowData"].get(column.get("sourceColumn") or column.get("key")) or ""),
            )
            column_index += 1
        for pivot_value in pivot_values:
            value = entry["values"].get(pivot_value)
            ws.cell(
                row=row_index,
                column=column_index,
                value=_format_display_value(value, value_column.get("format"), value_column.get("suffix")) if value else "-",
            )
            column_index += 1
        if config.get("showRowTotal", True):
            ws.cell(
                row=row_index,
                column=column_index,
                value=_format_display_value(entry["total"], value_column.get("format"), value_column.get("suffix")),
            )
        row_index += 1

    if config.get("showColumnTotal", True) and pivot_rows:
        column_index = 1
        for row_column_index in range(len(row_columns)):
            ws.cell(row=row_index, column=column_index, value="TOTAL" if row_column_index == 0 else "")
            column_index += 1
        for pivot_value in pivot_values:
            ws.cell(
                row=row_index,
                column=column_index,
                value=_format_display_value(col_totals[pivot_value], value_column.get("format"), value_column.get("suffix")),
            )
            column_index += 1
        if config.get("showRowTotal", True):
            ws.cell(
                row=row_index,
                column=column_index,
                value=_format_display_value(sum(col_totals.values()), value_column.get("format"), value_column.get("suffix")),
            )
        row_index += 1

    if not rows:
        _merge_row(ws, row_index, 1, total_columns, "No data rows.")
        row_index += 1

    return _write_footer(ws, definition, total_columns, row_index)


def _serialize_preview_rows(rows: List[Any], columns: List[Any]) -> List[Dict[str, Any]]:
    serialized_rows: List[Dict[str, Any]] = []
    normalized_columns = [str(column) for column in columns]
    for row in rows:
        if isinstance(row, dict):
            serialized_rows.append(row)
            continue
        serialized_rows.append({
            normalized_columns[index]: row[index]
            for index in range(min(len(normalized_columns), len(row)))
        })
    return serialized_rows


def _fetch_document_source_previews(
    db: Session,
    definition: Dict[str, Any],
) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    source_previews: Dict[str, Dict[str, Any]] = {}
    warnings: List[str] = []

    for raw_source in definition.get("dataSources") or []:
        if not isinstance(raw_source, dict):
            continue

        source_id = str(raw_source.get("id") or "").strip()
        source_kind = str(raw_source.get("kind") or "").strip().lower()
        if not source_id:
            continue
        if source_kind != "dataset_table":
            warnings.append(f"Source '{source_id}' uses unsupported kind '{source_kind or 'unknown'}' for export.")
            continue

        try:
            dataset_id = int(raw_source.get("datasetId") or 0)
            table_id = int(raw_source.get("tableId") or 0)
        except (TypeError, ValueError):
            warnings.append(f"Source '{source_id}' is missing a valid dataset/table binding.")
            continue

        if dataset_id <= 0 or table_id <= 0:
            warnings.append(f"Source '{source_id}' is missing a valid dataset/table binding.")
            continue

        datasource, target_table = _resolve_runtime_source(db, dataset_id, table_id)
        result = LiveQueryService.execute_preview_query(
            datasource=datasource,
            db_table=target_table,
            limit=_EXPORT_LIMIT,
            offset=0,
            filters=None,
        )
        columns = [str(column) for column in list(result.get("columns") or [])]
        rows = _serialize_preview_rows(list(result.get("rows") or []), columns)
        source_previews[source_id] = {
            "sourceId": source_id,
            "datasetId": dataset_id,
            "tableId": table_id,
            "columns": [
                {"name": column_name, "type": "string", "nullable": True}
                for column_name in columns
            ],
            "rows": rows,
            "total": len(rows),
        }

    return source_previews, warnings


def _format_document_preview_value(value: Any) -> Any:
    if value is None or value == "":
        return "-"
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        return f"{value:,.2f}"
    return str(value)


def _write_document_table(
    ws,
    row_index: int,
    depth: int,
    title: str,
    preview: Dict[str, Any],
) -> int:
    label = f"{'  ' * depth}{title}"
    title_cell = _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, label)
    title_cell.font = Font(bold=True)
    title_cell.fill = PatternFill(fill_type="solid", fgColor="E2E8F0")
    row_index += 1

    columns = [str(column) for column in list(preview.get("columns") or [])]
    rows = list(preview.get("rows") or [])
    if not columns:
        _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, "No preview rows available.")
        return row_index + 2

    for column_index, column_name in enumerate(columns[:_DOCUMENT_SHEET_MAX_COLUMNS], start=1):
        cell = ws.cell(row=row_index, column=column_index, value=column_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1E293B")
    row_index += 1

    if not rows:
        _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, "No preview rows available.")
        return row_index + 2

    for row in rows:
        for column_index, column_name in enumerate(columns[:_DOCUMENT_SHEET_MAX_COLUMNS], start=1):
            ws.cell(row=row_index, column=column_index, value=_format_document_preview_value(row.get(column_name)))
        row_index += 1

    return row_index + 1


def _write_document_block(
    ws,
    block: Dict[str, Any],
    block_previews: Dict[str, Dict[str, Any]],
    row_index: int,
    depth: int,
) -> int:
    block_type = str(block.get("type") or "text")
    block_name = str(block.get("name") or block_type.title())
    prefix = "  " * depth
    preview = block_previews.get(str(block.get("id") or ""), {})

    if block_type == "page":
        for child in block.get("children") or []:
            row_index = _write_document_block(ws, child, block_previews, row_index, depth)
        return row_index

    if block_type in {"section", "stack", "grid"}:
        section_cell = _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, f"{prefix}{block_name}")
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = PatternFill(fill_type="solid", fgColor="F1F5F9")
        row_index += 1
        for child in block.get("children") or []:
            row_index = _write_document_block(ws, child, block_previews, row_index, depth + 1)
        return row_index + 1

    if block_type == "text":
        content = str(block.get("content") or block_name)
        text_cell = _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, f"{prefix}{content}")
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")
        return row_index + 2

    if block_type == "table":
        return _write_document_table(
            ws,
            row_index,
            depth,
            str(block.get("title") or block_name),
            preview,
        )

    if block_type == "metric":
        ws.cell(row=row_index, column=1, value=f"{prefix}{block.get('title') or block_name}")
        value_cell = _merge_row(ws, row_index, 2, _DOCUMENT_SHEET_MAX_COLUMNS, _format_document_preview_value(preview.get("value")))
        value_cell.font = Font(bold=True, size=14)
        return row_index + 2

    if block_type == "input":
        ws.cell(row=row_index, column=1, value=f"{prefix}{block.get('label') or block_name}")
        _merge_row(ws, row_index, 2, _DOCUMENT_SHEET_MAX_COLUMNS, _format_document_preview_value(preview.get("value")))
        return row_index + 2

    if block_type == "repeater":
        items = [
            _format_document_preview_value(item)
            for item in list(preview.get("items") or [])
        ]
        ws.cell(row=row_index, column=1, value=f"{prefix}{block_name}")
        _merge_row(
            ws,
            row_index,
            2,
            _DOCUMENT_SHEET_MAX_COLUMNS,
            ", ".join(str(item) for item in items) if items else "-",
        )
        row_index += 2
        for child in block.get("children") or []:
            row_index = _write_document_block(ws, child, block_previews, row_index, depth + 1)
        return row_index

    if block_type == "signature":
        ws.cell(row=row_index, column=1, value=f"{prefix}{block_name}")
        _merge_row(ws, row_index + 1, 1, 4, "____________________")
        return row_index + 3

    if block_type == "page-break":
        divider = _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, "Page break")
        divider.font = Font(italic=True)
        return row_index + 2

    fallback = _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, f"{prefix}{block_name}")
    fallback.font = Font(bold=True)
    row_index += 1
    for child in block.get("children") or []:
        row_index = _write_document_block(ws, child, block_previews, row_index, depth + 1)
    return row_index + 1


def _write_document_layout(
    ws,
    template_name: str,
    definition: Dict[str, Any],
    runtime_preview: Dict[str, Any],
    warnings: List[str],
) -> None:
    ws.column_dimensions["A"].width = 28
    for column_name in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[column_name].width = 18

    title_cell = _merge_row(ws, 1, 1, _DOCUMENT_SHEET_MAX_COLUMNS, template_name)
    title_cell.font = Font(bold=True, size=16)
    subtitle_cell = _merge_row(
        ws,
        2,
        1,
        _DOCUMENT_SHEET_MAX_COLUMNS,
        f"Document export · {definition.get('page', {}).get('size', 'A4')} · {definition.get('page', {}).get('orientation', 'portrait')}",
    )
    subtitle_cell.font = Font(italic=True, size=10)

    row_index = 4
    if warnings:
        warning_cell = _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, "Warnings: " + " | ".join(warnings))
        warning_cell.fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
        row_index += 2

    row_index = _write_document_block(
        ws,
        definition.get("root") or {},
        runtime_preview.get("blocks") or {},
        row_index,
        0,
    )

    if row_index == 4:
        _merge_row(ws, row_index, 1, _DOCUMENT_SHEET_MAX_COLUMNS, "No document blocks configured.")


def export_template_to_excel(
    db: Session,
    template: Any,
    active_filters: List[Dict[str, Any]] | None,
    *,
    definition_override: Dict[str, Any] | None = None,
    filter_definitions: List[Dict[str, Any]] | None = None,
) -> bytes:
    """Render a legacy or document-engine template to an Excel workbook."""
    definition = definition_override if isinstance(definition_override, dict) else (template.blocks if isinstance(template.blocks, dict) else {})
    if is_template_document_definition(definition):
        normalized_definition = normalize_template_document(definition)
        source_previews, fetch_warnings = _fetch_document_source_previews(db, normalized_definition)
        runtime_preview = build_template_document_runtime_preview(normalized_definition, source_previews)
        warnings = [*fetch_warnings, *(runtime_preview.get("warnings") or [])]

        wb = Workbook()
        ws = wb.active
        ws.title = str(template.name or "Document")[:31]
        _write_document_layout(
            ws,
            str(template.name or "Document"),
            normalized_definition,
            runtime_preview,
            warnings,
        )

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    if definition.get("version") != 3:
        raise ValueError("Only TemplateDefinition v3 export is supported.")

    data_source = definition.get("dataSource") or {}
    dataset_id = int(data_source.get("datasetId") or 0)
    table_id = int(data_source.get("tableId") or 0)
    if dataset_id > 0 and table_id > 0:
        datasource, target_table = _resolve_runtime_source(db, dataset_id, table_id)
        preview_table, formula_errors = build_template_preview_table_proxy(target_table, list(definition.get("columns") or []))
        if formula_errors:
            details = "; ".join(f"{item['key']}: {item['error']}" for item in formula_errors)
            raise ValueError(f"Some template formulas are invalid. {details}")
        runtime_filters = resolve_template_runtime_filters(
            filter_definitions if filter_definitions is not None else (template.filters or []),
            active_filters or [],
            dataset_id=dataset_id,
            table_id=table_id,
        )
        result = LiveQueryService.execute_preview_query(
            datasource=datasource,
            db_table=preview_table,
            limit=_EXPORT_LIMIT,
            offset=0,
            filters=runtime_filters or None,
        )
        rows = list(result.get("rows") or [])
    else:
        rows = []

    wb = Workbook()
    ws = wb.active
    ws.title = str(template.name or "Report")[:31]

    layout = str(definition.get("layout") or "table")
    if layout == "card":
        _write_card_layout(ws, definition, rows, 1)
    elif layout == "cross-tab":
        _write_cross_tab_layout(ws, definition, rows, 1)
    else:
        _write_table_layout(ws, definition, rows, 1)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()