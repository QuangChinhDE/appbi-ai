"""
API router for data source endpoints.
"""
from fastapi import APIRouter, Depends, HTTPException, Request, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
from pydantic import BaseModel
import io
import csv as csv_module

from slowapi import Limiter
from slowapi.util import get_remote_address

from app.core import get_db
from app.core.dependencies import (
    get_current_user,
    require_permission,
    require_view_access,
    require_edit_access,
    require_full_access,
    get_effective_permission,
)
from app.core.permissions import _owned_or_shared, stamp_owner_emails
from app.models import DataSource, Dataset
from app.models.resource_share import ResourceType
from app.models.user import User
from app.schemas import (
    DataSourceCreate,
    DataSourceUpdate,
    DataSourceResponse,
    DataSourceTestRequest,
    DataSourceTestResponse,
    QueryExecuteRequest,
    QueryExecuteResponse,
    SqlValidateRequest,
    SqlValidateResponse,
)
from app.services import DataSourceCRUDService, DataSourceConnectionService
from app.core.logging import get_logger
from app.core.config import settings
from app.services.google_data_access_service import get_google_data_access_status

logger = get_logger(__name__)
router = APIRouter(prefix="/datasources", tags=["datasources"])
_limiter = Limiter(key_func=get_remote_address)


def _build_query_error_detail(exc: Exception) -> dict:
    """Return a user-facing query error payload without hiding the root cause."""
    message = " ".join(str(exc).split()).strip()
    return {
        "message": message or "Query execution failed. Please check your SQL and try again.",
    }


def _validate_datasource_connection_or_raise(ds_type: str, config: dict[str, Any]) -> None:
    """Validate datasource connectivity before persisting non-manual configs."""
    if ds_type == "manual":
        return

    success, message = DataSourceConnectionService.test_connection(ds_type, config)
    if success:
        return

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=message or "Connection failed",
    )


def _restore_sensitive_config_fields(
    config: dict[str, Any],
    existing_config: dict[str, Any] | None,
) -> dict[str, Any]:
    """
    Rehydrate masked/blank secret fields from a stored datasource config.

    This keeps validation and updates working when the frontend intentionally
    leaves sensitive inputs blank to mean "keep the stored value".
    """
    from app.core.crypto import MASKED_PLACEHOLDER, _SENSITIVE_FIELDS

    restored = dict(config or {})
    stored = dict(existing_config or {})
    for field in _SENSITIVE_FIELDS:
        if restored.get(field, None) in ("", None, MASKED_PLACEHOLDER) and stored.get(field):
            restored[field] = stored[field]
    return restored


def _normalize_google_oauth_config(
    config: dict[str, Any],
    *,
    current_user: User,
    existing_config: dict[str, Any] | None = None,
) -> dict[str, Any]:
    normalized = dict(config or {})
    if normalized.get("auth_mode") != "google_oauth":
        normalized.pop("google_oauth_user_id", None)
        normalized.pop("google_oauth_email", None)
        return normalized

    from app.core.crypto import decrypt_config

    existing = decrypt_config(existing_config or {})
    existing_user_id = str(existing.get("google_oauth_user_id") or "").strip()
    existing_email = str(existing.get("google_oauth_email") or "").strip().lower()
    desired_email = str(normalized.get("google_oauth_email") or "").strip().lower()

    if existing_user_id and existing_email and (not desired_email or desired_email == existing_email):
        normalized["google_oauth_user_id"] = existing_user_id
        normalized["google_oauth_email"] = existing_email
        return normalized

    status_payload = get_google_data_access_status(current_user)
    if not status_payload["configured"]:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "Google data access is not configured yet. Ask an admin to set "
                "AUTH_GOOGLE_CLIENT_SECRET and AUTH_GOOGLE_DATA_REDIRECT_URI."
            ),
        )
    if not status_payload["connected"] or not status_payload["email"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Connect your Google account inside AppBI before using Google OAuth "
                "for BigQuery or Google Sheets."
            ),
        )

    normalized["google_oauth_user_id"] = str(current_user.id)
    normalized["google_oauth_email"] = str(status_payload["email"]).strip().lower()
    return normalized


# ── Platform GCP credential info ──────────────────────────────────────────────

@router.get("/platform-gcp-info")
def get_platform_gcp_info(_: User = Depends(get_current_user)):
    """
    Returns whether a platform-level GCP service account is configured.
    If configured, also returns the email so users know which account to share with.
    """
    has_credential = bool((settings.GCP_SERVICE_ACCOUNT_JSON or "").strip())
    email = (settings.GCP_SERVICE_ACCOUNT_EMAIL or "").strip() or None
    return {
        "platform_credential_available": has_credential,
        "service_account_email": email,
    }


# ── Manual datasource: server-side file parsing ───────────────────────────────

def _infer_type(values: list) -> str:
    """Infer column type from a sample of values."""
    samples = [v for v in values if v is not None and str(v).strip() != ''][:20]
    if not samples:
        return 'string'
    num_count = sum(1 for v in samples if _is_number(v))
    if num_count == len(samples):
        return 'number'
    date_count = sum(1 for v in samples if _is_date_like(str(v)))
    if date_count >= len(samples) * 0.8:
        return 'date'
    return 'string'

def _is_number(v) -> bool:
    try:
        float(str(v).replace(',', ''))
        return True
    except (ValueError, TypeError):
        return False

def _is_date_like(s: str) -> bool:
    import re
    return bool(re.match(r'^\d{2,4}[-/]\d{1,2}[-/]\d{1,4}', s.strip()))

def _parse_excel_bytes(content: bytes) -> Dict[str, Any]:
    """Parse all sheets from an Excel file (.xlsx/.xls) using openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets: Dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            sheets[sheet_name] = {'columns': [], 'rows': []}
            continue
        # First non-empty row is the header
        header_row = [str(c).strip() if c is not None else '' for c in all_rows[0]]
        headers = [h if h else f'col{i+1}' for i, h in enumerate(header_row)]
        data_rows = all_rows[1:]
        rows = []
        for r in data_rows:
            row_dict = {}
            has_value = False
            for i, h in enumerate(headers):
                val = r[i] if i < len(r) else None
                # Convert to JSON-serialisable types
                if val is None:
                    row_dict[h] = ''
                elif isinstance(val, (int, float)):
                    row_dict[h] = val
                    has_value = True
                else:
                    str_val = str(val).strip()
                    row_dict[h] = str_val
                    if str_val:
                        has_value = True
            if has_value:
                rows.append(row_dict)
        # Build column metadata
        columns = [
            {'name': h, 'type': _infer_type([r.get(h) for r in rows])}
            for h in headers
        ]
        sheets[sheet_name] = {'columns': columns, 'rows': rows}
    wb.close()
    return sheets

def _parse_csv_bytes(content: bytes, filename: str) -> Dict[str, Any]:
    """Parse a CSV file and return as a single-sheet dict."""
    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = content.decode('latin-1')
    reader = csv_module.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    fieldnames = list(reader.fieldnames or [])
    columns = [
        {'name': h, 'type': _infer_type([r.get(h) for r in rows])}
        for h in fieldnames
    ]
    sheet_name = filename.rsplit('.', 1)[0] or 'Sheet1'
    return {sheet_name: {'columns': columns, 'rows': rows}}


@router.post("/manual/parse-file")
async def parse_manual_file(
    file: UploadFile = File(...),
    _: User = Depends(require_permission("data_sources", "edit")),
):
    """
    Parse an uploaded Excel (.xlsx/.xls) or CSV file server-side.
    Returns: { sheets: { sheetName: { columns: [...], rows: [...] } } }
    """
    allowed = {'.xlsx', '.xls', '.csv'}
    ext = '.' + (file.filename or '').rsplit('.', 1)[-1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Unsupported file type '{ext}'. Allowed: xlsx, xls, csv")

    content = await file.read()
    if len(content) > 250 * 1024 * 1024:  # 250 MB guard
        raise HTTPException(status_code=400, detail="File too large (max 250 MB)")

    try:
        if ext in ('.xlsx', '.xls'):
            sheets = _parse_excel_bytes(content)
        else:
            sheets = _parse_csv_bytes(content, file.filename or 'data')
    except Exception as e:
        logger.error(f"File parse error: {e}")
        raise HTTPException(status_code=422, detail=f"Failed to parse file: {str(e)}")

    total_rows = sum(len(v['rows']) for v in sheets.values())
    logger.info(f"Parsed '{file.filename}': {len(sheets)} sheet(s), {total_rows} total rows")
    return {'filename': file.filename, 'sheets': sheets}


@router.get("/", response_model=List[DataSourceResponse])
def list_data_sources(
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List data sources — filtered by ownership and shares."""
    sources = (
        _owned_or_shared(db, DataSource, ResourceType.DATASOURCE, current_user)
        .offset(skip)
        .limit(limit)
        .all()
    )
    for s in sources:
        s.user_permission = get_effective_permission(db, current_user, s, "data_sources")
    stamp_owner_emails(db, sources)
    return sources


@router.get("/{data_source_id}", response_model=DataSourceResponse)
def get_data_source(
    data_source_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get a data source by ID."""
    data_source = DataSourceCRUDService.get_by_id(db, data_source_id)
    if not data_source:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with ID {data_source_id} not found"
        )
    data_source.user_permission = require_view_access(db, current_user, data_source, "data_sources")
    stamp_owner_emails(db, [data_source])
    return data_source


@router.post("/", response_model=DataSourceResponse, status_code=status.HTTP_201_CREATED)
def create_data_source(
    data_source: DataSourceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("data_sources", "edit")),
):
    """Create a new data source."""
    try:
        data_source.config = _normalize_google_oauth_config(
            data_source.config,
            current_user=current_user,
        )
        _validate_datasource_connection_or_raise(data_source.type.value, data_source.config)
        created = DataSourceCRUDService.create(db, data_source, owner_id=current_user.id)
        created.user_permission = get_effective_permission(db, current_user, created, "data_sources")
        stamp_owner_emails(db, [created])
        return created
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.put("/{data_source_id}", response_model=DataSourceResponse)
def update_data_source(
    data_source_id: int,
    data_source_update: DataSourceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update a data source."""
    ds = db.query(DataSource).filter(DataSource.id == data_source_id).first()
    if not ds:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Data source with ID {data_source_id} not found")
    require_edit_access(db, current_user, ds, "data_sources")
    try:
        next_type = data_source_update.type.value if data_source_update.type is not None else ds.type.value
        if data_source_update.config is not None:
            restored_config = _restore_sensitive_config_fields(
                data_source_update.config,
                ds.config,
            )
            data_source_update.config = _normalize_google_oauth_config(
                restored_config,
                current_user=current_user,
                existing_config=ds.config,
            )
            _validate_datasource_connection_or_raise(next_type, data_source_update.config)
        data_source = DataSourceCRUDService.update(db, data_source_id, data_source_update)
        if data_source is not None:
            data_source.user_permission = get_effective_permission(db, current_user, data_source, "data_sources")
            stamp_owner_emails(db, [data_source])
        return data_source
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.delete("/{data_source_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_data_source(
    data_source_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a data source, blocked if any datasets still reference it."""
    datasource = db.query(DataSource).filter(DataSource.id == data_source_id).first()
    if not datasource:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with ID {data_source_id} not found"
        )
    require_full_access(db, current_user, datasource, "data_sources")

    blocking_datasets = db.query(Dataset).filter(
        Dataset.id.in_(
            db.query(Dataset.id)
            .join(Dataset.tables)
            .filter_by(datasource_id=data_source_id)
            .distinct()
        )
    ).all()

    constraints = [
        {"type": "dataset", "id": ds.id, "name": ds.name}
        for ds in blocking_datasets
    ]

    if constraints:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "message": f"Data source \"{datasource.name}\" đang được sử dụng và không thể xóa.",
                "constraints": constraints,
            },
        )

    success = DataSourceCRUDService.delete(db, data_source_id)
    if not success:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with ID {data_source_id} not found"
        )


@router.post("/test", response_model=DataSourceTestResponse)
def test_data_source_connection(
    request: DataSourceTestRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("data_sources", "view")),
):
    """Test a data source connection."""
    config = dict(request.config)

    # When editing an existing datasource, sensitive fields are cleared to ''
    # by the frontend (sanitizeConfigForForm strips the '__stored__' sentinel).
    # Re-fill them from the DB so the real (encrypted) credentials are used.
    if request.data_source_id is not None:
        db_ds = DataSourceCRUDService.get_by_id(db, request.data_source_id)
        if db_ds and db_ds.config:
            config = _restore_sensitive_config_fields(config, db_ds.config)

    config = _normalize_google_oauth_config(
        config,
        current_user=current_user,
        existing_config=db_ds.config if request.data_source_id is not None and db_ds else None,
    )

    success, message = DataSourceConnectionService.test_connection(
        request.type.value,
        config
    )
    return DataSourceTestResponse(success=success, message=message)


@router.post("/query", response_model=QueryExecuteResponse)
@_limiter.limit("20/minute")
def execute_query(
    request: Request,
    body: QueryExecuteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Execute an ad-hoc SQL query against a data source."""
    data_source = DataSourceCRUDService.get_by_id(db, body.data_source_id)
    if not data_source:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with ID {body.data_source_id} not found"
        )
    require_view_access(db, current_user, data_source, "data_sources")

    try:
        columns, data, execution_time_ms = DataSourceConnectionService.execute_query(
            data_source.type.value,
            data_source.config,
            body.sql_query,
            body.limit,
            timeout_seconds=body.timeout_seconds or 30,
        )
        
        return QueryExecuteResponse(
            columns=columns,
            data=data,
            row_count=len(data),
            execution_time_ms=execution_time_ms
        )
    except Exception as e:
        logger.exception("Query execution failed for datasource %s", body.data_source_id)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=_build_query_error_detail(e),
        )


@router.post("/validate-sql", response_model=SqlValidateResponse)
@_limiter.limit("30/minute")
def validate_sql(
    request: Request,
    body: SqlValidateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Validate a SQL query against a datasource without returning data.

    Sends the query wrapped in a dry-run construct (LIMIT 0, EXPLAIN, etc.)
    to the actual database engine so the user receives real error messages
    from their specific RDBMS dialect.
    """
    data_source = DataSourceCRUDService.get_by_id(db, body.data_source_id)
    if not data_source:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Data source with ID {body.data_source_id} not found",
        )
    require_view_access(db, current_user, data_source, "data_sources")

    ds_type = (
        data_source.type if isinstance(data_source.type, str) else data_source.type.value
    )

    # Client-side safety check first
    from app.services.query_validator import QueryValidator, QueryValidationError

    try:
        cleaned = QueryValidator.validate_and_clean(body.sql_query)
    except QueryValidationError as exc:
        return SqlValidateResponse(valid=False, error=str(exc), dialect=ds_type)

    # Dry-run: execute with LIMIT 0 to let the DB parse without returning rows
    try:
        DataSourceConnectionService.execute_query(
            ds_type,
            data_source.config,
            cleaned,
            limit=1,
            timeout_seconds=10,
        )
        return SqlValidateResponse(valid=True, error=None, dialect=ds_type)
    except Exception as exc:
        error_msg = " ".join(str(exc).split()).strip()
        return SqlValidateResponse(valid=False, error=error_msg, dialect=ds_type)


# ── Schema Browser ────────────────────────────────────────────────────────────

@router.get("/{data_source_id}/schema")
def get_schema_browser(
    data_source_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return schema tree: schemas → tables/views with row count estimates."""
    ds = DataSourceCRUDService.get_by_id(db, data_source_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Data source not found")
    require_view_access(db, current_user, ds, "data_sources")
    try:
        schemas = DataSourceConnectionService.get_schema_browser(ds.type.value, ds.config)
        return {"schemas": schemas}
    except Exception as e:
        logger.error(f"Schema browser failed for ds {data_source_id}: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve schema. Please check the connection.")


@router.get("/{data_source_id}/tables/{schema_name}/{table_name}")
def get_table_detail(
    data_source_id: int,
    schema_name: str,
    table_name: str,
    preview_rows: int = 5,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return column metadata (with PK/FK/IDX) and quick preview for a table."""
    ds = DataSourceCRUDService.get_by_id(db, data_source_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Data source not found")
    require_view_access(db, current_user, ds, "data_sources")
    try:
        detail = DataSourceConnectionService.get_table_detail(
            ds.type.value, ds.config, schema_name, table_name, preview_rows
        )
        return detail
    except Exception as e:
        logger.error(f"Table detail failed for {schema_name}.{table_name}: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve table details.")


@router.get("/{data_source_id}/tables/{schema_name}/{table_name}/watermarks")
def get_watermark_candidates(
    data_source_id: int,
    schema_name: str,
    table_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List columns usable as watermark (timestamp/date/integer) for incremental sync."""
    ds = DataSourceCRUDService.get_by_id(db, data_source_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Data source not found")
    require_view_access(db, current_user, ds, "data_sources")
    try:
        candidates = DataSourceConnectionService.get_watermark_candidates(
            ds.type.value, ds.config, schema_name, table_name
        )
        return {"columns": candidates}
    except Exception as e:
        logger.error(f"Watermark candidates failed for {schema_name}.{table_name}: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve watermark candidates.")


# ── Google Sheets write / structure endpoints ─────────────────────────────────
# These endpoints expose full CRUD on a connected Google Sheets datasource so
# that MCP tools (and the workboard builder) can manage sheet tabs and row data
# without requiring direct Google API access from the client.

def _require_gsheets_ds(data_source_id: int, db: Session, current_user: User):
    """Load + authorize a google_sheets datasource; raise 404/403/400 on error."""
    from app.models import DataSource
    from app.core.crypto import decrypt_config
    from app.services.google_sheets_connector import create_google_sheets_connector

    ds = DataSourceCRUDService.get_by_id(db, data_source_id)
    if not ds:
        raise HTTPException(status_code=404, detail="Data source not found")
    require_view_access(db, current_user, ds, "data_sources")

    ds_type = ds.type.value if hasattr(ds.type, "value") else str(ds.type or "")
    if ds_type != "google_sheets":
        raise HTTPException(
            status_code=400,
            detail=f"Data source {data_source_id} is type '{ds_type}', not 'google_sheets'.",
        )
    cfg = decrypt_config(ds.config)
    spreadsheet_id = (cfg.get("spreadsheet_id") or "").strip()
    if not spreadsheet_id:
        raise HTTPException(status_code=400, detail="Datasource missing spreadsheet_id")
    connector = create_google_sheets_connector(cfg)
    return connector, spreadsheet_id, ds


@router.get("/{data_source_id}/gsheets/sheets")
def list_gsheets_tabs(
    data_source_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all sheet tabs in the connected Google Spreadsheet."""
    connector, spreadsheet_id, _ = _require_gsheets_ds(data_source_id, db, current_user)
    try:
        sheets = connector.list_sheets(spreadsheet_id)
        return {"spreadsheet_id": spreadsheet_id, "sheets": sheets}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


class GSheetCreateRequest(BaseModel):
    sheet_name: str
    headers: Optional[List[str]] = None  # column names for row 1


@router.post("/{data_source_id}/gsheets/sheets", status_code=status.HTTP_201_CREATED)
def create_gsheets_tab(
    data_source_id: int,
    body: GSheetCreateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new sheet tab in the connected Google Spreadsheet.

    Optionally writes a header row (column names) as the first row so the
    sheet is immediately ready for workboard form submissions.
    """
    connector, spreadsheet_id, ds = _require_gsheets_ds(data_source_id, db, current_user)
    require_edit_access(db, current_user, ds, "data_sources")
    try:
        result = connector.create_sheet(spreadsheet_id, body.sheet_name, body.headers)
        return {"spreadsheet_id": spreadsheet_id, **result}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/{data_source_id}/gsheets/{sheet_name}/rows")
def read_gsheets_rows(
    data_source_id: int,
    sheet_name: str,
    limit: int = 200,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Read rows from a sheet tab. Returns columns + rows list."""
    connector, spreadsheet_id, _ = _require_gsheets_ds(data_source_id, db, current_user)
    try:
        data = connector.get_sheet_data(spreadsheet_id, sheet_name=sheet_name)
        rows = data.get("rows") or []
        if limit:
            rows = rows[:limit]
        return {
            "spreadsheet_id": spreadsheet_id,
            "sheet_name": sheet_name,
            "columns": data.get("columns") or [],
            "rows": rows,
            "row_count": len(rows),
        }
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


class GSheetAppendRequest(BaseModel):
    values: Dict[str, Any]  # {column_name: value}


@router.post("/{data_source_id}/gsheets/{sheet_name}/rows", status_code=status.HTTP_201_CREATED)
def append_gsheets_row(
    data_source_id: int,
    sheet_name: str,
    body: GSheetAppendRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Append a new row to a sheet tab.

    ``values`` must be a dict mapping column names (header row) to values.
    Columns not present in the payload receive an empty string.
    """
    connector, spreadsheet_id, ds = _require_gsheets_ds(data_source_id, db, current_user)
    require_edit_access(db, current_user, ds, "data_sources")
    try:
        row = connector.append_row(spreadsheet_id, sheet_name, body.values)
        return {"ok": True, "sheet_name": sheet_name, "row": row}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


class GSheetUpdateRequest(BaseModel):
    pk: Dict[str, Any]       # {pk_column: pk_value} — used to find the row
    values: Dict[str, Any]   # {column_name: new_value}


@router.patch("/{data_source_id}/gsheets/{sheet_name}/rows")
def update_gsheets_row(
    data_source_id: int,
    sheet_name: str,
    body: GSheetUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update a row in a sheet tab identified by a primary-key dict.

    ``pk`` identifies the row (e.g. ``{"id": "ROW-001"}``).
    ``values`` provides the columns to overwrite; other columns are unchanged.
    """
    connector, spreadsheet_id, ds = _require_gsheets_ds(data_source_id, db, current_user)
    require_edit_access(db, current_user, ds, "data_sources")
    try:
        row = connector.update_row_by_pk(spreadsheet_id, sheet_name, body.pk, body.values)
        return {"ok": True, "sheet_name": sheet_name, "row": row}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


class GSheetDeleteRequest(BaseModel):
    pk: Dict[str, Any]  # {pk_column: pk_value}


@router.delete("/{data_source_id}/gsheets/{sheet_name}/rows")
def delete_gsheets_row(
    data_source_id: int,
    sheet_name: str,
    body: GSheetDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a row from a sheet tab identified by a primary-key dict."""
    connector, spreadsheet_id, ds = _require_gsheets_ds(data_source_id, db, current_user)
    require_edit_access(db, current_user, ds, "data_sources")
    try:
        row_num = connector.delete_row_by_pk(spreadsheet_id, sheet_name, body.pk)
        return {"ok": True, "sheet_name": sheet_name, "deleted_row": row_num}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
