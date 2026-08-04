"""Mini-app screen runtime — the single read/write entry point for workboards.

A workboard is a mini-app: an ordered list of screens, each bound to its
own dataset table. This module resolves the screen first, then applies
the right read / write / RLS logic for its kind (``form`` / ``table`` /
``doc`` / ``dashboard``).

Phase-13 (2026-05-16) collapsed the previous ``list`` (read-only) and
``grid`` (editable) kinds into one ``table`` kind with a
``mode: readonly | editable`` flag. No backwards-compatibility shim —
the schema is the contract and there is only one runtime path.

Public API:
* :func:`get_screen` — fetch a Screen by id, raising 404 if missing.
* :func:`render_form_screen` — return form spec (fields + lookup options).
* :func:`render_table_screen` — paginated rows after RLS, with optional
  computed/lookup columns, totals, multi-header and row-merge.
* :func:`render_doc_screen` — block-based rendered doc payload.
* :func:`insert_screen_row` / :func:`update_screen_row` /
  :func:`delete_screen_row` — write paths.

Each helper takes a ``CallerIdentity`` so RLS is consistently applied.
"""
from __future__ import annotations

import hashlib
import itertools
import json
import numbers
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Set

from fastapi import HTTPException, status
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.models.dataset import DatasetTable
from app.models.models import DataSource
from app.modules.workboards.models import Workboard, WorkboardOpLog
from app.modules.workboards.roles import is_owner_role
from app.modules.workboards.services.geocode_service import (
    build_address,
    geocode_address,
)
from app.modules.workboards.schemas import (
    DataTableBlock,
    DataTablePivot,
    DataTableUnpivot,
    FormField,
    LayoutJson,
    Screen,
    ScreenGroup,
    ScreenRlsRule,
)
from app.modules.workboards.services.rls_service import (
    CallerIdentity,
    build_rls_filter,
    enforce_write_access,
    role_has_screen_grant as _role_has_screen_grant,
)
from app.modules.workboards.services.js_evaluator import (
    CompiledJs,
    JsCompileError,
    JsEvalError,
    compile_js_column,
    evaluate_js_cell,
)
from app.modules.workboards.services.write_service import (
    WorkboardWriteService,
)
from app.services.live_query_service import LiveQueryService
from app.services.google_sheets_cache import SheetsQuotaError

logger = get_logger(__name__)


def _op_actor_key(identity: CallerIdentity) -> str:
    if identity.appbi_user_id is not None:
        return f"appbi:{identity.appbi_user_id}"
    app_user = identity.app_user or {}
    return f"app-user:{app_user.get('username') or 'unknown'}"


def _op_request_fingerprint(
    screen_id: str,
    values: Dict[str, Any],
    relation_context: Optional[Dict[str, Any]],
) -> str:
    canonical = json.dumps(
        {
            "screen_id": screen_id,
            "values": values or {},
            "relation_context": relation_context or None,
        },
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        default=str,
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _replay_op_result(
    entry: WorkboardOpLog,
    *,
    workboard_id: int,
    screen_id: str,
    actor_key: str,
    request_fingerprint: str,
) -> Dict[str, Any]:
    if entry.workboard_id not in (None, workboard_id):
        raise HTTPException(status_code=409, detail="Operation ID belongs to another workboard.")
    if entry.screen_id not in (None, screen_id):
        raise HTTPException(status_code=409, detail="Operation ID belongs to another screen.")
    if entry.actor_key not in (None, actor_key):
        raise HTTPException(status_code=409, detail="Operation ID belongs to another user.")
    if entry.request_fingerprint not in (None, request_fingerprint):
        raise HTTPException(status_code=409, detail="Operation ID was reused with different values.")
    if not isinstance(entry.result_payload, dict):
        raise HTTPException(
            status_code=409,
            detail="Operation is already recorded but its result is not available yet.",
        )
    return {**entry.result_payload, "idempotent": True}


def _quota_503() -> "HTTPException":
    """Map a Sheets read-quota hit to an honest, retryable 503 instead of
    silently returning empty options/rows (the old behaviour that made forms
    look broken)."""
    return HTTPException(
        status_code=503,
        detail=(
            "Nguồn Google Sheets đang quá tải (giới hạn đọc 60 lần/phút). "
            "Vui lòng thử lại sau vài giây."
        ),
    )

_MAX_LOOKUP_ROWS = 500
_AGG_FNS = {"sum", "avg", "min", "max", "count"}
# Upper bound on rows pulled per table render so footer totals span the whole
# (filtered) table and pagination is sliced in memory. Beyond this the totals
# are over the first N rows and ``totals_partial`` is set true.
_TOTALS_ROW_CAP = 5000


# ── Generic helpers (formerly in runtime_service) ─────────────────────────

def _filter_dicts(filters: Optional[List[Any]]) -> List[Dict[str, Any]]:
    cleaned: List[Dict[str, Any]] = []
    for item in filters or []:
        if isinstance(item, dict):
            cleaned.append(item)
        elif hasattr(item, "model_dump"):
            cleaned.append(item.model_dump())
        elif hasattr(item, "dict"):
            cleaned.append(item.dict())
    return cleaned


def _parse_total_spec(spec: str) -> tuple[str, str]:
    """Parse ``"column"`` or ``"column:agg"`` → (column, agg). Default ``sum``."""
    if not isinstance(spec, str):
        return "", "sum"
    text = spec.strip()
    if not text:
        return "", "sum"
    if ":" in text:
        col, agg = text.split(":", 1)
        agg = agg.strip().lower()
        if agg not in _AGG_FNS:
            agg = "sum"
        return col.strip(), agg
    return text, "sum"


def _parse_locale_number(text: str) -> Optional[float]:
    """Parse a numeric string read from a Google Sheet, honouring the vi-VN
    number format (``.`` = thousands grouping, ``,`` = decimal separator).

    This is a DETERMINISTIC locale rule for the Workboard Sheets path — not a
    content guess. The sheet is read with the default ``FORMATTED_VALUE`` so a
    vi-VN-locale workbook returns decimals as comma strings ("1234,5") and
    grouped values with dots ("1.000.000"); plain ``float()`` drops every one
    of those, which silently excludes them from SUM/AVG totals. Examples::

        "1234,5"     -> 1234.5       "75.351.234,5" -> 75351234.5
        "1.000.000"  -> 1000000.0    "1,234"        -> 1.234
        "500000"     -> 500000.0     "(1.234,5)"    -> -1234.5

    Only the Sheets / manual VARCHAR path reaches here: BigQuery / Postgres
    numerics arrive as ``int``/``float`` and take the ``isinstance`` fast-path
    in the callers below, so this rule can never corrupt a native-numeric
    source. Returns ``None`` for anything that is not a plain number (same as
    the old ``float()`` behaviour — no new values are "rescued" into garbage).
    """
    s = text.strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")  # accounting negative
    if neg:
        s = s[1:-1].strip()
    s = s.replace(" ", "").replace(" ", "")  # NBSP + thin spaces
    if not s:
        return None
    # vi-VN: '.' groups thousands, ',' is the decimal mark. When a comma is
    # present it is the decimal point, so strip grouping dots then swap it for
    # a dot.
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        # A lone dot is thousands grouping ONLY when it forms valid vi-VN
        # groups: the first segment is 1-3 digits and every later segment is
        # EXACTLY 3 digits ("1.000.000" -> 1000000, "1.234" -> 1234). A dot
        # whose trailing segment isn't 3 digits is a decimal point, not
        # grouping ("98.0" -> 98.0, "88.8" -> 88.8, "125.4" -> 125.4) — this
        # is what a native numeric (Postgres/BQ) read back as a VARCHAR looks
        # like, and must NOT be inflated ×10 into SUM/AVG.
        parts = s.split(".")
        is_grouping = (
            len(parts) >= 2
            and all(p.isdigit() for p in parts)
            and 1 <= len(parts[0]) <= 3
            and all(len(p) == 3 for p in parts[1:])
        )
        if is_grouping:
            s = "".join(parts)
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val


def _coerce_number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # PostgreSQL numeric/decimal columns come back as Decimal (and other
    # backends may return Fraction). numbers.Number covers both without a
    # hard import of decimal here; str stays on the locale-aware parser.
    if isinstance(value, numbers.Number):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if isinstance(value, str):
        return _parse_locale_number(value)
    return None


def _compute_totals_row(
    totals: List[str],
    columns: List[str],
    rows: List[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    """Produce a footer aggregations payload for the doc-table block.

    Output shape: ``{rows: [{agg, label, values}, ...], single?: {col: number}}``.
    The FE renders one ``<tr>`` per ``rows`` entry; when ``single`` is
    present (each column has at most one agg) the FE can compact to a
    one-row footer.
    """
    if not totals:
        return None

    by_agg: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for spec in totals:
        col, agg = _parse_total_spec(spec)
        if not col or col not in columns:
            continue
        values = [_coerce_number(r.get(col)) for r in rows]
        numeric = [v for v in values if v is not None]
        if agg == "count":
            value: Any = sum(1 for r in rows if r.get(col) is not None)
        elif not numeric:
            value = sum(1 for r in rows if r.get(col) is not None)
        elif agg == "sum":
            value = sum(numeric)
        elif agg == "avg":
            value = sum(numeric) / len(numeric)
        elif agg == "min":
            value = min(numeric)
        elif agg == "max":
            value = max(numeric)
        else:
            continue
        bucket = by_agg.setdefault(agg, {})
        bucket[col] = value
        if agg not in order:
            order.append(agg)

    if not order:
        return None

    AGG_LABELS = {
        "sum": "Tổng",
        "avg": "TB",
        "count": "Đếm",
        "min": "Min",
        "max": "Max",
    }
    footer_rows = [
        {
            "agg": agg,
            "label": AGG_LABELS.get(agg, agg.upper()),
            "values": by_agg[agg],
        }
        for agg in order
    ]
    counts: Dict[str, int] = {}
    for fr in footer_rows:
        for c in fr["values"]:
            counts[c] = counts.get(c, 0) + 1
    out: Dict[str, Any] = {"rows": footer_rows}
    if all(v == 1 for v in counts.values()):
        single: Dict[str, Any] = {}
        for fr in footer_rows:
            single.update(fr["values"])
        out["single"] = single
    return out


def _compute_merges(
    rows: List[Dict[str, Any]],
    group_by: List[str],
    columns: List[str],
) -> List[Dict[str, Any]]:
    """Return rowspan recipes (``{column, row_start, row_span}``) for
    consecutive identical values in each group_by column.
    """
    if not rows or not group_by:
        return []
    valid = [c for c in group_by if c in columns]
    if not valid:
        return []
    out: List[Dict[str, Any]] = []
    for col in valid:
        run_start = 0
        run_value = rows[0].get(col)
        for i in range(1, len(rows)):
            current = rows[i].get(col)
            if current != run_value:
                if i - run_start > 1:
                    out.append({"column": col, "row_start": run_start, "row_span": i - run_start})
                run_start = i
                run_value = current
        if len(rows) - run_start > 1:
            out.append({
                "column": col,
                "row_start": run_start,
                "row_span": len(rows) - run_start,
            })
    return out


def _normalize_column_groups(
    columns: List[str],
    column_groups: Optional[List[Any]],
) -> List[Dict[str, Any]]:
    """Keep only valid, contiguous column groups in display order."""
    if not columns or not column_groups:
        return []
    order = {col: idx for idx, col in enumerate(columns)}
    assigned: set[str] = set()
    normalized: List[Dict[str, Any]] = []
    for raw in column_groups or []:
        if hasattr(raw, "model_dump"):
            item = raw.model_dump()
        elif isinstance(raw, dict):
            item = raw
        else:
            continue
        label = str(item.get("label") or "").strip()
        raw_columns = item.get("columns") or []
        if not label or not isinstance(raw_columns, list):
            continue
        cols: List[str] = []
        seen_local: set[str] = set()
        for raw_col in raw_columns:
            col = str(raw_col or "").strip()
            if (
                col
                and col in order
                and col not in assigned
                and col not in seen_local
            ):
                cols.append(col)
                seen_local.add(col)
        if len(cols) < 2:
            continue
        cols = sorted(cols, key=order.get)
        indices = [order[col] for col in cols]
        expected = list(range(indices[0], indices[0] + len(indices)))
        if indices != expected:
            continue
        normalized.append({"label": label, "columns": cols})
        assigned.update(cols)
    return normalized


def _build_substitution_map(
    workboard: Workboard,
    *,
    app_user: Optional[Dict[str, Any]] = None,
    shared: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    now = datetime.now(timezone.utc)
    return {
        "app_user": dict(app_user or {}),
        "shared": dict(shared or {}),
        "now": now.isoformat(),
        "today": now.date().isoformat(),
        "workboard": {"id": workboard.id, "name": workboard.name},
    }


def _substitute_string(text: str, mapping: Dict[str, Any]) -> str:
    if "{{" not in text:
        return text

    def _replace(match: "re.Match[str]") -> str:
        path = match.group(1).strip()
        cursor: Any = mapping
        for part in path.split("."):
            if isinstance(cursor, dict) and part in cursor:
                cursor = cursor[part]
            else:
                return match.group(0)
        return "" if cursor is None else str(cursor)

    return re.sub(r"\{\{\s*([a-zA-Z0-9_.\-]+)\s*\}\}", _replace, text)


def _substitute_strings_in_place(obj: Any, mapping: Dict[str, Any]) -> None:
    """Recursively replace ``{{path.to.value}}`` placeholders inside strings."""
    if isinstance(obj, dict):
        for key, value in obj.items():
            if isinstance(value, str):
                obj[key] = _substitute_string(value, mapping)
            else:
                _substitute_strings_in_place(value, mapping)
    elif isinstance(obj, list):
        for i, item in enumerate(obj):
            if isinstance(item, str):
                obj[i] = _substitute_string(item, mapping)
            else:
                _substitute_strings_in_place(item, mapping)


def _resolve_relationship_labels(
    db: Session,
    *,
    base_rows: List[Dict[str, Any]],
    base_label_column: Optional[str],
    base_join_column: str,
    hops: List[Any],
) -> Dict[Any, str]:
    """Walk a relationship chain and build {primary_value → final_label}.

    Each hop fetches its target table via LiveQueryService. The chain is
    bounded (4 hops) and silently degrades to an empty mapping on any
    failure — callers fall back to the single-hop label so the form never
    breaks.
    """
    if not base_rows or not hops:
        return {}
    MAX_HOPS = 4
    safe_hops = list(hops)[:MAX_HOPS]

    cursor: Dict[Any, Any] = {}
    for row in base_rows:
        primary_val = row.get(base_join_column)
        if primary_val is None:
            continue
        cursor[primary_val] = primary_val

    last_label_col: Optional[str] = base_label_column
    visited_tables: set[int] = set()
    for hop in safe_hops:
        if not cursor:
            return {}
        # Cycle guard: a chain that revisits a table (A→B→A) would remap the
        # cursor against already-seen keys and yield wrong labels. Stop at the
        # first repeat (MAX_HOPS already bounds length; this bounds semantics).
        if hop.table_id in visited_tables:
            break
        if hop.table_id:
            visited_tables.add(hop.table_id)
        table = _load_table(db, hop.table_id)
        if table is None:
            return {}
        datasource = _load_datasource(db, table)
        if datasource is None:
            return {}
        try:
            result = LiveQueryService.execute_preview_query(
                datasource, table, limit=_MAX_LOOKUP_ROWS, offset=0, filters=[],
            )
        except SheetsQuotaError:
            raise _quota_503()
        except Exception:
            logger.exception("Nested lookup hop failed (table_id=%s)", hop.table_id)
            return {}
        rows_by_key: Dict[Any, Dict[str, Any]] = {
            row.get(hop.value_column): row
            for row in (result.get("rows") or [])
            if row.get(hop.value_column) is not None
        }
        next_cursor: Dict[Any, Any] = {}
        for current_key, primary_val in cursor.items():
            row = rows_by_key.get(current_key)
            if row is None:
                continue
            forward_value = row.get(hop.label_column) if hop.label_column else None
            if forward_value is None:
                forward_value = current_key
            next_cursor[forward_value] = primary_val
        cursor = next_cursor
        last_label_col = hop.label_column

    if last_label_col is None:
        return {}
    return {primary_val: str(label_val) for label_val, primary_val in cursor.items()}


def _resolve_lookup_options(
    db: Session, field: FormField
) -> Optional[List[Dict[str, Any]]]:
    """Materialize ``[{label, value}]`` options for a form field's lookup.

    Returns ``None`` when the field has no lookup config (caller should
    skip), an empty list when the lookup is misconfigured / unresolvable,
    or the resolved options otherwise.
    """
    cfg = field.lookup
    if cfg is None:
        return None
    if cfg.kind == "static":
        return [
            {
                "label": (item.get("label") if isinstance(item, dict) else None) or "",
                "value": item.get("value") if isinstance(item, dict) else item,
            }
            for item in (cfg.values or [])
        ]
    if cfg.kind == "dataset_table" and cfg.table_id:
        table = _load_table(db, cfg.table_id)
        if table is None:
            return []
        datasource = _load_datasource(db, table)
        if datasource is None:
            return []
        try:
            result = LiveQueryService.execute_preview_query(
                datasource, table, limit=_MAX_LOOKUP_ROWS, offset=0, filters=[]
            )
        except SheetsQuotaError:
            raise _quota_503()
        except Exception:
            logger.exception(
                "Lookup for field '%s' failed (table_id=%s)",
                field.column,
                cfg.table_id,
            )
            return []
        value_col = cfg.value_column
        label_col = cfg.label_column or value_col
        if not value_col:
            return []
        base_rows = result.get("rows") or []

        def _attach_extras(opt: Dict[str, Any], row: Dict[str, Any]) -> Dict[str, Any]:
            """Additively enrich an option with map geometry and/or the
            cascading-filter key.

            select/lookup callers never read these keys, so this is safe for
            every field — the extra keys are simply ignored downstream.
            """
            if cfg.geometry_column:
                geo = row.get(cfg.geometry_column)
                if geo not in (None, ""):
                    opt["geometry"] = geo
            if cfg.lat_column and cfg.lng_column:
                lat, lng = row.get(cfg.lat_column), row.get(cfg.lng_column)
                if lat not in (None, "") and lng not in (None, ""):
                    opt["lat"] = lat
                    opt["lng"] = lng
            # Cascading select: carry the remote match value so the FE can keep
            # only the options whose `filter_column` == the parent field's value.
            if cfg.filter_column:
                opt["filter"] = row.get(cfg.filter_column)
            return opt

        if cfg.relationship_path:
            resolved_labels = _resolve_relationship_labels(
                db,
                base_rows=base_rows,
                base_label_column=label_col,
                base_join_column=cfg.relationship_path[0].value_column,
                hops=cfg.relationship_path,
            )
            return [
                _attach_extras(
                    {
                        "label": resolved_labels.get(row.get(value_col))
                        or str(row.get(label_col, "") or ""),
                        "value": row.get(value_col),
                    },
                    row,
                )
                for row in base_rows
            ]
        return [
            _attach_extras(
                {
                    "label": str(row.get(label_col, "") or ""),
                    "value": row.get(value_col),
                },
                row,
            )
            for row in base_rows
        ]
    return []


# ── Layout + screen lookup ────────────────────────────────────────────────

def parse_layout(workboard: Workboard, *, use_published: Optional[bool] = None) -> LayoutJson:
    """Parse a workboard's layout into the typed ``LayoutJson``.

    Reads the mutable DRAFT (``layout_json``) by default — what the builder,
    internal Preview, audit and webhooks want. The PUBLIC/LIVE runtime instead
    serves the immutable PUBLISHED snapshot (``published_layout_json``); it opts
    in via the transient ``_wb_use_published`` attribute that
    ``_resolve_workboard_for_workspace`` stamps on the workboard (mirrors the
    existing ``_cleared_screens`` transient-attr pattern), or via an explicit
    ``use_published`` argument. A never-published board has
    ``published_layout_json is None`` → the runtime resolver refuses to serve it
    before we ever reach here, so live callers never see an empty layout.
    """
    # The published-vs-draft decision lives in ONE place (runtime_config), so
    # every runtime/write/export path resolves the same stage. parse_layout is
    # the typed-layout convenience over it.
    from app.modules.workboards.services.runtime_config import effective_layout_raw

    raw = effective_layout_raw(workboard, published=use_published)
    try:
        return LayoutJson.model_validate(raw or {})
    except Exception:
        logger.exception("workboard %s has invalid layout", workboard.id)
        return LayoutJson()


def get_screen(layout: LayoutJson, screen_id: str) -> Screen:
    for screen in layout.screens:
        if screen.id == screen_id:
            return screen
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"Screen '{screen_id}' not found in this workboard.",
    )


def is_screen_visible_for(screen: Screen, identity: CallerIdentity) -> bool:
    if not identity.is_app_user:
        return True
    if is_owner_role(identity.role):
        return True
    if not screen.visible_for_roles:
        return True
    role = (identity.role or "").strip().lower()
    return any(r.strip().lower() == role for r in screen.visible_for_roles)


def role_has_screen_grant(screen: Screen, identity: CallerIdentity) -> bool:
    """Access-gate counterpart to :func:`is_screen_visible_for` (which is
    NAV-DISPLAY only). True when the caller's role has an RLS grant to the
    screen — so a screen hidden from the nav can still be opened when reached
    via an explicit row-action / after_submit navigation. Fail-closed for app
    users with no matching rule. See rls_service.role_has_screen_grant."""
    return _role_has_screen_grant(screen.rls, screen.rls_default, identity)


def is_group_visible_for(group: ScreenGroup, identity: CallerIdentity) -> bool:
    """Whether a screen-group (UI: "Workspace") shows in the nav for this
    identity. Same role semantics as :func:`is_screen_visible_for` (owner +
    internal/admin preview bypass; empty list = everyone). NAV-DISPLAY only —
    NOT an access gate; the screen-content endpoint enforces per-screen RLS.

    NOTE: ``group.visible_for_roles`` is RESERVED / not settable in the builder
    (Workspaces v1), so this filter is inert today (always returns True). See
    ScreenGroup in schemas.py for the divergence semantics before exposing it.
    """
    if not identity.is_app_user:
        return True
    if is_owner_role(identity.role):
        return True
    if not group.visible_for_roles:
        return True
    role = (identity.role or "").strip().lower()
    return any(r.strip().lower() == role for r in group.visible_for_roles)


# ── Table helpers ─────────────────────────────────────────────────────────

def _load_table(db: Session, table_id: int) -> Optional[DatasetTable]:
    if not table_id:
        return None
    return db.query(DatasetTable).filter(DatasetTable.id == table_id).first()


def _columns_cache_list(table: DatasetTable) -> List[Dict[str, Any]]:
    cache = table.columns_cache
    if isinstance(cache, dict):
        raw = cache.get("columns")
        if isinstance(raw, list):
            return [item for item in raw if isinstance(item, dict)]
    if isinstance(cache, list):
        return [item for item in cache if isinstance(item, dict)]
    return []


def _table_column_names(table: DatasetTable) -> set[str]:
    return {
        str(item.get("name") or "").strip()
        for item in _columns_cache_list(table)
        if item.get("name")
    }


def _load_datasource(db: Session, table: DatasetTable) -> Optional[DataSource]:
    if table is None:
        return None
    return db.query(DataSource).filter(DataSource.id == table.datasource_id).first()


def _resolve_read_target(db: Session, table: DatasetTable) -> tuple[Optional[DataSource], Any]:
    """Resolve the ``(datasource, query_table)`` pair used to READ a screen's rows.

    Physical / ``sql_query`` tables read directly against their own datasource.
    Derived (calculated) tables and the generated Date table have no
    ``datasource_id`` of their own — they reference sibling tables through
    ``dataset_table_<id>`` aliases. Route those through
    :func:`build_live_proxy_table_for_dataset_table`, which returns a live proxy
    (``source_kind='sql_query'`` carrying the fully-inlined CTE SQL) plus the
    underlying datasource. This lets a workboard ``table`` / ``doc`` screen
    render an aggregated or cross-table result exactly like the dataset preview
    does — including on a Google Sheets source, where the proxy SQL reads every
    referenced tab into DuckDB within a single query. Read-only by nature; write
    paths never call this (a derived table isn't editable).
    """
    if table is None:
        return None, None
    from app.services.dataset_table_sql_service import (
        is_derived_table,
        build_live_proxy_table_for_dataset_table,
    )
    from app.services.dataset_calendar_service import is_generated_calendar_table

    if is_derived_table(table) or is_generated_calendar_table(table):
        from app.models.dataset import Dataset

        dataset_obj = (
            db.query(Dataset).filter(Dataset.id == table.dataset_id).first()
        )
        if dataset_obj is not None:
            try:
                return build_live_proxy_table_for_dataset_table(db, dataset_obj, table)
            except Exception as exc:  # pragma: no cover - defensive
                logger.warning(
                    "workboard: derived-table proxy resolve failed for table %s: %s",
                    getattr(table, "id", None),
                    exc,
                )
                return None, None
    return _load_datasource(db, table), table


def _resolve_pos_catalog(db: Session, pos_cart: Any) -> Dict[str, Any]:
    """Load the product master for a POS scan-cart screen.

    Returns ``{match_column, label_column, price_column, group_column, rows}``.
    The catalog is reference data (products) shared by every operator, so no RLS
    is applied. Rows are read via :func:`_resolve_read_target` so the catalog may
    itself be a physical, sql or derived table. Never raises — a missing catalog
    degrades to an empty list (the scanner then just stores raw codes).
    """
    empty = {
        "match_column": getattr(pos_cart, "catalog_match_column", None),
        "label_column": getattr(pos_cart, "catalog_label_column", None),
        "price_column": getattr(pos_cart, "catalog_price_column", None),
        "group_column": getattr(pos_cart, "catalog_group_column", None),
        "rows": [],
    }
    try:
        catalog = _load_table(db, getattr(pos_cart, "catalog_table_id", 0))
        datasource, query_table = _resolve_read_target(db, catalog) if catalog else (None, None)
        if not catalog or not datasource or query_table is None:
            return empty
        result = LiveQueryService.execute_preview_query(
            datasource, query_table, limit=_TOTALS_ROW_CAP, offset=0, filters=[]
        )
        empty["rows"] = result.get("rows") or []
        return empty
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("workboard: POS catalog resolve failed: %s", exc)
        return empty


# ── Media size ceilings (single source of truth) ──────────────────────────
# Base64-into-JSONB media. Postgres JSONB is effectively unbounded, so we use
# a documented app ceiling; Google Sheets caps a single CELL at ~50,000 chars
# (base64 inflates 4/3 → ~35 KB safe). The effective cap is chosen per screen
# by :func:`media_cap_kb` from the datasource kind — NOT hardcoded per widget.
WORKBOARD_MEDIA_MAX_KB = 1024
WORKBOARD_MEDIA_MAX_KB_SHEETS = 35
# Every widget whose value is a base64 data-URI (or a JSON array of them) and
# must be size-capped. Previously only {file, image} were capped, so images/
# signature/audio silently bypassed the ceiling.
_MEDIA_WIDGETS = {"file", "image", "images", "signature", "audio", "video"}


def media_cap_kb(db: Session, workboard: Workboard, screen: Optional["Screen"] = None) -> int:
    """Storage-aware media size ceiling in KB for a screen's writes."""
    try:
        table_id = getattr(screen, "table_id", None) or workboard.primary_table_id
        table = _load_table(db, table_id) if table_id else None
        ds = _load_datasource(db, table) if table else None
        if ds is not None and getattr(ds, "type", None) is not None:
            tval = ds.type.value if hasattr(ds.type, "value") else str(ds.type)
            if tval == "google_sheets":
                return WORKBOARD_MEDIA_MAX_KB_SHEETS
    except Exception:  # pragma: no cover - defensive; fall back to safe default
        pass
    return WORKBOARD_MEDIA_MAX_KB


def _apply_field_conditions(
    screen: Screen,
    values: Dict[str, Any],
    identity: CallerIdentity,
    media_max_kb: int = WORKBOARD_MEDIA_MAX_KB,
) -> Dict[str, Any]:
    """Enforce field-level show_if / required_if / readonly_if at write time.

    Hidden fields (``show_if`` evaluates falsy) are stripped — the FE may
    have left a stale value in the payload after the user toggled some
    other field. Required-if-true fields with empty values raise 422.
    """
    from app.modules.workboards.services.expr_eval import evaluate, evaluate_truthy

    if screen.form is None:
        return dict(values or {})

    ctx = {
        "row": dict(values or {}),
        "app_user": dict(identity.app_user or {}),
        "shared": {},
    }
    cleaned = dict(values or {})
    violations: List[str] = []
    # widget='computed' fields are DERIVED — recomputed server-side after the
    # main loop (never trust the client's submitted value). Collected here.
    computed_fields: List[Any] = []
    # Storage-aware hard ceiling for base64 media (see WORKBOARD_MEDIA_MAX_KB).
    # ``media_max_kb`` is chosen per screen from the datasource kind; the builder
    # can tighten it further per-field via FormField.max_file_kb.
    _HARD_FILE_KB_CAP = int(media_max_kb or WORKBOARD_MEDIA_MAX_KB)
    # Pages whose ``show_if`` is falsy are skipped in the wizard — their fields
    # must NOT be required or written (mirror field-level show_if). Build the
    # set of hidden page ids once.
    hidden_pages: set = set()
    for _pg in (screen.form.pages or []):
        _expr = getattr(_pg, "show_if", None)
        if _expr and not evaluate_truthy(_expr, ctx, default=True):
            hidden_pages.add(_pg.id)
    for field in screen.form.fields:
        col = field.column
        if hidden_pages and getattr(field, "page", None) in hidden_pages:
            # Field lives on a skipped page — drop any stale value, never require.
            cleaned.pop(col, None)
            continue
        if getattr(field, "readonly", False) or getattr(field, "widget", None) == "qr":
            # Static readonly fields — and the display-only `qr` widget — are
            # never written. Drop submitted values so callers cannot override
            # system columns such as generated PKs (or the QR's source column).
            cleaned.pop(col, None)
            continue
        if getattr(field, "widget", None) in _MEDIA_WIDGETS:
            raw_value = cleaned.get(col)
            # A media cell is either a single data-URL string, or (images) a
            # JSON array of data-URL strings — possibly still a JSON string.
            parts: List[str] = []
            if isinstance(raw_value, str) and raw_value:
                if raw_value.startswith("["):
                    try:
                        decoded = json.loads(raw_value)
                        parts = [p for p in decoded if isinstance(p, str)]
                    except (ValueError, TypeError):
                        parts = [raw_value]
                else:
                    parts = [raw_value]
            elif isinstance(raw_value, list):
                parts = [p for p in raw_value if isinstance(p, str)]
            if parts:
                # Base64 expands by 4/3; size_kb ≈ len * 3 / 4 / 1024. Sum every
                # item so a multi-image cell can't exceed the cell budget.
                total_len = 0
                for item in parts:
                    total_len += len(
                        item.split(",", 1)[1]
                        if item.startswith("data:") and "," in item
                        else item
                    )
                size_kb = (total_len * 3) // 4 // 1024
                builder_cap = int(getattr(field, "max_file_kb", None) or 0)
                effective_cap = (
                    min(builder_cap, _HARD_FILE_KB_CAP) if builder_cap > 0
                    else _HARD_FILE_KB_CAP
                )
                if size_kb > effective_cap:
                    label = field.label or col
                    violations.append(
                        f"Tệp '{label}' lớn hơn giới hạn {effective_cap} KB "
                        f"(thực tế ≈ {size_kb} KB)."
                    )
                    cleaned.pop(col, None)
                    continue
        # ── Type coercion for rich numeric / list widgets ──────────────
        # The FE normally emits the storable form, but coerce defensively so a
        # locale-formatted string ("1.234,5") or a JSON-array string can never
        # land in a numeric/array column raw. Numbers reuse the vi-VN-aware
        # parser (dot-decimal safe — see _parse_locale_number).
        _widget = getattr(field, "widget", None)
        if _widget in ("currency", "percent", "rating", "slider", "duration"):
            _v = cleaned.get(col)
            if isinstance(_v, str) and _v.strip() != "":
                _n = _parse_locale_number(_v)
                if _n is not None:
                    cleaned[col] = _n
        elif _widget == "enum_list":
            # Store as a JSON STRING so a text/jsonb cell write never receives a
            # Python list (the connector would emit a PG array literal → type
            # mismatch). The FE already sends a JSON string; normalise a stray
            # list defensively.
            _v = cleaned.get(col)
            if isinstance(_v, list):
                cleaned[col] = json.dumps(_v, ensure_ascii=False)
        if getattr(field, "computed_from_dataset", None):
            # Field is auto-filled by a dataset-side transformation (calculated
            # column, lookup, etc.). The dataset is the source of truth, so any
            # value the FE submits here would clobber the computed result on
            # the next fetch. Drop it — same contract as `readonly` static
            # fields, but the cause is "computed upstream" not "system PK".
            cleaned.pop(col, None)
            continue
        show_if_expr = getattr(field, "show_if", None)
        if show_if_expr and not evaluate_truthy(show_if_expr, ctx, default=True):
            cleaned.pop(col, None)
            continue
        readonly_if_expr = getattr(field, "readonly_if", None)
        if readonly_if_expr and evaluate_truthy(readonly_if_expr, ctx, default=False):
            # Drop the column so the existing DB value (or default) wins.
            cleaned.pop(col, None)
            continue
        if _widget == "computed":
            # DERIVED, display-only value. NEVER trust the client's submitted
            # number — a tampered payload could store anything. Strip it now;
            # the authoritative value is recomputed from `formula` in the
            # dedicated fixpoint pass below.
            cleaned.pop(col, None)
            computed_fields.append(field)
            continue
        required_if_expr = getattr(field, "required_if", None)
        is_required_static = bool(getattr(field, "required", False))
        is_required = is_required_static
        if required_if_expr:
            is_required = bool(evaluate_truthy(required_if_expr, ctx, default=False))
        if is_required:
            v = cleaned.get(col)
            if v is None or (isinstance(v, str) and v.strip() == ""):
                violations.append(f"Trường '{field.label or col}' là bắt buộc.")
        # valid_if runs AFTER required — an empty optional field skips it.
        # Refresh ctx.row with the latest cleaned values so expressions that
        # reference sibling fields see the values being submitted, not the
        # untouched copy from the top of the loop.
        valid_if_expr = getattr(field, "valid_if", None)
        if valid_if_expr:
            value = cleaned.get(col)
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue
            ctx["row"] = dict(cleaned)
            if not evaluate_truthy(valid_if_expr, ctx, default=True):
                custom_msg = getattr(field, "valid_if_error", None)
                label = field.label or col
                violations.append(
                    custom_msg.strip() if custom_msg and custom_msg.strip()
                    else f"Trường '{label}' không thoả điều kiện kiểm tra."
                )
    # ── Server-authoritative recompute of widget='computed' fields ──────────
    # Derive each value from `formula` (strict mode → a non-numeric SUM_SPLIT
    # segment rejects the save). Iterate to a fixpoint so one computed field
    # may reference another (mirrors the TABLE pipeline's topological pass).
    if computed_fields:
        from app.modules.workboards.services.expr_eval import evaluate_detailed

        _cctx_base = {
            "app_user": dict(identity.app_user or {}),
            "shared": {},
        }
        for _ in range(len(computed_fields) + 1):
            changed = False
            _row_now = dict(cleaned)
            for f in computed_fields:
                _formula = (getattr(f, "formula", None) or "").strip()
                if not _formula:
                    continue
                _res = evaluate_detailed(_formula, {"row": _row_now, **_cctx_base})
                if _res.get("error") is None:
                    _val = _res.get("value")
                    if cleaned.get(f.column) != _val:
                        cleaned[f.column] = _val
                        changed = True
            if not changed:
                break
        # Final settled pass: write results + surface strict errors as 422.
        _row_final = dict(cleaned)
        for f in computed_fields:
            _formula = (getattr(f, "formula", None) or "").strip()
            if not _formula:
                # Draft/empty formula — no server value; leave column unwritten.
                continue
            _res = evaluate_detailed(_formula, {"row": _row_final, **_cctx_base})
            if _res.get("error"):
                _label = getattr(f, "label", None) or f.column
                _msg = (_res.get("error") or {}).get("message") or "lỗi biểu thức"
                violations.append(f"Trường '{_label}' không tính được: {_msg}.")
            else:
                cleaned[f.column] = _res.get("value")

    if violations:
        raise HTTPException(
            status_code=422,
            detail={"message": "Validation failed", "violations": violations},
        )
    return cleaned


def _status_fields(screen: Screen) -> List[Any]:
    """Form fields with widget='status' that carry a StatusConfig."""
    if screen.form is None:
        return []
    return [
        f
        for f in screen.form.fields
        if getattr(f, "widget", None) == "status" and getattr(f, "status_config", None)
    ]


def _geocode_config_for_screen(screen: Screen) -> Any:
    if screen.kind == "form" and screen.form is not None:
        return getattr(screen.form, "geocode", None)
    if screen.kind == "table" and screen.table is not None:
        return getattr(screen.table, "geocode", None)
    return None


def _apply_geocode(
    screen: Screen,
    values: Dict[str, Any],
    *,
    previous_row: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Auto-fill lat/lng from an address when a screen opts in (form/table
    ``geocode`` config). Best-effort by design: a provider miss must not block
    the write — expose ``status_column`` when operators need to review
    unresolved rows. Only runs on single-row writes (form submit / edit); bulk
    writes are intentionally not geocoded to avoid N sequential provider calls.
    """
    cfg = _geocode_config_for_screen(screen)
    if cfg is None or not bool(getattr(cfg, "enabled", True)):
        return values
    lat_col = getattr(cfg, "lat_column", None)
    lng_col = getattr(cfg, "lng_column", None)
    if not lat_col or not lng_col:
        return values

    def _blank(v: Any) -> bool:
        return v is None or (isinstance(v, str) and not v.strip())

    out = dict(values or {})
    combined = dict(previous_row or {})
    combined.update(out)
    address_col = getattr(cfg, "address_column", None)
    template = getattr(cfg, "address_template", None)
    # On insert (no previous row) with a plain address column: only geocode
    # when that address is actually part of this write.
    if previous_row is None and address_col and address_col not in out and not template:
        return out
    overwrite = bool(getattr(cfg, "overwrite_existing", False))
    if not overwrite and not (_blank(combined.get(lat_col)) or _blank(combined.get(lng_col))):
        return out

    status_col = getattr(cfg, "status_column", None)
    label_col = getattr(cfg, "provider_label_column", None)
    address = build_address(cfg, combined)
    if not address:
        if status_col:
            out[status_col] = "Thiếu địa chỉ"
        return out

    result = geocode_address(cfg, address)
    if result is None:
        if status_col:
            out[status_col] = "Không tìm thấy tọa độ"
        return out

    out[lat_col] = result.lat
    out[lng_col] = result.lng
    if status_col:
        out[status_col] = "Đã tự sinh tọa độ"
    if label_col and result.label:
        out[label_col] = result.label
    return out


def _fetch_current_row(
    db: Session, workboard: Workboard, screen: Screen, pk: Dict[str, Any]
) -> Optional[Dict[str, Any]]:
    """Load the row targeted by ``pk`` on the screen's table (pre-update
    snapshot) so status-transition rules can compare previous -> new."""
    if not pk or screen.table_id is None:
        return None
    table = _load_table(db, screen.table_id)
    datasource = _load_datasource(db, table) if table else None
    if not table or not datasource:
        return None
    filters = [{"field": k, "operator": "eq", "value": v} for k, v in pk.items()]
    result = LiveQueryService.execute_preview_query(
        datasource, table, limit=1, offset=0, filters=filters
    )
    rows = result.get("rows") or []
    return rows[0] if rows else None


def _resolve_relation_bound_values(
    db: Session,
    workboard: Workboard,
    layout: LayoutJson,
    child_screen: Screen,
    relation_context: Optional[Dict[str, Any]],
    identity: CallerIdentity,
    *,
    enforce_cardinality: bool = False,
) -> Dict[str, Any]:
    """Validate a parent-child relation context and return server-bound values.

    The client may name the relation and parent key value, but it never decides
    which child column is written. That comes from the parent form's
    RelatedRecordConfig after parent-row access has been verified.
    """

    if not relation_context:
        return {}
    if not isinstance(relation_context, dict):
        raise HTTPException(status_code=400, detail="Invalid relation context.")

    relation_id = str(relation_context.get("relation_id") or "").strip()
    parent_screen_id = str(relation_context.get("parent_screen_id") or "").strip()
    parent_key_value = relation_context.get("parent_key_value")
    if not relation_id or not parent_screen_id:
        raise HTTPException(status_code=400, detail="Relation context is incomplete.")
    if parent_key_value is None or (
        isinstance(parent_key_value, str) and not parent_key_value.strip()
    ):
        raise HTTPException(status_code=400, detail="Parent key value is required.")

    parent_screen = get_screen(layout, parent_screen_id)
    if parent_screen.kind != "form" or parent_screen.form is None:
        raise HTTPException(status_code=400, detail="Relation parent must be a form screen.")

    relation = None
    for cfg in parent_screen.form.related_records or []:
        if cfg.id == relation_id:
            relation = cfg
            break
    if relation is None:
        raise HTTPException(status_code=400, detail="Related-record config was not found.")
    if relation.child_screen_id != child_screen.id:
        raise HTTPException(status_code=403, detail="Child screen is not part of this relation.")
    if relation.parent_key_column not in set(parent_screen.primary_key_columns or []):
        raise HTTPException(
            status_code=400,
            detail="Relation parent key must be one of the parent screen's primary_key_columns.",
        )

    if not parent_screen.table_id or not child_screen.table_id:
        raise HTTPException(status_code=400, detail="Relation screens must be bound to tables.")

    parent_table = _load_table(db, parent_screen.table_id)
    child_table = _load_table(db, child_screen.table_id)
    if not parent_table or not child_table:
        raise HTTPException(status_code=400, detail="Relation table binding is missing.")

    parent_columns = _table_column_names(parent_table)
    child_columns = _table_column_names(child_table)
    if relation.parent_key_column not in parent_columns:
        raise HTTPException(status_code=400, detail="Relation parent key column is missing.")
    if relation.child_foreign_key_column not in child_columns:
        raise HTTPException(status_code=400, detail="Relation child FK column is missing.")

    datasource = _load_datasource(db, parent_table)
    if datasource is None:
        raise HTTPException(status_code=400, detail="Relation parent datasource is missing.")

    rls_filters, allowed = build_rls_filter(
        parent_screen.rls, parent_screen.rls_default, identity
    )
    if not allowed:
        raise HTTPException(status_code=403, detail="You don't have access to the parent row.")

    filters = [
        {
            "field": relation.parent_key_column,
            "operator": "eq",
            "value": parent_key_value,
        }
    ] + rls_filters
    result = LiveQueryService.execute_preview_query(
        datasource, parent_table, limit=1, offset=0, filters=filters
    )
    if not (result.get("rows") or []):
        raise HTTPException(status_code=403, detail="You don't have access to the parent row.")

    if enforce_cardinality and not relation.allow_multiple:
        child_datasource = _load_datasource(db, child_table)
        if child_datasource is None:
            raise HTTPException(status_code=400, detail="Relation child datasource is missing.")
        existing = LiveQueryService.execute_preview_query(
            child_datasource,
            child_table,
            limit=1,
            offset=0,
            filters=[
                {
                    "field": relation.child_foreign_key_column,
                    "operator": "eq",
                    "value": parent_key_value,
                }
            ],
        )
        if existing.get("rows") or []:
            raise HTTPException(
                status_code=409,
                detail="This relation allows only one child record for each parent.",
            )

    return {relation.child_foreign_key_column: parent_key_value}


def open_related_records_context(
    db: Session,
    workboard: Workboard,
    source_screen: Screen,
    *,
    action_id: str,
    pk: Dict[str, Any],
    identity: CallerIdentity,
) -> Dict[str, Any]:
    """Resolve a row action into a server-verified relation context.

    The client supplies only the stored action id and row PK. Relation ownership,
    the parent key value, child target, and policy flags all come from the
    published layout after both the source-table and parent-form RLS checks pass.
    """

    if source_screen.kind != "table" or source_screen.table is None or not source_screen.table_id:
        raise HTTPException(status_code=400, detail="Open Related Records requires a table screen.")
    action = next(
        (item for item in (source_screen.table.row_actions or []) if item.id == action_id),
        None,
    )
    if action is None or action.action_type != "open_related_records":
        raise HTTPException(status_code=404, detail="Open Related Records action was not found.")
    if identity.is_app_user and not is_owner_role(identity.role) and action.visible_for_roles:
        role = (identity.role or "").strip().lower()
        if not any(item.strip().lower() == role for item in action.visible_for_roles):
            raise HTTPException(status_code=403, detail="You don't have access to that action.")
    relation_id = str(action.relation_id or "").strip()
    if not relation_id:
        raise HTTPException(status_code=400, detail="The row action has no relation_id.")

    layout = parse_layout(workboard)
    candidates: List[Screen] = []
    for candidate in layout.screens:
        if (
            candidate.kind == "form"
            and candidate.form is not None
            and candidate.table_id == source_screen.table_id
            and (not action.parent_screen_id or candidate.id == action.parent_screen_id)
            and any(item.id == relation_id for item in (candidate.form.related_records or []))
        ):
            candidates.append(candidate)
    if len(candidates) != 1:
        raise HTTPException(
            status_code=400,
            detail=(
                "The relation_id must resolve to exactly one parent form bound "
                "to this table."
            ),
        )
    parent_screen = candidates[0]
    relation = next(
        item for item in (parent_screen.form.related_records or []) if item.id == relation_id
    )
    if relation.parent_key_column not in set(parent_screen.primary_key_columns or []):
        raise HTTPException(
            status_code=400,
            detail="Relation parent key must be one of the parent screen's primary_key_columns.",
        )

    required_pk = list(source_screen.primary_key_columns or [])
    if not required_pk or any(column not in pk for column in required_pk):
        raise HTTPException(status_code=400, detail="A complete row primary key is required.")
    source_table = _load_table(db, source_screen.table_id)
    datasource = _load_datasource(db, source_table) if source_table else None
    if not source_table or not datasource:
        raise HTTPException(status_code=400, detail="Source table binding is missing.")
    rls_filters, allowed = build_rls_filter(
        source_screen.rls, source_screen.rls_default, identity
    )
    if not allowed:
        raise HTTPException(status_code=403, detail="You don't have access to that row.")
    filters = [
        {"field": column, "operator": "eq", "value": pk[column]}
        for column in required_pk
    ] + rls_filters
    result = LiveQueryService.execute_preview_query(
        datasource, source_table, limit=1, offset=0, filters=filters
    )
    rows = result.get("rows") or []
    if not rows:
        raise HTTPException(status_code=403, detail="You don't have access to that row.")
    parent_row = rows[0]
    parent_key_value = parent_row.get(relation.parent_key_column)
    if parent_key_value is None or (
        isinstance(parent_key_value, str) and not parent_key_value.strip()
    ):
        raise HTTPException(status_code=409, detail="The selected parent row has no relation key.")

    child_screen = get_screen(layout, relation.child_screen_id)
    if not role_has_screen_grant(child_screen, identity):
        raise HTTPException(status_code=403, detail="You don't have access to the child screen.")
    _resolve_relation_bound_values(
        db,
        workboard,
        layout,
        child_screen,
        {
            "relation_id": relation.id,
            "parent_screen_id": parent_screen.id,
            "parent_key_value": parent_key_value,
        },
        identity,
    )
    return {
        "child_screen_id": child_screen.id,
        "parent_values": parent_row,
        "relation_context": {
            "relation_id": relation.id,
            "relation_label": relation.label,
            "parent_screen_id": parent_screen.id,
            "child_screen_id": child_screen.id,
            "parent_key_column": relation.parent_key_column,
            "parent_key_value": parent_key_value,
            "child_foreign_key_column": relation.child_foreign_key_column,
            "finish_screen_id": relation.finish_screen_id,
            "show_existing": relation.show_existing,
            "allow_multiple": relation.allow_multiple,
            "keep_parent_context": relation.keep_parent_context,
        },
    }


def render_related_records(
    db: Session,
    workboard: Workboard,
    *,
    parent_screen_id: str,
    relation_id: str,
    parent_key_value: Any,
    identity: CallerIdentity,
) -> Dict[str, Any]:
    """Return child rows for a validated parent-child relation context."""

    layout = parse_layout(workboard)
    parent_screen = get_screen(layout, parent_screen_id)
    if parent_screen.kind != "form" or parent_screen.form is None:
        raise HTTPException(status_code=400, detail="Relation parent must be a form screen.")

    relation = None
    for cfg in parent_screen.form.related_records or []:
        if cfg.id == relation_id:
            relation = cfg
            break
    if relation is None:
        raise HTTPException(status_code=400, detail="Related-record config was not found.")

    child_screen = get_screen(layout, relation.child_screen_id)
    _resolve_relation_bound_values(
        db,
        workboard,
        layout,
        child_screen,
        {
            "relation_id": relation_id,
            "parent_screen_id": parent_screen_id,
            "parent_key_value": parent_key_value,
        },
        identity,
    )

    if child_screen.table_id is None:
        raise HTTPException(status_code=400, detail="Relation child screen has no table.")
    child_table = _load_table(db, child_screen.table_id)
    datasource = _load_datasource(db, child_table) if child_table else None
    if not child_table or not datasource:
        raise HTTPException(status_code=400, detail="Relation child datasource is missing.")

    rls_filters, allowed = build_rls_filter(
        child_screen.rls, child_screen.rls_default, identity
    )
    if not allowed:
        return {"columns": [], "rows": [], "total_count": 0}

    filters = [
        {
            "field": relation.child_foreign_key_column,
            "operator": "eq",
            "value": parent_key_value,
        }
    ] + rls_filters
    result = LiveQueryService.execute_preview_query(
        datasource, child_table, limit=500, offset=0, filters=filters
    )
    rows = result.get("rows") or []
    all_columns = result.get("columns") or list(rows[0].keys() if rows else [])
    edit_columns: List[str] = []
    if child_screen.kind == "form" and child_screen.form is not None:
        edit_columns = [
            f.column for f in child_screen.form.fields
            if f.column in all_columns and f.column != relation.child_foreign_key_column
        ]
    if relation.display_columns:
        columns = [c for c in relation.display_columns if c in all_columns]
    elif edit_columns:
        columns = edit_columns
    elif child_screen.kind == "table" and child_screen.table is not None:
        columns = [
            c for c in child_screen.table.columns
            if c in all_columns and c != relation.child_foreign_key_column
        ]
    else:
        columns = [c for c in all_columns if c != relation.child_foreign_key_column]
    pk_cols = list(child_screen.primary_key_columns or [])
    row_keys = list(dict.fromkeys([*pk_cols, *columns, *edit_columns]))
    shaped_rows = [{col: row.get(col) for col in row_keys} for row in rows]
    return {
        "columns": columns,
        "primary_key_columns": pk_cols,
        "rows": shaped_rows,
        "total_count": len(rows),
    }


def _enforce_status_rules(
    status_fields: List[Any],
    values: Dict[str, Any],
    identity: CallerIdentity,
    *,
    previous_row: Optional[Dict[str, Any]] = None,
) -> None:
    """Server-side status lifecycle guard.

    The FE ``editable_by_roles`` gate is advisory only (bypassable via a
    crafted API call), so re-check it here, plus the new ``allowed_transitions``
    map. For each status field actually being written:
      * ``editable_by_roles`` — the caller's role must be allowed to change it.
      * ``allowed_transitions`` — the previous->new pair must be permitted
        (only checked on update, where a previous value exists).
    AppBI staff and the ``owner`` role bypass (same contract as RLS).
    """
    if not status_fields:
        return
    if not identity.is_app_user or is_owner_role(identity.role):
        return
    role = (identity.role or "").strip().lower()
    for field in status_fields:
        col = field.column
        if col not in values:
            continue  # not being written (e.g. RLS writable_columns stripped it)
        new_s = "" if values.get(col) is None else str(values.get(col))
        prev_raw = previous_row.get(col) if previous_row else None
        prev_s = "" if prev_raw is None else str(prev_raw)
        if previous_row is not None and new_s == prev_s:
            continue  # unchanged — never blocked
        cfg = field.status_config
        editable_roles = [r.strip().lower() for r in (cfg.editable_by_roles or [])]
        if editable_roles and role not in editable_roles:
            raise HTTPException(
                status_code=403,
                detail=f"Vai trò của bạn không được phép đổi trạng thái '{field.label or col}'.",
            )
        transitions = cfg.allowed_transitions or {}
        if transitions and previous_row is not None and prev_s:
            allowed = transitions.get(prev_s)
            if allowed is not None:
                allowed_s = [str(x) for x in allowed]
                if new_s not in allowed_s:
                    nice = ", ".join(allowed_s) if allowed_s else "(trạng thái kết thúc)"
                    raise HTTPException(
                        status_code=403,
                        detail=(
                            f"Không thể chuyển '{field.label or col}' từ '{prev_s}' "
                            f"sang '{new_s}'. Cho phép: {nice}."
                        ),
                    )


def _enforce_row_lock(
    cfg: Any,
    previous_row: Optional[Dict[str, Any]],
    identity: CallerIdentity,
    *,
    op: str = "update",
) -> None:
    """Server-side per-row edit/delete lock for table screens.

    Re-evaluates ``row_lock.lock_if`` against the row's CURRENT stored values
    (``previous_row``) — NOT the incoming payload — so a user cannot unlock a
    row by editing the lock column in the same request. When the row is locked
    and the caller's role is not in ``editable_by_roles``, block with 403.
    AppBI staff and the ``owner`` role always bypass (same contract as RLS and
    the status guard, so a locked record is never permanently un-fixable). For
    a delete, only enforce when ``lock_delete`` is set.
    """
    if cfg is None or previous_row is None:
        return
    if op == "delete" and not bool(getattr(cfg, "lock_delete", True)):
        return
    if not identity.is_app_user or is_owner_role(identity.role):
        return
    lock_if = (getattr(cfg, "lock_if", None) or "").strip()
    if not lock_if:
        return
    from app.modules.workboards.services.expr_eval import evaluate_truthy

    ctx = {
        "row": dict(previous_row),
        "app_user": dict(identity.app_user or {}),
        "shared": {},
    }
    # default=False → a malformed expression fails OPEN (row not locked), same
    # fail-soft stance as FormField.readonly_if. The builder dry-run validates
    # the expression, so this only bites on hand-crafted bad layouts.
    if not evaluate_truthy(lock_if, ctx, default=False):
        return  # row not locked for this row's data
    role = (identity.role or "").strip().lower()
    allowed = [r.strip().lower() for r in (getattr(cfg, "editable_by_roles", None) or [])]
    if role in allowed:
        return  # role explicitly allowed to edit/delete locked rows
    msg = (getattr(cfg, "message", None) or "").strip() or (
        "Dòng này đã bị khóa — bạn không có quyền "
        + ("xóa." if op == "delete" else "chỉnh sửa.")
    )
    raise HTTPException(status_code=403, detail=msg)


def _resolve_initial_values(
    initial: Dict[str, Any],
    *,
    identity: CallerIdentity,
    shared_context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Substitute ``{{app_user.x}}`` / ``{{shared.col}}`` / ``{{today}}``
    in form initial values so the FE can prefill the form without having
    to know about placeholders itself."""
    from datetime import datetime, timezone

    now = datetime.now(timezone.utc)
    out: Dict[str, Any] = {}
    for col, raw in (initial or {}).items():
        if not isinstance(raw, str) or "{{" not in raw:
            out[col] = raw
            continue
        text = raw.strip()
        # Cheap exact-match resolution first.
        if text == "{{today}}":
            out[col] = now.date().isoformat()
            continue
        if text == "{{now}}":
            out[col] = now.isoformat()
            continue
        if text.startswith("{{app_user.") and text.endswith("}}"):
            key = text[len("{{app_user.") : -2].strip()
            if identity.app_user:
                if key == "username":
                    out[col] = identity.app_user.get("username")
                else:
                    out[col] = identity.app_user.get(key)
            continue
        if text.startswith("{{shared.") and text.endswith("}}"):
            key = text[len("{{shared.") : -2].strip()
            if shared_context:
                out[col] = shared_context.get(key)
            continue
        # Otherwise leave as-is for the runtime layer to handle.
        out[col] = raw
    return out


# ── Form screen ───────────────────────────────────────────────────────────

def render_form_screen(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    *,
    identity: CallerIdentity,
    shared_context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    if screen.kind != "form" or screen.form is None:
        raise HTTPException(status_code=400, detail="Screen is not a form.")

    # Resolve lookup options the same way the legacy form view does, but
    # restricted to this screen's fields.
    lookups: Dict[str, List[Dict[str, Any]]] = {}
    for field in screen.form.fields:
        opts = _resolve_lookup_options(db, field)
        if opts is not None:
            lookups[field.column] = opts

    initial = _resolve_initial_values(
        screen.form.initial_values,
        identity=identity,
        shared_context=shared_context,
    )

    # Per-field `default` values support the same {{app_user.x}} / {{today}} /
    # {{shared.x}} placeholders as `initial_values`, but were shipped verbatim
    # (model_dump) so the FE showed the literal "{{app_user.username}}". Resolve
    # each field default through the same substituter before emitting.
    def _emit_field(field: Any) -> Dict[str, Any]:
        fd = field.model_dump()
        raw = fd.get("default")
        if isinstance(raw, str) and "{{" in raw:
            fd["default"] = _resolve_initial_values(
                {"_": raw}, identity=identity, shared_context=shared_context
            ).get("_")
        return fd

    layout = parse_layout(workboard)
    auto_number_columns = [
        cfg.column for cfg in (layout.auto_number_columns or []) if cfg.column
    ]
    # Per-column scope metadata so the FE can tell the user WHICH fields must be
    # filled first for a scoped sequence to generate (otherwise a scoped rule
    # silently produces no number — confusing). Additive; back-compat preserved.
    auto_number_meta = {
        cfg.column: {
            "scope_columns": list(getattr(cfg, "scope_columns", None) or []),
            "date_column": getattr(cfg, "date_column", None),
            "missing_scope_behavior": getattr(cfg, "missing_scope_behavior", "empty"),
        }
        for cfg in (layout.auto_number_columns or [])
        if cfg.column
    }

    return {
        "screen_id": screen.id,
        "kind": "form",
        "title": screen.title,
        "icon": screen.icon,
        "description": screen.description,
        "table_id": screen.table_id,
        "primary_key_columns": list(screen.primary_key_columns or []),
        "submit_label": screen.form.submit_label,
        "fields": [_emit_field(field) for field in screen.form.fields],
        "lookups": lookups,
        "initial_values": initial,
        "after_submit": (
            screen.form.after_submit.model_dump()
            if screen.form.after_submit is not None
            else None
        ),
        "related_records": [
            relation.model_dump() for relation in (screen.form.related_records or [])
        ],
        "pages": [p.model_dump() for p in (screen.form.pages or [])],
        "sections": list(screen.form.sections or []),
        # Expose ONLY whether photo-capture/OCR is on — never the token/model.
        "ocr": {"enabled": bool(getattr(screen.form.ocr, "enabled", False))}
        if screen.form.ocr is not None else {"enabled": False},
        # Columns the server will auto-fill on insert when blank. The FE
        # treats these as readonly + shows a hint so users don't think the
        # form is broken when typing into them is ignored.
        "auto_number_columns": auto_number_columns,
        "auto_number_meta": auto_number_meta,
        # When set, the FE captures device GPS at submit and writes "lat,lng"
        # into this column (anti-fraud geo-audit).
        "geo_stamp_column": screen.form.geo_stamp_column,
    }


def _form_split_field(screen: Screen) -> Optional[Any]:
    """The form's fan-out field (widget='enum_list' with split_to_rows), if any.

    The schema validator guarantees at most one, so return the first match.
    """
    if screen.kind != "form" or screen.form is None:
        return None
    for f in screen.form.fields:
        if getattr(f, "split_to_rows", False) and getattr(f, "widget", None) == "enum_list":
            return f
    return None


def _parse_split_values(raw: Any) -> List[Any]:
    """Normalise an enum_list submit value into a list of non-empty scalars.

    The FE sends enum_list as a JSON-array STRING (``'["A","B"]'``); tolerate a
    raw Python list or a bare scalar too. Blank/None entries are dropped.
    """
    if raw is None:
        return []
    if isinstance(raw, list):
        vals = raw
    elif isinstance(raw, str):
        s = raw.strip()
        if not s:
            return []
        if s.startswith("["):
            try:
                parsed = json.loads(s)
                vals = parsed if isinstance(parsed, list) else [s]
            except (ValueError, TypeError):
                vals = [s]
        else:
            vals = [s]
    else:
        vals = [raw]
    out: List[Any] = []
    for v in vals:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        out.append(v)
    return out


def _plan_split_submit(
    values: Dict[str, Any],
    split_field: Any,
    client_op_id: Optional[str],
) -> Tuple[str, Any]:
    """Plan a multi-select fan-out. Pure (no DB) so it can be unit-tested.

    Returns one of:
      * ``("fanout", [(child_values, child_op_id), ...])`` — >1 selected values;
        each child copies every field and overrides the split column with one
        scalar value, with a deterministic ``"{op}:{i}"`` op-id.
      * ``("single", scalar)`` — exactly one value; caller stores the scalar.
      * ``("none", None)`` — nothing selected; caller inserts as-is.
    """
    vals = _parse_split_values((values or {}).get(split_field.column))
    if len(vals) > 1:
        base = dict(values or {})
        children = [
            (
                {**base, split_field.column: v},
                (f"{client_op_id}:{i}" if client_op_id else None),
            )
            for i, v in enumerate(vals)
        ]
        return ("fanout", children)
    if len(vals) == 1:
        return ("single", vals[0])
    return ("none", None)


def insert_screen_row(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    values: Dict[str, Any],
    *,
    identity: CallerIdentity,
    client_op_id: Optional[str] = None,
    relation_context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    # Form and table screens are writable. Form runs field-level
    # conditional rules (show_if / required_if) first; table screens use
    # their own column-set rules (editable_columns + computed/lookup
    # stripping).
    if screen.table_id is None or screen.kind not in ("form", "table"):
        raise HTTPException(status_code=400, detail="Screen is not writable.")
    if screen.kind == "form" and screen.form is None:
        raise HTTPException(status_code=400, detail="Form screen has no spec.")
    if screen.kind == "table":
        if screen.table is None:
            raise HTTPException(status_code=400, detail="Table screen has no spec.")
        if not screen.table.allow_add_row:
            raise HTTPException(
                status_code=403, detail="Adding rows is disabled on this table."
            )
        if not (screen.table.editable_columns or []):
            raise HTTPException(
                status_code=403,
                detail=(
                    "This table has no editable columns — add at least one "
                    "to editable_columns before turning on allow_add_row."
                ),
            )

    # ── Multi-select fan-out: one submit → one row per selected value ──────
    # A form field flagged split_to_rows explodes into N rows here (before any
    # op-log / write), recursing once per value. Each child carries a scalar
    # for the split column (so it won't re-expand) and a deterministic child
    # op-id so offline replay stays idempotent. Every other field is copied.
    if screen.kind == "form":
        _split_field = _form_split_field(screen)
        if _split_field is not None:
            _mode, _payload = _plan_split_submit(values, _split_field, client_op_id)
            if _mode == "fanout":
                _results = [
                    insert_screen_row(
                        db,
                        workboard,
                        screen,
                        _child_values,
                        identity=identity,
                        client_op_id=_child_op,
                        relation_context=relation_context,
                    )
                    for _child_values, _child_op in _payload
                ]
                return {
                    "action": "insert_split",
                    "affected_rows": len(_results),
                    "results": _results,
                    # Keep row/pk for after_submit compatibility (first child).
                    "row": _results[0].get("row", {}) if _results else {},
                    "pk": _results[0].get("pk", {}) if _results else {},
                }
            if _mode == "single":
                # Single pick → store the scalar (not a 1-element JSON array), so
                # the column holds one value per row consistently.
                values = {**(values or {}), _split_field.column: _payload}

    op_entry: Optional[WorkboardOpLog] = None
    if client_op_id is not None:
        normalized_op_id = client_op_id.strip()
        if not normalized_op_id or len(normalized_op_id) > 64:
            raise HTTPException(status_code=400, detail="client_op_id must be 1 to 64 characters.")
        actor_key = _op_actor_key(identity)
        request_fingerprint = _op_request_fingerprint(screen.id, values, relation_context)
        existing = db.get(WorkboardOpLog, normalized_op_id)
        if existing is not None:
            return _replay_op_result(
                existing,
                workboard_id=workboard.id,
                screen_id=screen.id,
                actor_key=actor_key,
                request_fingerprint=request_fingerprint,
            )
        op_entry = WorkboardOpLog(
            op_id=normalized_op_id,
            workboard_id=workboard.id,
            screen_id=screen.id,
            actor_key=actor_key,
            request_fingerprint=request_fingerprint,
        )
        try:
            db.add(op_entry)
            db.flush()
        except IntegrityError:
            # A concurrent replay may have committed while this request was
            # waiting on the PK. Resolve it before touching the datasource.
            db.rollback()
            existing = db.get(WorkboardOpLog, normalized_op_id)
            if existing is None:
                raise HTTPException(status_code=409, detail="Operation replay could not be resolved.")
            return _replay_op_result(
                existing,
                workboard_id=workboard.id,
                screen_id=screen.id,
                actor_key=actor_key,
                request_fingerprint=request_fingerprint,
            )

    if screen.kind == "form":
        cleaned_pre = _apply_field_conditions(
            screen, values, identity, media_max_kb=media_cap_kb(db, workboard, screen)
        )
    else:
        # Table: merge builder default_values (with placeholders) under
        # user values. Computed/lookup columns are never writeable from
        # the client — strip them first so a hand-crafted payload can't
        # slip a fake `total` past us.
        derived = {c.name for c in (screen.table.computed_columns or [])} | {
            l.name for l in (screen.table.lookup_columns or [])
        }
        merged = _resolve_table_defaults(screen, identity)
        merged.update({
            k: v for k, v in (values or {}).items() if k not in derived
        })
        missing = [
            col
            for col in (screen.table.required_columns if screen.table else [])
            if merged.get(col) in (None, "") and col not in derived
        ]
        if missing:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": f"Required columns missing: {', '.join(missing)}",
                    "violations": [{"column": c, "rule": "required"} for c in missing],
                },
            )
        cleaned_pre = merged

    cleaned = enforce_write_access(
        screen.rls, screen.rls_default, identity, op="insert", row_values=cleaned_pre
    )
    relation_bound = _resolve_relation_bound_values(
        db,
        workboard,
        parse_layout(workboard),
        screen,
        relation_context,
        identity,
        enforce_cardinality=True,
    )
    if relation_bound:
        cleaned.update(relation_bound)
    # Server-side enrichment: fill lat/lng from the address if the screen opts
    # in via geocode config (no-op otherwise).
    cleaned = _apply_geocode(screen, cleaned)
    if screen.kind == "form":
        # Value-level status guard on top of RLS column masking (insert = no prev).
        _enforce_status_rules(_status_fields(screen), cleaned, identity, previous_row=None)
    result = WorkboardWriteService.insert_row(
        db,
        workboard,
        cleaned,
        None,
        target_table_id=screen.table_id,
        primary_key_columns=list(screen.primary_key_columns or []),
    )
    if op_entry is not None:
        op_entry.result_payload = result
        db.flush()
    return result


def insert_screen_rows(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    rows: List[Dict[str, Any]],
    *,
    identity: CallerIdentity,
) -> Dict[str, Any]:
    """Insert many rows while preserving the single-row write contract.

    Table screens can be prepared once and handed to the write service's
    datasource-aware batch path. Non-table screens deliberately fall back to
    the normal one-by-one path because form validation can depend on each
    field's dynamic conditions.
    """
    if not rows:
        return {"action": "insert_many", "affected_rows": 0, "results": []}
    if screen.kind != "table":
        results = [
            insert_screen_row(db, workboard, screen, row, identity=identity)
            for row in rows
        ]
        return {
            "action": "insert_many",
            "affected_rows": len(results),
            "results": results,
        }
    if screen.table_id is None or screen.table is None:
        raise HTTPException(status_code=400, detail="Table screen has no backing table.")
    if not screen.table.allow_add_row:
        raise HTTPException(status_code=403, detail="Adding rows is disabled on this table.")
    if not (screen.table.editable_columns or []):
        raise HTTPException(
            status_code=403,
            detail=(
                "This table has no editable columns — add at least one "
                "to editable_columns before turning on allow_add_row."
            ),
        )

    derived = {c.name for c in (screen.table.computed_columns or [])} | {
        l.name for l in (screen.table.lookup_columns or [])
    }
    prepared: List[Dict[str, Any]] = []
    for values in rows:
        merged = _resolve_table_defaults(screen, identity)
        merged.update({
            k: v for k, v in (values or {}).items() if k not in derived
        })
        missing = [
            col
            for col in (screen.table.required_columns if screen.table else [])
            if merged.get(col) in (None, "") and col not in derived
        ]
        if missing:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": f"Required columns missing: {', '.join(missing)}",
                    "violations": [{"column": c, "rule": "required"} for c in missing],
                },
            )
        prepared.append(
            enforce_write_access(
                screen.rls,
                screen.rls_default,
                identity,
                op="insert",
                row_values=merged,
            )
        )

    return WorkboardWriteService.insert_rows(
        db,
        workboard,
        prepared,
        None,
        target_table_id=screen.table_id,
        primary_key_columns=list(screen.primary_key_columns or []),
    )


def update_screen_rows(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    updates: List[Dict[str, Any]],
    *,
    identity: CallerIdentity,
    enforce_editable: bool = True,
) -> Dict[str, Any]:
    """Update many table rows with a datasource-aware batch path when safe.

    ``enforce_editable`` gates the *inline-editable* column filter. It is True for
    user-driven edits (a column absent from ``editable_columns`` is dropped, so a
    UI-readonly cell can't be hand-edited). Server-driven callers — e.g. a
    bulk-action ``update_selected`` step linking a freshly-created parent code onto
    the selected rows — pass False: the columns are author-declared in the recipe,
    not user input, so a UI-readonly link column (the common case) must still be
    writable. Computed/lookup (``derived``) columns and per-role RLS
    (``writable_columns``/``can_update``) are STILL enforced either way.
    """
    if not updates:
        return {"action": "update_many", "affected_rows": 0, "results": []}
    if screen.kind != "table":
        results = [
            update_screen_row(
                db,
                workboard,
                screen,
                item.get("pk") if isinstance(item, dict) else {},
                item.get("values") if isinstance(item, dict) else {},
                identity=identity,
                enforce_editable=enforce_editable,
            )
            for item in updates
        ]
        return {"action": "update_many", "affected_rows": len(results), "results": results}
    if screen.table_id is None or screen.table is None:
        raise HTTPException(status_code=400, detail="Table screen has no spec.")

    rls_filters, allowed = build_rls_filter(screen.rls, screen.rls_default, identity)
    if not allowed:
        raise HTTPException(status_code=403, detail="You don't have access to those rows.")
    # A row_lock needs each row's current values evaluated individually, so take
    # the per-row path (which fetches + gates each row) rather than the batch
    # fast-path — same reason RLS filters force it.
    if rls_filters or getattr(screen.table, "row_lock", None) is not None:
        results = [
            update_screen_row(
                db,
                workboard,
                screen,
                item.get("pk") if isinstance(item, dict) else {},
                item.get("values") if isinstance(item, dict) else {},
                identity=identity,
                enforce_editable=enforce_editable,
            )
            for item in updates
        ]
        return {"action": "update_many", "affected_rows": len(results), "results": results}

    inline_editable = set((screen.table.editable_columns or []))
    panel = screen.table.detail_panel
    panel_editable: set[str] = set()
    if panel and panel.enabled:
        panel_editable = set((panel.editable_columns or []))
    editable = inline_editable | panel_editable
    derived = {c.name for c in (screen.table.computed_columns or [])} | {
        l.name for l in (screen.table.lookup_columns or [])
    }
    editable -= derived

    prepared: List[Dict[str, Any]] = []
    for item in updates:
        pk = item.get("pk") if isinstance(item, dict) else None
        values = item.get("values") if isinstance(item, dict) else None
        if not isinstance(pk, dict) or not isinstance(values, dict):
            raise HTTPException(status_code=400, detail="Each update needs pk and values.")
        cleaned_pre = {
            k: v
            for k, v in values.items()
            if k not in derived and (not enforce_editable or not editable or k in editable)
        }
        if not cleaned_pre:
            raise HTTPException(status_code=400, detail="No editable columns in payload.")
        cleaned = enforce_write_access(
            screen.rls,
            screen.rls_default,
            identity,
            op="update",
            row_values=cleaned_pre,
        )
        prepared.append({
            "pk": pk,
            "values": cleaned,
            "lock_token": item.get("lock_token") if isinstance(item, dict) else None,
        })

    return WorkboardWriteService.update_rows(
        db,
        workboard,
        prepared,
        None,
        target_table_id=screen.table_id,
        primary_key_columns=list(screen.primary_key_columns or []),
    )


def update_screen_row(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    pk: Dict[str, Any],
    values: Dict[str, Any],
    *,
    identity: CallerIdentity,
    enforce_editable: bool = True,
) -> Dict[str, Any]:
    # ``enforce_editable`` — see update_screen_rows. False lets a server-driven
    # bulk step write author-declared columns that are UI-readonly (e.g. a link
    # column); derived-column + RLS enforcement below still apply.
    if screen.table_id is None or screen.kind not in ("form", "table"):
        raise HTTPException(status_code=400, detail="Screen is not writable.")
    if screen.kind == "form" and screen.form is None:
        raise HTTPException(status_code=400, detail="Form screen has no spec.")
    if screen.kind == "table" and screen.table is None:
        raise HTTPException(status_code=400, detail="Table screen has no spec.")

    if screen.kind == "form":
        cleaned_pre = _apply_field_conditions(
            screen, values, identity, media_max_kb=media_cap_kb(db, workboard, screen)
        )
    else:
        # Table: a column must be flagged in either the inline grid
        # ``editable_columns`` OR the detail panel's ``editable_columns``.
        # Computed/lookup columns are NEVER writeable regardless of how
        # they appear in those lists. RLS further constrains by role
        # ``writable_columns`` / ``readonly_columns``.
        inline_editable = set((screen.table.editable_columns or []))
        panel = screen.table.detail_panel
        panel_editable: set[str] = set()
        if panel and panel.enabled:
            panel_editable = set((panel.editable_columns or []))
        editable = (inline_editable | panel_editable)
        derived = {c.name for c in (screen.table.computed_columns or [])} | {
            l.name for l in (screen.table.lookup_columns or [])
        }
        editable -= derived
        cleaned_pre = {
            k: v
            for k, v in (values or {}).items()
            if k not in derived and (not enforce_editable or not editable or k in editable)
        }
        if not cleaned_pre:
            raise HTTPException(
                status_code=400,
                detail="No editable columns in payload.",
            )
    cleaned = enforce_write_access(
        screen.rls, screen.rls_default, identity, op="update", row_values=cleaned_pre
    )
    # Server-side geocode enrichment. Only fetch the previous row when the
    # screen opts into geocoding, and pass it so existing coordinates are not
    # re-resolved (overwrite_existing=False).
    if _geocode_config_for_screen(screen) is not None:
        _geo_prev = _fetch_current_row(db, workboard, screen, pk)
        cleaned = _apply_geocode(screen, cleaned, previous_row=_geo_prev)
    if screen.kind == "form":
        # Value-level status guard: compare the previous row's status to the new
        # value so allowed_transitions + editable_by_roles are enforced server-side.
        _sf = _status_fields(screen)
        if _sf:
            _prev = _fetch_current_row(db, workboard, screen, pk)
            _enforce_status_rules(_sf, cleaned, identity, previous_row=_prev)
    if (
        screen.kind == "table"
        and screen.table is not None
        and getattr(screen.table, "row_lock", None) is not None
    ):
        # Per-row edit lock: block updates to a LOCKED row unless the caller's
        # role is allow-listed (owner/staff bypass). Evaluated on the row's
        # current stored values so the lock column can't be self-unlocked.
        _lock_prev = _fetch_current_row(db, workboard, screen, pk)
        _enforce_row_lock(screen.table.row_lock, _lock_prev, identity, op="update")

    # Make sure the targeted row passes RLS before touching it.
    rls_filters, allowed = build_rls_filter(
        screen.rls, screen.rls_default, identity
    )
    if not allowed:
        raise HTTPException(status_code=403, detail="You don't have access to that row.")
    if rls_filters:
        # Existence check on the screen's bound table — bypass the
        # render_table_screen kind check because update screens are forms,
        # but they share the same RLS rule set.
        table = _load_table(db, screen.table_id)
        datasource = _load_datasource(db, table) if table else None
        if not table or not datasource:
            raise HTTPException(status_code=400, detail="Screen table missing.")
        check_filters = [
            {"field": k, "operator": "eq", "value": v} for k, v in (pk or {}).items()
        ] + rls_filters
        result = LiveQueryService.execute_preview_query(
            datasource, table, limit=1, offset=0, filters=check_filters
        )
        if not (result.get("rows") or []):
            raise HTTPException(
                status_code=403, detail="You don't have access to that row."
            )

    return WorkboardWriteService.update_row(
        db,
        workboard,
        pk,
        cleaned,
        None,
        target_table_id=screen.table_id,
        primary_key_columns=list(screen.primary_key_columns or []),
    )


# ── Table screen helpers ────────────────────────────────────────────────

def _resolve_table_defaults(
    screen: Screen, identity: CallerIdentity
) -> Dict[str, Any]:
    """Return table ``default_values`` with placeholders substituted.

    Reuses :func:`_resolve_initial_values` so a table behaves the same as
    a form for ``{{app_user.x}}`` / ``{{today}}`` / ``{{now}}`` defaults.
    """
    if screen.table is None:
        return {}
    return _resolve_initial_values(
        dict(screen.table.default_values or {}), identity=identity
    )


def _query_relation_rows(
    datasource: DataSource,
    table: DatasetTable,
    filters: List[Dict[str, Any]],
    *,
    max_rows: int = 5000,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    offset = 0
    page_size = 500
    while True:
        limit = min(page_size, max_rows + 1 - len(rows))
        result = LiveQueryService.execute_preview_query(
            datasource,
            table,
            limit=limit,
            offset=offset,
            filters=filters,
        )
        page = result.get("rows") or []
        rows.extend(page)
        if len(rows) > max_rows:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Relation delete affects more than 5,000 child rows. "
                    "Narrow or archive the children before deleting the parent."
                ),
            )
        if len(page) < limit:
            return rows
        offset += len(page)


def _apply_parent_relation_delete_behaviors(
    db: Session,
    workboard: Workboard,
    layout: LayoutJson,
    *,
    parent_table_id: int,
    parent_row: Dict[str, Any],
    identity: CallerIdentity,
    delete_stack: Set[tuple[int, str]],
) -> None:
    plans: List[Dict[str, Any]] = []
    seen_relations: Set[tuple[int, str, str, str]] = set()
    for parent_screen in layout.screens:
        if (
            parent_screen.kind != "form"
            or parent_screen.form is None
            or parent_screen.table_id != parent_table_id
        ):
            continue
        parent_pk = set(parent_screen.primary_key_columns or [])
        for relation in parent_screen.form.related_records or []:
            if relation.parent_key_column not in parent_pk:
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Relation '{relation.id}' has an unsafe parent key. "
                        "Select one of the parent screen's primary_key_columns."
                    ),
                )
            parent_key_value = parent_row.get(relation.parent_key_column)
            if parent_key_value is None or (
                isinstance(parent_key_value, str) and not parent_key_value.strip()
            ):
                continue
            child_screen = get_screen(layout, relation.child_screen_id)
            if not child_screen.table_id:
                raise HTTPException(
                    status_code=409,
                    detail=f"Relation '{relation.id}' has no child table binding.",
                )
            relation_signature = (
                int(child_screen.table_id),
                relation.child_foreign_key_column,
                json.dumps(parent_key_value, sort_keys=True, default=str),
                relation.delete_behavior,
            )
            if relation_signature in seen_relations:
                continue
            seen_relations.add(relation_signature)
            child_table = _load_table(db, child_screen.table_id)
            child_datasource = _load_datasource(db, child_table) if child_table else None
            if not child_table or not child_datasource:
                raise HTTPException(
                    status_code=409,
                    detail=f"Relation '{relation.id}' child datasource is missing.",
                )
            relation_filter = [
                {
                    "field": relation.child_foreign_key_column,
                    "operator": "eq",
                    "value": parent_key_value,
                }
            ]
            child_rows = _query_relation_rows(
                child_datasource, child_table, relation_filter
            )
            if not child_rows:
                continue
            if relation.delete_behavior == "restrict":
                raise HTTPException(
                    status_code=409,
                    detail=(
                        f"Cannot delete this parent while relation '{relation.id}' "
                        f"still has {len(child_rows)} child record(s)."
                    ),
                )

            operation = "delete" if relation.delete_behavior == "cascade" else "update"
            unlink_values = (
                {relation.child_foreign_key_column: None}
                if operation == "update"
                else {}
            )
            clean_unlink = enforce_write_access(
                child_screen.rls,
                child_screen.rls_default,
                identity,
                op=operation,
                row_values=unlink_values,
            )
            rls_filters, allowed = build_rls_filter(
                child_screen.rls, child_screen.rls_default, identity
            )
            if not allowed:
                raise HTTPException(
                    status_code=403,
                    detail=f"You don't have access to modify relation '{relation.id}'.",
                )
            child_pk = list(child_screen.primary_key_columns or [])
            if not child_pk or any(
                any(column not in row for column in child_pk) for row in child_rows
            ):
                raise HTTPException(
                    status_code=409,
                    detail=f"Relation '{relation.id}' child screen needs a complete primary key.",
                )
            if rls_filters:
                accessible_rows = _query_relation_rows(
                    child_datasource,
                    child_table,
                    relation_filter + rls_filters,
                )
                accessible_keys = {
                    tuple(json.dumps(row.get(col), default=str) for col in child_pk)
                    for row in accessible_rows
                }
                child_keys = {
                    tuple(json.dumps(row.get(col), default=str) for col in child_pk)
                    for row in child_rows
                }
                if accessible_keys != child_keys:
                    raise HTTPException(
                        status_code=403,
                        detail=(
                            f"Some child records in relation '{relation.id}' are "
                            "outside your row-level access."
                        ),
                    )
            plans.append(
                {
                    "behavior": relation.delete_behavior,
                    "child_screen": child_screen,
                    "child_rows": child_rows,
                    "child_pk": child_pk,
                    "unlink_values": clean_unlink,
                }
            )

    for plan in plans:
        child_screen = plan["child_screen"]
        for child_row in plan["child_rows"]:
            child_pk_values = {
                column: child_row.get(column) for column in plan["child_pk"]
            }
            if plan["behavior"] == "unlink":
                WorkboardWriteService.update_row(
                    db,
                    workboard,
                    child_pk_values,
                    plan["unlink_values"],
                    None,
                    target_table_id=child_screen.table_id,
                    primary_key_columns=plan["child_pk"],
                )
                continue

            stack_key = (
                int(child_screen.table_id),
                json.dumps(child_pk_values, sort_keys=True, default=str),
            )
            if stack_key in delete_stack:
                raise HTTPException(
                    status_code=409,
                    detail="Cyclic cascade relation detected; use restrict or unlink.",
                )
            delete_stack.add(stack_key)
            try:
                _apply_parent_relation_delete_behaviors(
                    db,
                    workboard,
                    layout,
                    parent_table_id=int(child_screen.table_id),
                    parent_row=child_row,
                    identity=identity,
                    delete_stack=delete_stack,
                )
                WorkboardWriteService.delete_row(
                    db,
                    workboard,
                    child_pk_values,
                    None,
                    target_table_id=child_screen.table_id,
                    primary_key_columns=plan["child_pk"],
                )
            finally:
                delete_stack.discard(stack_key)


def delete_screen_row(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    pk: Dict[str, Any],
    *,
    identity: CallerIdentity,
) -> Dict[str, Any]:
    """Delete a single row through a table screen, honouring RLS.

    Mirrors :func:`update_screen_row`: enforce ``can_delete`` per role, then
    confirm the row passes the read-filter rules before issuing the DELETE
    so a viewer can't delete rows outside their RLS scope.
    """
    if screen.table_id is None or screen.kind != "table":
        raise HTTPException(
            status_code=400, detail="Only table screens support row deletion."
        )
    if screen.table is None or not screen.table.allow_delete_row:
        raise HTTPException(
            status_code=403, detail="Deleting rows is disabled on this table."
        )

    # Enforce ``can_delete`` via the shared RLS pipeline. ``row_values={}``
    # because delete has no payload; ``enforce_write_access`` will raise if
    # the matching rule blocks the op.
    enforce_write_access(
        screen.rls, screen.rls_default, identity, op="delete", row_values={}
    )

    rls_filters, allowed = build_rls_filter(
        screen.rls, screen.rls_default, identity
    )
    if not allowed:
        raise HTTPException(status_code=403, detail="You don't have access to that row.")
    primary_key_columns = list(screen.primary_key_columns or [])
    if not primary_key_columns or any(column not in (pk or {}) for column in primary_key_columns):
        raise HTTPException(status_code=400, detail="A complete row primary key is required.")
    table = _load_table(db, screen.table_id)
    datasource = _load_datasource(db, table) if table else None
    if not table or not datasource:
        raise HTTPException(status_code=400, detail="Screen table missing.")
    check_filters = [
        {"field": column, "operator": "eq", "value": pk[column]}
        for column in primary_key_columns
    ] + rls_filters
    result = LiveQueryService.execute_preview_query(
        datasource, table, limit=1, offset=0, filters=check_filters
    )
    parent_rows = result.get("rows") or []
    if not parent_rows:
        raise HTTPException(
            status_code=403, detail="You don't have access to that row."
        )
    # Per-row lock: block deleting a locked row (unless lock_delete is off or the
    # caller's role is allow-listed / owner). Reuses the row just fetched above.
    if getattr(screen.table, "row_lock", None) is not None:
        _enforce_row_lock(screen.table.row_lock, parent_rows[0], identity, op="delete")
    layout = parse_layout(workboard)
    root_key = (
        int(screen.table_id),
        json.dumps(pk, sort_keys=True, default=str),
    )
    _apply_parent_relation_delete_behaviors(
        db,
        workboard,
        layout,
        parent_table_id=int(screen.table_id),
        parent_row=parent_rows[0],
        identity=identity,
        delete_stack={root_key},
    )

    return WorkboardWriteService.delete_row(
        db,
        workboard,
        pk or {},
        None,
        target_table_id=screen.table_id,
        primary_key_columns=primary_key_columns,
    )


def fetch_table_row_for_panel(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    pk: Dict[str, Any],
    *,
    identity: CallerIdentity,
) -> Dict[str, Any]:
    """Fetch one row by PK, with lookup columns resolved, for the detail
    panel. RLS is enforced — a row outside the viewer's scope returns 403.

    Returns ``{row, columns, editable_columns, sections, computed_columns,
    lookup_columns, column_labels, primary_key_columns, title}`` so the FE
    can render the panel without re-reading the layout. Computed columns
    are evaluated per-row only (cross-row aggregates resolve against the
    single-row scope so COL_SUM degenerates to a sum over [row]).
    """
    if screen.kind != "table" or screen.table is None or screen.table_id is None:
        raise HTTPException(status_code=400, detail="Screen is not a table.")
    panel = screen.table.detail_panel
    if not panel or not panel.enabled:
        raise HTTPException(status_code=400, detail="Detail panel is disabled.")

    table = _load_table(db, screen.table_id)
    datasource = _load_datasource(db, table) if table else None
    if not table or not datasource:
        raise HTTPException(status_code=400, detail="Screen table missing.")

    rls_filters, allowed = build_rls_filter(
        screen.rls, screen.rls_default, identity
    )
    if not allowed:
        raise HTTPException(status_code=403, detail="You don't have access to that row.")

    check_filters = [
        {"field": k, "operator": "eq", "value": v} for k, v in (pk or {}).items()
    ] + rls_filters
    result = LiveQueryService.execute_preview_query(
        datasource, table, limit=1, offset=0, filters=check_filters
    )
    rows = result.get("rows") or []
    if not rows:
        raise HTTPException(status_code=403, detail="You don't have access to that row.")
    raw_row = rows[0]

    table_spec = screen.table
    all_db_columns: List[str] = result.get("columns") or list(raw_row.keys())
    panel_columns = list(panel.columns or table_spec.columns or all_db_columns)
    # Include lookup local match columns + PK so they're always available
    # in ``row`` even when the panel hides them.
    extra_keys = list(screen.primary_key_columns or [])
    for lookup in table_spec.lookup_columns or []:
        if lookup.match_column_local in all_db_columns:
            extra_keys.append(lookup.match_column_local)
    for rollup in getattr(table_spec, "rollup_columns", None) or []:
        if rollup.match_column_local in all_db_columns:
            extra_keys.append(rollup.match_column_local)

    row_keys = list(dict.fromkeys([
        *(c for c in panel_columns if c in all_db_columns),
        *extra_keys,
    ]))
    row: Dict[str, Any] = {col: raw_row.get(col) for col in row_keys}

    # Lookup + roll-up resolution for this single row.
    lookup_maps = _resolve_table_lookups(db, table_spec, [row])
    rollup_maps = _resolve_table_rollups(db, table_spec, [row])
    for lookup in table_spec.lookup_columns or []:
        match_value = row.get(lookup.match_column_local)
        row[lookup.name] = lookup_maps.get(lookup.name, {}).get(match_value)
    for rollup in getattr(table_spec, "rollup_columns", None) or []:
        match_value = row.get(rollup.match_column_local)
        row[rollup.name] = rollup_maps.get(rollup.name, {}).get(match_value)

    # Per-row computed columns (cross-row aggregates degrade to single-row).
    if table_spec.computed_columns:
        # Detail panel always shows a single row, so cross-row aggregates
        # in JS degrade to operating over [row] — that's fine for the
        # panel view (it's meant to show one record's details).
        for col in table_spec.computed_columns:
            if not (col.formula or "").strip():
                if col.name not in row:
                    row[col.name] = None
                continue
            try:
                compiled_js = compile_js_column(col.name, col.formula)
            except JsCompileError as exc:
                row[col.name] = f"#ERR: {exc}"
                continue
            try:
                row[col.name] = evaluate_js_cell(compiled_js, row, [row], 0)
            except JsEvalError as exc:
                row[col.name] = f"#ERR: {exc}"

    return {
        "row": row,
        "primary_key_columns": list(screen.primary_key_columns or []),
        "columns": panel_columns,
        "editable_columns": list(panel.editable_columns or []),
        "sections": dict(panel.sections or {}),
        "title": panel.title or screen.title,
        "column_labels": dict(screen.column_labels or {}),
        "column_metadata": {
            k: v.model_dump() if hasattr(v, "model_dump") else v
            for k, v in (table_spec.column_metadata or {}).items()
        },
        "computed_columns": [c.model_dump() for c in (table_spec.computed_columns or [])],
        "lookup_columns": [l.model_dump() for l in (table_spec.lookup_columns or [])],
    }


def _resolve_table_lookups(
    db: Session,
    table_spec: Any,
    rows: List[Dict[str, Any]],
) -> Dict[str, Dict[Any, Any]]:
    """Batch-resolve every lookup column on the table.

    Returns a mapping ``{lookup_name: {match_value: return_value}}`` so the
    caller can fan the values out across rows without re-querying. One
    ``SELECT match_col, return_col FROM linked WHERE match_col IN (...)``
    per lookup column, regardless of how many rows are on the page.
    """
    out: Dict[str, Dict[Any, Any]] = {}
    lookups = list(getattr(table_spec, "lookup_columns", None) or [])
    if not lookups or not rows:
        return out

    for lookup in lookups:
        # Skip incomplete lookups (builder draft state). The schema allows
        # empty match/remote/return columns and from_table_id=0 so autosave
        # doesn't reject an in-progress edit; the runtime simply leaves the
        # cell empty until the builder finishes wiring it up.
        if (
            not lookup.from_table_id
            or not lookup.match_column_local
            or not lookup.match_column_remote
            or not lookup.return_column
        ):
            out[lookup.name] = {}
            continue
        local_col = lookup.match_column_local
        match_values = [
            row.get(local_col)
            for row in rows
            if row.get(local_col) not in (None, "")
        ]
        if not match_values:
            out[lookup.name] = {}
            continue
        # De-dup; the linked-table query only needs distinct match values.
        distinct_values = list({v for v in match_values})

        linked_table = _load_table(db, lookup.from_table_id)
        if not linked_table:
            out[lookup.name] = {}
            continue
        linked_ds = _load_datasource(db, linked_table)
        if not linked_ds:
            out[lookup.name] = {}
            continue

        try:
            result = LiveQueryService.execute_preview_query(
                linked_ds,
                linked_table,
                limit=max(len(distinct_values) * 2, 100),
                offset=0,
                filters=[
                    {
                        "field": lookup.match_column_remote,
                        "operator": "in",
                        "value": distinct_values,
                    }
                ],
            )
        except SheetsQuotaError:
            raise _quota_503()
        except Exception:  # pragma: no cover - defensive, lookup must never crash render
            logger.exception(
                "Grid lookup '%s' failed for screen lookup_table=%s",
                lookup.name,
                lookup.from_table_id,
            )
            out[lookup.name] = {}
            continue

        mapping: Dict[Any, Any] = {}
        for row in result.get("rows") or []:
            key = row.get(lookup.match_column_remote)
            if key is None:
                continue
            mapping[key] = row.get(lookup.return_column)
        out[lookup.name] = mapping
    return out


def _resolve_table_rollups(
    db: Session,
    table_spec: Any,
    rows: List[Dict[str, Any]],
) -> Dict[str, Dict[Any, Any]]:
    """Reverse-reference roll-ups: aggregate child rows per parent key.

    The inverse of :func:`_resolve_table_lookups`. For each roll-up column,
    fetch ALL child rows whose ``match_column_remote`` is IN the page's parent
    keys (one batched ``IN`` query, no GROUP-BY transport needed), bucket by
    key and reduce with ``agg`` in Python. Returns
    ``{rollup_name: {parent_key: aggregated_value}}``.

    Child fetch is capped at 1000 rows (LiveQueryService clamp); a roll-up
    over more child rows than that for the visible parents is approximate.
    """
    out: Dict[str, Dict[Any, Any]] = {}
    rollups = list(getattr(table_spec, "rollup_columns", None) or [])
    if not rollups or not rows:
        return out

    for rollup in rollups:
        agg = (getattr(rollup, "agg", "count") or "count").lower()
        needs_value = agg != "count"
        if (
            not rollup.from_table_id
            or not rollup.match_column_local
            or not rollup.match_column_remote
            or (needs_value and not rollup.value_column)
        ):
            out[rollup.name] = {}
            continue
        local_col = rollup.match_column_local
        distinct_values = list({
            row.get(local_col) for row in rows if row.get(local_col) not in (None, "")
        })
        if not distinct_values:
            out[rollup.name] = {}
            continue
        linked_table = _load_table(db, rollup.from_table_id)
        linked_ds = _load_datasource(db, linked_table) if linked_table else None
        if not linked_table or not linked_ds:
            out[rollup.name] = {}
            continue
        try:
            result = LiveQueryService.execute_preview_query(
                linked_ds,
                linked_table,
                limit=1000,
                offset=0,
                filters=[
                    {
                        "field": rollup.match_column_remote,
                        "operator": "in",
                        "value": distinct_values,
                    }
                ],
            )
        except SheetsQuotaError:
            raise _quota_503()
        except Exception:  # pragma: no cover - defensive
            logger.exception(
                "Grid roll-up '%s' failed for from_table=%s",
                rollup.name,
                rollup.from_table_id,
            )
            out[rollup.name] = {}
            continue

        # Bucket child rows by parent key.
        buckets: Dict[Any, List[Dict[str, Any]]] = {}
        for child in result.get("rows") or []:
            key = child.get(rollup.match_column_remote)
            if key is None:
                continue
            buckets.setdefault(key, []).append(child)

        mapping: Dict[Any, Any] = {}
        for key, children in buckets.items():
            if agg == "count":
                mapping[key] = len(children)
                continue
            nums = [
                n for n in (_coerce_total(c.get(rollup.value_column)) for c in children)
                if n is not None
            ]
            if not nums:
                mapping[key] = None
            elif agg == "sum":
                mapping[key] = sum(nums)
            elif agg in ("avg", "average"):
                mapping[key] = sum(nums) / len(nums)
            elif agg == "min":
                mapping[key] = min(nums)
            elif agg == "max":
                mapping[key] = max(nums)
            else:
                mapping[key] = None
        out[rollup.name] = mapping
    return out


def _coerce_total(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # vi-VN-aware: a Sheets cell read as "1234,5" / "1.000.000" must still
        # land in the SUM. Native int/float (BQ/Postgres) handled above.
        return _parse_locale_number(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _compute_table_totals(
    table_spec: Any,
    rows: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Apply :class:`TableScreenSpec.totals` to the rows on the page.

    Returns ``{column_name: aggregated_value}`` so the runtime can render
    a footer row keyed by column. ``count`` runs over non-empty cells;
    everything else coerces to ``float``.
    """
    totals_spec: Dict[str, str] = dict(getattr(table_spec, "totals", None) or {})
    if not totals_spec or not rows:
        return {}
    out: Dict[str, Any] = {}
    for column, kind in totals_spec.items():
        kind = (kind or "").lower()
        if kind == "count":
            out[column] = sum(
                1 for row in rows if row.get(column) not in (None, "")
            )
            continue
        nums: List[float] = []
        for row in rows:
            n = _coerce_total(row.get(column))
            if n is not None:
                nums.append(n)
        if not nums:
            out[column] = None
            continue
        if kind == "sum":
            out[column] = sum(nums)
        elif kind in ("avg", "average"):
            out[column] = sum(nums) / len(nums)
        elif kind == "min":
            out[column] = min(nums)
        elif kind == "max":
            out[column] = max(nums)
        else:
            out[column] = None
    return out


def _compute_stat_tiles(
    table_spec: Any,
    rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Aggregate :class:`TableScreenSpec.stat_tiles` across ``rows``.

    Returns ``[{label, value, format, unit}]`` — the FE renders one KPI card
    per entry above the table/gallery. Reuses the same coercion as footer
    totals, over the same (RLS-filtered, capped) row set.
    """
    tiles = list(getattr(table_spec, "stat_tiles", None) or [])
    if not tiles:
        return []
    out: List[Dict[str, Any]] = []
    for tile in tiles:
        column = getattr(tile, "column", None)
        agg = (getattr(tile, "agg", "sum") or "sum").lower()
        value: Any
        if not column:
            value = None
        elif agg == "count":
            value = sum(1 for row in rows if row.get(column) not in (None, ""))
        else:
            nums = [
                n for n in (_coerce_total(row.get(column)) for row in rows)
                if n is not None
            ]
            if not nums:
                value = None
            elif agg == "sum":
                value = sum(nums)
            elif agg in ("avg", "average"):
                value = sum(nums) / len(nums)
            elif agg == "min":
                value = min(nums)
            elif agg == "max":
                value = max(nums)
            else:
                value = None
        out.append({
            "label": getattr(tile, "label", "") or "",
            "value": value,
            "format": getattr(tile, "format", None),
            "unit": getattr(tile, "unit", None),
        })
    return out


def render_table_screen(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    *,
    identity: CallerIdentity,
    page: int = 1,
    page_size: Optional[int] = None,
    extra_filters: Optional[List[Dict[str, Any]]] = None,
    shared_context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Render a table screen — paginated rows + computed/lookup cells +
    totals + multi-header + row-merge.

    Evaluation order per row:

    1. Pull regular columns from the source table (RLS-filtered).
    2. Resolve every lookup column with one batched query per linked table.
    3. Evaluate computed columns in topological order so a formula can
       reference another computed column.
    4. Aggregate the totals map over the resulting (filtered) page.
    5. Compute multi-header (``column_groups``) and row-merge spans
       (``group_by``) for FE rendering.

    The response payload is consumed by both the inline grid (cells +
    totals + merges) and the side detail panel (a single row + the panel
    config; the panel itself is materialised client-side).
    """
    if screen.kind != "table" or screen.table is None or screen.table_id is None:
        raise HTTPException(status_code=400, detail="Screen is not a table.")

    table_spec = screen.table

    table = _load_table(db, screen.table_id)
    datasource, query_table = _resolve_read_target(db, table) if table else (None, None)
    if not table or not datasource or query_table is None:
        return {"columns": [], "rows": [], "page": page, "page_size": page_size}

    rls_filters, allowed = build_rls_filter(
        screen.rls, screen.rls_default, identity
    )
    if not allowed:
        return {"columns": [], "rows": [], "page": page, "page_size": page_size}

    page = max(int(page or 1), 1)
    configured_page_size = getattr(table_spec, "page_size", None) or 50
    page_size = min(max(int(page_size or configured_page_size or 50), 1), 500)
    offset = (page - 1) * page_size
    context_filters: List[Dict[str, Any]] = []
    shared = shared_context or {}
    for cf in getattr(table_spec, "context_filters", None) or []:
        raw = shared.get(cf.from_shared)
        present = raw is not None and not (isinstance(raw, str) and not raw.strip())
        if not present:
            if cf.required:
                return {
                    "columns": list(table_spec.columns or []),
                    "rows": [],
                    "page": page,
                    "page_size": page_size,
                    "total_rows": 0,
                    "context_applied": True,
                }
            continue
        context_filters.append({"field": cf.column, "operator": "eq", "value": raw})

    merged = context_filters + _filter_dicts(extra_filters) + rls_filters
    sort_column = getattr(table_spec, "default_sort_column", None)
    sort_direction = getattr(table_spec, "default_sort_direction", None) or "desc"

    computed_names = {c.name for c in (table_spec.computed_columns or [])}
    lookup_names = {l.name for l in (table_spec.lookup_columns or [])}
    rollup_names = {r.name for r in (getattr(table_spec, "rollup_columns", None) or [])}
    if sort_column and sort_column in (computed_names | lookup_names | rollup_names):
        # Sorting by a derived column requires post-eval sort, which would
        # silently break pagination. Drop to default ordering.
        sort_column = None

    # Fetch the WHOLE filtered+sorted set (capped) once, then slice the page in
    # memory. This makes footer totals span the entire table (not just the
    # visible page), keeps page ordering stable across pages, and lets cross-row
    # computed columns aggregate over the full set. Cheap for Sheets (already
    # cached + materialised into DuckDB). _TOTALS_ROW_CAP bounds worst case.
    result = LiveQueryService.execute_preview_query(
        datasource,
        query_table,
        limit=_TOTALS_ROW_CAP + 1,
        offset=0,
        filters=merged,
        sort_column=sort_column,
        sort_direction=sort_direction,
    )
    all_db_columns: List[str] = result.get("columns") or []
    pk_cols = list(screen.primary_key_columns or [])

    declared_columns = list(table_spec.columns or all_db_columns)
    selected_columns = [
        c
        for c in declared_columns
        if c in all_db_columns or c in computed_names or c in lookup_names or c in rollup_names
    ] or all_db_columns

    # PK columns stay in the row payload so the runtime can issue PATCH /
    # DELETE / detail-panel-fetch keyed by PK even when hidden from the grid.
    row_keys = list({*pk_cols, *(c for c in selected_columns if c in all_db_columns)})
    for lookup in table_spec.lookup_columns or []:
        if lookup.match_column_local in all_db_columns:
            row_keys.append(lookup.match_column_local)
    for rollup in getattr(table_spec, "rollup_columns", None) or []:
        if rollup.match_column_local in all_db_columns:
            row_keys.append(rollup.match_column_local)
    # Include any extra column referenced only by the detail panel so the
    # current-page rows already carry the data needed by the side panel.
    panel = table_spec.detail_panel
    if panel and panel.enabled:
        for col in (panel.columns or selected_columns):
            if col in all_db_columns and col not in row_keys:
                row_keys.append(col)
    # KPI stat tiles aggregate a column across the loaded rows — make sure that
    # column is fetched even when it isn't a visible grid column.
    for tile in getattr(table_spec, "stat_tiles", None) or []:
        tcol = getattr(tile, "column", None)
        if tcol in all_db_columns and tcol not in row_keys:
            row_keys.append(tcol)
    row_keys = list(dict.fromkeys(row_keys))

    base_rows: List[Dict[str, Any]] = [
        {column: row.get(column) for column in row_keys}
        for row in (result.get("rows") or [])
    ]

    # ── Lookup + roll-up resolution ──────────────────────────────────
    lookup_maps = _resolve_table_lookups(db, table_spec, base_rows)
    rollup_maps = _resolve_table_rollups(db, table_spec, base_rows)
    for row in base_rows:
        for lookup in table_spec.lookup_columns or []:
            match_value = row.get(lookup.match_column_local)
            row[lookup.name] = lookup_maps.get(lookup.name, {}).get(match_value)
        for rollup in getattr(table_spec, "rollup_columns", None) or []:
            match_value = row.get(rollup.match_column_local)
            row[rollup.name] = rollup_maps.get(rollup.name, {}).get(match_value)

    # ── Computed columns (JS sandbox, one pass) ──────────────────────
    # Each column compiles ONCE, then evaluates once per row. Compile
    # errors AND per-row eval errors surface as ``#ERR: ...`` in the
    # cell so the rest of the grid stays readable.
    if table_spec.computed_columns:
        js_compiled: Dict[str, CompiledJs] = {}
        js_compile_err: Dict[str, str] = {}
        js_draft: set[str] = set()
        for col in table_spec.computed_columns:
            if not (col.formula or "").strip():
                js_draft.add(col.name)
                continue
            try:
                js_compiled[col.name] = compile_js_column(col.name, col.formula)
            except JsCompileError as exc:
                js_compile_err[col.name] = str(exc)

        for index, row in enumerate(base_rows):
            for col in table_spec.computed_columns:
                name = col.name
                if name in js_draft:
                    if name not in row:
                        row[name] = None
                    continue
                if name in js_compile_err:
                    row[name] = f"#ERR: {js_compile_err[name]}"
                    continue
                compiled_js = js_compiled.get(name)
                if compiled_js is None:
                    row[name] = "#ERR: js not compiled"
                    continue
                try:
                    row[name] = evaluate_js_cell(compiled_js, row, base_rows, index)
                except JsEvalError as exc:
                    row[name] = f"#ERR: {exc}"

    # ── Footer totals over the WHOLE filtered set (not just the page) ──
    totals_partial = len(base_rows) > _TOTALS_ROW_CAP
    if totals_partial:
        base_rows = base_rows[:_TOTALS_ROW_CAP]
    total_count = len(base_rows)
    totals_row = _compute_table_totals(table_spec, base_rows) or None
    stat_tiles = _compute_stat_tiles(table_spec, base_rows)

    # Distinct option lists for "single select" filters, derived from the whole
    # filtered set already in memory (no extra query). Lets the runtime render a
    # real dropdown instead of a free-text box for select-kind slicers.
    filter_options: Dict[str, List[str]] = {}
    for _f in (table_spec.filters or []):
        if getattr(_f, "kind", None) != "select":
            continue
        col = getattr(_f, "column", None)
        if not col or col in filter_options:
            continue
        seen: set[str] = set()
        values: List[str] = []
        for r in base_rows:
            raw = r.get(col)
            if raw is None:
                continue
            s = str(raw).strip()
            if not s or s in seen:
                continue
            seen.add(s)
            values.append(s)
            if len(values) >= 500:
                break
        filter_options[col] = sorted(values)

    # ── Slice the requested page out of the full set ─────────────────
    page_rows = base_rows[offset:offset + page_size]

    # ── Multi-header + row merges (over the visible page) ────────────
    column_groups = _normalize_column_groups(selected_columns, table_spec.column_groups)
    merges = _compute_merges(page_rows, list(table_spec.group_by or []), selected_columns)

    payload = {
        "columns": selected_columns,
        "primary_key_columns": pk_cols,
        "rows": page_rows,
        "page": page,
        "page_size": page_size,
        "total_count": total_count,
        "totals_partial": totals_partial,
        "table_view": table_spec.model_dump(),
        "totals_row": totals_row,
        "stat_tiles": stat_tiles,
        "filter_options": filter_options,
        "column_groups": column_groups,
        "merges": merges,
        "column_labels": screen.column_labels or {},
    }
    # POS scan-cart screens carry the resolved product catalog so the FE
    # scanner resolves codes → {name, unit, price} instantly, client-side.
    if getattr(table_spec, "pos_cart", None) is not None:
        payload["pos_catalog"] = _resolve_pos_catalog(db, table_spec.pos_cart)
    return payload


# ── Doc screen ────────────────────────────────────────────────────────────

def _apply_data_table_transform(
    columns: List[str],
    rows: List[Dict[str, Any]],
    transform: Optional[Any],
) -> tuple[List[str], List[Dict[str, Any]], Optional[Dict[str, Any]]]:
    """Apply pivot/unpivot in memory.

    Returns ``(columns, rows, extra)`` where ``extra`` may contain
    ``column_groups`` and ``column_labels`` when a multi-level pivot is
    performed. The input is not mutated.
    Raises ``HTTPException(422)`` when the transform refers to missing
    columns or when a pivot would explode beyond ``max_columns``.
    """
    if transform is None:
        return columns, rows, None

    if isinstance(transform, DataTableUnpivot):
        id_cols = [c for c in transform.id_columns if c in columns]
        value_cols = [c for c in transform.value_columns if c in columns]
        if not value_cols:
            raise HTTPException(
                status_code=422,
                detail="Unpivot: none of value_columns exist in the source.",
            )
        var_name = transform.var_name
        value_name = transform.value_name
        if var_name in id_cols or value_name in id_cols or var_name == value_name:
            raise HTTPException(
                status_code=422,
                detail="Unpivot var_name/value_name must not collide with id_columns or each other.",
            )
        new_cols = id_cols + [var_name, value_name]
        new_rows: List[Dict[str, Any]] = []
        for row in rows:
            id_part = {c: row.get(c) for c in id_cols}
            for vc in value_cols:
                cell = row.get(vc)
                if transform.drop_nulls and cell is None:
                    continue
                new_row = dict(id_part)
                new_row[var_name] = vc
                new_row[value_name] = cell
                new_rows.append(new_row)
        return new_cols, new_rows, None

    if isinstance(transform, DataTablePivot):
        idx_cols = [c for c in transform.index if c in columns]
        if len(idx_cols) != len(transform.index):
            missing = [c for c in transform.index if c not in columns]
            raise HTTPException(
                status_code=422,
                detail=f"Pivot index columns missing from source: {missing}",
            )
        col_dims: List[str] = (
            [transform.columns]
            if isinstance(transform.columns, str)
            else list(transform.columns)
        )
        for _dim in col_dims:
            if _dim not in columns:
                raise HTTPException(
                    status_code=422,
                    detail=f"Pivot columns key '{_dim}' not in source.",
                )
        if transform.values not in columns:
            raise HTTPException(
                status_code=422,
                detail=f"Pivot values column '{transform.values}' not in source.",
            )

        agg = transform.agg

        def _aggregate(cells: List[Any]) -> Any:
            if not cells:
                return transform.fill_value
            if agg == "count":
                return sum(1 for v in cells if v is not None)
            if agg == "first":
                return cells[0]
            numeric = [_coerce_number(v) for v in cells]
            numeric = [v for v in numeric if v is not None]
            if not numeric:
                return transform.fill_value
            if agg == "sum":
                return sum(numeric)
            if agg == "avg":
                return sum(numeric) / len(numeric)
            if agg == "min":
                return min(numeric)
            if agg == "max":
                return max(numeric)
            return transform.fill_value

        if len(col_dims) == 1:
            # ── Single-dimension pivot (original behaviour) ─────────────
            pivot_keys_order: List[str] = []
            seen_keys: set[str] = set()
            # ``buckets[index_tuple][pivot_key]`` accumulates raw cell values.
            buckets: Dict[tuple, Dict[str, List[Any]]] = {}
            index_order: List[tuple] = []
            first_seen: Dict[tuple, Dict[str, Any]] = {}

            for row in rows:
                key_raw = row.get(col_dims[0])
                # Bucket NULL/blank pivot keys under "(blank)" instead of
                # DROPPING the row — otherwise those rows' values silently
                # vanish from the matrix (data loss the user can't see).
                pivot_key = "(blank)" if key_raw is None or key_raw == "" else str(key_raw)
                if pivot_key not in seen_keys:
                    if len(pivot_keys_order) >= transform.max_columns:
                        raise HTTPException(
                            status_code=422,
                            detail=(
                                f"Pivot exceeded max_columns={transform.max_columns} "
                                f"distinct values of '{col_dims[0]}'. "
                                "Raise max_columns or pre-filter the source."
                            ),
                        )
                    seen_keys.add(pivot_key)
                    pivot_keys_order.append(pivot_key)
                idx_tuple = tuple(row.get(c) for c in idx_cols)
                if idx_tuple not in buckets:
                    buckets[idx_tuple] = {}
                    index_order.append(idx_tuple)
                    first_seen[idx_tuple] = {c: row.get(c) for c in idx_cols}
                buckets[idx_tuple].setdefault(pivot_key, []).append(row.get(transform.values))

            new_cols = idx_cols + pivot_keys_order
            new_rows: List[Dict[str, Any]] = []
            for idx_tuple in index_order:
                row_out = dict(first_seen[idx_tuple])
                cells_map = buckets[idx_tuple]
                for pk in pivot_keys_order:
                    row_out[pk] = _aggregate(cells_map.get(pk, []))
                new_rows.append(row_out)
            return new_cols, new_rows, None

        # ── Multi-dimension pivot (generates 2-level column headers) ────
        # Composite key separator — unlikely to appear in real data values.
        _SEP = "__|__"

        # Gather ordered distinct values for each column dimension.
        dim_vals: List[List[str]] = [[] for _ in col_dims]
        dim_seen_sets: List[set] = [set() for _ in col_dims]
        for row in rows:
            for i, _dim in enumerate(col_dims):
                v = row.get(_dim)
                k = "(blank)" if v is None or v == "" else str(v)
                if k not in dim_seen_sets[i]:
                    dim_seen_sets[i].add(k)
                    dim_vals[i].append(k)

        # Build ordered composite keys (Cartesian product in dim order).
        composite_keys: List[str] = []
        composite_key_combos: List[tuple] = []
        composite_set: set = set()
        for combo in itertools.product(*dim_vals):
            ck = _SEP.join(combo)
            if ck not in composite_set:
                if len(composite_keys) >= transform.max_columns:
                    raise HTTPException(
                        status_code=422,
                        detail=(
                            f"Pivot exceeded max_columns={transform.max_columns}. "
                            "Raise max_columns or pre-filter the source."
                        ),
                    )
                composite_set.add(ck)
                composite_keys.append(ck)
                composite_key_combos.append(combo)

        # Collect buckets: index_tuple → composite_key → [values].
        buckets_ml: Dict[tuple, Dict[str, List[Any]]] = {}
        index_order_ml: List[tuple] = []
        first_seen_ml: Dict[tuple, Dict[str, Any]] = {}
        for row in rows:
            parts: List[str] = []
            for _dim in col_dims:
                v = row.get(_dim)
                parts.append("(blank)" if v is None or v == "" else str(v))
            ck = _SEP.join(parts)
            idx_tuple = tuple(row.get(c) for c in idx_cols)
            if idx_tuple not in buckets_ml:
                buckets_ml[idx_tuple] = {}
                index_order_ml.append(idx_tuple)
                first_seen_ml[idx_tuple] = {c: row.get(c) for c in idx_cols}
            buckets_ml[idx_tuple].setdefault(ck, []).append(row.get(transform.values))

        new_cols_ml = idx_cols + composite_keys
        new_rows_ml: List[Dict[str, Any]] = []
        for idx_tuple in index_order_ml:
            row_out = dict(first_seen_ml[idx_tuple])
            cells_map_ml = buckets_ml[idx_tuple]
            for ck in composite_keys:
                row_out[ck] = _aggregate(cells_map_ml.get(ck, []))
            new_rows_ml.append(row_out)

        # Build column_groups: one group per first-dimension value.
        extra_groups: List[Dict[str, Any]] = []
        for first_val in dim_vals[0]:
            group_cols = [
                ck
                for ck, combo in zip(composite_keys, composite_key_combos)
                if combo[0] == first_val
            ]
            if group_cols:
                extra_groups.append({"label": first_val, "columns": group_cols})

        # column_labels: composite key → last dimension value (display label).
        col_labels: Dict[str, str] = {
            ck: combo[-1]
            for ck, combo in zip(composite_keys, composite_key_combos)
        }

        return new_cols_ml, new_rows_ml, {
            "column_groups": extra_groups,
            "column_labels": col_labels,
        }

    return columns, rows, None


def render_doc_screen(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    *,
    identity: CallerIdentity,
    app_user_payload: Optional[Dict[str, Any]] = None,
    shared_context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    if screen.kind != "doc" or screen.doc is None:
        raise HTTPException(status_code=400, detail="Screen is not a doc.")

    substitution = _build_substitution_map(
        workboard, app_user=app_user_payload, shared=shared_context
    )
    rendered_blocks: List[Dict[str, Any]] = []
    screen_column_labels = screen.column_labels or {}
    for block in screen.doc.blocks:
        payload = block.model_dump()
        if isinstance(block, DataTableBlock):
            payload["data"] = _resolve_doc_data_block(
                db, workboard, screen, block,
                identity=identity, shared_context=shared_context,
            )
        # Inject screen-level column_labels so frontend can show friendly names
        if screen_column_labels and not payload.get("column_labels"):
            payload["column_labels"] = screen_column_labels
        _substitute_strings_in_place(payload, substitution)
        rendered_blocks.append(payload)

    return {
        "screen_id": screen.id,
        "kind": "doc",
        "title": screen.title,
        "page": screen.doc.page.model_dump(),
        "blocks": rendered_blocks,
        "context": substitution,
        "print_template": _doc_print_template(workboard),
    }


def _doc_print_template(workboard: Workboard) -> Optional[Dict[str, Any]]:
    """The reusable letterhead for doc print/export, read from the layout.

    Stored under ``layout_json.print_template`` (LayoutJson ignores it as an
    extra key, so it round-trips as a plain dict). Returns None when unset or
    disabled so the FE simply omits the letterhead band.
    """
    # Live doc render/print/export must use the PUBLISHED letterhead, not the
    # mutable draft. Route through the stage resolver (published for Live via the
    # _wb_use_published flag, draft for Preview) instead of reading layout_json.
    try:
        from app.modules.workboards.services.runtime_config import effective_layout_raw

        raw = effective_layout_raw(workboard)
        pt = raw.get("print_template") if isinstance(raw, dict) else None
        if isinstance(pt, dict) and pt.get("enabled", True):
            return pt
    except Exception:  # pragma: no cover - defensive
        pass
    return None


def _resolve_doc_data_block(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    block: DataTableBlock,
    *,
    identity: CallerIdentity,
    shared_context: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    # Doc blocks default to ``primary`` (the screen's table) but also
    # support ``lookup:<table_id>`` so a single doc screen can join over
    # multiple tables (e.g. shifts + hourly_output for the overview).
    if block.source.startswith("lookup:"):
        try:
            table_id = int(block.source.split(":", 1)[1])
        except (ValueError, IndexError):
            return {"columns": [], "rows": []}
    else:
        table_id = screen.table_id or 0
    if not table_id:
        return {"columns": [], "rows": []}

    table = _load_table(db, table_id)
    datasource, query_table = _resolve_read_target(db, table) if table else (None, None)
    if not table or not datasource or query_table is None:
        return {"columns": [], "rows": []}

    filters: List[Dict[str, Any]] = []
    rls_filters, allowed = build_rls_filter(
        screen.rls, screen.rls_default, identity
    )
    if not allowed:
        return {"columns": [], "rows": []}
    if rls_filters:
        table_columns = _table_column_names(table)
        if any(str(item.get("field") or "") not in table_columns for item in rls_filters):
            return {"columns": [], "rows": []}
        filters = filters + rls_filters

    # Context filters bind a column to a runtime shared-context value so a
    # per-record document (a printable phiếu) shows ONLY that record. A
    # required binding whose value is missing yields NO rows (never a full
    # dump); an optional one is simply skipped.
    shared = shared_context or {}
    for cf in getattr(block, "context_filters", None) or []:
        raw = shared.get(cf.from_shared)
        present = raw is not None and not (isinstance(raw, str) and not raw.strip())
        if not present:
            if cf.required:
                return {"columns": [], "rows": []}
            continue
        filters = filters + [{"field": cf.column, "operator": "eq", "value": raw}]

    result = LiveQueryService.execute_preview_query(
        datasource, query_table, limit=block.max_rows, offset=0, filters=filters
    )
    all_columns: List[str] = result.get("columns") or []
    rows: List[Dict[str, Any]] = result.get("rows") or []
    # Pivot/unpivot first so column projection, group_by merges and totals
    # operate on the *reported* shape — not the raw source shape.
    _pivot_extra: Optional[Dict[str, Any]] = None
    if block.transform is not None:
        all_columns, rows, _pivot_extra = _apply_data_table_transform(
            all_columns, rows, block.transform
        )
    selected = [c for c in (block.columns or []) if c in all_columns] or all_columns
    if selected != all_columns:
        rows = [{c: row.get(c) for c in selected} for row in rows]
    if block.group_by:
        group_cols = [c for c in block.group_by if c in selected]
        if group_cols:
            rows = sorted(
                rows,
                key=lambda r: tuple(
                    ("" if r.get(c) is None else str(r.get(c))) for c in group_cols
                ),
            )
    payload: Dict[str, Any] = {"columns": selected, "rows": rows}
    _cg_source = (_pivot_extra or {}).get("column_groups") or block.column_groups
    column_groups = _normalize_column_groups(selected, _cg_source)
    if column_groups:
        payload["column_groups"] = column_groups
    _auto_labels = (_pivot_extra or {}).get("column_labels")
    if _auto_labels:
        payload["column_labels"] = _auto_labels
    # Surface the builder-side `column_metadata` (label override, width, format,
    # align, per-column total override, merge flag) so the doc runtime can
    # honour the formatting the builder configured. Was previously only
    # additive — clients now opt in by reading payload.column_metadata.
    column_metadata_raw = block.column_metadata or {}
    if column_metadata_raw:
        forwarded_meta: Dict[str, Any] = {}
        for col_name, meta in column_metadata_raw.items():
            if col_name not in selected:
                continue
            if hasattr(meta, "model_dump"):
                meta_dict = meta.model_dump(exclude_none=True)
            elif isinstance(meta, dict):
                meta_dict = {k: v for k, v in meta.items() if v is not None}
            else:
                continue
            if meta_dict:
                forwarded_meta[col_name] = meta_dict
        if forwarded_meta:
            payload["column_metadata"] = forwarded_meta
            # Pre-fold metadata.label into the column_labels map so existing
            # FE table renderers (which already read column_labels) display
            # the override without needing to know about column_metadata.
            existing_labels = dict(payload.get("column_labels") or {})
            for col_name, meta_dict in forwarded_meta.items():
                label_override = meta_dict.get("label")
                if isinstance(label_override, str) and label_override and col_name not in existing_labels:
                    existing_labels[col_name] = label_override
            if existing_labels:
                payload["column_labels"] = existing_labels
    merges = _compute_merges(rows, block.group_by, selected)
    if merges:
        payload["merges"] = merges
    footer = _compute_totals_row(block.totals, selected, rows)
    if footer:
        payload["footer_row"] = footer
    return payload


# ── Doc data-table export ─────────────────────────────────────────────────

def export_doc_data_block_to_excel(
    db: Session,
    workboard: Workboard,
    screen: Screen,
    block_index: int,
    *,
    identity: CallerIdentity,
    shared_context: Optional[Dict[str, Any]] = None,
) -> tuple[bytes, str]:
    """Render the *displayed* table of a doc ``data_table`` block to XLSX.

    Returns ``(content, filename)``. Only blocks with
    ``allow_export_excel=True`` are exportable; the same RLS + shared-context
    filters as on-screen rendering apply because we go through
    :func:`_resolve_doc_data_block`.
    """
    if screen.kind != "doc" or screen.doc is None:
        raise HTTPException(status_code=400, detail="Screen is not a doc.")
    blocks = screen.doc.blocks
    if block_index < 0 or block_index >= len(blocks):
        raise HTTPException(status_code=404, detail="Block not found.")
    block = blocks[block_index]
    if not isinstance(block, DataTableBlock):
        raise HTTPException(status_code=400, detail="Block is not a data_table.")
    if not block.allow_export_excel:
        raise HTTPException(status_code=403, detail="Excel export is disabled for this block.")

    data = _resolve_doc_data_block(
        db, workboard, screen, block,
        identity=identity, shared_context=shared_context,
    )
    columns: List[str] = data.get("columns") or []
    rows: List[Dict[str, Any]] = data.get("rows") or []

    content = _build_templated_excel(
        workboard=workboard,
        screen=screen,
        block=block,
        columns=columns,
        rows=rows,
        column_labels=data.get("column_labels") or {},
        shared_context=shared_context or {},
    )
    safe_screen = re.sub(r"[^A-Za-z0-9_.-]+", "_", screen.id) or "screen"
    safe_wb = re.sub(r"[^A-Za-z0-9_.-]+", "_", workboard.slug or workboard.name or "workboard")
    filename = f"{safe_wb}__{safe_screen}__block-{block_index + 1}.xlsx"
    return content, filename


_XLSX_NUMFMT = {
    "currency": '#,##0 "₫"',
    "number": "#,##0.####",
    "integer": "#,##0",
    "percent": "0.0%",
}


def _build_templated_excel(
    *,
    workboard: Workboard,
    screen: Screen,
    block: DataTableBlock,
    columns: List[str],
    rows: List[Dict[str, Any]],
    column_labels: Dict[str, str],
    shared_context: Dict[str, Any],
) -> bytes:
    """Build a letterhead-templated XLSX for a doc data_table.

    Layout: company letterhead (from the reusable print_template) → report
    title → export date + carried filters → an optional grouped header row
    (``column_groups``) → the styled table with per-column number formats →
    a bold totals row (``block.totals``). This is the "Excel có sẵn biểu mẫu"
    output — matches the on-screen document.
    """
    import io
    from datetime import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    pt = _doc_print_template(workboard) or {}
    accent = str(pt.get("accent_color") or "0F766E").lstrip("#") or "0F766E"

    meta_by_col = {k: v for k, v in (block.column_metadata or {}).items()}

    def _label(col: str) -> str:
        m = meta_by_col.get(col)
        if m is not None and getattr(m, "label", None):
            return str(m.label)
        return column_labels.get(col) or col

    def _fmt(col: str) -> Optional[str]:
        m = meta_by_col.get(col)
        return str(getattr(m, "format", None)) if m is not None and getattr(m, "format", None) else None

    numeric_formats = {"currency", "number", "integer", "percent"}

    def _num(value: Any) -> Optional[float]:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = re.sub(r"[^0-9.\-]", "", str(value))
        try:
            return float(s) if s not in ("", "-", ".") else None
        except ValueError:
            return None

    ncols = max(len(columns), 1)
    wb = Workbook()
    ws = wb.active
    ws.title = (re.sub(r"[\[\]\*/\\?:]", " ", (block.title or screen.title or "Báo cáo"))[:31]) or "Báo cáo"

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    accent_fill = PatternFill("solid", fgColor=accent)
    head_font = Font(bold=True, color="FFFFFF", size=11)

    r = 1

    def _merge_line(text: str, *, bold=False, size=11, color="111827", align="left"):
        nonlocal r
        ws.cell(r, 1, text)
        cell = ws.cell(r, 1)
        cell.font = Font(bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if ncols > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        r += 1

    # ── Letterhead ──────────────────────────────────────────────────
    if pt.get("company_name"):
        _merge_line(str(pt["company_name"]), bold=True, size=14, color=accent)
    sub_bits: List[str] = []
    if pt.get("address"):
        sub_bits.append(f"Địa chỉ: {pt['address']}")
    if pt.get("tax_code"):
        sub_bits.append(f"MST: {pt['tax_code']}")
    if pt.get("hotline"):
        sub_bits.append(f"Hotline: {pt['hotline']}")
    if pt.get("email"):
        sub_bits.append(pt["email"])
    for line in sub_bits:
        _merge_line(line, size=10, color="6B7280")
    if pt.get("company_name") or sub_bits:
        r += 1  # spacer

    # ── Report title + meta ─────────────────────────────────────────
    _merge_line(str(block.title or screen.title or "Báo cáo").upper(), bold=True, size=13, align="center")
    _merge_line(f"Ngày xuất: {_dt.now():%d/%m/%Y %H:%M}", size=9, color="6B7280", align="center")
    carried = [f"{k}: {v}" for k, v in (shared_context or {}).items() if v not in (None, "")]
    if carried:
        _merge_line(" · ".join(carried), size=9, color="6B7280", align="center")
    r += 1  # spacer

    # ── Optional grouped header (column_groups) ─────────────────────
    groups = list(block.column_groups or [])
    if groups:
        col_pos = {c: i for i, c in enumerate(columns)}
        for g in groups:
            gcols = [c for c in (g.columns or []) if c in col_pos]
            if not gcols:
                continue
            idxs = sorted(col_pos[c] for c in gcols)
            start, end = idxs[0] + 1, idxs[-1] + 1
            cell = ws.cell(r, start, g.label or "")
            cell.font = head_font
            cell.fill = accent_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if end > start:
                ws.merge_cells(start_row=r, start_column=start, end_row=r, end_column=end)
                for cc in range(start, end + 1):
                    ws.cell(r, cc).border = border
                    ws.cell(r, cc).fill = accent_fill
        r += 1

    # ── Table header row ────────────────────────────────────────────
    header_row = r
    for j, col in enumerate(columns):
        cell = ws.cell(r, j + 1, _label(col))
        cell.font = head_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    r += 1

    # ── Data rows ───────────────────────────────────────────────────
    for row in rows:
        for j, col in enumerate(columns):
            fmt = _fmt(col)
            raw = row.get(col)
            cell = ws.cell(r, j + 1)
            if fmt in numeric_formats:
                num = _num(raw)
                cell.value = num
                cell.number_format = _XLSX_NUMFMT.get(fmt, "#,##0")
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = "" if raw is None else str(raw)
                cell.alignment = Alignment(horizontal="left")
            cell.border = border
        r += 1

    # ── Totals row ──────────────────────────────────────────────────
    total_cols = set(block.totals or [])
    if total_cols and rows:
        for j, col in enumerate(columns):
            cell = ws.cell(r, j + 1)
            cell.font = Font(bold=True)
            cell.border = border
            if j == 0:
                cell.value = "TỔNG"
            elif col in total_cols:
                cell.value = sum((_num(row.get(col)) or 0) for row in rows)
                cell.number_format = _XLSX_NUMFMT.get(_fmt(col) or "integer", "#,##0")
                cell.alignment = Alignment(horizontal="right")
        r += 1

    # ── Column widths ───────────────────────────────────────────────
    for j, col in enumerate(columns):
        header_len = len(_label(col))
        sample = max((len(str(row.get(col) or "")) for row in rows[:50]), default=0)
        width = min(max(header_len + 2, sample + 2, 10), 42)
        ws.column_dimensions[get_column_letter(j + 1)].width = width
    ws.freeze_panes = ws.cell(header_row + 1, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Public app shell payload ──────────────────────────────────────────────

def _default_experience() -> Dict[str, Any]:
    """Fully-populated default presentation contract, so the runtime always
    receives concrete tokens (no null-handling needed client-side)."""
    return {
        "schema_version": 1,
        "preset": None,
        "theme": {
            "primary": "#2563eb", "success": "#16a34a", "warning": "#f59e0b",
            "danger": "#ef4444", "info": "#3b82f6", "neutral": "#6b7280",
            "background": "#f8fafc", "surface": "#ffffff", "border": "#e5e7eb",
            "text": "#111827", "font_family": "system", "heading_weight": "semibold",
            "body_weight": "regular", "type_scale": 100, "density": "cozy",
            "radius": "small", "elevation": "small", "motion": "standard",
            "mode": "auto", "app_background": None,
        },
        "shell": {
            "sticky_header": True, "show_search": False, "show_logo": True,
            "content_width": "full_bleed", "content_width_px": None,
            "page_padding": "cozy", "footer_enabled": False, "background": "gray",
        },
        "navigation": {
            "desktop_kind": "sidebar", "mobile_kind": "bottom_nav",
            "sidebar_width": 224, "default_collapsed": False, "show_icons": True,
            "show_labels": True, "active_style": "pill", "breadcrumbs": False,
        },
        "feedback": {
            "loading": "spinner", "empty_style": "message", "success": "inline",
            "confirmation": "modal", "error_retry": True, "motion_ms": 160,
        },
        # Runtime-only metadata. ``overrides`` is the exact author-authored
        # block, allowing the frontend to distinguish inheritance from an
        # explicit token (important for legacy dark mode and draft preview).
        "explicit": False,
        "overrides": {},
    }


def resolve_experience(layout: LayoutJson) -> Dict[str, Any]:
    """Effective presentation = defaults ← legacy(branding/mini_app_nav) ←
    explicit ``experience``. Keeps old boards visually identical when they have
    no ``experience`` block (the adapter maps their branding/nav look)."""
    eff = _default_experience()
    b = getattr(layout, "branding", None)
    if b is not None:
        if getattr(b, "primary_color", None):
            eff["theme"]["primary"] = b.primary_color
        if getattr(b, "accent_color", None):
            eff["theme"]["info"] = b.accent_color
        if getattr(b, "theme", None):
            eff["theme"]["mode"] = b.theme
        if getattr(b, "font_family", None):
            eff["theme"]["font_family"] = b.font_family
    nav = getattr(layout, "mini_app_nav", None)
    if nav is not None:
        if getattr(nav, "desktop_kind", None):
            eff["navigation"]["desktop_kind"] = nav.desktop_kind
        if getattr(nav, "mobile_kind", None):
            eff["navigation"]["mobile_kind"] = nav.mobile_kind
    exp = getattr(layout, "experience", None)
    if exp is not None:
        eff["explicit"] = True
        eff["overrides"] = exp.model_dump(exclude_none=True)
        eff["schema_version"] = getattr(exp, "schema_version", 1) or 1
        if getattr(exp, "preset", None):
            eff["preset"] = exp.preset
        for section in ("theme", "shell", "navigation", "feedback"):
            sec = getattr(exp, section, None)
            if sec is not None:
                for key, value in sec.model_dump(exclude_none=True).items():
                    eff[section][key] = value
    return eff


def render_app_shell(
    workboard: Workboard,
    identity: CallerIdentity,
    hidden_screen_ids: Optional[Set[str]] = None,
    db: Optional[Session] = None,
) -> Dict[str, Any]:
    """Initial payload the public runtime fetches on entry to a workboard.

    Returns a slim dict with everything the FE needs to render the app
    shell (header + nav) without bouncing back to the API to discover
    screens. Per-screen content stays lazy-loaded.

    ``hidden_screen_ids`` are screens hidden ON THE PUBLIC LINK (the Cổng menu
    item's ``hidden_screen_ids``) — dropped from the nav/groups/screen list here
    while the builder layout stays intact. ``None`` = hide nothing.
    """
    hidden = hidden_screen_ids or set()
    layout = parse_layout(workboard)
    visible_screens = [
        s
        for s in layout.screens
        if is_screen_visible_for(s, identity) and s.id not in hidden
    ]
    nav_items = list(layout.mini_app_nav.items)
    if not nav_items:
        nav_items = [s.id for s in visible_screens if s.show_in_nav]
    # Filter nav items to those actually visible to this identity.
    visible_ids = {s.id for s in visible_screens}
    nav_items = [sid for sid in nav_items if sid in visible_ids]

    # Screen-groups (UI: "Workspaces"). Additive: empty => flat nav above.
    # Drop groups hidden by role, members that are RLS-hidden or deleted
    # (not in visible_ids), de-dupe within a group, and skip now-empty groups.
    visible_groups: List[Dict[str, Any]] = []
    for grp in (layout.screen_groups or []):
        if not is_group_visible_for(grp, identity):
            continue
        seen: set[str] = set()
        member_ids: List[str] = []
        for sid in grp.screen_ids:
            if sid in visible_ids and sid not in seen:
                seen.add(sid)
                member_ids.append(sid)
        if not member_ids:
            continue
        visible_groups.append(
            {"id": grp.id, "label": grp.label, "icon": grp.icon, "screen_ids": member_ids}
        )

    return {
        "workboard": {
            "id": workboard.id,
            "name": workboard.name,
            "slug": workboard.slug,
            "icon": workboard.icon,
            "description": workboard.description,
        },
        "branding": layout.branding.model_dump(),
        "experience": resolve_experience(layout),
        "media_max_kb": (
            media_cap_kb(db, workboard) if db is not None else WORKBOARD_MEDIA_MAX_KB
        ),
        "nav": {
            **layout.mini_app_nav.model_dump(),
            "items": nav_items,
        },
        "screens": [
            {
                "id": s.id,
                "kind": s.kind,
                "title": s.title,
                "icon": s.icon,
                "description": s.description,
                "show_in_nav": s.show_in_nav,
            }
            for s in visible_screens
        ],
        "screen_groups": visible_groups,
        # Surface the caller's role so the FE can hide per-action
        # buttons whose `visible_for_roles` excludes the current user.
        # Internal/admin previews come in as identity.role=None — the FE
        # treats None as "see everything".
        "viewer": {
            "role": identity.role,
            "username": (
                (identity.app_user or {}).get("username")
                if identity.app_user
                else identity.appbi_user_id
            ),
        },
    }
