"""
Workboard doc export service — converts a rendered doc-view (produced by
``WorkboardRuntimeService.render_doc``) into HTML / PDF / XLSX bytes.

The renderer is intentionally minimal: each of the seven supported block
types maps to a small, dependency-light primitive in each output format.
Anything fancier (custom fonts, complex layouts, page numbering hooks)
is deliberately out of scope for the MVP.
"""
from __future__ import annotations

import io
from html import escape as html_escape
from typing import Any, Dict, List

from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, LETTER, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _build_merge_index(
    merges: List[Dict[str, Any]] | None,
) -> tuple[Dict[tuple[int, str], int], set[tuple[int, str]]]:
    """Translate the runtime ``merges`` payload into per-cell directives.

    Returns ``(rowspan_map, hidden_cells)`` where:
      * ``rowspan_map[(row_idx, col)] = span`` is the rowspan for a cell
        that is the *first* cell of a merged group;
      * ``hidden_cells`` lists ``(row_idx, col)`` cells that fall inside a
        merged group but are not the first row — they should be skipped
        entirely in HTML/PDF and merged via openpyxl in Excel.
    """
    rowspan_map: Dict[tuple[int, str], int] = {}
    hidden_cells: set[tuple[int, str]] = set()
    for spec in merges or []:
        col = spec.get("column")
        start = int(spec.get("row_start", 0) or 0)
        span = int(spec.get("row_span", 0) or 0)
        if not col or span < 2:
            continue
        rowspan_map[(start, col)] = span
        for offset in range(1, span):
            hidden_cells.add((start + offset, col))
    return rowspan_map, hidden_cells


def _normalize_footer_rows(
    footer: Any, columns: List[str]
) -> List[Dict[str, Any]]:
    """Accept both the new ``{rows:[{agg,label,values}], single?}`` shape and
    the legacy flat ``{col: value}`` dict.
    """
    if isinstance(footer, dict) and isinstance(footer.get("rows"), list):
        return [
            {
                "agg": fr.get("agg") or "",
                "label": fr.get("label") or fr.get("agg") or "",
                "values": fr.get("values") or {},
            }
            for fr in footer["rows"]
        ]
    if isinstance(footer, dict):
        return [{"agg": "sum", "label": "Tổng", "values": footer}]
    return []


def _format_total_cell(value: Any) -> str:
    """Render an aggregated value for a footer cell.

    Floats with no fractional part collapse to integers; everything else
    rounds to 2 decimals so HTML/PDF output stays compact. Non-numeric
    values pass through as ``str``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}"
    return str(value)


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

def to_html(rendered_doc: Dict[str, Any]) -> str:
    title = html_escape(rendered_doc.get("title") or "Document")
    blocks = rendered_doc.get("blocks") or []
    body_parts: List[str] = []
    for block in blocks:
        body_parts.append(_block_to_html(block))
    body = "\n".join(body_parts)
    return f"""<!DOCTYPE html>
<html lang=\"en\">
<head>
<meta charset=\"utf-8\" />
<title>{title}</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
         color: #111; padding: 24px; max-width: 960px; margin: 0 auto; }}
  h1 {{ margin: 0 0 4px 0; }}
  h2 {{ font-size: 16px; margin: 16px 0 8px 0; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0; }}
  th, td {{ border: 1px solid #ddd; padding: 6px 8px; font-size: 13px; vertical-align: top; }}
  th {{ background: #f6f7f9; text-align: left; }}
  tfoot td {{ background: #fafbfc; font-weight: 600; }}
  .wb-header {{ text-align: center; margin-bottom: 16px; }}
  .wb-header.left {{ text-align: left; }}
  .wb-header.right {{ text-align: right; }}
  .wb-kv-grid {{ display: grid; grid-template-columns: repeat(var(--cols, 2), 1fr); gap: 8px 16px; margin: 8px 0; }}
  .wb-kv-grid .label {{ color: #666; font-size: 12px; text-transform: uppercase; }}
  .wb-kv-grid .value {{ font-size: 14px; }}
  .wb-text.center {{ text-align: center; }}
  .wb-text.right {{ text-align: right; }}
  .wb-spacer {{ display: block; }}
  .wb-signature {{ display: flex; justify-content: space-between; margin: 32px 0; gap: 16px; }}
  .wb-signature .slot {{ flex: 1; text-align: center; border-top: 1px solid #999; padding-top: 6px; }}
  .wb-footer {{ display: flex; justify-content: space-between; margin-top: 24px;
                color: #666; font-size: 12px; }}
</style>
</head>
<body>
{body}
</body>
</html>
"""


def _block_to_html(block: Dict[str, Any]) -> str:
    block_type = block.get("type")
    if block_type == "header":
        align = html_escape(block.get("align") or "center")
        title = html_escape(block.get("title") or "")
        subtitle = block.get("subtitle")
        logo = block.get("logo_url")
        parts: List[str] = [f'<div class="wb-header {align}">']
        if logo:
            parts.append(f'<img src="{html_escape(logo)}" alt="logo" style="max-height:48px;"/>')
        if title:
            parts.append(f"<h1>{title}</h1>")
        if subtitle:
            parts.append(f'<div style="color:#666">{html_escape(subtitle)}</div>')
        parts.append("</div>")
        return "\n".join(parts)
    if block_type == "kv_grid":
        cols = int(block.get("columns") or 2)
        items = block.get("items") or []
        cells: List[str] = []
        for item in items:
            label = html_escape(item.get("label") or "")
            value = html_escape(item.get("value") or "")
            cells.append(f'<div><div class="label">{label}</div><div class="value">{value}</div></div>')
        return f'<div class="wb-kv-grid" style="--cols:{cols}">{"".join(cells)}</div>'
    if block_type == "data_table":
        title = block.get("title")
        data = block.get("data") or {}
        columns: List[str] = data.get("columns") or block.get("columns") or []
        rows: List[Dict[str, Any]] = data.get("rows") or []
        footer_row = data.get("footer_row") or {}
        merges = data.get("merges") or []
        rowspan_map, hidden = _build_merge_index(merges)
        thead = "".join(f"<th>{html_escape(str(c))}</th>" for c in columns)
        tbody_parts: List[str] = []
        for r_idx, row in enumerate(rows):
            cells_html = []
            for c in columns:
                if (r_idx, c) in hidden:
                    continue
                span = rowspan_map.get((r_idx, c))
                rs_attr = f' rowspan="{span}"' if span else ""
                value = "" if row.get(c) is None else str(row.get(c))
                cells_html.append(f"<td{rs_attr}>{html_escape(value)}</td>")
            tbody_parts.append(f"<tr>{''.join(cells_html)}</tr>")
        tfoot_html = ""
        if footer_row:
            footer_rows_data = _normalize_footer_rows(footer_row, columns)
            tfoot_inner: List[str] = []
            for fr in footer_rows_data:
                cells: List[str] = []
                for ci, c in enumerate(columns):
                    val = (fr.get("values") or {}).get(c)
                    if val is None and ci == 0 and fr.get("label"):
                        cells.append(
                            f'<td class="agg-label">{html_escape(str(fr.get("label")))}</td>'
                        )
                    else:
                        cells.append(f"<td>{html_escape(_format_total_cell(val))}</td>")
                tfoot_inner.append(f"<tr>{''.join(cells)}</tr>")
            tfoot_html = f"<tfoot>{''.join(tfoot_inner)}</tfoot>"
        title_html = f"<h2>{html_escape(title)}</h2>" if title else ""
        return (
            f"{title_html}<table><thead><tr>{thead}</tr></thead>"
            f"<tbody>{''.join(tbody_parts)}</tbody>{tfoot_html}</table>"
        )
    if block_type == "text":
        align = html_escape(block.get("align") or "left")
        content = block.get("content") or ""
        if block.get("markdown"):
            # Minimal markdown safety: escape, then convert newlines to <br/>.
            content = html_escape(content).replace("\n", "<br/>")
        else:
            content = html_escape(content)
        return f'<div class="wb-text {align}">{content}</div>'
    if block_type == "spacer":
        height = int(block.get("height_mm") or 10)
        return f'<div class="wb-spacer" style="height:{height}mm"></div>'
    if block_type == "signature":
        slots = block.get("slots") or []
        slot_html = "".join(
            f'<div class="slot"><div>{html_escape(s.get("label") or "")}</div>'
            f'<div style="color:#666;font-size:12px">{html_escape(s.get("role") or "")}</div></div>'
            for s in slots
        )
        return f'<div class="wb-signature">{slot_html}</div>'
    if block_type == "footer":
        left = html_escape(block.get("left") or "")
        center = html_escape(block.get("center") or "")
        right = html_escape(block.get("right") or "")
        return f'<div class="wb-footer"><div>{left}</div><div>{center}</div><div>{right}</div></div>'
    return ""


# ---------------------------------------------------------------------------
# PDF (ReportLab platypus)
# ---------------------------------------------------------------------------

_PAGE_SIZES = {"A4": A4, "A3": A3, "Letter": LETTER}


def to_pdf(rendered_doc: Dict[str, Any]) -> bytes:
    page_meta = rendered_doc.get("page") or {}
    page_size = _PAGE_SIZES.get(page_meta.get("size") or "A4", A4)
    if page_meta.get("orientation") == "landscape":
        page_size = landscape(page_size)
    margin = float(page_meta.get("margin_mm") or 15) * mm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title=rendered_doc.get("title") or "Document",
    )
    styles = getSampleStyleSheet()
    story: List[Any] = []
    for block in rendered_doc.get("blocks") or []:
        story.extend(_block_to_pdf_flowables(block, styles))

    if not story:
        story.append(Paragraph("(empty document)", styles["Normal"]))
    doc.build(story)
    return buffer.getvalue()


def _block_to_pdf_flowables(block: Dict[str, Any], styles) -> List[Any]:
    block_type = block.get("type")
    flow: List[Any] = []
    if block_type == "header":
        title_style = ParagraphStyle(
            "WbHeader",
            parent=styles["Title"],
            alignment={"left": 0, "center": 1, "right": 2}.get(block.get("align") or "center", 1),
        )
        if block.get("title"):
            flow.append(Paragraph(html_escape(block.get("title") or ""), title_style))
        if block.get("subtitle"):
            flow.append(Paragraph(html_escape(block.get("subtitle") or ""), styles["Normal"]))
        flow.append(Spacer(1, 6 * mm))
        return flow
    if block_type == "kv_grid":
        items = block.get("items") or []
        cols = int(block.get("columns") or 2)
        # Build rows of (label, value) pairs grouped into `cols` pairs per row.
        pairs = [(item.get("label") or "", item.get("value") or "") for item in items]
        # Each grid cell becomes 2 columns: label + value.
        grid_rows: List[List[str]] = []
        per_row = cols
        for i in range(0, len(pairs), per_row):
            chunk = pairs[i:i + per_row]
            row: List[str] = []
            for label, value in chunk:
                row.append(f"<b>{html_escape(str(label))}</b><br/>{html_escape(str(value))}")
            while len(row) < per_row:
                row.append("")
            grid_rows.append(row)
        if grid_rows:
            cell_paragraphs = [
                [Paragraph(c, styles["BodyText"]) for c in row] for row in grid_rows
            ]
            tbl = Table(cell_paragraphs, hAlign="LEFT")
            tbl.setStyle(
                TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ])
            )
            flow.append(tbl)
            flow.append(Spacer(1, 4 * mm))
        return flow
    if block_type == "data_table":
        if block.get("title"):
            flow.append(Paragraph(f"<b>{html_escape(block.get('title') or '')}</b>", styles["Heading3"]))
        data = block.get("data") or {}
        columns: List[str] = data.get("columns") or block.get("columns") or []
        rows: List[Dict[str, Any]] = data.get("rows") or []
        footer_row = data.get("footer_row") or {}
        merges = data.get("merges") or []
        if columns:
            col_index = {c: i for i, c in enumerate(columns)}
            table_data: List[List[str]] = [[str(c) for c in columns]]
            for row in rows:
                table_data.append(["" if row.get(c) is None else str(row.get(c)) for c in columns])
            footer_indices: List[int] = []
            if footer_row:
                for fr in _normalize_footer_rows(footer_row, columns):
                    cells: List[str] = []
                    for ci, c in enumerate(columns):
                        val = (fr.get("values") or {}).get(c)
                        if val is None and ci == 0 and fr.get("label"):
                            cells.append(str(fr.get("label")))
                        else:
                            cells.append(_format_total_cell(val))
                    table_data.append(cells)
                    footer_indices.append(len(table_data) - 1)
            footer_idx = footer_indices[0] if footer_indices else None
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f1f4")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
            # Add SPAN ranges for merged cells. ReportLab uses (col, row)
            # tuples; row indexes here are body-relative offset by +1 for
            # the header row.
            for spec in merges:
                col = spec.get("column")
                start = int(spec.get("row_start", 0) or 0)
                span = int(spec.get("row_span", 0) or 0)
                if col not in col_index or span < 2:
                    continue
                ci = col_index[col]
                style_cmds.append(("SPAN", (ci, start + 1), (ci, start + span)))
            for f_pos, f_idx in enumerate(footer_indices):
                style_cmds.append(
                    ("BACKGROUND", (0, f_idx), (-1, f_idx), colors.HexColor("#fafbfc"))
                )
                style_cmds.append(
                    ("FONTNAME", (0, f_idx), (-1, f_idx), "Helvetica-Bold")
                )
                if f_pos == 0:
                    style_cmds.append(
                        ("LINEABOVE", (0, f_idx), (-1, f_idx), 0.75, colors.HexColor("#888888"))
                    )
            tbl = Table(table_data, repeatRows=1)
            tbl.setStyle(TableStyle(style_cmds))
            flow.append(tbl)
            flow.append(Spacer(1, 4 * mm))
        return flow
    if block_type == "text":
        align = block.get("align") or "left"
        para_style = ParagraphStyle(
            "WbText",
            parent=styles["BodyText"],
            alignment={"left": 0, "center": 1, "right": 2}.get(align, 0),
        )
        text = html_escape(block.get("content") or "").replace("\n", "<br/>")
        flow.append(Paragraph(text, para_style))
        return flow
    if block_type == "spacer":
        flow.append(Spacer(1, float(block.get("height_mm") or 10) * mm))
        return flow
    if block_type == "signature":
        slots = block.get("slots") or []
        if slots:
            row = [
                Paragraph(
                    f"<para alignment='center'>__________________________<br/>"
                    f"<b>{html_escape(s.get('label') or '')}</b><br/>"
                    f"<font size=8 color='#666'>{html_escape(s.get('role') or '')}</font></para>",
                    styles["BodyText"],
                )
                for s in slots
            ]
            tbl = Table([row], colWidths=[None] * len(row))
            tbl.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "BOTTOM")]))
            flow.append(Spacer(1, 12 * mm))
            flow.append(tbl)
            flow.append(Spacer(1, 4 * mm))
        return flow
    if block_type == "footer":
        left = html_escape(block.get("left") or "")
        center = html_escape(block.get("center") or "")
        right = html_escape(block.get("right") or "")
        if any([left, center, right]):
            tbl = Table([[
                Paragraph(left, styles["BodyText"]),
                Paragraph(f"<para alignment='center'>{center}</para>", styles["BodyText"]),
                Paragraph(f"<para alignment='right'>{right}</para>", styles["BodyText"]),
            ]])
            tbl.setStyle(TableStyle([
                ("LINEABOVE", (0, 0), (-1, 0), 0.25, colors.HexColor("#cccccc")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]))
            flow.append(Spacer(1, 8 * mm))
            flow.append(tbl)
        return flow
    return flow


# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------

def to_excel(rendered_doc: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (rendered_doc.get("title") or "Document")[:31] or "Document"
    cursor_row = 1
    bold = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="EEF1F5")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    for block in rendered_doc.get("blocks") or []:
        block_type = block.get("type")
        if block_type == "header":
            if block.get("title"):
                cell = ws.cell(row=cursor_row, column=1, value=block.get("title"))
                cell.font = title_font
                cell.alignment = center
                ws.merge_cells(start_row=cursor_row, start_column=1, end_row=cursor_row, end_column=6)
                cursor_row += 1
            if block.get("subtitle"):
                cell = ws.cell(row=cursor_row, column=1, value=block.get("subtitle"))
                cell.alignment = center
                ws.merge_cells(start_row=cursor_row, start_column=1, end_row=cursor_row, end_column=6)
                cursor_row += 1
            cursor_row += 1
        elif block_type == "kv_grid":
            items = block.get("items") or []
            cols_per_row = int(block.get("columns") or 2)
            for chunk_start in range(0, len(items), cols_per_row):
                chunk = items[chunk_start:chunk_start + cols_per_row]
                col_idx = 1
                for item in chunk:
                    label_cell = ws.cell(row=cursor_row, column=col_idx, value=item.get("label") or "")
                    label_cell.font = bold
                    ws.cell(row=cursor_row, column=col_idx + 1, value=item.get("value") or "")
                    col_idx += 2
                cursor_row += 1
            cursor_row += 1
        elif block_type == "data_table":
            if block.get("title"):
                cell = ws.cell(row=cursor_row, column=1, value=block.get("title"))
                cell.font = bold
                cursor_row += 1
            data = block.get("data") or {}
            columns: List[str] = data.get("columns") or block.get("columns") or []
            rows: List[Dict[str, Any]] = data.get("rows") or []
            footer_row = data.get("footer_row") or {}
            merges = data.get("merges") or []
            if columns:
                col_index = {c: i for i, c in enumerate(columns, start=1)}
                for c_idx, col in enumerate(columns, start=1):
                    cell = ws.cell(row=cursor_row, column=c_idx, value=str(col))
                    cell.font = bold
                    cell.fill = header_fill
                    cell.alignment = wrap
                cursor_row += 1
                first_data_row = cursor_row
                for row in rows:
                    for c_idx, col in enumerate(columns, start=1):
                        ws.cell(row=cursor_row, column=c_idx, value=row.get(col))
                    cursor_row += 1
                # Apply Excel merge_cells for each rowspan group.
                for spec in merges:
                    col = spec.get("column")
                    start = int(spec.get("row_start", 0) or 0)
                    span = int(spec.get("row_span", 0) or 0)
                    if col not in col_index or span < 2:
                        continue
                    ci = col_index[col]
                    excel_start = first_data_row + start
                    excel_end = excel_start + span - 1
                    ws.merge_cells(
                        start_row=excel_start, start_column=ci,
                        end_row=excel_end, end_column=ci,
                    )
                    ws.cell(row=excel_start, column=ci).alignment = Alignment(
                        vertical="center", wrap_text=True
                    )
                if footer_row:
                    footer_fill = PatternFill("solid", fgColor="FAFBFC")
                    for fr in _normalize_footer_rows(footer_row, columns):
                        for c_idx, col in enumerate(columns, start=1):
                            val = (fr.get("values") or {}).get(col)
                            if val is None and c_idx == 1 and fr.get("label"):
                                val = fr.get("label")
                            cell = ws.cell(row=cursor_row, column=c_idx, value=val)
                            cell.font = bold
                            cell.fill = footer_fill
                        cursor_row += 1
                # Auto width hint
                for c_idx, col in enumerate(columns, start=1):
                    ws.column_dimensions[get_column_letter(c_idx)].width = max(
                        ws.column_dimensions[get_column_letter(c_idx)].width or 10,
                        min(40, max(12, len(str(col)) + 2)),
                    )
            cursor_row += 1
        elif block_type == "text":
            cell = ws.cell(row=cursor_row, column=1, value=block.get("content") or "")
            cell.alignment = wrap
            ws.merge_cells(start_row=cursor_row, start_column=1, end_row=cursor_row, end_column=6)
            cursor_row += 2
        elif block_type == "spacer":
            cursor_row += 1
        elif block_type == "signature":
            slots = block.get("slots") or []
            for c_idx, slot in enumerate(slots, start=1):
                cell = ws.cell(row=cursor_row, column=c_idx, value="__________________")
                cell.alignment = center
            cursor_row += 1
            for c_idx, slot in enumerate(slots, start=1):
                cell = ws.cell(row=cursor_row, column=c_idx, value=slot.get("label") or "")
                cell.font = bold
                cell.alignment = center
            cursor_row += 1
            for c_idx, slot in enumerate(slots, start=1):
                cell = ws.cell(row=cursor_row, column=c_idx, value=slot.get("role") or "")
                cell.alignment = center
            cursor_row += 2
        elif block_type == "footer":
            footer_text = " | ".join(
                part for part in (
                    block.get("left"), block.get("center"), block.get("right"),
                ) if part
            )
            if footer_text:
                cell = ws.cell(row=cursor_row, column=1, value=footer_text)
                cell.font = Font(italic=True, color="666666")
                ws.merge_cells(start_row=cursor_row, start_column=1, end_row=cursor_row, end_column=6)
                cursor_row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
